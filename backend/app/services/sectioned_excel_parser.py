from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any


# LOSSQ_SECTIONED_EXCEL_LOSS_RUN_SERVICE_V1
def lossq_sectioned_excel_loss_run_repair_v1(
    file_path: str,
    parsed_claims: list[dict[str, Any]] | None = None,
    parsed_profile: dict[str, Any] | None = None,
    direct_profile: dict[str, Any] | None = None,
):
    """
    Universal sectioned Excel loss run parser.

    Purpose:
    - Handles XLSX/XLSM files with label/value account rows, exposure rows,
      and a Claims Detail table.
    - Supports English/French bilingual labels.
    - Does not hardcode a customer, carrier, policy, or file name.
    - Returns stronger claims/profile only when the uploaded workbook contains
      a recognizable sectioned loss-run structure.
    """

    parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
    parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
    direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

    lower_path = str(file_path or "").lower()
    if not lower_path.endswith((".xlsx", ".xlsm")):
        return parsed_claims, parsed_profile, direct_profile

    def clean(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        return re.sub(r"\s+", " ", str(value).replace("\ufeff", "").strip()).strip(" :-|/")

    def key(value: Any) -> str:
        return re.sub(r"[^a-z0-9À-ÿ]+", " ", clean(value).lower()).strip()

    # LOSSQ_SECTIONED_EXCEL_PERIOD_AU_TO_AND_DISPLAY_CASE_V1
    def display_name_case(value: Any) -> str:
        raw = clean(value)
        if not raw:
            return ""

        letters = [char for char in raw if char.isalpha()]
        uppercase_letters = [char for char in letters if char.isupper()]

        if not letters:
            return raw

        # Only clean obvious all-caps document headers. Do not alter normal names.
        if len(uppercase_letters) / max(len(letters), 1) < 0.70:
            return raw

        keep_upper = {
            "AIG",
            "CNA",
            "USLI",
            "BHSI",
            "AXA",
            "QBE",
            "HDI",
            "RLI",
            "D&O",
            "E&O",
            "GL",
            "WC",
            "BOP",
            "EPLI",
            "IBC",
            "BAC",
            "CAD",
            "USA",
            "UK",
        }

        lower_words = {"and", "or", "of", "the", "de", "des", "du", "la", "le", "les", "et"}

        def fix_piece(piece: str) -> str:
            tokens = piece.split(" ")
            fixed = []

            for index, token in enumerate(tokens):
                if not token:
                    continue

                clean_token = token.strip()
                upper_token = clean_token.upper()

                if upper_token in keep_upper or "&" in clean_token:
                    fixed.append(upper_token)
                    continue

                next_value = clean_token[:1].upper() + clean_token[1:].lower()

                if index > 0 and next_value.lower() in lower_words:
                    next_value = next_value.lower()

                fixed.append(next_value)

            return " ".join(fixed)

        return " / ".join(fix_piece(piece.strip()) for piece in raw.split("/"))


    # LOSSQ_SECTIONED_EXCEL_COMPACT_DO_EO_LIMIT_LABELS_V1
    def compact_coverage_label(value: Any) -> str:
        raw = clean(value)

        if not raw:
            return ""

        if re.search(r"(?i)\bD\s*&\s*O\b|administrateurs|directors", raw):
            return "D&O"

        if re.search(r"(?i)\bE\s*&\s*O\b|professionnelle|professional|erreurs|omissions", raw):
            return "E&O"

        return raw

    def money_float(value: Any) -> float:
        raw = clean(value)
        if not raw or raw in {"—", "-", "N/A", "n/a"}:
            return 0.0

        negative = raw.startswith("(") and raw.endswith(")")
        raw = re.sub(r"[^0-9.\-]", "", raw)

        try:
            amount = float(raw or 0)
            return -amount if negative else amount
        except Exception:
            return 0.0

    def money_text(value: Any) -> str:
        amount = money_float(value)
        if amount == int(amount):
            return str(int(amount))
        return str(amount)

    def normalize_status(value: Any) -> str:
        raw = clean(value).lower()
        if "open" in raw or "ouvert" in raw or "ouverte" in raw:
            return "Open"
        if "closed" in raw or "fermé" in raw or "fermee" in raw or "clos" in raw or "clôturé" in raw:
            return "Closed"
        return clean(value)

    def parse_province(value: Any):
        raw = clean(value)
        compact = re.sub(r"[^a-z0-9À-ÿ]+", "", raw.lower())

        provinces = {
            "alberta": ("AB", "Alberta", "Alberta Superintendent of Insurance"),
            "ab": ("AB", "Alberta", "Alberta Superintendent of Insurance"),
            "britishcolumbia": ("BC", "British Columbia", "BCFSA"),
            "bc": ("BC", "British Columbia", "BCFSA"),
            "manitoba": ("MB", "Manitoba", "FIRB"),
            "mb": ("MB", "Manitoba", "FIRB"),
            "newbrunswick": ("NB", "New Brunswick", "FCNB"),
            "nb": ("NB", "New Brunswick", "FCNB"),
            "newfoundlandandlabrador": ("NL", "Newfoundland and Labrador", "Digital Government and Service NL"),
            "nl": ("NL", "Newfoundland and Labrador", "Digital Government and Service NL"),
            "novascotia": ("NS", "Nova Scotia", "Nova Scotia Office of the Superintendent of Insurance"),
            "ns": ("NS", "Nova Scotia", "Nova Scotia Office of the Superintendent of Insurance"),
            "northwestterritories": ("NT", "Northwest Territories", "Northwest Territories Superintendent of Insurance"),
            "nt": ("NT", "Northwest Territories", "Northwest Territories Superintendent of Insurance"),
            "nunavut": ("NU", "Nunavut", "Nunavut Superintendent of Insurance"),
            "nu": ("NU", "Nunavut", "Nunavut Superintendent of Insurance"),
            "ontario": ("ON", "Ontario", "FSRA"),
            "on": ("ON", "Ontario", "FSRA"),
            "princeedwardisland": ("PE", "Prince Edward Island", "PEI Superintendent of Insurance"),
            "pei": ("PE", "Prince Edward Island", "PEI Superintendent of Insurance"),
            "pe": ("PE", "Prince Edward Island", "PEI Superintendent of Insurance"),
            "quebec": ("QC", "Québec", "AMF"),
            "québec": ("QC", "Québec", "AMF"),
            "qc": ("QC", "Québec", "AMF"),
            "saskatchewan": ("SK", "Saskatchewan", "Saskatchewan Superintendent of Insurance"),
            "sk": ("SK", "Saskatchewan", "Saskatchewan Superintendent of Insurance"),
            "yukon": ("YT", "Yukon", "Yukon Superintendent of Insurance"),
            "yt": ("YT", "Yukon", "Yukon Superintendent of Insurance"),
        }

        return provinces.get(compact, ("", raw, ""))

    def parse_period(value: Any):
        raw = clean(value)
        match = re.search(
            r"(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\s*(?:au\s*/\s*to|au\s+to|au|to|through|thru|-|–)\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
            raw,
            flags=re.I,
        )
        if not match:
            return "", ""
        return clean(match.group(1)), clean(match.group(2))

    def looks_like_claim_number(value: Any) -> bool:
        raw = clean(value).upper()
        if not raw or raw in {"TOTAL", "N/A", "NA", "-", "—"}:
            return False
        return bool(re.search(r"[A-Z]{1,8}[-_]\d{2,}", raw) or re.search(r"\d{4,}", raw))

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True)
    except Exception as exc:
        print("LOSSQ_SECTIONED_EXCEL_LOSS_RUN_SERVICE_READ_ERROR_V1:", str(exc)[:200])
        return parsed_claims, parsed_profile, direct_profile

    all_rows: list[list[str]] = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [clean(cell) for cell in row]
            if any(values):
                all_rows.append(values)

    if not all_rows:
        return parsed_claims, parsed_profile, direct_profile

    first_text = "\n".join(" ".join(row) for row in all_rows[:35])

    if not re.search(r"(?i)(claims detail|détail des sinistres|detail des sinistres|sinistre|claim #)", first_text):
        return parsed_claims, parsed_profile, direct_profile

    carrier = ""
    business_name = ""
    policy_number = ""
    province_code = ""
    province_name = ""
    regulator = ""
    currency = ""
    line_of_business = ""
    effective_date = ""
    expiration_date = ""
    evaluation_date = ""

    for row in all_rows[:12]:
        line = clean(" ".join(cell for cell in row if cell))
        if re.search(r"(?i)\b(insurance|assurance|assurances|mutual|casualty|indemnity|underwriters|general insurance)\b", line):
            if not re.search(r"(?i)\b(loss run|relevé des sinistres|releve des sinistres|claims detail)\b", line):
                carrier = line
                break

    for row in all_rows:
        label = key(row[0] if len(row) > 0 else "")
        value = clean(row[1] if len(row) > 1 else "")

        if not label or not value:
            continue

        if not business_name and re.search(r"(assuré|assure|insured|named insured)", label):
            business_name = value
            continue

        if not policy_number and re.search(r"(policy|police)", label):
            policy_number = value
            continue

        if not province_code and re.search(r"(province|state)", label):
            province_code, province_name, regulator = parse_province(value)
            continue

        if not currency and re.search(r"(currency|devise)", label):
            currency = value.upper()
            continue

        if not line_of_business and re.search(r"(ibc|bac|line of business|ligne)", label):
            line_of_business = value
            continue

        if not evaluation_date and re.search(r"(report date|date du rapport|valuation|evaluation|évaluation)", label):
            evaluation_date = value
            continue

        if re.search(r"(period|période|periode)", label):
            start, end = parse_period(value)
            effective_date = effective_date or start
            expiration_date = expiration_date or end
            continue

    if not carrier:
        for row in all_rows[-8:]:
            line = clean(" ".join(cell for cell in row if cell))
            match = re.search(r"(?i)(préparé par|prepare par|prepared by)\s*[:#-]\s*(.+)$", line)
            if match:
                carrier = re.sub(r"(?i)\s+[–-]\s+(lignes commerciales|commercial lines).*$", "", clean(match.group(2))).strip()
                break

    header_index = None
    headers: list[str] = []

    for index, row in enumerate(all_rows):
        joined = " ".join(row)
        if re.search(r"(?i)(claim #|sinistre)", joined) and re.search(r"(?i)(loss date|date sinistre|reported|déclaration|declaration|incurred|engagé|engage)", joined):
            header_index = index
            headers = row
            break

    claims: list[dict[str, Any]] = []

    if header_index is not None and headers:
        normalized_headers = [key(item) for item in headers]

        def col(*patterns: str) -> int:
            for idx, item in enumerate(normalized_headers):
                for pattern in patterns:
                    if re.search(pattern, item, flags=re.I):
                        return idx
            return -1

        c_claim = col(r"claim", r"sinistre")
        c_loss = col(r"loss date", r"date sinistre")
        c_report = col(r"reported", r"déclaration", r"declaration")
        c_claimant = col(r"claimant", r"réclamant", r"reclamant")
        c_desc = col(r"description")
        c_cov = col(r"coverage", r"couverture")
        c_status = col(r"status", r"statut")
        c_paid = col(r"paid", r"payé", r"paye")
        c_reserve = col(r"reserve", r"réserve")
        c_incurred = col(r"incurred", r"engagé", r"engage")

        for row in all_rows[header_index + 1 :]:
            if not any(row):
                continue

            if clean(row[0] if row else "").upper() == "TOTAL":
                continue

            joined = " ".join(row)
            if re.search(r"(?i)(notes|remarques|prepared by|préparé par|confidential|confidentiel)", joined):
                break

            claim_number = clean(row[c_claim]) if 0 <= c_claim < len(row) else ""

            if not looks_like_claim_number(claim_number):
                continue

            paid = money_text(row[c_paid]) if 0 <= c_paid < len(row) else "0"
            reserve = money_text(row[c_reserve]) if 0 <= c_reserve < len(row) else "0"
            incurred = money_text(row[c_incurred]) if 0 <= c_incurred < len(row) else ""

            if not incurred:
                incurred = money_text(money_float(paid) + money_float(reserve))

            coverage = clean(row[c_cov]) if 0 <= c_cov < len(row) else line_of_business

            claims.append({
                "claim_number": claim_number,
                "policy_number": policy_number,
                "line_of_business": coverage or line_of_business,
                "coverage": coverage or line_of_business,
                "loss_date": clean(row[c_loss]) if 0 <= c_loss < len(row) else "",
                "reported_date": clean(row[c_report]) if 0 <= c_report < len(row) else "",
                "claimant": clean(row[c_claimant]) if 0 <= c_claimant < len(row) else "",
                "description": clean(row[c_desc]) if 0 <= c_desc < len(row) else "",
                "status": normalize_status(row[c_status]) if 0 <= c_status < len(row) else "",
                "paid": paid,
                "paid_amount": paid,
                "reserve": reserve,
                "reserve_amount": reserve,
                "total_incurred": incurred,
                "incurred": incurred,
                "currency": currency,
                "state": province_code,
                "province": province_name or province_code,
                "carrier_name": carrier,
                "writing_carrier": carrier,
                "source_parser": "sectioned_excel_loss_run_service",
            })

    real_existing = [
        claim for claim in parsed_claims
        if isinstance(claim, dict) and looks_like_claim_number(claim.get("claim_number") or claim.get("claim_no"))
    ]

    if claims and len(claims) >= len(real_existing):
        parsed_claims = claims

    coverage_claim_counts: dict[str, int] = {}
    coverage_totals: dict[str, float] = {}

    for claim in parsed_claims:
        if not isinstance(claim, dict):
            continue
        cov = clean(claim.get("coverage") or claim.get("line_of_business") or line_of_business) or "Policy"
        coverage_claim_counts[cov] = coverage_claim_counts.get(cov, 0) + 1
        coverage_totals[cov] = coverage_totals.get(cov, 0.0) + money_float(claim.get("total_incurred") or claim.get("incurred"))

    # LOSSQ_SECTIONED_EXCEL_EXPOSURE_INPUTS_V1
    exposure_inputs: dict[str, Any] = {}
    exposure_rows: list[dict[str, Any]] = []
    policy_rows = []

    in_exposure = False
    exposure_header_seen = False
    exposure_headers: list[str] = []
    total_premium_from_total_row = ""

    # LOSSQ_SECTIONED_EXCEL_POLICY_LIMIT_NONZERO_AND_PHYSICIAN_ALIAS_V1
    def exposure_header_col(*patterns: str) -> int:
        for idx, header in enumerate(exposure_headers):
            header_key = key(header)
            for pattern in patterns:
                if re.search(pattern, header_key, flags=re.I):
                    return idx
        return -1

    def exposure_header_cols(*patterns: str) -> list[int]:
        indexes: list[int] = []

        for idx, header in enumerate(exposure_headers):
            header_key = key(header)
            for pattern in patterns:
                if re.search(pattern, header_key, flags=re.I):
                    indexes.append(idx)
                    break

        return indexes

    def exposure_money_display(value: Any) -> str:
        amount = money_float(value)
        if not amount:
            return ""
        return str(int(amount)) if amount == int(amount) else str(amount)

    def first_nonzero_exposure_money(row_values: list[str], indexes: list[int]) -> str:
        for idx in indexes:
            if 0 <= idx < len(row_values):
                value = exposure_money_display(row_values[idx])
                if value and money_float(value) > 0:
                    return value

        return ""


    for row in all_rows:
        joined = " ".join(row)

        if re.search(r"(?i)(exposure data|données d'exposition|donnees d exposition)", joined):
            in_exposure = True
            continue

        if in_exposure and re.search(r"(?i)(claims detail|détail des sinistres|detail des sinistres)", joined):
            break

        if not in_exposure:
            continue

        if re.search(r"(?i)(category|catégorie|categorie)", joined):
            exposure_header_seen = True
            exposure_headers = row
            continue

        if not exposure_header_seen:
            continue

        coverage_name = clean(row[0] if row else "")

        if not coverage_name:
            continue

        if coverage_name.upper() == "TOTAL":
            numeric_values = [clean(cell) for cell in row[1:] if clean(cell) and re.search(r"\d", clean(cell))]
            if numeric_values:
                total_premium_from_total_row = exposure_money_display(numeric_values[-1])
            continue

        if re.search(r"(?i)(included|inclus|defense costs|frais de défense|frais de defense)", coverage_name):
            continue

        c_physicians = exposure_header_col(r"physician", r"médecin", r"medecin")
        c_employees = exposure_header_col(r"employee", r"employ")
        c_revenue = exposure_header_col(r"revenue", r"revenu")
        c_limit = exposure_header_col(r"limit", r"limite")
        c_limit_cols = exposure_header_cols(r"limit", r"limite")
        c_premium = exposure_header_col(r"premium", r"prime")

        physicians = clean(row[c_physicians]) if 0 <= c_physicians < len(row) else ""
        employees = clean(row[c_employees]) if 0 <= c_employees < len(row) else ""
        revenue = exposure_money_display(row[c_revenue]) if 0 <= c_revenue < len(row) else ""

        # Use the first non-zero limit value across all limit columns.
        # This prevents E&O rows from showing 0 when the workbook has separate D&O / E&O limit columns.
        limit = first_nonzero_exposure_money(row, c_limit_cols) or (exposure_money_display(row[c_limit]) if 0 <= c_limit < len(row) else "")

        premium = exposure_money_display(row[c_premium]) if 0 <= c_premium < len(row) else ""

        numeric_values = [clean(cell) for cell in row[1:] if clean(cell) and re.search(r"\d", clean(cell))]

        if not premium and numeric_values:
            premium = exposure_money_display(numeric_values[-1])

        matching_claims = 0
        matching_total = 0.0

        for cov, count in coverage_claim_counts.items():
            if coverage_name.lower().find(cov.lower()) >= 0 or cov.lower().find(coverage_name.lower()) >= 0:
                matching_claims += count
                matching_total += coverage_totals.get(cov, 0.0)

        if not matching_claims:
            if re.search(r"(?i)d&o|directors|administrateurs", coverage_name):
                for cov, count in coverage_claim_counts.items():
                    if re.search(r"(?i)d&o|directors|administrateurs", cov):
                        matching_claims += count
                        matching_total += coverage_totals.get(cov, 0.0)

            if re.search(r"(?i)e&o|professional|professionnelle|erreurs|omissions", coverage_name):
                for cov, count in coverage_claim_counts.items():
                    if re.search(r"(?i)e&o|professional|professionnelle|erreurs|omissions", cov):
                        matching_claims += count
                        matching_total += coverage_totals.get(cov, 0.0)

        exposure_rows.append({
            "coverage": coverage_name,
            "physicians": physicians,
            "employees": employees,
            "revenue": revenue,
            "limit": limit,
            "premium": premium,
        })

        policy_rows.append({
            "policy_number": policy_number,
            "policyNumber": policy_number,
            # LOSSQ_SECTIONED_EXCEL_COMPACT_POLICY_ROW_LABELS_V1
            "line_of_business": compact_coverage_label(coverage_name),
            "lineOfBusiness": compact_coverage_label(coverage_name),
            "policy_type": compact_coverage_label(coverage_name),
            "policyType": compact_coverage_label(coverage_name),
            "coverage": compact_coverage_label(coverage_name),
            "coverage_full": coverage_name,
            "coverageFull": coverage_name,
            "carrier": carrier,
            "carrier_name": carrier,
            "carrierName": carrier,
            "writing_carrier": carrier,
            "writingCarrier": carrier,
            "effective_date": effective_date,
            "effectiveDate": effective_date,
            "effective": effective_date,
            "expiration_date": expiration_date,
            "expirationDate": expiration_date,
            "expiration": expiration_date,
            "evaluation_date": evaluation_date,
            "evaluationDate": evaluation_date,
            "state": province_code,
            "stateProvince": province_code,
            "province": province_name or province_code,
            "province_code": province_code,
            "currency": currency,
            "physicians": physicians,
            "employees": employees,
            "revenue": revenue,
            "limit": limit,
            "policy_limit": limit,
            "premium": premium,
            "claims": matching_claims,
            "claim_count": matching_claims,
            "total_incurred": matching_total,
        })

    if exposure_rows:
        lines = [row.get("coverage") for row in exposure_rows if row.get("coverage")]
        premiums = [money_float(row.get("premium")) for row in exposure_rows if row.get("premium")]
        revenues = [money_float(row.get("revenue")) for row in exposure_rows if row.get("revenue")]
        employees_values = [money_float(row.get("employees")) for row in exposure_rows if row.get("employees")]
        physician_values = [money_float(row.get("physicians")) for row in exposure_rows if row.get("physicians")]

        total_premium = money_float(total_premium_from_total_row) or sum(premiums)
        max_revenue = max(revenues) if revenues else 0.0
        max_employees = max(employees_values) if employees_values else 0.0
        max_physicians = max(physician_values) if physician_values else 0.0

        limit_parts = []
        for row in exposure_rows:
            if row.get("coverage") and row.get("limit"):
                limit_parts.append(f"{compact_coverage_label(row.get('coverage'))}: {row.get('limit')}")

        primary_lob = " / ".join(compact_coverage_label(line) for line in lines if compact_coverage_label(line)) if lines else line_of_business

        exposure_inputs.update({
            "detected_lines": len(lines),
            "lines_detected": len(lines),
            "line_count": len(lines),
            "primary_line_of_business": primary_lob,
            "primaryLineOfBusiness": primary_lob,
            "line_of_business": primary_lob,
            "lineOfBusiness": primary_lob,
            "state": province_code,
            "State": province_code,
            "state_province": province_code,
            "stateProvince": province_code,
            "State / Province": province_code,
            "province": province_name or province_code,
            "province_code": province_code,
            "currency": currency,
            "Currency": currency,
            "policy_limits": "; ".join(limit_parts),
            "policyLimits": "; ".join(limit_parts),
            "Policy Limits": "; ".join(limit_parts),
            "current_premium": str(int(total_premium)) if total_premium else "",
            "currentPremium": str(int(total_premium)) if total_premium else "",
            "Current Premium": str(int(total_premium)) if total_premium else "",
            "expiring_premium": str(int(total_premium)) if total_premium else "",
            "expiringPremium": str(int(total_premium)) if total_premium else "",
            "Expiring Premium": str(int(total_premium)) if total_premium else "",
            "revenue": str(int(max_revenue)) if max_revenue else "",
            "annual_revenue": str(int(max_revenue)) if max_revenue else "",
            "revenue_sales": str(int(max_revenue)) if max_revenue else "",
            "revenueSales": str(int(max_revenue)) if max_revenue else "",
            "Revenue / Sales": str(int(max_revenue)) if max_revenue else "",
            "professional_revenue": str(int(max_revenue)) if max_revenue else "",
            "professionalRevenue": str(int(max_revenue)) if max_revenue else "",
            "employee_count": str(int(max_employees)) if max_employees else "",
            "employeeCount": str(int(max_employees)) if max_employees else "",
            "Employee Count": str(int(max_employees)) if max_employees else "",
            "physician_count": str(int(max_physicians)) if max_physicians else "",
            "physicianCount": str(int(max_physicians)) if max_physicians else "",
            "physicians": str(int(max_physicians)) if max_physicians else "",
            "Physician Count": str(int(max_physicians)) if max_physicians else "",
            "Physicians": str(int(max_physicians)) if max_physicians else "",
            "physician_value": str(int(max_physicians)) if max_physicians else "",
            "physicianValue": str(int(max_physicians)) if max_physicians else "",
            "Physician Value": str(int(max_physicians)) if max_physicians else "",
            "exposure_rows": exposure_rows,
            "exposureRows": exposure_rows,
        })

    if not policy_rows and policy_number:
        total_incurred = sum(money_float(claim.get("total_incurred")) for claim in parsed_claims if isinstance(claim, dict))
        policy_rows = [{
            "policy_number": policy_number,
            "policyNumber": policy_number,
            "line_of_business": line_of_business,
            "lineOfBusiness": line_of_business,
            "policy_type": line_of_business,
            "policyType": line_of_business,
            "coverage": line_of_business,
            "carrier": carrier,
            "carrier_name": carrier,
            "carrierName": carrier,
            "writing_carrier": carrier,
            "writingCarrier": carrier,
            "effective_date": effective_date,
            "effectiveDate": effective_date,
            "expiration_date": expiration_date,
            "expirationDate": expiration_date,
            "evaluation_date": evaluation_date,
            "evaluationDate": evaluation_date,
            "state": province_code,
            "stateProvince": province_code,
            "province": province_name or province_code,
            "province_code": province_code,
            "currency": currency,
            "claims": len(parsed_claims),
            "claim_count": len(parsed_claims),
            "total_incurred": total_incurred,
        }]

    # LOSSQ_SECTIONED_EXCEL_DISPLAY_CASE_APPLY_V1
    if carrier:
        carrier = display_name_case(carrier)

    if business_name:
        business_name = display_name_case(business_name)

    # LOSSQ_SECTIONED_EXCEL_FINAL_EXPOSURE_OVERLAY_V2
    # Final overlay from the workbook exposure table. This forces the account
    # exposure fields to use the exposure rows, not only the IBC/BAC header line.
    try:
        exposure_inputs
    except NameError:
        exposure_inputs = {}

    exposure_rows_overlay: list[dict[str, Any]] = []
    policy_rows_overlay: list[dict[str, Any]] = []
    exposure_headers_overlay: list[str] = []
    total_premium_overlay = ""

    in_exposure_overlay = False

    def exposure_overlay_header_indexes(*patterns: str) -> list[int]:
        indexes: list[int] = []

        for idx, header in enumerate(exposure_headers_overlay):
            header_key = key(header)

            for pattern in patterns:
                if re.search(pattern, header_key, flags=re.I):
                    indexes.append(idx)
                    break

        return indexes

    def exposure_overlay_first_value(row_values: list[str], indexes: list[int]) -> str:
        for idx in indexes:
            if 0 <= idx < len(row_values):
                value = clean(row_values[idx])
                if value and value not in {"—", "-", "N/A", "n/a"}:
                    return value

        return ""

    def exposure_overlay_first_money(row_values: list[str], indexes: list[int]) -> str:
        for idx in indexes:
            if 0 <= idx < len(row_values):
                amount = money_float(row_values[idx])
                if amount > 0:
                    return str(int(amount)) if amount == int(amount) else str(amount)

        return ""

    for row in all_rows:
        joined = " ".join(row)

        if re.search(r"(?i)(exposure data|données d'exposition|donnees d exposition)", joined):
            in_exposure_overlay = True
            exposure_headers_overlay = []
            continue

        if in_exposure_overlay and re.search(r"(?i)(claims detail|détail des sinistres|detail des sinistres)", joined):
            break

        if not in_exposure_overlay:
            continue

        if not exposure_headers_overlay and re.search(r"(?i)(category|catégorie|categorie)", joined):
            exposure_headers_overlay = row
            continue

        if not exposure_headers_overlay:
            continue

        coverage_name = clean(row[0] if row else "")

        if not coverage_name:
            continue

        c_premium_cols = exposure_overlay_header_indexes(r"premium", r"prime")

        if coverage_name.upper() == "TOTAL":
            total_premium_overlay = exposure_overlay_first_money(row, c_premium_cols)
            continue

        if re.search(r"(?i)(included|inclus|defense costs|frais de défense|frais de defense)", coverage_name):
            continue

        c_physician_cols = exposure_overlay_header_indexes(r"physician", r"médecin", r"medecin")
        c_employee_cols = exposure_overlay_header_indexes(r"employee", r"employ")
        c_revenue_cols = exposure_overlay_header_indexes(r"revenue", r"revenu")
        c_limit_cols = exposure_overlay_header_indexes(r"limit", r"limite")

        physicians = exposure_overlay_first_value(row, c_physician_cols)
        employees = exposure_overlay_first_value(row, c_employee_cols)
        revenue = exposure_overlay_first_money(row, c_revenue_cols)
        limit = exposure_overlay_first_money(row, c_limit_cols)
        premium = exposure_overlay_first_money(row, c_premium_cols)

        exposure_rows_overlay.append({
            "coverage": coverage_name,
            "physicians": physicians,
            "physician_count": physicians,
            "employees": employees,
            "employee_count": employees,
            "revenue": revenue,
            "limit": limit,
            "policy_limit": limit,
            "premium": premium,
        })

    if exposure_rows_overlay:
        exposure_lines = [row.get("coverage") for row in exposure_rows_overlay if row.get("coverage")]
        primary_lob_overlay = " / ".join(compact_coverage_label(line) for line in exposure_lines if compact_coverage_label(line))

        limit_parts = [
            f"{compact_coverage_label(row.get('coverage'))}: {row.get('limit')}"
            for row in exposure_rows_overlay
            if row.get("coverage") and row.get("limit") and money_float(row.get("limit")) > 0
        ]

        physician_values = [
            money_float(row.get("physicians"))
            for row in exposure_rows_overlay
            if row.get("physicians") and money_float(row.get("physicians")) > 0
        ]

        employee_values = [
            money_float(row.get("employees"))
            for row in exposure_rows_overlay
            if row.get("employees") and money_float(row.get("employees")) > 0
        ]

        revenue_values = [
            money_float(row.get("revenue"))
            for row in exposure_rows_overlay
            if row.get("revenue") and money_float(row.get("revenue")) > 0
        ]

        premium_values = [
            money_float(row.get("premium"))
            for row in exposure_rows_overlay
            if row.get("premium") and money_float(row.get("premium")) > 0
        ]

        physician_count_overlay = max(physician_values) if physician_values else 0.0
        employee_count_overlay = max(employee_values) if employee_values else 0.0
        revenue_overlay = max(revenue_values) if revenue_values else 0.0
        total_premium_value_overlay = money_float(total_premium_overlay) or sum(premium_values)
        policy_limits_overlay = "; ".join(limit_parts)

        if primary_lob_overlay:
            line_of_business = primary_lob_overlay

        exposure_inputs.update({
            "detected_lines": len(exposure_lines),
            "lines_detected": len(exposure_lines),
            "line_count": len(exposure_lines),
            "primary_line_of_business": primary_lob_overlay,
            "primaryLineOfBusiness": primary_lob_overlay,
            "line_of_business": primary_lob_overlay,
            "lineOfBusiness": primary_lob_overlay,
            "policy_limits": policy_limits_overlay,
            "policyLimits": policy_limits_overlay,
            "Policy Limits": policy_limits_overlay,
            "limits": policy_limits_overlay,
            "coverage_limit": policy_limits_overlay,
            "coverageLimit": policy_limits_overlay,
            "current_premium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "currentPremium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "Current Premium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "expiring_premium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "expiringPremium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "Expiring Premium": str(int(total_premium_value_overlay)) if total_premium_value_overlay else "",
            "revenue": str(int(revenue_overlay)) if revenue_overlay else "",
            "annual_revenue": str(int(revenue_overlay)) if revenue_overlay else "",
            "revenue_sales": str(int(revenue_overlay)) if revenue_overlay else "",
            "revenueSales": str(int(revenue_overlay)) if revenue_overlay else "",
            "Revenue / Sales": str(int(revenue_overlay)) if revenue_overlay else "",
            "professional_revenue": str(int(revenue_overlay)) if revenue_overlay else "",
            "professionalRevenue": str(int(revenue_overlay)) if revenue_overlay else "",
            "employee_count": str(int(employee_count_overlay)) if employee_count_overlay else "",
            "employeeCount": str(int(employee_count_overlay)) if employee_count_overlay else "",
            "Employee Count": str(int(employee_count_overlay)) if employee_count_overlay else "",
            "physician_count": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "physicianCount": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "physicians": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "Physician Count": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "physician_value": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "physicianValue": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "Physician Value": str(int(physician_count_overlay)) if physician_count_overlay else "",
            "exposure_rows": exposure_rows_overlay,
            "exposureRows": exposure_rows_overlay,
        })

        policy_rows_overlay = []

        for row in exposure_rows_overlay:
            coverage_name = clean(row.get("coverage"))
            limit = clean(row.get("limit"))
            premium = clean(row.get("premium"))

            matching_claims = 0
            matching_total = 0.0

            for claim in parsed_claims:
                if not isinstance(claim, dict):
                    continue

                claim_coverage = clean(claim.get("coverage") or claim.get("line_of_business"))

                is_match = False

                if coverage_name and claim_coverage:
                    is_match = coverage_name.lower().find(claim_coverage.lower()) >= 0 or claim_coverage.lower().find(coverage_name.lower()) >= 0

                if not is_match and re.search(r"(?i)d&o|directors|administrateurs", coverage_name) and re.search(r"(?i)d&o|directors|administrateurs", claim_coverage):
                    is_match = True

                if not is_match and re.search(r"(?i)e&o|professional|professionnelle|erreurs|omissions", coverage_name) and re.search(r"(?i)e&o|professional|professionnelle|erreurs|omissions", claim_coverage):
                    is_match = True

                if is_match:
                    matching_claims += 1
                    matching_total += money_float(claim.get("total_incurred") or claim.get("incurred"))

            policy_rows_overlay.append({
                "policy_number": policy_number,
                "policyNumber": policy_number,
                # LOSSQ_SECTIONED_EXCEL_COMPACT_OVERLAY_POLICY_ROW_LABELS_V1
                "line_of_business": compact_coverage_label(coverage_name),
                "lineOfBusiness": compact_coverage_label(coverage_name),
                "policy_type": compact_coverage_label(coverage_name),
                "policyType": compact_coverage_label(coverage_name),
                "coverage": compact_coverage_label(coverage_name),
                "coverage_full": coverage_name,
                "coverageFull": coverage_name,
                "carrier": carrier,
                "carrier_name": carrier,
                "carrierName": carrier,
                "writing_carrier": carrier,
                "writingCarrier": carrier,
                "effective_date": effective_date,
                "effectiveDate": effective_date,
                "expiration_date": expiration_date,
                "expirationDate": expiration_date,
                "evaluation_date": evaluation_date,
                "evaluationDate": evaluation_date,
                "state": province_code,
                "stateProvince": province_code,
                "province": province_name or province_code,
                "province_code": province_code,
                "currency": currency,
                "physicians": row.get("physicians"),
                "physician_count": row.get("physicians"),
                "physicianCount": row.get("physicians"),
                "employees": row.get("employees"),
                "employee_count": row.get("employees"),
                "revenue": row.get("revenue"),
                "limit": limit,
                "policy_limit": limit,
                "policyLimit": limit,
                "premium": premium,
                "claims": matching_claims,
                "claim_count": matching_claims,
                "total_incurred": matching_total,
            })

        if policy_rows_overlay:
            policy_rows = policy_rows_overlay

    is_canada = bool(province_code or currency == "CAD" or re.search(r"(?i)canada|cad|québec|quebec|ontario|manitoba|ibc|bac", first_text))

    def apply_profile(target: dict[str, Any]):
        if not isinstance(target, dict):
            return

        if business_name:
            target["business_name"] = business_name
            target["insured_name"] = business_name
            target["insured"] = business_name
            target["company_name"] = business_name
            target["account_name"] = business_name

        if carrier:
            target["carrier_name"] = carrier
            target["carrierName"] = carrier
            target["carrier"] = carrier
            target["writing_carrier"] = carrier
            target["writingCarrier"] = carrier
            target["document_writing_carrier"] = carrier
            target["documentWritingCarrier"] = carrier
            target["uploaded_writing_carrier"] = carrier
            target["uploadedWritingCarrier"] = carrier

        if policy_number:
            target["policy_number"] = policy_number
            target["policyNumber"] = policy_number
            target["main_policy_number"] = policy_number
            target["mainPolicyNumber"] = policy_number

        if line_of_business:
            target["line_of_business"] = line_of_business
            target["lineOfBusiness"] = line_of_business
            target["policy_type"] = line_of_business
            target["policyType"] = line_of_business

        if effective_date:
            target["effective_date"] = effective_date
            target["effectiveDate"] = effective_date
            target["effective"] = effective_date

        if expiration_date:
            target["expiration_date"] = expiration_date
            target["expirationDate"] = expiration_date
            target["expiration"] = expiration_date

        if evaluation_date:
            target["evaluation_date"] = evaluation_date
            target["evaluationDate"] = evaluation_date
            target["valuation_date"] = evaluation_date
            target["valuationDate"] = evaluation_date
            target["report_date"] = evaluation_date
            target["reportDate"] = evaluation_date

        if currency:
            target["currency"] = currency
            target["default_currency"] = currency
            target["defaultCurrency"] = currency

        if province_code:
            target["state"] = province_code
            target["primary_state"] = province_code
            target["primaryState"] = province_code
            target["state_province"] = province_code
            target["stateProvince"] = province_code
            target["province_code"] = province_code
            target["provinceCode"] = province_code
            target["province"] = province_name or province_code
            target["province_name"] = province_name or province_code
            target["provinceName"] = province_name or province_code

        if is_canada:
            target["country"] = "Canada"
            target["market"] = "Canada"
            target["country_market"] = "Canada"
            target["countryMarket"] = "Canada"
            target["market_country"] = "Canada"
            target["marketCountry"] = "Canada"
            target["date_format"] = "DD/MM/YYYY"
            target["dateFormat"] = "DD/MM/YYYY"
            target["market_date_format"] = "DD/MM/YYYY"
            target["marketDateFormat"] = "DD/MM/YYYY"

        if regulator:
            target["regulator"] = regulator
            target["insurance_regulator"] = regulator
            target["insuranceRegulator"] = regulator
            target["market_regulator"] = regulator
            target["marketRegulator"] = regulator

        market_context = target.get("market_context") or target.get("marketContext")
        if not isinstance(market_context, dict):
            market_context = {}

        if is_canada:
            market_context["country"] = "Canada"
            market_context["market"] = "Canada"
            market_context["date_format"] = "DD/MM/YYYY"
            market_context["dateFormat"] = "DD/MM/YYYY"

        if province_code:
            market_context["state"] = province_code
            market_context["state_province"] = province_code
            market_context["stateProvince"] = province_code
            market_context["province"] = province_code
            market_context["province_code"] = province_code
            market_context["provinceCode"] = province_code
            market_context["province_name"] = province_name or province_code

        if currency:
            market_context["currency"] = currency

        if regulator:
            market_context["regulator"] = regulator

        target["market_context"] = market_context
        target["marketContext"] = market_context

        if policy_rows:
            target["policies"] = policy_rows
            target["policy_schedule"] = policy_rows
            target["policySchedule"] = policy_rows

        # LOSSQ_SECTIONED_EXCEL_PROFILE_EXPOSURE_ROOT_APPLY_V2
        if isinstance(exposure_inputs, dict) and exposure_inputs:
            existing_exposure = target.get("exposure_inputs") or target.get("exposureInputs")
            if not isinstance(existing_exposure, dict):
                existing_exposure = {}

            merged_exposure = dict(existing_exposure)
            merged_exposure.update({
                k: v
                for k, v in exposure_inputs.items()
                if v not in ("", None, [], {})
            })

            target["exposure_inputs"] = merged_exposure
            target["exposureInputs"] = merged_exposure
            target["exposures"] = merged_exposure
            target["manual_exposure_inputs"] = merged_exposure
            target["manualExposureInputs"] = merged_exposure

            for exposure_key, exposure_value in merged_exposure.items():
                if exposure_value not in ("", None, [], {}):
                    target[exposure_key] = exposure_value

        # LOSSQ_SECTIONED_EXCEL_APPLY_EXPOSURE_INPUTS_V1
        if exposure_inputs:
            existing_exposure = target.get("exposure_inputs") or target.get("exposureInputs")
            if not isinstance(existing_exposure, dict):
                existing_exposure = {}

            merged_exposure = dict(existing_exposure)
            merged_exposure.update({k: v for k, v in exposure_inputs.items() if v not in ("", None, [], {})})

            target["exposure_inputs"] = merged_exposure
            target["exposureInputs"] = merged_exposure
            target["exposures"] = merged_exposure
            target["manual_exposure_inputs"] = merged_exposure
            target["manualExposureInputs"] = merged_exposure

            for key_name, key_value in merged_exposure.items():
                if key_value not in ("", None, [], {}):
                    target[key_name] = key_value

        if parsed_claims:
            target["claims"] = parsed_claims
            target["parsed_claims"] = parsed_claims

    apply_profile(parsed_profile)
    apply_profile(direct_profile)

    print("LOSSQ_SECTIONED_EXCEL_LOSS_RUN_SERVICE_V1:", {
        "business_name": business_name,
        "carrier": carrier,
        "policy_number": policy_number,
        "province": province_code,
        "currency": currency,
        "effective_date": effective_date,
        "expiration_date": expiration_date,
        "evaluation_date": evaluation_date,
        "claims": len(parsed_claims),
        "policy_rows": len(policy_rows),
    })

    return parsed_claims, parsed_profile, direct_profile
