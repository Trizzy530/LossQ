import csv
from fastapi import HTTPException, APIRouter, UploadFile, File, Depends, Form
from sqlalchemy.orm import Session
from sqlalchemy import text, inspect, func
import shutil
import os
import json
from datetime import datetime
from typing import List, Any

# LOSSQ_MARKET_INTELLIGENCE_IMPORT_V1
try:
  from app.services.market_intelligence import lossq_normalize_market_profile, lossq_market_intelligence_summary
except Exception as lossq_market_intelligence_import_error:
  lossq_normalize_market_profile = None
  lossq_market_intelligence_summary = None
  print("LOSSQ_MARKET_INTELLIGENCE_IMPORT_FAILED_V1", str(lossq_market_intelligence_import_error))

from app.database import SessionLocal
from app.models.claim import Claim
from app.models.upload_history import UploadHistory
from app.models.account_profile import AccountProfile
import re
from app.services.audit import record_audit_event
from app.services.loss_run_pipeline import parse_loss_run_file
from app.services.universal_profile import extract_universal_profile_from_text
# LOSSQ_CANADA_UPLOAD_SUPPORT_IMPORT_V3
try:
  from app.services.canada_loss_run_support import (
    enhance_claim_for_canada as lossq_canada_enhance_claim_for_canada,
    enhance_profile_for_canada as lossq_canada_enhance_profile_for_canada,
  )
except Exception:
  lossq_canada_enhance_claim_for_canada = None
  lossq_canada_enhance_profile_for_canada = None
import traceback
from app.role_utils import require_permission
from app.services.row_policy_preservation import preserve_row_policy_fields

try:
  from app.services.excel_parser_service import parse_claims_from_excel
except Exception:
  parse_claims_from_excel = None


# LOSSQ_UPLOAD_SECURITY_PHASE_2_V1
MAX_UPLOAD_SIZE_MB = int(os.getenv("MAX_UPLOAD_SIZE_MB", "25"))
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024

ALLOWED_UPLOAD_EXTENSIONS = {
  ".pdf",
  ".csv",
  ".xlsx",
  ".xls",
  ".png",
  ".jpg",
  ".jpeg",
  ".txt",
}

ALLOWED_UPLOAD_CONTENT_TYPES = {
  "application/pdf",
  "text/csv",
  "text/plain",
  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
  "application/vnd.ms-excel",
  "image/png",
  "image/jpeg",
  "application/octet-stream", # Some browsers send this for CSV/XLSX/PDF.
}

BLOCKED_UPLOAD_EXTENSIONS = {
  ".exe",
  ".bat",
  ".cmd",
  ".com",
  ".scr",
  ".js",
  ".vbs",
  ".ps1",
  ".sh",
  ".php",
  ".py",
  ".jar",
  ".msi",
  ".dll",
  ".html",
  ".htm",
  ".svg",
}


def sanitize_upload_filename(filename: str):
  filename = str(filename or "upload").strip()
  filename = filename.replace("\\", "_").replace("/", "_")
  filename = re.sub(r"[^A-Za-z0-9._ -]", "_", filename)
  filename = re.sub(r"\s+", "_", filename)
  filename = filename.strip("._- ")

  if not filename:
    filename = "upload"

  if len(filename) > 140:
    stem, ext = os.path.splitext(filename)
    filename = f"{stem[:120]}{ext}"

  return filename


async def validate_upload_file_security(file):
  filename = sanitize_upload_filename(getattr(file, "filename", "") or "")
  content_type = str(getattr(file, "content_type", "") or "").lower().strip()
  _, ext = os.path.splitext(filename.lower())

  if not ext:
    raise HTTPException(
      status_code=400,
      detail="Upload blocked. File must include a valid extension.",
    )

  if ext in BLOCKED_UPLOAD_EXTENSIONS:
    raise HTTPException(
      status_code=400,
      detail="Upload blocked. This file type is not allowed.",
    )

  if ext not in ALLOWED_UPLOAD_EXTENSIONS:
    raise HTTPException(
      status_code=400,
      detail="Upload blocked. Allowed file types are PDF, CSV, XLSX, XLS, PNG, JPG, JPEG, and TXT.",
    )

  if content_type and content_type not in ALLOWED_UPLOAD_CONTENT_TYPES:
    # Do not over-block octet-stream cases, but block clearly dangerous browser-reported types.
    if not content_type.startswith("application/octet-stream"):
      raise HTTPException(
        status_code=400,
        detail="Upload blocked. The uploaded file content type is not allowed.",
      )

  # Check file size without permanently consuming the stream.
  try:
    await file.seek(0)
    content = await file.read()
    size = len(content or b"")

    if size <= 0:
      raise HTTPException(
        status_code=400,
        detail="Upload blocked. The file appears to be empty.",
      )

    if size > MAX_UPLOAD_SIZE_BYTES:
      raise HTTPException(
        status_code=413,
        detail=f"Upload blocked. File size must be {MAX_UPLOAD_SIZE_MB}MB or less.",
      )

    await file.seek(0)
  except HTTPException:
    raise
  except Exception:
    try:
      await file.seek(0)
    except Exception:
      pass
    raise HTTPException(
      status_code=400,
      detail="Upload blocked. The file could not be validated safely.",
    )

  return filename


# LOSSQ_FILTER_CLAIM_MODEL_FIELDS_BEFORE_SAVE_V1
def lossq_filter_claim_model_fields(data: dict):
  """Keep only fields that exist on the Claim SQLAlchemy model before Claim(**data)."""
  if not isinstance(data, dict):
    return {}

  try:
    allowed_fields = set(Claim.__table__.columns.keys())
  except Exception:
    allowed_fields = {
      "id",
      "organization_id",
      "account_profile_id",
      "claim_number",
      "policy_number",
      "carrier_name",
      "line_of_business",
      "claim_type",
      "status",
      "date_of_loss",
      "date_reported",
      "date_closed",
      "paid_amount",
      "reserve_amount",
      "total_incurred",
      "description",
      "claimant_name",
      "litigation",
      "fraud_flag",
      "risk_flag",
      "created_at",
      "updated_at",
    }

  cleaned = {}
  removed = {}

  for key, value in data.items():
    if key in allowed_fields:
      cleaned[key] = value
    else:
      removed[key] = value

  if removed:
    print("LOSSQ_CLAIM_FIELD_FILTER_REMOVED:", sorted(list(removed.keys())))

  return cleaned


# LOSSQ_ROW_LEVEL_POLICY_SAVE_PRESERVATION_V1
def lossq_preserve_row_policy_before_save(normalized: dict, raw_claim: dict, fallback_policy_number: str = ""):
  """
  Preserve each claim row's own policy number and policy type/line before Claim(**normalized).
  This prevents account/main policy from overwriting every claim row.
  """
  if not isinstance(normalized, dict):
    normalized = {}

  if not isinstance(raw_claim, dict):
    raw_claim = {}

  def clean(value):
    return clean_profile_value(value)

  row_policy = (
    clean(raw_claim.get("policy_number"))
    or clean(raw_claim.get("Policy Number"))
    or clean(raw_claim.get("policy_no"))
    or clean(raw_claim.get("Policy No"))
    or clean(raw_claim.get("policy"))
    or clean(raw_claim.get("Policy"))
  )

  row_line = (
    clean(raw_claim.get("policy_type"))
    or clean(raw_claim.get("Policy Type"))
    or clean(raw_claim.get("line_of_business"))
    or clean(raw_claim.get("Line of Business"))
    or clean(raw_claim.get("claim_type"))
    or clean(raw_claim.get("Coverage"))
    or clean(raw_claim.get("coverage"))
    or clean(raw_claim.get("Line"))
    or clean(raw_claim.get("line"))
  )

  row_status = (
    clean(raw_claim.get("status"))
    or clean(raw_claim.get("Status"))
    or clean(raw_claim.get("claim_status"))
    or clean(raw_claim.get("Claim Status"))
  )

  if row_policy and not is_bad_policy_key_for_upload(row_policy):
    normalized["policy_number"] = row_policy
  elif not clean(normalized.get("policy_number")):
    fallback = clean(fallback_policy_number)
    if fallback and not is_bad_policy_key_for_upload(fallback):
      normalized["policy_number"] = fallback

  if row_line:
    normalized["line_of_business"] = row_line
    normalized["claim_type"] = row_line

  if row_status:
    normalized["status"] = row_status

  return normalized


# LOSSQ_CLEAN_STANDARD_CSV_ROW_POLICY_OVERRIDE_V1


# LOSSQ_AGENCY_HEADER_FIRST_EXTRACTION_V1
def lossq_header_agency_from_csv(file_path):
  """
  Extract Producing Agency / Agency / Producer / Broker from clean CSV column values.
  This prevents reading the next header cell such as Policy Number as the agency.
  """
  try:
    if not str(file_path or "").lower().endswith(".csv"):
      return ""

    import csv
    import re

    def clean(value):
      return re.sub(r"\s+", " ", str(value or "").strip())

    def key(value):
      return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

    agency_keys = {
      "producingagency",
      "agency",
      "agencyname",
      "producer",
      "broker",
      "brokerage",
      "producingbroker",
      "brokeragency",
    }

    bad_values = {
      "policy number",
      "policy no",
      "policy type",
      "coverage",
      "line",
      "line of business",
      "effective date",
      "expiration date",
      "claim number",
      "claim no",
      "status",
      "paid",
      "reserve",
      "total incurred",
      "carrier",
      "writing carrier",
      "account name",
      "named insured",
      "insured",
    }

    with open(file_path, "r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
      reader = csv.DictReader(handle)

      if not reader.fieldnames:
        return ""

      agency_fields = [
        field for field in reader.fieldnames
        if key(field) in agency_keys
      ]

      if not agency_fields:
        return ""

      for row in reader:
        for field in agency_fields:
          value = clean((row or {}).get(field, ""))
          if value and value.lower() not in bad_values and key(value) not in agency_keys:
            return value

    return ""
  except Exception as exc:
    print("LOSSQ_AGENCY_HEADER_FIRST_EXTRACTION_ERROR:", str(exc)[:200])
    return ""


# LOSSQ_UNIVERSAL_PRODUCING_AGENCY_EXTRACTION_V1
def lossq_universal_agency_from_csv(file_path):
  """
  Extract producing agency/broker/producer from common CSV layouts:
  - Clean tabular columns: Producing Agency, Agency, Producer, Broker
  - Label-pair rows: Agency, Summit Table Risk Advisors
  - Messy section rows: Producing Agency / Broker / Brokerage
  """
  try:
    if not str(file_path or "").lower().endswith(".csv"):
      return ""

    import csv
    import re

    def clean(value):
      return re.sub(r"\s+", " ", str(value or "").strip())

    def key(value):
      return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

    agency_keys = {
      "producingagency",
      "agency",
      "agencyname",
      "producer",
      "broker",
      "brokerage",
      "producingbroker",
      "brokeragency",
    }

    with open(file_path, "r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
      rows = list(csv.reader(handle))

    for row in rows[:80]:
      cleaned_row = [clean(cell) for cell in row]

      for idx, cell in enumerate(cleaned_row):
        if key(cell) in agency_keys:
          for value in cleaned_row[idx + 1:]:
            value_key = key(value)
            if value and value_key not in agency_keys:
              return value

    # Header-style extraction.
    with open(file_path, "r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
      reader = csv.DictReader(handle)
      for row in reader:
        for name, value in (row or {}).items():
          if key(name) in agency_keys and clean(value):
            return clean(value)

    return ""
  except Exception as exc:
    print("LOSSQ_UNIVERSAL_AGENCY_EXTRACTION_ERROR:", str(exc)[:200])
    return ""


def lossq_clean_standard_csv_override(file_path, parsed_claims=None, parsed_profile=None):
  # LOSSQ_CLEAN_STANDARD_CSV_OVERRIDE_SAFE_DEFAULTS_V1
  # LOSSQ_FLAT_CSV_PROFILE_STATUS_ZERO_CLAIM_POLICY_V1
  # LOSSQ_FLAT_CSV_PREAMBLE_TABLE_HEADER_V2
  # Universal flat CSV repair:
  # - Handles metadata preamble rows before the real CSV table header.
  # - Uses Account Name / Business Name as insured instead of filename fallback.
  # - Preserves every policy row, including zero-claim Umbrella / Excess rows.
  # - Saves Claim Status into both status and claim_status.
  # - Treats blank/no-direct-claim rows as policy schedule rows, not claims.
  import csv
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  raw_rows = []
  for encoding in ("utf-8-sig", "utf-8", "latin-1"):
    try:
      with open(file_path, "r", newline="", encoding=encoding, errors="ignore") as f:
        raw_rows = list(csv.reader(f))
      break
    except Exception:
      raw_rows = []

  if not raw_rows:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def good_profile_value(value):
    value = clean(value)
    bad = {
      "",
      "-",
      "n/a",
      "na",
      "none",
      "unknown",
      "account",
      "account number",
      "account no",
      "policy",
      "policy number",
      "policy no",
      "claim",
      "claim number",
      "claim no",
      "business name",
      "named insured",
      "insured",
      "carrier",
      "writing carrier",
    }
    return value if value.lower() not in bad else ""

  # Find the real table header row. Some carrier CSVs have preamble rows first:
  # Account Name, Account Number, Carrier, Policy Period, Valuation Date, then a blank row,
  # then Business Name, Account Number, Policy Number, Claim Number, Claim Status, etc.
  header_index = None
  for idx, row in enumerate(raw_rows):
    row_keys = {key(cell) for cell in row}
    has_policy = bool({"policynumber", "policyno", "policy"} & row_keys)
    has_claim_or_policy_type = bool({"claimnumber", "claimno", "claim", "policytype", "lineofbusiness", "coverage"} & row_keys)
    has_business = bool({"businessname", "accountname", "namedinsured", "insured"} & row_keys)
    if has_policy and (has_claim_or_policy_type or has_business):
      header_index = idx
      break

  if header_index is None:
    return parsed_claims, parsed_profile

  preamble = {}
  for row in raw_rows[:header_index]:
    nonempty = [clean(cell) for cell in row if clean(cell)]
    if len(nonempty) >= 2:
      preamble[key(nonempty[0])] = nonempty[1]

  headers = [clean(cell) for cell in raw_rows[header_index]]
  rows = []
  for raw in raw_rows[header_index + 1:]:
    if not any(clean(cell) for cell in raw):
      continue
    row = {}
    for i, header in enumerate(headers):
      if not header:
        continue
      row[header] = raw[i] if i < len(raw) else ""
    rows.append(row)

  if not rows:
    return parsed_claims, parsed_profile

  def get(row, *aliases):
    if not isinstance(row, dict):
      return ""

    wanted = {key(alias) for alias in aliases if clean(alias)}
    for raw_key, raw_value in row.items():
      if key(raw_key) in wanted:
        value = clean(raw_value)
        if value:
          return value
    return ""

  def preamble_value(*aliases):
    for alias in aliases:
      value = good_profile_value(preamble.get(key(alias)))
      if value:
        return value
    return ""

  def first_value(*aliases):
    value = preamble_value(*aliases)
    if value:
      return value

    for row in rows:
      value = good_profile_value(get(row, *aliases))
      if value:
        return value
    return ""

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    raw = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      return float(raw or 0)
    except Exception:
      return 0.0

  def split_policy_period(value):
    raw = clean(value)
    if not raw:
      return "", ""
    parts = re.split(r"\s*(?:-|to|through|thru)\s*", raw, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) == 2:
      return clean(parts[0]), clean(parts[1])
    return "", ""

  def status_value(value, paid=0.0, reserve=0.0, incurred=0.0):
    raw = clean(value)
    low = raw.lower()

    if low in {"open", "opened", "active", "pending", "reopened", "reopen", "in progress"}:
      return "Open"

    if low in {"closed", "close", "final", "settled", "resolved"}:
      return "Closed"

    if low in {
      "no direct claim activity",
      "no claim activity",
      "no claims",
      "none",
      "n/a",
      "na",
      "-",
    }:
      return ""

    if reserve > 0:
      return "Open"

    if incurred > 0 or paid > 0:
      return "Closed"

    return raw.title() if raw else ""

  def policy_line_from_number(value, fallback=""):
    upper = clean(value).upper()
    fallback = clean(fallback)

    if fallback and fallback.lower() not in {"unknown", "n/a", "none", "-"}:
      return fallback

    if upper.startswith(("GL", "CGL")):
      return "General Liability"
    if upper.startswith(("LIQ", "LQ")):
      return "Liquor Liability"
    if upper.startswith(("WC", "WORK")):
      return "Workers Compensation"
    if upper.startswith(("BOP", "PROP", "CP")):
      return "Businessowners Policy"
    if upper.startswith(("UMB", "UM", "EXC", "XS")):
      return "Umbrella"
    if upper.startswith(("AUTO", "CA", "AL")):
      return "Commercial Auto"
    if upper.startswith(("CARGO", "MTC")):
      return "Cargo"
    if upper.startswith(("CY", "CYB")):
      return "Cyber Liability"

    return fallback or "Commercial Policy"

  def is_no_claim_row(row, claim_number):
    text = " ".join(clean(v).lower() for v in (row or {}).values())
    if clean(claim_number):
      return False
    return bool(
      re.search(
        r"\b(no\s+direct\s+claim\s+activity|no\s+claim\s+activity|no\s+claims?|none\s+reported|zero\s+claims?)\b",
        text,
      )
    )

  business_name = first_value("Account Name", "Business Name", "Named Insured", "Insured", "Insured Name", "Company Name")
  carrier_name = first_value("Carrier", "Writing Carrier", "Insurance Carrier", "Carrier Name")
  producing_agency = first_value("Producing Agency", "Agency", "Agency Name", "Producer", "Producer Name", "Broker", "Broker Name")
  account_number = first_value("Account Number", "Account No", "Account #", "Customer Number", "Client Number")
  effective_date = first_value("Effective Date", "Policy Effective Date", "Policy Effective", "Eff Date")
  expiration_date = first_value("Expiration Date", "Policy Expiration Date", "Policy Expiration", "Exp Date")
  evaluation_date = first_value("Evaluation Date", "Valuation Date", "As Of Date", "Loss Run Date")

  period_start, period_end = split_policy_period(preamble_value("Policy Period", "Policy Term", "Coverage Period"))
  if not effective_date and period_start:
    effective_date = period_start
  if not expiration_date and period_end:
    expiration_date = period_end

  if business_name:
    parsed_profile["business_name"] = business_name
    parsed_profile["insured_name"] = business_name
    parsed_profile["named_insured"] = business_name
    parsed_profile["account_name"] = business_name

  if carrier_name:
    parsed_profile["carrier_name"] = carrier_name
    parsed_profile["writing_carrier"] = carrier_name
    parsed_profile["carrier"] = carrier_name

  if producing_agency:
    parsed_profile["producing_agency"] = producing_agency
    parsed_profile["agency_name"] = producing_agency
    parsed_profile["producer"] = producing_agency

  if account_number:
    parsed_profile["account_number"] = account_number
    parsed_profile["customer_number"] = account_number

  if effective_date:
    parsed_profile["effective_date"] = effective_date

  if expiration_date:
    parsed_profile["expiration_date"] = expiration_date

  if evaluation_date:
    parsed_profile["evaluation_date"] = evaluation_date
    parsed_profile["valuation_date"] = evaluation_date

  # LOSSQ_FLAT_CSV_PREAMBLE_EXPOSURE_INPUTS_V1
  # Extract exposure/rating-basis values from clean flat CSV rows that have
  # metadata preamble above the actual table header. No customer/carrier hardcoding.
  def first_table_value(*aliases):
    for row in rows:
      value = get(row, *aliases)
      value = good_profile_value(value)
      if value:
        return value
    return ""

  def money_number(value):
    raw = clean(value)
    if not raw:
      return 0.0
    raw = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      return float(raw or 0)
    except Exception:
      return 0.0

  def clean_numeric_text(value):
    raw = clean(value)
    if not raw:
      return ""
    numeric = raw.replace("$", "").replace(",", "").strip()
    return numeric

  def money_display(value):
    amount = money_number(value)
    if amount <= 0:
      return ""
    return str(int(amount)) if float(amount).is_integer() else f"{amount:.2f}"

  revenue_value = first_table_value(
    "Annual Revenue",
    "Revenue",
    "Sales",
    "Gross Sales",
    "Receipts",
    "Total Revenue",
    "Revenue / Sales",
  )
  payroll_value = first_table_value("Payroll", "Annual Payroll", "Total Payroll")
  employee_count_value = first_table_value("Employee Count", "Employees", "Number of Employees")
  location_count_value = first_table_value("Location Count", "Locations", "Number of Locations")
  liquor_sales_value = first_table_value("Liquor Sales", "Alcohol Sales", "Beer Wine Liquor Sales")

  premium_by_policy = {}
  for exposure_row in rows:
    exposure_policy = get(exposure_row, "Policy Number", "Policy No", "Policy #", "policy_number", "policy")
    exposure_premium = get(
      exposure_row,
      "Current Premium",
      "Annual Premium",
      "Premium",
      "Written Premium",
      "Policy Premium",
    )
    exposure_policy_key = clean(exposure_policy).upper()
    exposure_premium_amount = money_number(exposure_premium)
    if exposure_policy_key and exposure_premium_amount > 0:
      premium_by_policy[exposure_policy_key] = exposure_premium_amount

  current_premium_total = sum(premium_by_policy.values())

  exposure_inputs = parsed_profile.get("exposure_inputs")
  if not isinstance(exposure_inputs, dict):
    exposure_inputs = {}

  exposure_rows = parsed_profile.get("exposures")
  if not isinstance(exposure_rows, list):
    exposure_rows = []

  def set_exposure(field, value, label=None, money_field=False):
    value = clean(value)
    if not value:
      return

    if money_field:
      normalized = money_display(value)
    else:
      normalized = clean_numeric_text(value) or value

    if not normalized:
      return

    parsed_profile[field] = normalized

    if label:
      exposure_inputs[label] = normalized

    exposure_rows.append({
      "field": field,
      "label": label or field,
      "value": normalized,
      "source": "flat_csv_preamble_table",
    })

  set_exposure("revenue", revenue_value, "Revenue / Sales", money_field=True)
  set_exposure("sales", revenue_value, "Sales", money_field=True)
  set_exposure("annual_revenue", revenue_value, "Annual Revenue", money_field=True)
  set_exposure("payroll", payroll_value, "Payroll", money_field=True)
  set_exposure("employee_count", employee_count_value, "Employee Count")
  set_exposure("location_count", location_count_value, "Location Count")
  set_exposure("liquor_sales", liquor_sales_value, "Liquor Sales", money_field=True)

  # LOSSQ_FLAT_CSV_EXPOSURE_LOCATION_LIQUOR_TIV_FIX_V1
  # Preserve common frontend/report aliases for restaurant/location/liquor exposures.
  if parsed_profile.get("location_count"):
    parsed_profile["locations"] = parsed_profile.get("location_count")
    parsed_profile["locationCount"] = parsed_profile.get("location_count")
    exposure_inputs["Locations"] = parsed_profile.get("location_count")

  if parsed_profile.get("liquor_sales"):
    parsed_profile["alcohol_sales"] = parsed_profile.get("liquor_sales")
    parsed_profile["liquorSales"] = parsed_profile.get("liquor_sales")
    exposure_inputs["Alcohol Sales"] = parsed_profile.get("liquor_sales")

  # Property TIV should not be populated from date fragments like "01".
  for tiv_key in ["property_tiv", "tiv"]:
    raw_tiv = clean(parsed_profile.get(tiv_key))
    raw_tiv_number = raw_tiv.replace("$", "").replace(",", "").strip()
    try:
      tiv_amount = float(raw_tiv_number or 0)
    except Exception:
      tiv_amount = 0

    if raw_tiv and tiv_amount > 0 and tiv_amount < 1000:
      parsed_profile.pop(tiv_key, None)

  if current_premium_total > 0:
    current_premium_value = str(int(current_premium_total)) if float(current_premium_total).is_integer() else f"{current_premium_total:.2f}"
    parsed_profile["current_premium"] = current_premium_value
    exposure_inputs["Current Premium"] = current_premium_value
    exposure_rows.append({
      "field": "current_premium",
      "label": "Current Premium",
      "value": current_premium_value,
      "source": "flat_csv_preamble_table",
    })

  exposure_basis_parts = []
  if parsed_profile.get("revenue"):
    exposure_basis_parts.append(f"Revenue: {parsed_profile.get('revenue')}")
  if parsed_profile.get("payroll"):
    exposure_basis_parts.append(f"Payroll: {parsed_profile.get('payroll')}")
  if parsed_profile.get("employee_count"):
    exposure_basis_parts.append(f"Employees: {parsed_profile.get('employee_count')}")
  if parsed_profile.get("location_count"):
    exposure_basis_parts.append(f"Locations: {parsed_profile.get('location_count')}")
  if parsed_profile.get("liquor_sales"):
    exposure_basis_parts.append(f"Liquor Sales: {parsed_profile.get('liquor_sales')}")

  if exposure_basis_parts:
    parsed_profile["exposure_basis"] = " | ".join(exposure_basis_parts)
    exposure_inputs["Exposure Basis"] = parsed_profile["exposure_basis"]

  if exposure_inputs:
    parsed_profile["exposure_inputs"] = exposure_inputs

  if exposure_rows:
    parsed_profile["exposures"] = exposure_rows

  policies_by_key = {}
  claims = []

  for row in rows:
    policy_number = get(row, "Policy Number", "Policy No", "Policy #", "policy_number", "policy")
    if not policy_number:
      continue

    policy_key = policy_number.upper()
    policy_type = policy_line_from_number(
      policy_number,
      get(row, "Policy Type", "Line of Business", "Coverage", "Coverage Line", "LOB", "line_of_business", "claim_type"),
    )

    row_carrier = get(row, "Carrier", "Writing Carrier", "Insurance Carrier", "Carrier Name") or carrier_name
    row_effective = get(row, "Effective Date", "Policy Effective Date", "Policy Effective", "Eff Date") or effective_date
    row_expiration = get(row, "Expiration Date", "Policy Expiration Date", "Policy Expiration", "Exp Date") or expiration_date
    premium = get(row, "Premium", "Annual Premium", "Written Premium", "Current Premium", "Policy Premium")

    if policy_key not in policies_by_key:
      policies_by_key[policy_key] = {
        "policy_number": policy_number,
        "policy_type": policy_type,
        "line_of_business": policy_type,
        "coverage": policy_type,
        "carrier": row_carrier,
        "carrier_name": row_carrier,
        "writing_carrier": row_carrier,
        "effective_date": row_effective,
        "expiration_date": row_expiration,
        "premium": premium,
        "current_premium": money_display(premium),
        "claim_count": 0,
        "total_incurred": 0.0,
      }

    claim_number = get(row, "Claim Number", "Claim #", "Claim No", "claim_number", "claim_no", "claim")
    no_claim_policy_only = is_no_claim_row(row, claim_number)

    if no_claim_policy_only or not claim_number:
      continue

    paid = money(get(row, "Paid", "Paid Amount", "Total Paid", "Loss Paid"))
    reserve = money(get(row, "Reserve", "Reserve Amount", "Outstanding Reserve", "Case Reserve", "Total Reserve"))
    incurred = money(get(row, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred", "Total"))

    if incurred <= 0 and (paid > 0 or reserve > 0):
      incurred = paid + reserve

    status = status_value(get(row, "Claim Status", "Status", "Open Closed", "Open/Closed", "claim_status"), paid, reserve, incurred)

    claim = {
      "business_name": business_name,
      "named_insured": business_name,
      "insured_name": business_name,
      "carrier_name": row_carrier,
      "writing_carrier": row_carrier,
      "producing_agency": producing_agency,
      "policy_number": policy_number,
      "policy": policy_number,
      "policy_type": policy_type,
      "line_of_business": policy_type,
      "claim_type": policy_type,
      "coverage": policy_type,
      "effective_date": row_effective,
      "expiration_date": row_expiration,
      "evaluation_date": get(row, "Evaluation Date", "Valuation Date", "As Of Date") or evaluation_date,
      "claim_number": claim_number,
      "claim_no": claim_number,
      "date_of_loss": get(row, "Date of Loss", "Loss Date", "Accident Date"),
      "loss_date": get(row, "Date of Loss", "Loss Date", "Accident Date"),
      "date_reported": get(row, "Date Reported", "Reported Date", "Report Date"),
      "date_closed": get(row, "Date Closed", "Closed Date"),
      "status": status,
      "claim_status": status,
      "cause_of_loss": get(row, "Cause of Loss", "Loss Cause", "Cause"),
      "description": get(row, "Claim Notes", "Loss Notes", "Notes", "Narrative", "Claim Description", "Description", "Loss Description", "Cause of Loss"),
      "paid_amount": paid,
      "paid": paid,
      "reserve_amount": reserve,
      "reserve": reserve,
      "total_incurred": incurred,
      "incurred": incurred,
      "total_amount": incurred,
      "litigation": get(row, "Litigation", "Litigated", "Attorney Involvement", "Counsel", "Suit Filed", "Lawsuit", "Legal Status"),
      "litigation_status": get(row, "Litigation Status", "Legal Status", "Suit Status", "Lawsuit Status"),
      "attorney_assigned": get(row, "Attorney Assigned", "Attorney", "Attorney Name", "Attorney Involvement", "Counsel", "Claimant Counsel", "Plaintiff Attorney", "Defense Counsel", "Represented", "Claimant Represented"),
      "suit_filed": get(row, "Suit Filed", "Lawsuit Filed", "Complaint Filed"),
      "venue_state": get(row, "Venue State", "Venue", "Jurisdiction", "Jurisdiction/State", "Loss State", "State"),
      "flag": get(row, "Flag", "Flags", "Red Flag", "Red Flags", "Claim Flag", "Alert", "Concern"),
      "account_number": account_number,
      "customer_number": account_number,
    }

    claims.append(claim)

    policies_by_key[policy_key]["claim_count"] = int(policies_by_key[policy_key].get("claim_count") or 0) + 1
    policies_by_key[policy_key]["total_incurred"] = float(policies_by_key[policy_key].get("total_incurred") or 0) + incurred

  policies = list(policies_by_key.values())

  if policies:
    parsed_profile["policies"] = policies
    parsed_profile["policy_schedule"] = policies
    parsed_profile["policy_numbers"] = [item.get("policy_number") for item in policies if item.get("policy_number")]
    parsed_profile["policy_number"] = parsed_profile.get("policy_number") or policies[0].get("policy_number")
    parsed_profile["main_policy"] = parsed_profile.get("main_policy") or policies[0].get("policy_number")

  if claims:
    parsed_claims = claims
    parsed_profile["claims"] = claims
    parsed_profile["parsed_claims"] = claims

  print("LOSSQ_FLAT_CSV_PREAMBLE_TABLE_HEADER_V2:", {
    "business_name": parsed_profile.get("business_name"),
    "claims": len(claims),
    "policies": len(policies),
    "policy_numbers": [p.get("policy_number") for p in policies],
    "statuses": [c.get("status") for c in claims[:10]],
  })

  return parsed_claims, parsed_profile



# LOSSQ_AUTHORITATIVE_FLAT_CSV_FINAL_REPAIR_V1
def lossq_authoritative_flat_csv_snapshot_v1(file_path):
  """
  Final authoritative reader for clean flat CSV loss runs.

  This does not hardcode carrier/account/sample names. It uses actual CSV headers:
  Account Name / Business Name, Policy Number, Policy Type, Claim Number,
  Claim Status, paid/reserve/incurred fields, and zero-claim policy rows.
  """
  import csv
  import re

  if not str(file_path or "").lower().endswith(".csv"):
    return {}

  rows = []
  for encoding in ("utf-8-sig", "utf-8", "latin-1"):
    try:
      with open(file_path, "r", encoding=encoding, errors="ignore", newline="") as handle:
        rows = list(csv.DictReader(handle))
      break
    except Exception:
      rows = []

  if not rows:
    return {}

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def get(row, *aliases):
    wanted = {norm(a) for a in aliases if clean(a)}
    for k, v in (row or {}).items():
      if norm(k) in wanted:
        value = clean(v)
        if value:
          return value
    return ""

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    raw = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      return float(raw or 0)
    except Exception:
      return 0.0

  def good_profile_value(value):
    value = clean(value)
    bad = {
      "",
      "-",
      "n/a",
      "na",
      "none",
      "unknown",
      "account number",
      "account no",
      "account",
      "policy number",
      "policy no",
      "policy",
      "claim number",
      "claim no",
      "business name",
      "named insured",
      "insured",
      "carrier",
      "writing carrier",
    }
    return value if value.lower() not in bad else ""

  def first_value(*aliases):
    for row in rows:
      value = good_profile_value(get(row, *aliases))
      if value:
        return value
    return ""

  def line_from_policy(policy_number, fallback=""):
    policy = clean(policy_number).upper()
    fallback = clean(fallback)
    if fallback and fallback.lower() not in {"unknown", "n/a", "none", "-"}:
      return fallback

    if policy.startswith(("GL", "CGL")):
      return "General Liability"
    if policy.startswith(("LIQ", "LQ")):
      return "Liquor Liability"
    if policy.startswith(("WC", "WORK")):
      return "Workers Compensation"
    if policy.startswith(("BOP", "PROP", "CP")):
      return "Businessowners Policy"
    if policy.startswith(("UMB", "UM", "EXC", "XS")):
      return "Umbrella"
    if policy.startswith(("AUTO", "CA", "AL")):
      return "Commercial Auto"
    if policy.startswith(("CARGO", "MTC")):
      return "Cargo"
    if policy.startswith(("CY", "CYB")):
      return "Cyber Liability"

    return fallback or "Commercial Policy"

  def status_from_row(row, paid, reserve, incurred):
    raw = clean(get(row, "Claim Status", "Status", "Open Closed", "Open/Closed", "claim_status"))
    low = raw.lower()

    if low in {"open", "opened", "active", "pending", "reopened", "reopen", "in progress"}:
      return "Open"

    if low in {"closed", "close", "final", "settled", "resolved"}:
      return "Closed"

    if low in {"no direct claim activity", "no claim activity", "no claims", "none", "n/a", "na", "-"}:
      return ""

    if reserve > 0:
      return "Open"

    if incurred > 0 or paid > 0:
      return "Closed"

    return raw.title() if raw else ""

  def no_claim_policy_row(row, claim_number):
    if clean(claim_number):
      return False
    joined = " ".join(clean(v).lower() for v in (row or {}).values())
    return bool(re.search(r"\b(no\s+direct\s+claim\s+activity|no\s+claim\s+activity|no\s+claims?|none\s+reported|zero\s+claims?)\b", joined))

  header_keys = {norm(k) for k in (rows[0] or {}).keys()}
  if not ({"policynumber", "policyno", "policy"} & header_keys):
    return {}

  business_name = first_value("Account Name", "Business Name", "Named Insured", "Insured", "Insured Name", "Company Name")
  carrier_name = first_value("Carrier", "Writing Carrier", "Insurance Carrier", "Carrier Name")
  producing_agency = first_value("Producing Agency", "Agency", "Agency Name", "Producer", "Producer Name", "Broker", "Broker Name")
  account_number = first_value("Account Number", "Account No", "Account #", "Customer Number", "Client Number")
  effective_date = first_value("Effective Date", "Policy Effective Date", "Policy Effective", "Eff Date")
  expiration_date = first_value("Expiration Date", "Policy Expiration Date", "Policy Expiration", "Exp Date")
  evaluation_date = first_value("Evaluation Date", "Valuation Date", "As Of Date", "Loss Run Date")

  profile = {}

  if business_name:
    profile["business_name"] = business_name
    profile["insured_name"] = business_name
    profile["named_insured"] = business_name
    profile["account_name"] = business_name

  if carrier_name:
    profile["carrier_name"] = carrier_name
    profile["writing_carrier"] = carrier_name
    profile["carrier"] = carrier_name

  if producing_agency:
    profile["producing_agency"] = producing_agency
    profile["agency_name"] = producing_agency
    profile["producer"] = producing_agency

  if account_number:
    profile["account_number"] = account_number
    profile["customer_number"] = account_number

  if effective_date:
    profile["effective_date"] = effective_date

  if expiration_date:
    profile["expiration_date"] = expiration_date

  if evaluation_date:
    profile["evaluation_date"] = evaluation_date
    profile["valuation_date"] = evaluation_date

  policies_by_key = {}
  claims = []

  for row in rows:
    policy_number = get(row, "Policy Number", "Policy No", "Policy #", "policy_number", "policy")
    if not policy_number:
      continue

    policy_key = clean(policy_number).upper()
    policy_type = line_from_policy(
      policy_number,
      get(row, "Policy Type", "Line of Business", "Coverage", "Coverage Line", "LOB", "line_of_business", "claim_type"),
    )
    row_carrier = get(row, "Carrier", "Writing Carrier", "Insurance Carrier", "Carrier Name") or carrier_name
    row_effective = get(row, "Effective Date", "Policy Effective Date", "Policy Effective", "Eff Date") or effective_date
    row_expiration = get(row, "Expiration Date", "Policy Expiration Date", "Policy Expiration", "Exp Date") or expiration_date

    if policy_key not in policies_by_key:
      policies_by_key[policy_key] = {
        "policy_number": policy_number,
        "policy_type": policy_type,
        "line_of_business": policy_type,
        "coverage": policy_type,
        "carrier": row_carrier,
        "carrier_name": row_carrier,
        "writing_carrier": row_carrier,
        "effective_date": row_effective,
        "expiration_date": row_expiration,
        "claim_count": 0,
        "total_incurred": 0.0,
      }

    claim_number = get(row, "Claim Number", "Claim #", "Claim No", "claim_number", "claim_no", "claim")
    if no_claim_policy_row(row, claim_number) or not claim_number:
      continue

    paid = money(get(row, "Paid", "Paid Amount", "Total Paid", "Loss Paid"))
    reserve = money(get(row, "Reserve", "Reserve Amount", "Outstanding Reserve", "Case Reserve", "Total Reserve"))
    incurred = money(get(row, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred", "Total"))
    if incurred <= 0 and (paid > 0 or reserve > 0):
      incurred = paid + reserve

    status = status_from_row(row, paid, reserve, incurred)

    claim = {
      "business_name": business_name,
      "named_insured": business_name,
      "insured_name": business_name,
      "carrier_name": row_carrier,
      "writing_carrier": row_carrier,
      "producing_agency": producing_agency,
      "policy_number": policy_number,
      "policy": policy_number,
      "policy_type": policy_type,
      "line_of_business": policy_type,
      "claim_type": policy_type,
      "coverage": policy_type,
      "effective_date": row_effective,
      "expiration_date": row_expiration,
      "evaluation_date": get(row, "Evaluation Date", "Valuation Date", "As Of Date") or evaluation_date,
      "claim_number": claim_number,
      "claim_no": claim_number,
      "date_of_loss": get(row, "Date of Loss", "Loss Date", "Accident Date"),
      "loss_date": get(row, "Date of Loss", "Loss Date", "Accident Date"),
      "date_reported": get(row, "Date Reported", "Reported Date", "Report Date"),
      "date_closed": get(row, "Date Closed", "Closed Date"),
      "status": status,
      "claim_status": status,
      "cause_of_loss": get(row, "Cause of Loss", "Loss Cause", "Cause"),
      "description": get(row, "Claim Notes", "Loss Notes", "Notes", "Narrative", "Claim Description", "Description", "Loss Description", "Cause of Loss"),
      "paid_amount": paid,
      "paid": paid,
      "reserve_amount": reserve,
      "reserve": reserve,
      "total_incurred": incurred,
      "incurred": incurred,
      "total_amount": incurred,
      "litigation": get(row, "Litigation", "Litigated", "Attorney Involvement", "Counsel", "Suit Filed", "Lawsuit", "Legal Status"),
      "litigation_status": get(row, "Litigation Status", "Legal Status", "Suit Status", "Lawsuit Status"),
      "attorney_assigned": get(row, "Attorney Assigned", "Attorney", "Attorney Name", "Attorney Involvement", "Counsel", "Claimant Counsel", "Plaintiff Attorney", "Defense Counsel", "Represented", "Claimant Represented"),
      "suit_filed": get(row, "Suit Filed", "Lawsuit Filed", "Complaint Filed"),
      "venue_state": get(row, "Venue State", "Venue", "Jurisdiction", "Jurisdiction/State", "Loss State", "State"),
      "flag": get(row, "Flag", "Flags", "Red Flag", "Red Flags", "Claim Flag", "Alert", "Concern"),
      "account_number": account_number,
      "customer_number": account_number,
    }

    claims.append(claim)
    policies_by_key[policy_key]["claim_count"] = int(policies_by_key[policy_key].get("claim_count") or 0) + 1
    policies_by_key[policy_key]["total_incurred"] = float(policies_by_key[policy_key].get("total_incurred") or 0) + incurred

  policies = list(policies_by_key.values())
  policy_numbers = [p.get("policy_number") for p in policies if p.get("policy_number")]

  if policies:
    profile["policies"] = policies
    profile["policy_schedule"] = policies
    profile["policy_numbers"] = policy_numbers
    profile["policy_number"] = policy_numbers[0] if policy_numbers else profile.get("policy_number")
    profile["main_policy"] = policy_numbers[0] if policy_numbers else profile.get("main_policy")

  return {
    "profile": profile,
    "claims": claims,
    "policies": policies,
    "policy_numbers": policy_numbers,
  }


def lossq_apply_authoritative_flat_csv_snapshot_v1(file_path, parsed_claims=None, parsed_profile=None, profile_data=None):
  # LOSSQ_AUTHORITATIVE_FLAT_CSV_FINAL_REPAIR_V1
  snapshot = lossq_authoritative_flat_csv_snapshot_v1(file_path)
  if not snapshot:
    return parsed_claims, parsed_profile, profile_data

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  profile_data = profile_data if isinstance(profile_data, dict) else profile_data

  profile = snapshot.get("profile") if isinstance(snapshot.get("profile"), dict) else {}
  snapshot_claims = snapshot.get("claims") if isinstance(snapshot.get("claims"), list) else []

  if isinstance(parsed_profile, dict):
    for key, value in profile.items():
      if value not in ("", None, [], {}):
        parsed_profile[key] = value

  if isinstance(profile_data, dict):
    for key, value in profile.items():
      if value not in ("", None, [], {}):
        profile_data[key] = value

  if snapshot_claims:
    existing_by_claim = {}
    for claim in parsed_claims:
      if isinstance(claim, dict) and claim.get("claim_number"):
        existing_by_claim[str(claim.get("claim_number")).strip().upper()] = claim

    merged = []
    seen = set()

    for csv_claim in snapshot_claims:
      key = str(csv_claim.get("claim_number") or "").strip().upper()
      if not key:
        continue

      base = existing_by_claim.get(key, {})
      if isinstance(base, dict):
        next_claim = dict(base)
        for field, value in csv_claim.items():
          if value not in ("", None, [], {}):
            next_claim[field] = value
      else:
        next_claim = dict(csv_claim)

      merged.append(next_claim)
      seen.add(key)

    for claim in parsed_claims:
      if isinstance(claim, dict):
        key = str(claim.get("claim_number") or "").strip().upper()
        if key and key not in seen:
          merged.append(claim)

    parsed_claims = merged
    parsed_profile["claims"] = parsed_claims
    parsed_profile["parsed_claims"] = parsed_claims

  print("LOSSQ_AUTHORITATIVE_FLAT_CSV_FINAL_REPAIR_V1:", {
    "business_name": parsed_profile.get("business_name") if isinstance(parsed_profile, dict) else None,
    "claims": len(parsed_claims or []),
    "policies": len(profile.get("policies") or []),
    "policy_numbers": profile.get("policy_numbers"),
    "statuses": [c.get("status") for c in (parsed_claims or [])[:10] if isinstance(c, dict)],
  })

  return parsed_claims, parsed_profile, profile_data



def lossq_apply_row_values_at_final_save(normalized: dict, raw_claim: dict):
  """
  Final safety layer before Claim(**normalized).
  Row-level claim values must win over account/main-policy values.
  This prevents all claims from being saved under the first/main policy.
  """
  if not isinstance(normalized, dict):
    normalized = {}

  if not isinstance(raw_claim, dict):
    return normalized

  def clean(value):
    return clean_profile_value(value)

  def get_any(*names):
    lower_map = {str(k or "").strip().lower(): v for k, v in raw_claim.items()}
    for name in names:
      key = str(name or "").strip().lower()
      if key in lower_map:
        value = clean(lower_map.get(key))
        if value:
          return value
    return ""

  row_policy_number = get_any(
    "policy_number",
    "policy number",
    "policy no",
    "policy_no",
    "policy",
    "main policy",
    "account number",
  )

  row_policy_type = get_any(
    "policy_type",
    "policy type",
    "line_of_business",
    "line of business",
    "coverage",
    "coverage line",
    "claim_type",
    "claim type",
    "line",
    "lob",
  )

  row_status = get_any(
    "status",
    "claim status",
    "claim_status",
  )

  row_claim_number = get_any(
    "claim_number",
    "claim number",
    "claim #",
    "claim no",
    "claim_no",
  )

  row_paid = get_any("paid_amount", "paid", "paid amount", "total paid")
  row_reserve = get_any("reserve_amount", "reserve", "reserve amount", "outstanding reserve")
  row_total = get_any("total_incurred", "total incurred", "incurred", "total")

  if row_policy_number and not is_bad_policy_key_for_upload(row_policy_number):
    normalized["policy_number"] = row_policy_number

  if row_policy_type:
    normalized["line_of_business"] = row_policy_type
    normalized["claim_type"] = row_policy_type

  if row_status:
    normalized["status"] = row_status

  if row_claim_number:
    normalized["claim_number"] = row_claim_number

  if row_paid:
    normalized["paid_amount"] = row_paid

  if row_reserve:
    normalized["reserve_amount"] = row_reserve

  if row_total:
    normalized["total_incurred"] = row_total

  return normalized


router = APIRouter(prefix="/upload", tags=["Upload"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# LOSSQ_BETA_UPLOAD_LIMIT_ENFORCEMENT_V1
def lossq_beta_upload_usage_guard(db: Session, current_user: dict, incoming_file_count: int = 1):
  org_id = None
  if isinstance(current_user, dict):
    org_id = current_user.get("organization_id")

  if not org_id:
    return

  try:
    org_row = db.execute(
      text(
        """
        SELECT plan, subscription_status, upload_limit
        FROM organizations
        WHERE id = :org_id
        """
      ),
      {"org_id": org_id},
    ).mappings().first()
  except Exception:
    return

  if not org_row:
    return

  plan = str(org_row.get("plan") or "").strip().lower()
  subscription_status = str(org_row.get("subscription_status") or "").strip().lower()

  is_beta = plan in {"beta", "beta_access", "early_access"} or subscription_status.startswith("beta")

  if not is_beta:
    return

  try:
    upload_limit = int(org_row.get("upload_limit") or 10)
  except Exception:
    upload_limit = 10

  if upload_limit <= 0:
    upload_limit = 10

  try:
    uploads_used = (
      db.query(UploadHistory)
     .filter(UploadHistory.organization_id == org_id)
     .count()
    )
  except Exception:
    uploads_used = 0

  incoming_file_count = max(int(incoming_file_count or 1), 1)
  projected_uploads = uploads_used + incoming_file_count

  if projected_uploads > upload_limit:
    remaining = max(upload_limit - uploads_used, 0)
    raise HTTPException(
      status_code=403,
      detail=(
        f"Beta upload limit reached. This beta account has used "
        f"{uploads_used} of {upload_limit} uploads. "
        f"{remaining} upload(s) remaining."
      ),
    )

def get_db():
  db = SessionLocal()
  try:
    yield db
  finally:
    db.close()


# LOSSQ_EXTRACT_EXPOSURE_FROM_PARSED_ROWS_V1
def extract_exposure_inputs_from_parsed_rows(rows):
  """Extract exposure/premium fields from parsed CSV/XLSX/PDF row dictionaries."""
  import re

  profile = {}

  def clean(value):
    return str(value or "").replace("\ufeff", "").replace("", "").strip()

  def norm_key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())

  def is_bad_value(value):
    v = clean(value)
    if not v:
      return True
    if re.fullmatch(r"(19|20)\d{2}", v):
      return True
    if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", v):
      return True
    if re.fullmatch(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", v):
      return True
    return False

  def money_value(value):
    v = clean(value)
    if is_bad_value(v):
      return ""
    match = re.search(r"\$?\s*[0-9][0-9,]*(?:\.\d{2})?", v)
    if not match:
      return ""
    found = match.group(0).replace(" ", "")
    numeric = found.replace("$", "").replace(",", "")
    if is_bad_value(numeric):
      return ""
    return found

  def count_value(value):
    v = clean(value)
    if is_bad_value(v):
      return ""
    match = re.search(r"\b[0-9][0-9,]*\b", v)
    if not match:
      return ""
    found = match.group(0).replace(",", "")
    if is_bad_value(found):
      return ""
    return found

  field_map = {
    "currentpremium": "current_premium",
    "annualpremium": "current_premium",
    "writtenpremium": "current_premium",
    "totalpremium": "current_premium",
    "premium": "current_premium",

    "expiringpremium": "expiring_premium",
    "priorpremium": "expiring_premium",
    "previouspremium": "expiring_premium",

    "targetrenewalpremium": "target_renewal_premium",
    "renewalpremium": "target_renewal_premium",
    "estimatedrenewalpremium": "target_renewal_premium",

    "policylimits": "limits",
    "limits": "limits",
    "coveragelimit": "coverage_limit",
    "deductible": "deductible",
    "retention": "retention",
    "sir": "retention",

    "payroll": "payroll",
    "annualpayroll": "payroll",
    "estimatedpayroll": "payroll",

    "revenue": "revenue",
    "annualrevenue": "revenue",
    "revenuesales": "revenue",
    "sales": "sales",
    "grosssales": "sales",
    "receipts": "receipts",
    "grossreceipts": "receipts",

    "employeecount": "employee_count",
    "employees": "employee_count",
    "numberofemployees": "employee_count",

    "vehiclecount": "vehicle_count",
    "vehicles": "vehicle_count",
    "powerunits": "vehicle_count",

    "drivercount": "driver_count",
    "drivers": "driver_count",

    "propertytiv": "property_tiv",
    "totalinsuredvalue": "property_tiv",
    "tiv": "tiv",

    "buildingvalue": "building_value",
    "buildinglimit": "building_value",
    "contentsvalue": "contents_value",
    "businesspersonalproperty": "contents_value",
    "bpp": "contents_value",

    "squarefootage": "square_footage",
    "sqft": "square_footage",
    "locationcount": "location_count",
    "locations": "location_count",
    "unitcount": "unit_count",
    "units": "unit_count",

    "cargolimit": "cargo_limit",
    "umbrellalimit": "umbrella_limit",
    "excesslimit": "umbrella_limit",

    "experiencemod": "experience_mod",
    "mod": "mod",
    "exposurechangepercent": "exposure_change_percent",
    "cyberrevenue": "cyber_revenue",
    "professionalrevenue": "professional_revenue",
    "exposurebasis": "exposure_basis",
  }

  money_fields = {
    "current_premium",
    "expiring_premium",
    "target_renewal_premium",
    "limits",
    "coverage_limit",
    "deductible",
    "retention",
    "payroll",
    "revenue",
    "sales",
    "receipts",
    "property_tiv",
    "tiv",
    "building_value",
    "contents_value",
    "cargo_limit",
    "umbrella_limit",
    "cyber_revenue",
    "professional_revenue",
  }

  count_fields = {
    "employee_count",
    "vehicle_count",
    "driver_count",
    "square_footage",
    "location_count",
    "unit_count",
  }

  def set_field(field, value):
    if not field:
      return

    if field in money_fields:
      value = money_value(value)
    elif field in count_fields:
      value = count_value(value)
    else:
      value = clean(value)

    if value and not profile.get(field):
      profile[field] = value

  if not isinstance(rows, list):
    return {}

  for row in rows:
    if not isinstance(row, dict):
      continue

    for key, value in row.items():
      mapped = field_map.get(norm_key(key))
      if mapped:
        set_field(mapped, value)

    # Some parsers store label/value pairs instead of normal columns.
    label = (
      row.get("label")
      or row.get("field")
      or row.get("metric")
      or row.get("name")
      or row.get("exposure_label")
      or row.get("exposure_type")
    )
    value = (
      row.get("value")
      or row.get("amount")
      or row.get("exposure_value")
      or row.get("exposure")
      or row.get("current_value")
    )

    if label and value:
      mapped = field_map.get(norm_key(label))
      if mapped:
        set_field(mapped, value)

    # One fully populated row is enough because exposure columns repeat on every CSV claim row.
    if len(profile.keys()) >= 5:
      break

  basis_parts = []
  if profile.get("payroll"):
    basis_parts.append(f"Payroll: {profile['payroll']}")
  if profile.get("revenue"):
    basis_parts.append(f"Revenue: {profile['revenue']}")
  if profile.get("vehicle_count"):
    basis_parts.append(f"Vehicles: {profile['vehicle_count']}")
  if profile.get("driver_count"):
    basis_parts.append(f"Drivers: {profile['driver_count']}")
  if profile.get("employee_count"):
    basis_parts.append(f"Employees: {profile['employee_count']}")
  if profile.get("property_tiv"):
    basis_parts.append(f"Property TIV: {profile['property_tiv']}")

  if basis_parts and not profile.get("exposure_basis"):
    profile["exposure_basis"] = " | ".join(basis_parts)

  return profile


def extract_exposure_inputs_from_raw_text(raw_text: str):
  # LOSSQ_ENABLE_AUTO_EXPOSURE_EXTRACTION_V3
  # Universal exposure extractor for labeled CSV, XLSX text, PDF text, premium worksheets, and policy schedules.
  import csv
  import io
  import re

  text_value = str(raw_text or "")
  profile = {}

  def clean(value):
    return str(value or "").replace("\ufeff", "").replace("", "").strip()

  def norm_key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())

  def is_bad_value(value):
    v = clean(value)
    if not v:
      return True
    # Do not treat policy years or dates as exposure values.
    if re.fullmatch(r"(19|20)\d{2}", v):
      return True
    if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", v):
      return True
    if re.fullmatch(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", v):
      return True
    return False

  def money_value(value):
    v = clean(value)
    if is_bad_value(v):
      return ""
    match = re.search(r"\$?\s*[0-9][0-9,]*(?:\.\d{2})?", v)
    if not match:
      return ""
    found = match.group(0).replace(" ", "")
    if is_bad_value(found.replace("$", "").replace(",", "")):
      return ""
    return found

  def count_value(value):
    v = clean(value)
    if is_bad_value(v):
      return ""
    match = re.search(r"\b[0-9][0-9,]*\b", v)
    if not match:
      return ""
    found = match.group(0).replace(",", "")
    if is_bad_value(found):
      return ""
    return found

  field_map = {
    "currentpremium": "current_premium",
    "annualpremium": "current_premium",
    "writtenpremium": "current_premium",
    "totalpremium": "current_premium",
    "premium": "current_premium",

    "expiringpremium": "expiring_premium",
    "priorpremium": "expiring_premium",
    "previouspremium": "expiring_premium",

    "targetrenewalpremium": "target_renewal_premium",
    "renewalpremium": "target_renewal_premium",
    "estimatedrenewalpremium": "target_renewal_premium",

    "primarylineofbusiness": "line_of_business",
    "lineofbusiness": "line_of_business",
    "lob": "line_of_business",
    "policytype": "line_of_business",
    "coverage": "line_of_business",

    "state": "state",
    "primarystate": "state",
    "classcode": "class_code",
    "classcodes": "class_codes",

    "policylimits": "limits",
    "limits": "limits",
    "coveragelimit": "coverage_limit",
    "deductible": "deductible",
    "retention": "retention",
    "sir": "retention",

    "payroll": "payroll",
    "annualpayroll": "payroll",
    "estimatedpayroll": "payroll",

    "revenue": "revenue",
    "annualrevenue": "revenue",
    "sales": "sales",
    "grosssales": "sales",
    "revenuesales": "revenue",
    "receipts": "receipts",
    "grossreceipts": "receipts",

    "employeecount": "employee_count",
    "employees": "employee_count",
    "numberofemployees": "employee_count",

    "vehiclecount": "vehicle_count",
    "vehicles": "vehicle_count",
    "powerunits": "vehicle_count",

    "drivercount": "driver_count",
    "drivers": "driver_count",

    "propertytiv": "property_tiv",
    "totalinsuredvalue": "property_tiv",
    "tiv": "tiv",

    "buildingvalue": "building_value",
    "buildinglimit": "building_value",
    "contentsvalue": "contents_value",
    "businesspersonalproperty": "contents_value",
    "bpp": "contents_value",

    "squarefootage": "square_footage",
    "sqft": "square_footage",
    "locationcount": "location_count",
    "locations": "location_count",
    "unitcount": "unit_count",
    "units": "unit_count",

    "cargolimit": "cargo_limit",
    "umbrellalimit": "umbrella_limit",
    "excesslimit": "umbrella_limit",

    "experiencemod": "experience_mod",
    "mod": "mod",
    "exposurechangepercent": "exposure_change_percent",
    "cyberrevenue": "cyber_revenue",
    "professionalrevenue": "professional_revenue",
    "exposurebasis": "exposure_basis",
  }

  money_fields = {
    "current_premium",
    "expiring_premium",
    "target_renewal_premium",
    "limits",
    "coverage_limit",
    "deductible",
    "retention",
    "payroll",
    "revenue",
    "sales",
    "receipts",
    "property_tiv",
    "tiv",
    "building_value",
    "contents_value",
    "cargo_limit",
    "umbrella_limit",
    "cyber_revenue",
    "professional_revenue",
  }

  count_fields = {
    "employee_count",
    "vehicle_count",
    "driver_count",
    "square_footage",
    "location_count",
    "unit_count",
  }

  def set_field(field, value):
    if not field or field not in field_map.values():
      return

    if field in money_fields:
      value = money_value(value)
    elif field in count_fields:
      value = count_value(value)
    else:
      value = clean(value)

    if value and not profile.get(field):
      profile[field] = value

  def apply_pair(key, value):
    mapped = field_map.get(norm_key(key))
    if mapped:
      set_field(mapped, value)

  # CSV-style extraction: headers on first line, values on following rows.
  try:
    sample = text_value.strip()
    if "," in sample and "\n" in sample:
      reader = csv.DictReader(io.StringIO(sample))
      for row in reader:
        for key, value in dict(row or {}).items():
          apply_pair(key, value)
        # One good row is enough because exposure columns repeat per claim row.
        if profile:
          break
  except Exception:
    pass

  # Label/value extraction from text lines and worksheet-style rows.
  for line in text_value.splitlines():
    if not line.strip():
      continue

    if ":" in line:
      left, right = line.split(":", 1)
      apply_pair(left, right)

    if "," in line:
      parts = [p.strip() for p in line.split(",")]
      if len(parts) >= 2:
        for i in range(len(parts) - 1):
          apply_pair(parts[i], parts[i + 1])

  # Regex fallback for labels embedded in text.
  label_aliases = {
    "current_premium": ["current premium", "annual premium", "written premium", "total premium"],
    "expiring_premium": ["expiring premium", "prior premium", "previous premium"],
    "target_renewal_premium": ["target renewal premium", "renewal premium", "estimated renewal premium"],
    "payroll": ["annual payroll", "estimated payroll", "payroll"],
    "revenue": ["annual revenue", "revenue"],
    "sales": ["gross sales", "sales"],
    "receipts": ["gross receipts", "receipts"],
    "employee_count": ["employee count", "number of employees", "employees"],
    "vehicle_count": ["vehicle count", "vehicles", "power units"],
    "driver_count": ["driver count", "drivers"],
    "property_tiv": ["property tiv", "total insured value"],
    "tiv": ["tiv"],
    "coverage_limit": ["coverage limit", "policy limit"],
    "limits": ["policy limits", "limits"],
    "deductible": ["deductible"],
    "umbrella_limit": ["umbrella limit", "excess limit"],
    "cyber_revenue": ["cyber revenue"],
    "professional_revenue": ["professional revenue"],
    "experience_mod": ["experience mod", "mod"],
  }

  for field, labels in label_aliases.items():
    if profile.get(field):
      continue
    for label in labels:
      pattern = re.compile(
        re.escape(label) + r"[^$0-9A-Za-z]{0,50}(\$?\s*[0-9][0-9,]*(?:\.\d{2})?|[A-Za-z][A-Za-z0-9./%-]{1,80})",
        re.IGNORECASE,
      )
      match = pattern.search(text_value)
      if match:
        set_field(field, match.group(1))
        if profile.get(field):
          break

  basis_parts = []
  if profile.get("payroll"):
    basis_parts.append(f"Payroll: {profile['payroll']}")
  if profile.get("revenue"):
    basis_parts.append(f"Revenue: {profile['revenue']}")
  if profile.get("vehicle_count"):
    basis_parts.append(f"Vehicles: {profile['vehicle_count']}")
  if profile.get("driver_count"):
    basis_parts.append(f"Drivers: {profile['driver_count']}")
  if profile.get("employee_count"):
    basis_parts.append(f"Employees: {profile['employee_count']}")
  if profile.get("property_tiv"):
    basis_parts.append(f"Property TIV: {profile['property_tiv']}")

  if basis_parts and not profile.get("exposure_basis"):
    profile["exposure_basis"] = " | ".join(basis_parts)

  return profile


def _lossq_live_clean_cell(value):
  return re.sub(r"\s+", " ", str(value or "").strip())

def _lossq_live_money_to_float(value):
  raw = _lossq_live_clean_cell(value)
  if not raw:
    return 0.0
  raw = raw.replace("$", "").replace(",", "").replace("%", "").strip()
  try:
    return float(raw)
  except Exception:
    return 0.0

def _lossq_live_date_to_iso(value):
  raw = _lossq_live_clean_cell(value)
  if not raw:
    return ""

  raw = raw.replace("\\", "/").replace(".", "/").replace("-", "/")
  raw = re.sub(r"\s+", "", raw)

  m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", raw)
  if m:
    month, day, year = m.groups()
    year = int(year)
    if year < 100:
      year += 2000 if year < 50 else 1900
    try:
      return f"{year:04d}-{int(month):02d}-{int(day):02d}"
    except Exception:
      return ""

  m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", raw)
  if m:
    year, month, day = m.groups()
    try:
      return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    except Exception:
      return ""

  return raw

def _lossq_live_is_policy_number(value):
  raw = _lossq_live_clean_cell(value).upper()
  if not raw:
    return False

  blocked = {
    "POLICY NUMBER", "POLICY", "ACCOUNT INFORMATION", "POLICY SCHEDULE",
    "CLAIM DETAIL", "LOSS SUMMARY", "UNDERWRITING NOTES", "N/A", "NONE", "UNKNOWN",
  }
  if raw in blocked:
    return False

  # LOSSQ_LIVE_UNIVERSAL_POLICY_ID_V1
  if re.search(r"[A-Z0-9]{2,}[-_][A-Z0-9]{2,}[-_](19|20)\d{2}[-_][A-Z0-9]{2,}", raw):
    return True

  if re.search(r"[A-Z]{2,10}[-_](19|20)\d{2}[-_][A-Z0-9]{2,}", raw):
    return True

  return False

def _lossq_live_is_claim_number(value):
  raw = _lossq_live_clean_cell(value).upper()
  if not raw:
    return False

  blocked = {
    "NOTE", "NOTES", "LOSS SUMMARY", "METRIC", "TOTAL CLAIMS", "OPEN CLAIMS",
    "CLOSED CLAIMS", "TOTAL PAID", "TOTAL RESERVE", "TOTAL INCURRED",
    "LARGEST LOSS", "LITIGATED CLAIMS", "CLAIMS WITH ATTORNEY INVOLVEMENT",
    "UNDERWRITING NOTES", "CLAIM NUMBER", "POLICY NUMBER", "DESCRIPTION",
  }
  if raw in blocked:
    return False

  if not re.search(r"\d", raw):
    return False

  # LOSSQ_LIVE_UNIVERSAL_CLAIM_ID_V1
  if re.search(r"[A-Z0-9]{2,}[-_][A-Z0-9]{2,}[-_]\d{4,8}", raw):
    return True

  if re.search(r"[A-Z0-9]{2,}[-_][A-Z0-9]{2,}[-_](19|20)\d{2}[-_][A-Z0-9]{2,}", raw):
    return True

  compact = re.sub(r"[^A-Z0-9]", "", raw)
  if len(compact) >= 6 and re.search(r"[A-Z]", compact) and re.search(r"\d", compact):
    return True

  return False

def _lossq_live_read_section_csv_rows(file_path):
  rows = []
  for encoding in ("utf-8-sig", "utf-8", "latin-1"):
    try:
      with open(file_path, "r", newline="", encoding=encoding) as f:
        rows = [row for row in csv.reader(f)]
      break
    except Exception:
      rows = []

  cleaned = []
  for row in rows:
    cleaned.append([_lossq_live_clean_cell(cell) for cell in row])

  return cleaned

def _lossq_live_extract_section_based_csv(file_path):
  rows = _lossq_live_read_section_csv_rows(file_path)
  print("LOSSQ_SECTION_CSV_ENTERED:", {"file_path": str(file_path), "rows": len(rows)})
  for idx, raw_debug_row in enumerate(rows[:25]):
    print("LOSSQ_SECTION_CSV_RAW_ROW:", {"idx": idx, "row": raw_debug_row})

  if not rows:
    return [], {}

  section_names = {
    "account information": "account",
    "policy schedule": "policies",
    "exposure inputs": "exposures",
    "exposure information": "exposures",
    "exposure / policy information": "policies",
    "exposure and policy information": "policies",
    "premium worksheet": "policies",
    "policy information": "policies",
    "claim detail": "claims",
    "loss summary": "summary",
    "underwriting notes": "notes",
  }

  current_section = ""
  account = {}
  exposures = {}
  loss_summary = {}
  policies = []
  claims = []

  policy_header_seen = False
  claim_header_seen = False
  exposure_header_seen = False
  summary_header_seen = False

  for row in rows:
    nonempty = [cell for cell in row if _lossq_live_clean_cell(cell)]
    if not nonempty:
      continue

    first = _lossq_live_clean_cell(nonempty[0])
    first_lower = first.lower()

    if first_lower in section_names:
      current_section = section_names[first_lower]
      policy_header_seen = False
      claim_header_seen = False
      exposure_header_seen = False
      summary_header_seen = False
      continue

    if current_section == "account":
      if len(nonempty) >= 2:
        key = _lossq_live_clean_cell(nonempty[0]).lower()
        value = _lossq_live_clean_cell(nonempty[1])

        if key in {"carrier"}:
          account["carrier_name"] = value
          account["carrier"] = value
        elif key in {"valuation date", "evaluation date"}:
          account["evaluation_date"] = _lossq_live_date_to_iso(value)
        elif key in {"named insured", "insured", "business name"}:
          account["business_name"] = value
          account["insured_name"] = value
          account["named_insured"] = value
        elif key in {"account"}:
          # LOSSQ_ACCOUNT_LABEL_BUSINESS_NAME_V1
          raw_account_value = str(value or "").strip()
          upper_account_value = raw_account_value.upper()
          looks_like_id = _lossq_live_is_policy_number(raw_account_value) or bool(re.search(r"\b[A-Z0-9]{2,}[-_][A-Z0-9]{2,}", upper_account_value))
          if not looks_like_id:
            account["business_name"] = raw_account_value
            account["insured_name"] = raw_account_value
            account["named_insured"] = raw_account_value
          else:
            account["account_number"] = raw_account_value
            account["customer_number"] = raw_account_value
        elif key in {"account number"}:
          account["account_number"] = value
          account["customer_number"] = value
        elif key in {"producer / producing agency", "producer", "producing agency", "agency"}:
          account["agency_name"] = value
          account["producing_agency"] = value
          account["producer"] = value
        elif key in {"producer number"}:
          account["producer_number"] = value
        elif key in {"effective date", "effective", "policy effective", "policy effective date", "policy start", "policy start date", "period start", "period from", "term start", "inception date"}:
          account["effective_date"] = _lossq_live_date_to_iso(value)
          account["effective"] = account["effective_date"]
        elif key in {"expiration date", "expiration", "expiry date", "policy expiration", "policy expiration date", "policy expiry", "policy expiry date", "policy end", "policy end date", "period end", "period to", "term end"}:
          account["expiration_date"] = _lossq_live_date_to_iso(value)
          account["expiration"] = account["expiration_date"]
        elif key in {"main policy number", "main policy", "policy number"}:
          account["policy_number"] = value
        elif key in {"writing carrier"}:
          account["writing_carrier"] = value
          account["carrier_name"] = value or account.get("carrier_name", "")
      continue

    if current_section == "policies":
      # LOSSQ_ACCOUNT_CARRIER_BAD_VALUE_CLEANUP_V1
      # Do not let table headers like Exposure Value become carrier values.
      bad_carrier_values = {"exposure value", "exposure basis", "premium", "annual premium", "policy number", "line of business"}
      for carrier_key in ["carrier_name", "writing_carrier", "carrier"]:
        if str(account.get(carrier_key) or "").strip().lower() in bad_carrier_values:
          account[carrier_key] = ""

      # LOSSQ_UNIVERSAL_POLICY_SCHEDULE_HEADER_MAP_V1
      def _policy_header_key(v):
        return " ".join(_lossq_live_clean_cell(v).lower().replace("/", " ").replace("_", " ").replace("#", "number").split())

      lower_row = [_policy_header_key(cell) for cell in nonempty]

      policy_header_aliases = {"policy number", "policy no", "policy num", "policy id", "policy"}
      line_header_aliases = {"line of business", "line", "coverage line", "coverage", "policy type", "lob"}
      effective_header_aliases = {"effective date", "effective", "eff date", "eff", "policy effective", "policy effective date", "period start", "period from", "term start"}
      expiration_header_aliases = {"expiration date", "expiration", "exp date", "exp", "expiry date", "policy expiration", "policy expiration date", "period end", "period to", "term end"}
      policy_period_aliases = {"policy period", "policy term", "period", "coverage period", "policy dates", "date range"}
      carrier_header_aliases = {"carrier", "writing carrier", "insurer", "company"}
      premium_header_aliases = {"premium", "annual premium", "current premium", "written premium"}
      exposure_basis_aliases = {"exposure basis", "basis", "exposure"}
      exposure_value_aliases = {"exposure value", "exposure amount", "value", "basis value"}
      expiring_premium_aliases = {"expiring premium", "prior premium", "current term premium"}
      target_renewal_aliases = {"target renewal premium", "target premium", "renewal premium"}
      payroll_aliases = {"payroll", "annual payroll", "estimated payroll"}
      revenue_aliases = {"revenue", "revenue sales", "revenue / sales", "sales", "gross sales", "receipts", "gross receipts"}
      employee_count_aliases = {"employee count", "employees", "number of employees"}
      vehicle_count_aliases = {"vehicle count", "vehicles", "number of vehicles", "power units"}
      driver_count_aliases = {"driver count", "drivers", "number of drivers"}
      property_tiv_aliases = {"property tiv", "tiv", "total insured value", "total insurable value", "property value"}
      building_value_aliases = {"building value", "building limit"}
      contents_value_aliases = {"contents value", "bpp", "business personal property"}

      if any(h in policy_header_aliases for h in lower_row) and any(h in line_header_aliases for h in lower_row):
        policy_header_seen = True
        policy_header_map = {h: idx for idx, h in enumerate(lower_row)}
        continue

      if not policy_header_seen:
        continue

      header_map = locals().get("policy_header_map", {})

      def _get_by_alias(row_values, aliases, fallback_index=None):
        for alias in aliases:
          if alias in header_map:
            idx = header_map[alias]
            return _lossq_live_clean_cell(row_values[idx]) if idx < len(row_values) else ""
        if fallback_index is not None and fallback_index < len(row_values):
          return _lossq_live_clean_cell(row_values[fallback_index])
        return ""

      # LOSSQ_POLICY_PERIOD_RANGE_SECTION_CSV_V2
      # Do not let fallback indexes shift Carrier into Effective Date.
      policy_number = _get_by_alias(row, policy_header_aliases, 0).upper()
      if not _lossq_live_is_policy_number(policy_number):
        continue

      lob = _get_by_alias(row, line_header_aliases, 1)
      carrier = _get_by_alias(row, carrier_header_aliases, 2) or account.get("writing_carrier") or account.get("carrier_name") or ""

      effective_raw = _get_by_alias(row, effective_header_aliases, None)
      expiration_raw = _get_by_alias(row, expiration_header_aliases, None)
      effective = _lossq_live_date_to_iso(effective_raw) if re.search(r"\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", str(effective_raw or "")) else ""
      expiration = _lossq_live_date_to_iso(expiration_raw) if re.search(r"\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", str(expiration_raw or "")) else ""

      if not effective or not expiration:
        period_value = _get_by_alias(row, policy_period_aliases, None)
        period_dates = re.findall(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}", str(period_value or ""))
        if len(period_dates) >= 2:
          effective = effective or _lossq_live_date_to_iso(period_dates[0])
          expiration = expiration or _lossq_live_date_to_iso(period_dates[1])

      current_premium = _get_by_alias(row, premium_header_aliases, 5)
      exposure_basis = _get_by_alias(row, exposure_basis_aliases, 6)
      exposure_value = _get_by_alias(row, exposure_value_aliases, None)
      expiring_premium = _get_by_alias(row, expiring_premium_aliases, None)
      target_renewal = _get_by_alias(row, target_renewal_aliases, None)
      payroll = _get_by_alias(row, payroll_aliases, None)
      revenue = _get_by_alias(row, revenue_aliases, None)
      employee_count = _get_by_alias(row, employee_count_aliases, None)
      vehicle_count = _get_by_alias(row, vehicle_count_aliases, None)
      driver_count = _get_by_alias(row, driver_count_aliases, None)
      property_tiv = _get_by_alias(row, property_tiv_aliases, None)
      building_value = _get_by_alias(row, building_value_aliases, None)
      contents_value = _get_by_alias(row, contents_value_aliases, None)

      policy = {
        "line_of_business": lob,
        "policy_type": lob,
        "coverage": lob,
        "policy_number": policy_number,
        "carrier": carrier,
        "carrier_name": carrier,
        "effective_date": effective,
        "effective": effective,
        "effectiveDate": effective,
        "expiration_date": expiration,
        "expiration": expiration,
        "expirationDate": expiration,
        "exposure_basis": exposure_basis,
        "exposure_value": exposure_value,
        "current_premium": current_premium,
        "premium": current_premium,
        "expiring_premium": expiring_premium,
        "target_renewal_premium": target_renewal,
        "payroll": payroll,
        "revenue": revenue,
        "sales": revenue,
        "receipts": revenue,
        "employee_count": employee_count,
        "vehicle_count": vehicle_count,
        "driver_count": driver_count,
        "property_tiv": property_tiv,
        "building_value": building_value,
        "contents_value": contents_value,
      }
      policies.append(policy)
      continue

    if current_section == "exposures":
      lower_row = [cell.lower() for cell in nonempty]
      if "field" in lower_row and "value" in lower_row:
        exposure_header_seen = True
        continue

      if not exposure_header_seen:
        continue

      if len(nonempty) >= 2:
        exposures[_lossq_live_clean_cell(nonempty[0])] = _lossq_live_clean_cell(nonempty[1])
      continue

    if current_section == "claims":
      lower_row = [_lossq_live_clean_cell(cell).lower() for cell in nonempty]

      # LOSSQ_SECTION_CSV_CLAIM_DETAIL_HEADER_MAP_V1
      # Universal section-based CSV claim parser. Do not rely on one fixed
      # column order because real loss runs may include Date Reported,
      # Date Closed, Litigation, Flag, Cause, etc.
      if "claim number" in lower_row and "policy number" in lower_row:
        claim_header_seen = True
        claim_headers = [_lossq_live_clean_cell(cell).lower() for cell in row]
        continue

      if not claim_header_seen:
        continue

      if not row or len(row) < 2:
        continue

      def claim_value(*names):
        for name in names:
          key = str(name or "").strip().lower()
          if key in claim_headers:
            idx = claim_headers.index(key)
            if idx < len(row):
              return _lossq_live_clean_cell(row[idx])
        return ""

      claim_number = claim_value("claim number", "claim no", "claim #", "claim id", "claim")
      policy_number = claim_value("policy number", "policy no", "policy #", "policy")
      lob = claim_value("line of business", "coverage", "line", "claim type", "policy type")
      status = claim_value("status", "claim status")
      loss_date = claim_value("date of loss", "loss date", "dol")
      reported_date = claim_value("date reported", "reported date", "report date")
      closed_date = claim_value("date closed", "closed date")
      paid_raw = claim_value("paid", "paid amount", "total paid", "loss paid")
      reserve_raw = claim_value("reserve", "reserves", "reserve amount", "total reserve")
      total_raw = claim_value("total incurred", "incurred", "gross incurred", "net incurred", "total")
      description = claim_value("description", "loss description", "claim description", "cause of loss", "cause")
      litigation = claim_value("litigation", "litigated", "suit")
      flag = claim_value("flag", "risk flag", "severity flag")

      if _lossq_live_is_claim_number(claim_number) and _lossq_live_is_policy_number(policy_number):
        claim_number = _lossq_live_clean_cell(claim_number).upper()
        policy_number = _lossq_live_clean_cell(policy_number).upper()
        paid = _lossq_live_money_to_float(paid_raw)
        reserve = _lossq_live_money_to_float(reserve_raw)
        total = _lossq_live_money_to_float(total_raw)

        # If carrier file omits total incurred, calculate safe total.
        if not total and (paid or reserve):
          total = paid + reserve

        claim = {
          "claim_number": claim_number,
          "policy_number": policy_number,
          "policy": policy_number,
          "line_of_business": lob,
          "claim_type": lob,
          "date_of_loss": _lossq_live_date_to_iso(loss_date),
          "loss_date": _lossq_live_date_to_iso(loss_date),
          "date_reported": _lossq_live_date_to_iso(reported_date),
          "reported_date": _lossq_live_date_to_iso(reported_date),
          "date_closed": _lossq_live_date_to_iso(closed_date),
          "closed_date": _lossq_live_date_to_iso(closed_date),
          "status": _lossq_live_clean_cell(status).title() or "Open",
          "paid": paid,
          "paid_amount": paid,
          "reserve": reserve,
          "reserve_amount": reserve,
          "total_incurred": total,
          "total_amount": total,
          "total_net_loss": total,
          "description": description,
          "loss_description": description,
          "litigation": litigation,
          "flag": flag,
        }
        claims.append(claim)
      continue

    if current_section == "summary":
      lower_row = [cell.lower() for cell in nonempty]
      if "metric" in lower_row and "value" in lower_row:
        summary_header_seen = True
        continue

      if not summary_header_seen:
        continue

      if len(nonempty) >= 2:
        loss_summary[_lossq_live_clean_cell(nonempty[0])] = _lossq_live_clean_cell(nonempty[1])
      continue

  if policies:
    account["policies"] = policies
    account["policy_schedule"] = policies

    # Prefer explicit main policy; otherwise use first policy.
    if not _lossq_live_is_policy_number(account.get("policy_number")):
      account["policy_number"] = policies[0].get("policy_number", "")

    # Use matching policy dates for main policy.
    main_policy = account.get("policy_number", "")
    matched_main = next((p for p in policies if p.get("policy_number") == main_policy), policies[0])
    account["effective_date"] = account.get("effective_date") or matched_main.get("effective_date", "")
    account["expiration_date"] = account.get("expiration_date") or matched_main.get("expiration_date", "")
    account["effective"] = account["effective_date"]
    account["expiration"] = account["expiration_date"]

  # LOSSQ_SECTION_CSV_EXPOSURE_POLICY_INFO_ROLLUP_V1
  # Universal rollup for CSV sections such as EXPOSURE / POLICY INFORMATION,
  # POLICY INFORMATION, PREMIUM WORKSHEET, or POLICY SCHEDULE.
  def _lossq_section_money(value):
    try:
      raw = str(value or "").replace("$", "").replace(",", "").strip()
      if raw in {"", "-", "None", "none", "null"}:
        return 0.0
      return float(raw)
    except Exception:
      return 0.0

  def _lossq_section_fmt_money(value):
    try:
      amount = float(value or 0)
    except Exception:
      amount = 0.0
    if amount <= 0:
      return ""
    if amount.is_integer():
      return str(int(amount))
    return f"{amount:.2f}"

  if policies:
    current_total = sum(_lossq_section_money(p.get("current_premium") or p.get("premium")) for p in policies)
    expiring_total = sum(_lossq_section_money(p.get("expiring_premium")) for p in policies)
    target_total = sum(_lossq_section_money(p.get("target_renewal_premium")) for p in policies)

    if current_total and not exposures.get("Current Premium"):
      exposures["Current Premium"] = _lossq_section_fmt_money(current_total)
    if expiring_total and not exposures.get("Expiring Premium"):
      exposures["Expiring Premium"] = _lossq_section_fmt_money(expiring_total)
    if target_total and not exposures.get("Target Renewal Premium"):
      exposures["Target Renewal Premium"] = _lossq_section_fmt_money(target_total)

    for policy in policies:
      basis = str(policy.get("exposure_basis") or "").strip()
      value = str(policy.get("exposure_value") or "").strip()
      basis_key = basis.lower()

      if not value:
        continue

      if "payroll" in basis_key and not exposures.get("Payroll"):
        exposures["Payroll"] = value
      elif ("revenue" in basis_key or "sales" in basis_key or "receipts" in basis_key) and not exposures.get("Revenue / Sales"):
        exposures["Revenue / Sales"] = value
      elif "employee" in basis_key and not exposures.get("Employee Count"):
        exposures["Employee Count"] = value
      elif "vehicle" in basis_key and not exposures.get("Vehicle Count"):
        exposures["Vehicle Count"] = value
      elif "driver" in basis_key and not exposures.get("Driver Count"):
        exposures["Driver Count"] = value
      elif ("tiv" in basis_key or "insured value" in basis_key or "property" in basis_key) and not exposures.get("Property TIV"):
        exposures["Property TIV"] = value

  # LOSSQ_SECTION_CSV_PRECISE_POLICY_EXPOSURE_ROLLUP_V1
  # Tight extraction: only use values from named columns or correctly paired
  # Exposure Basis / Exposure Value on the same policy row. Never use nearby
  # header labels as values.
  def _lossq_precise_clean_number(value):
    raw = str(value or "").replace("$", "").replace(",", "").strip()
    if raw in {"", "-", "None", "none", "null"}:
      return ""
    if not re.search(r"\d", raw):
      return ""
    return raw

  def _lossq_precise_money_float(value):
    raw = _lossq_precise_clean_number(value)
    if not raw:
      return 0.0
    try:
      return float(raw)
    except Exception:
      return 0.0

  def _lossq_precise_money_text(value):
    try:
      amount = float(value or 0)
    except Exception:
      amount = 0.0
    if amount <= 0:
      return ""
    return str(int(amount)) if amount.is_integer() else f"{amount:.2f}"

  def _lossq_set_first_exposure(label, value):
    cleaned = _lossq_precise_clean_number(value)
    if cleaned and not exposures.get(label):
      exposures[label] = cleaned

  if policies:
    current_total = sum(_lossq_precise_money_float(p.get("current_premium") or p.get("premium")) for p in policies)
    expiring_total = sum(_lossq_precise_money_float(p.get("expiring_premium")) for p in policies)
    target_total = sum(_lossq_precise_money_float(p.get("target_renewal_premium")) for p in policies)

    if current_total:
      exposures["Current Premium"] = exposures.get("Current Premium") or _lossq_precise_money_text(current_total)
    if expiring_total:
      exposures["Expiring Premium"] = exposures.get("Expiring Premium") or _lossq_precise_money_text(expiring_total)
    if target_total:
      exposures["Target Renewal Premium"] = exposures.get("Target Renewal Premium") or _lossq_precise_money_text(target_total)

    for policy in policies:
      _lossq_set_first_exposure("Payroll", policy.get("payroll"))
      _lossq_set_first_exposure("Revenue / Sales", policy.get("revenue") or policy.get("sales") or policy.get("receipts"))
      _lossq_set_first_exposure("Employee Count", policy.get("employee_count"))
      _lossq_set_first_exposure("Vehicle Count", policy.get("vehicle_count"))
      _lossq_set_first_exposure("Driver Count", policy.get("driver_count"))
      _lossq_set_first_exposure("Property TIV", policy.get("property_tiv"))
      _lossq_set_first_exposure("Building Value", policy.get("building_value"))
      _lossq_set_first_exposure("Contents Value", policy.get("contents_value"))

      basis = str(policy.get("exposure_basis") or "").strip().lower()
      exposure_value = _lossq_precise_clean_number(policy.get("exposure_value"))

      if not exposure_value:
        continue

      if "payroll" in basis:
        _lossq_set_first_exposure("Payroll", exposure_value)
      elif "employee" in basis:
        _lossq_set_first_exposure("Employee Count", exposure_value)
      elif "vehicle" in basis or "power unit" in basis:
        _lossq_set_first_exposure("Vehicle Count", exposure_value)
      elif "driver" in basis:
        _lossq_set_first_exposure("Driver Count", exposure_value)
      elif "tiv" in basis or "insured value" in basis or "property value" in basis:
        _lossq_set_first_exposure("Property TIV", exposure_value)
      elif "building" in basis or "contents" in basis:
        _lossq_set_first_exposure("Property TIV", exposure_value)
      elif "revenue" in basis or "sales" in basis or "receipt" in basis:
        _lossq_set_first_exposure("Revenue / Sales", exposure_value)

  if exposures:
    account["exposure_inputs"] = exposures
    account["exposures"] = exposures
    account["current_premium"] = exposures.get("Current Premium", "")
    account["expiring_premium"] = exposures.get("Expiring Premium", "")
    account["target_renewal_premium"] = exposures.get("Target Renewal Premium", "")
    account["payroll"] = exposures.get("Payroll", "")
    account["revenue"] = exposures.get("Revenue / Sales", "")
    account["sales"] = exposures.get("Revenue / Sales", "")
    account["receipts"] = exposures.get("Revenue / Sales", "")
    account["employee_count"] = exposures.get("Employee Count", "")
    account["vehicle_count"] = exposures.get("Vehicle Count", "")
    account["driver_count"] = exposures.get("Driver Count", "")
    account["property_tiv"] = exposures.get("Property TIV", "")
    account["building_value"] = exposures.get("Building Value", "")
    account["contents_value"] = exposures.get("Contents Value", "")

  if loss_summary:
    account["loss_summary"] = loss_summary

  if claims or policies or exposures:
    account["lossq_section_based_csv_detected"] = True
    account["extraction_status"] = "passed" if claims and policies else "needs_attention"
    account["extraction_score"] = 95 if claims and policies else 75
    account["requires_review"] = False if claims and policies else True

  print("LOSSQ_SECTION_CSV_RETURN_COUNTS:", {"claims": len(claims), "policies": len(policies), "exposures": len(exposures)})
  # LOSSQ_FINAL_PRE_SECTION_ACCOUNT_PROFILE_FALLBACK_V1
  # Final universal fallback for carrier CSVs that place account fields before formal sections.
  if not account.get("business_name"):
    for raw_row in rows[:25]:
      cells = [_lossq_live_clean_cell(c) for c in raw_row if _lossq_live_clean_cell(c)]
      if len(cells) < 2:
        continue
      raw_key = cells[0].lower().replace(".", "").strip()
      raw_value = cells[1].strip()
      if raw_key in {"account", "account name", "insured", "named insured", "business name"}:
        account["business_name"] = raw_value
        account["insured_name"] = raw_value
        account["named_insured"] = raw_value
        print("LOSSQ_FINAL_ACCOUNT_NAME_FROM_PRE_SECTION:", raw_value)
        break

  return claims, account


# LOSSQ_SECTION_CSV_HEADER_FALLBACK_V1
def _lossq_header_fallback_parse_section_csv(file_path):
  rows = _lossq_live_read_section_csv_rows(file_path)
  claims = []
  policies = []

  def clean(v):
    return _lossq_live_clean_cell(v)

  def key(v):
    return clean(v).lower().replace("/", " ").replace("_", " ").strip()

  for idx, row in enumerate(rows):
    header = [key(c) for c in row]

    if "policy number" in header and ("coverage line" in header or "line of business" in header or "coverage" in header):
      for data in rows[idx + 1:]:
        if not data or not any(clean(c) for c in data):
          break
        if len(data) < 2:
          continue

        policy_number = clean(data[0]).upper()
        if not _lossq_live_is_policy_number(policy_number):
          continue

        lob = clean(data[1]) if len(data) > 1 else ""
        carrier = clean(data[2]) if len(data) > 2 else ""
        effective = _lossq_live_date_to_iso(data[3]) if len(data) > 3 else ""
        expiration = _lossq_live_date_to_iso(data[4]) if len(data) > 4 else ""
        premium = clean(data[5]) if len(data) > 5 else ""

        policies.append({
          "policy_number": policy_number,
          "line_of_business": lob,
          "policy_type": lob,
          "coverage": lob,
          "carrier": carrier,
          "carrier_name": carrier,
          "effective_date": effective,
          "effective": effective,
          "expiration_date": expiration,
          "expiration": expiration,
          "current_premium": premium,
          "premium": premium,
        })

    if "claim number" in header and "policy number" in header:
      for data in rows[idx + 1:]:
        if not data or not any(clean(c) for c in data):
          break

        row_map = {}
        for h_i, h in enumerate(header):
          row_map[h] = clean(data[h_i]) if h_i < len(data) else ""

        claim_number = row_map.get("claim number", "").upper()
        policy_number = row_map.get("policy number", "").upper()

        if not _lossq_live_is_claim_number(claim_number) or not _lossq_live_is_policy_number(policy_number):
          continue

        paid = _lossq_live_money_to_float(row_map.get("paid", ""))
        reserve = _lossq_live_money_to_float(row_map.get("reserve", ""))
        total = _lossq_live_money_to_float(row_map.get("total incurred", ""))
        if not total and (paid or reserve):
          total = paid + reserve

        lob = row_map.get("line of business", "") or row_map.get("coverage", "")
        status = row_map.get("status", "") or "Open"
        loss_date = _lossq_live_date_to_iso(row_map.get("date of loss", ""))
        reported_date = _lossq_live_date_to_iso(row_map.get("date reported", ""))
        closed_date = _lossq_live_date_to_iso(row_map.get("date closed", ""))
        description = row_map.get("description", "")

        claims.append({
          "claim_number": claim_number,
          "policy_number": policy_number,
          "policy": policy_number,
          "line_of_business": lob,
          "claim_type": lob,
          "date_of_loss": loss_date,
          "loss_date": loss_date,
          "date_reported": reported_date,
          "reported_date": reported_date,
          "date_closed": closed_date,
          "closed_date": closed_date,
          "status": status.title(),
          "paid": paid,
          "paid_amount": paid,
          "reserve": reserve,
          "reserve_amount": reserve,
          "total_incurred": total,
          "total_amount": total,
          "total_net_loss": total,
          "description": description,
          "loss_description": description,
          "litigation": row_map.get("litigation", ""),
          "flag": row_map.get("flag", ""),
        })

  profile = {}
  if policies:
    profile["policies"] = policies
    profile["policy_schedule"] = policies
    profile["policy_number"] = policies[0].get("policy_number", "")
    profile["effective_date"] = policies[0].get("effective_date", "")
    profile["expiration_date"] = policies[0].get("expiration_date", "")

  print("LOSSQ_SECTION_CSV_HEADER_FALLBACK_COUNTS:", {"claims": len(claims), "policies": len(policies)})
  return claims, profile

def lossq_live_repair_section_csv_upload(file_path, parsed_claims, parsed_profile):
  """
  If the uploaded file is a section-based CSV, override the old row parser
  so Notes, Loss Summary, Metric, and exposure rows do not become claims.
  """
  filename = str(file_path or "").lower()
  if not filename.endswith(".csv"):
    return parsed_claims, parsed_profile

  section_claims, section_profile = _lossq_live_extract_section_based_csv(file_path)
  # LOSSQ_SECTION_CSV_EMPTY_CLAIMS_FALLBACK_V1
  # If section extraction produced zero claims, use universal header fallback.
  if not section_claims:
    fallback_claims, fallback_profile = _lossq_header_fallback_parse_section_csv(file_path)
    if fallback_claims or fallback_profile:
      print("LOSSQ_SECTION_CSV_USING_HEADER_FALLBACK:", {"claims": len(fallback_claims), "profile_keys": list(fallback_profile.keys())})
      section_claims = fallback_claims or section_claims
      if fallback_profile:
        section_profile.update({k: v for k, v in fallback_profile.items() if v not in ("", None, [], {})})
    elif not section_profile:
      return parsed_claims, parsed_profile

  if not isinstance(parsed_profile, dict):
    parsed_profile = {}

  merged_profile = dict(parsed_profile)
  merged_profile.update({k: v for k, v in section_profile.items() if v not in ("", None, [], {})})

  if section_claims:
    parsed_claims = section_claims
    merged_profile["claims"] = section_claims
    merged_profile["parsed_claims"] = section_claims

  # LOSSQ_APPLY_EXPOSURE_INPUTS_TO_GENERIC_PARSE_RESULT_V1
  raw_text_for_exposure = str(
    merged_profile.get("raw_text_preview")
    or merged_profile.get("raw_text")
    or parsed_profile.get("raw_text_preview")
    or parsed_profile.get("raw_text")
    or ""
  )
  exposure_inputs = {}
  exposure_inputs.update(extract_exposure_inputs_from_raw_text(raw_text_for_exposure) or {})
  exposure_inputs.update(extract_exposure_inputs_from_parsed_rows(parsed_claims) or {})

  if exposure_inputs:
    merged_profile.update({k: v for k, v in exposure_inputs.items() if v not in ("", None, [], {})})
    merged_profile["exposure_inputs"] = exposure_inputs
    merged_profile["exposures"] = exposure_inputs

  return parsed_claims, merged_profile


# LOSSQ_PDF_CLAIM_DETAIL_NUMBER_REPAIR_V1
def lossq_pdf_clean_text(value):
  return re.sub(r"\s+", " ", str(value or "")).strip()


def lossq_pdf_claim_number_is_policy_derived(claim_number, policy_number="", line_of_business=""):
  claim = lossq_beta_norm_key(claim_number)
  policy = lossq_beta_norm_key(policy_number)

  if not claim:
    return True

  if policy and claim == policy:
    return True

  claim_compact = re.sub(r"[^A-Z0-9]", "", claim)
  policy_compact = re.sub(r"[^A-Z0-9]", "", policy)

  if policy_compact and (claim_compact in policy_compact or policy_compact in claim_compact):
    return True

  line_tokens = (
    "GL", "WC", "AUTO", "AU", "PROP", "PR", "CP", "BOP", "CY", "CYBER",
    "UMB", "EXCESS", "EPLI", "EPL", "DO", "DNO", "EO", "PL", "IM",
    "CRIME", "FID", "FIDUCIARY", "CARGO", "MTC", "LIAB", "ABUSE",
    "MOLESTATION", "GAR", "GARAGE"
  )

  # Generated/policy-derived examples:
  # CP-2025, UMB-2025, GL-2025-4701-GENERAL, WC-2025-4703-WORKERS.
  generated_pattern = r"^(" + "|".join(line_tokens) + r")[-_ ]?(19|20)\d{2}([-_ ][A-Z0-9]+){0,4}$"
  if re.match(generated_pattern, claim):
    return True

  return False


def lossq_pdf_extract_claim_detail_rows_from_text(raw_text):
  text_value = str(raw_text or "")
  if not text_value.strip():
    return []

  upper_text = text_value.upper()
  start_index = upper_text.find("CLAIM DETAIL")
  if start_index >= 0:
    text_value = text_value[start_index:]

  end_index = text_value.upper().find("UNDERWRITING NOTES")
  if end_index >= 0:
    claim_section = text_value[:end_index]
  else:
    claim_section = text_value

  claim_id_pattern = r"[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){1,6}"
  policy_id_pattern = r"[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,8}"

  row_start_pattern = re.compile(
    rf"(?m)^\s*(?P<claim>{claim_id_pattern})\s+(?P<policy>{policy_id_pattern})\s+"
  )

  starts = list(row_start_pattern.finditer(claim_section))
  if not starts:
    return []

  money_pattern = r"\$?\(?\d[\d,]*(?:\.\d+)?\)?"
  date_pattern = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"

  claims = []

  for index, match in enumerate(starts):
    next_start = starts[index + 1].start() if index + 1 < len(starts) else len(claim_section)

    claim_number = lossq_pdf_clean_text(match.group("claim"))
    policy_number = lossq_pdf_clean_text(match.group("policy"))
    rest = claim_section[match.end():next_start]
    rest = lossq_pdf_clean_text(rest)

    if not claim_number or not policy_number:
      continue

    status_match = re.search(
      rf"(?P<line>.*?)(?P<status>Open|Closed|Reopened|Pending|Denied|Reported)\s+(?P<loss>{date_pattern})\s+(?P<reported>{date_pattern})(?:\s+(?P<closed>{date_pattern}))?\s+(?P<paid>{money_pattern})\s+(?P<reserve>{money_pattern})\s+(?P<total>{money_pattern})\s*(?P<description>.*)$",
      rest,
      re.IGNORECASE | re.DOTALL,
    )

    if not status_match:
      continue

    line_of_business = lossq_pdf_clean_text(status_match.group("line"))
    status = lossq_pdf_clean_text(status_match.group("status")).title()
    loss_date = lossq_pdf_clean_text(status_match.group("loss"))
    reported_date = lossq_pdf_clean_text(status_match.group("reported"))
    closed_date = lossq_pdf_clean_text(status_match.group("closed"))

    paid = lossq_beta_money_to_float(status_match.group("paid"))
    reserve = lossq_beta_money_to_float(status_match.group("reserve"))
    total = lossq_beta_money_to_float(status_match.group("total"))

    if total <= 0 and (paid > 0 or reserve > 0):
      total = paid + reserve

    description = lossq_pdf_clean_text(status_match.group("description"))

    litigation = False
    litigation_match = re.search(r"\b(Yes|No)\b", description, re.IGNORECASE)
    if litigation_match:
      litigation = litigation_match.group(1).lower() == "yes"
      description = lossq_pdf_clean_text(description[:litigation_match.start()])

    if not line_of_business:
      continue

    if total <= 0 and paid <= 0 and reserve <= 0:
      continue

    claims.append({
      "claim_number": claim_number,
      "claim_id": claim_number,
      "policy_number": policy_number,
      "line_of_business": line_of_business,
      "claim_type": line_of_business,
      "policy_type": line_of_business,
      "status": status,
      "date_of_loss": loss_date,
      "loss_date": loss_date,
      "date_reported": reported_date,
      "reported_date": reported_date,
      "date_closed": closed_date,
      "closed_date": closed_date,
      "paid": paid,
      "paid_amount": paid,
      "reserve": reserve,
      "reserve_amount": reserve,
      "total_incurred": total,
      "total_amount": total,
      "description": description,
      "loss_description": description,
      "litigation": litigation,
    })

  return claims


def lossq_pdf_amounts_close(left, right):
  try:
    return abs(float(left or 0) - float(right or 0)) <= 2.0
  except Exception:
    return False


def lossq_repair_pdf_claims_from_raw_text(raw_text, parsed_claims):
  raw_claims = lossq_pdf_extract_claim_detail_rows_from_text(raw_text)
  if not raw_claims:
    return parsed_claims

  existing_claims = [dict(item) for item in (parsed_claims or []) if isinstance(item, dict)]
  repaired = []
  used_existing = set()

  for raw_claim in raw_claims:
    match_index = None

    raw_policy = lossq_beta_norm_key(raw_claim.get("policy_number"))
    raw_paid = lossq_beta_money_to_float(raw_claim.get("paid_amount"))
    raw_reserve = lossq_beta_money_to_float(raw_claim.get("reserve_amount"))
    raw_total = lossq_beta_money_to_float(raw_claim.get("total_incurred"))

    for idx, existing in enumerate(existing_claims):
      if idx in used_existing:
        continue

      existing_policy = lossq_beta_norm_key(existing.get("policy_number"))
      existing_paid = lossq_beta_money_to_float(existing.get("paid_amount"))
      existing_reserve = lossq_beta_money_to_float(existing.get("reserve_amount"))
      existing_total = lossq_beta_money_to_float(existing.get("total_incurred"))

      same_policy = bool(raw_policy and raw_policy == existing_policy)
      same_amounts = (
        lossq_pdf_amounts_close(raw_paid, existing_paid)
        and lossq_pdf_amounts_close(raw_reserve, existing_reserve)
        and lossq_pdf_amounts_close(raw_total, existing_total)
      )

      if same_policy and same_amounts:
        match_index = idx
        break

    if match_index is not None:
      used_existing.add(match_index)
      merged = dict(existing_claims[match_index])

      # Trust the actual row-start claim number from the PDF claim detail table.
      for key, value in raw_claim.items():
        if value not in ("", None, [], {}):
          merged[key] = value

      repaired.append(merged)
    else:
      repaired.append(raw_claim)

  # Keep any unmatched existing claim only if it does not look generated/policy-derived.
  for idx, existing in enumerate(existing_claims):
    if idx in used_existing:
      continue

    if not lossq_pdf_claim_number_is_policy_derived(
      existing.get("claim_number"),
      existing.get("policy_number"),
      existing.get("line_of_business") or existing.get("claim_type"),
    ):
      repaired.append(existing)

  if repaired and len(repaired) >= len(existing_claims):
    print("LOSSQ_PDF_CLAIM_NUMBER_REPAIR_APPLIED:", {
      "before": len(existing_claims),
      "after": len(repaired),
      "claim_numbers": [claim.get("claim_number") for claim in repaired],
      "total_incurred": sum(lossq_beta_money_to_float(claim.get("total_incurred")) for claim in repaired),
    })
    return repaired

  return parsed_claims


# LOSSQ_DIRECT_FILE_EXPOSURE_CAPTURE_V1
def lossq_extract_exposure_inputs_directly_from_file(file_path: str):
  """
  Universal direct file exposure extractor.
  Reads CSV/XLSX rows before the claim parser strips non-claim exposure columns.
  """
  exposure_inputs = {}

  try:
    lower_path = str(file_path or "").lower()

    if lower_path.endswith(".csv"):
      rows = []
      for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
          with open(file_path, "r", newline="", encoding=encoding, errors="ignore") as handle:
            rows = list(csv.DictReader(handle))
          break
        except Exception:
          rows = []

      if rows:
        exposure_inputs.update(extract_exposure_inputs_from_parsed_rows(rows) or {})

    elif lower_path.endswith((".xlsx", ".xls")):
      try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True)
        rows = []

        for sheet in workbook.worksheets:
          values = list(sheet.iter_rows(values_only=True))
          if not values:
            continue

          # Header row format.
          header = [str(cell or "").strip() for cell in values[0]]
          if any(header):
            for raw_row in values[1:]:
              row = {}
              for index, header_name in enumerate(header):
                if not header_name:
                  continue
                row[header_name] = raw_row[index] if index < len(raw_row) else ""
              if row:
                rows.append(row)

          # Label/value rows.
          for raw_row in values:
            clean_cells = [cell for cell in raw_row if cell not in ("", None)]
            if len(clean_cells) >= 2:
              rows.append({
                "label": clean_cells[0],
                "value": clean_cells[1],
              })

        if rows:
          exposure_inputs.update(extract_exposure_inputs_from_parsed_rows(rows) or {})
      except Exception as exc:
        print("LOSSQ_XLSX_DIRECT_EXPOSURE_CAPTURE_ERROR:", str(exc)[:200])

  except Exception as exc:
    print("LOSSQ_DIRECT_FILE_EXPOSURE_CAPTURE_ERROR:", str(exc)[:200])

  exposure_inputs = {
    k: v for k, v in (exposure_inputs or {}).items()
    if v not in ("", None, [], {})
  }

  if exposure_inputs:
    print("LOSSQ_DIRECT_FILE_EXPOSURE_CAPTURED:", exposure_inputs)

  return exposure_inputs


# LOSSQ_CLEAN_FLAT_CSV_PARSER_V1
def lossq_parse_clean_flat_csv_v1(file_path: str):
  """
  Universal parser for clean flat CSV loss runs where every claim row contains
  account/profile, policy, exposure, and claim columns.
  This prevents clean carrier spreadsheets from falling into older parser paths.
  """
  import csv
  import re

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())

  def money_text(value):
    raw = clean(value).replace("$", "").replace(",", "")
    if raw in {"", "-", "None", "none", "null"}:
      return ""
    return raw

  def money_float(value):
    raw = money_text(value)
    if not raw:
      return 0.0
    try:
      return float(raw)
    except Exception:
      return 0.0

  def get(row, *aliases):
    lookup = {key(k): v for k, v in (row or {}).items()}
    for alias in aliases:
      value = lookup.get(key(alias))
      if clean(value):
        return clean(value)
    return ""

  rows = []
  for encoding in ("utf-8-sig", "utf-8", "latin-1"):
    try:
      with open(file_path, "r", newline="", encoding=encoding, errors="ignore") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row or {}) for row in reader]
      break
    except Exception:
      rows = []

  if not rows:
    return [], {}

  has_claim_number = any(key(h) in {"claimnumber", "claimno", "claimid", "claim"} for h in (rows[0] or {}).keys())
  has_policy_number = any(key(h) in {"policynumber", "policyno", "policy"} for h in (rows[0] or {}).keys())

  if not has_claim_number or not has_policy_number:
    return [], {}

  profile = {}
  claims = []
  policies_by_key = {}

  exposure_inputs = {}

  def set_first(label, value):
    cleaned = money_text(value)
    if not cleaned:
      return

    existing = exposure_inputs.get(label)
    if not existing or str(existing).strip() in {"0", "0.0", "0.00"}:
      exposure_inputs[label] = cleaned

  for row in rows:
    business_name = get(row, "Business Name", "Named Insured", "Insured", "Account Name")
    account_number = get(row, "Account Number", "Customer Number", "Client Number")
    agency_name = get(row, "Producing Agency", "Agency", "Broker")
    carrier = get(row, "Carrier", "Writing Carrier", "Insurer", "Company")
    policy_number = get(row, "Policy Number", "Policy No", "Policy")
    lob = get(row, "Line of Business", "LOB", "Coverage", "Policy Type")
    policy_period = get(row, "Policy Period", "Policy Term", "Coverage Period")
    effective_date = get(row, "Effective Date", "Effective")
    expiration_date = get(row, "Expiration Date", "Expiration", "Expiry")
    state = get(row, "State", "Primary State")

    if business_name and not profile.get("business_name"):
      profile["business_name"] = business_name
      profile["insured"] = business_name
      profile["named_insured"] = business_name

    if account_number and not profile.get("account_number"):
      profile["account_number"] = account_number
      profile["customer_number"] = account_number

    if agency_name and not profile.get("agency_name"):
      profile["agency_name"] = agency_name

    if carrier and not profile.get("carrier_name"):
      profile["carrier_name"] = carrier
      profile["writing_carrier"] = carrier

    if policy_number and not profile.get("policy_number"):
      profile["policy_number"] = policy_number

    if effective_date and not profile.get("effective_date"):
      profile["effective_date"] = effective_date
      profile["effective"] = effective_date

    if expiration_date and not profile.get("expiration_date"):
      profile["expiration_date"] = expiration_date
      profile["expiration"] = expiration_date

    if state and not profile.get("state"):
      profile["state"] = state

    set_first("Payroll", get(row, "Payroll", "Annual Payroll", "Estimated Payroll"))
    set_first("Revenue / Sales", get(row, "Revenue / Sales", "Revenue", "Sales", "Gross Sales", "Receipts", "Gross Receipts"))
    set_first("Employee Count", get(row, "Employee Count", "Employees", "Number of Employees"))
    set_first("Vehicle Count", get(row, "Vehicle Count", "Vehicles", "Number of Vehicles", "Power Units"))
    set_first("Driver Count", get(row, "Driver Count", "Drivers", "Number of Drivers"))
    set_first("Property TIV", get(row, "Property TIV", "TIV", "Total Insured Value", "Total Insurable Value", "Property Value"))
    set_first("Building Value", get(row, "Building Value", "Building Limit"))
    set_first("Contents Value", get(row, "Contents Value", "BPP", "Business Personal Property"))

    if policy_number:
      policy_key = policy_number.upper()
      if policy_key not in policies_by_key:
        policies_by_key[policy_key] = {
          "policy_number": policy_number,
          "line_of_business": lob,
          "policy_type": lob,
          "coverage": lob,
          "carrier": carrier,
          "carrier_name": carrier,
          "writing_carrier": carrier,
          "policy_period": policy_period,
          "effective_date": effective_date,
          "effective": effective_date,
          "expiration_date": expiration_date,
          "expiration": expiration_date,
          "state": state,
          "current_premium": money_text(get(row, "Current Premium", "Annual Premium", "Written Premium", "Premium")),
          "premium": money_text(get(row, "Current Premium", "Annual Premium", "Written Premium", "Premium")),
          "expiring_premium": money_text(get(row, "Expiring Premium", "Prior Premium", "Previous Premium")),
          "target_renewal_premium": money_text(get(row, "Target Renewal Premium", "Renewal Premium", "Estimated Renewal Premium")),
          "payroll": money_text(get(row, "Payroll", "Annual Payroll", "Estimated Payroll")),
          "revenue": money_text(get(row, "Revenue / Sales", "Revenue", "Sales", "Gross Sales", "Receipts")),
          "employee_count": money_text(get(row, "Employee Count", "Employees", "Number of Employees")),
          "vehicle_count": money_text(get(row, "Vehicle Count", "Vehicles", "Number of Vehicles", "Power Units")),
          "driver_count": money_text(get(row, "Driver Count", "Drivers", "Number of Drivers")),
          "property_tiv": money_text(get(row, "Property TIV", "TIV", "Total Insured Value", "Property Value")),
        }

    claim_number = get(row, "Claim Number", "Claim No", "Claim ID", "Claim")
    if not claim_number:
      continue

    paid = money_text(get(row, "Paid", "Paid Amount", "Total Paid"))
    reserve = money_text(get(row, "Reserve", "Outstanding Reserve", "Total Reserve"))
    incurred = money_text(get(row, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred"))
    if not incurred:
      total = money_float(paid) + money_float(reserve)
      incurred = str(int(total)) if total.is_integer() else f"{total:.2f}"

    claims.append({
      "claim_number": claim_number,
      "claim_no": claim_number,
      "policy_number": policy_number,
      "line_of_business": lob,
      "coverage": lob,
      "date_of_loss": get(row, "Date of Loss", "Loss Date"),
      "loss_date": get(row, "Date of Loss", "Loss Date"),
      "date_reported": get(row, "Date Reported", "Reported Date"),
      "date_closed": get(row, "Date Closed", "Closed Date"),
      "status": get(row, "Status", "Claim Status") or ("Closed" if get(row, "Date Closed", "Closed Date") else "Open"),
      "paid": paid,
      "paid_amount": paid,
      "reserve": reserve,
      "total_incurred": incurred,
      "incurred": incurred,
      "cause_of_loss": get(row, "Cause of Loss", "Cause"),
      "description": get(row, "Claim Notes", "Loss Notes", "Notes", "Narrative", "Claim Description", "Description", "Loss Description"),
      "litigation": get(row, "Litigation", "Litigated"),
      "flag": get(row, "Flag", "Alert"),
    })

  policies = list(policies_by_key.values())

  current_total = sum(money_float(p.get("current_premium")) for p in policies)
  expiring_total = sum(money_float(p.get("expiring_premium")) for p in policies)
  target_total = sum(money_float(p.get("target_renewal_premium")) for p in policies)

  if current_total:
    exposure_inputs["Current Premium"] = str(int(current_total)) if current_total.is_integer() else f"{current_total:.2f}"
  if expiring_total:
    exposure_inputs["Expiring Premium"] = str(int(expiring_total)) if expiring_total.is_integer() else f"{expiring_total:.2f}"
  if target_total:
    exposure_inputs["Target Renewal Premium"] = str(int(target_total)) if target_total.is_integer() else f"{target_total:.2f}"

  if exposure_inputs:
    profile["exposure_inputs"] = exposure_inputs
    profile["exposures"] = exposure_inputs
    profile["current_premium"] = exposure_inputs.get("Current Premium", "")
    profile["expiring_premium"] = exposure_inputs.get("Expiring Premium", "")
    profile["target_renewal_premium"] = exposure_inputs.get("Target Renewal Premium", "")
    profile["payroll"] = exposure_inputs.get("Payroll", "")
    profile["revenue"] = exposure_inputs.get("Revenue / Sales", "")
    profile["sales"] = exposure_inputs.get("Revenue / Sales", "")
    profile["receipts"] = exposure_inputs.get("Revenue / Sales", "")
    profile["employee_count"] = exposure_inputs.get("Employee Count", "")
    profile["vehicle_count"] = exposure_inputs.get("Vehicle Count", "")
    profile["driver_count"] = exposure_inputs.get("Driver Count", "")
    profile["property_tiv"] = exposure_inputs.get("Property TIV", "")
    profile["building_value"] = exposure_inputs.get("Building Value", "")
    profile["contents_value"] = exposure_inputs.get("Contents Value", "")

  profile["policies"] = policies
  profile["claims"] = claims
  profile["parsed_claims"] = claims
  profile["lossq_clean_flat_csv_detected"] = True
  profile["extraction_status"] = "passed"
  profile["extraction_score"] = 95
  profile["requires_review"] = False

  print("LOSSQ_CLEAN_FLAT_CSV_PARSED:", {
    "claims": len(claims),
    "policies": len(policies),
    "exposures": exposure_inputs,
  })

  return claims, profile




# LOSSQ_POLICY_SCHEDULE_DISPLAY_CLEANUP_V2
def lossq_clean_policy_schedule_display_names_v2(claims, profile):
    """
    Cleans policy schedule display rows after extraction.
    Removes account-number rows from policy cards and fills Unknown coverage names
    from clear policy prefixes such as GL, WC, BOP, UMB, GAR, DOL, etc.
    """
    import json
    import re

    def clean(value):
        return re.sub(r"\s+", " ", str(value or "").strip())

    def compact(value):
        return re.sub(r"[^A-Z0-9]", "", clean(value).upper())

    def is_unknown(value):
        value = clean(value).upper()
        return value in {"", "UNKNOWN", "N/A", "NA", "NONE", "NULL", "NOT SET", "POLICY", "COVERAGE"}

    def looks_like_account_number(value):
        raw = clean(value).upper()
        key = compact(value)
        if not key:
            return False

        if key.startswith(("ACCT", "ACCOUNT", "CUST", "CUSTOMER", "CLIENT")):
            return True

        if "ACCT" in key or "ACCOUNT" in key or "CUSTOMER" in key or "CLIENT" in key:
            if not re.match(r"^(GL|WC|BOP|PROP|CP|UMB|UM|GAR|DOL|AUTO|CA|PL|EPLI|CY|CARGO|IM|DO|DNO|LIQ)", key):
                return True

        if re.search(r"\b(ACCOUNT|ACCT|CUSTOMER|CLIENT)\b", raw):
            return True

        return False

    def infer_lob(policy_number, current_lob=""):
        current_lob = clean(current_lob)
        if not is_unknown(current_lob):
            return current_lob

        key = compact(policy_number)

        prefix_map = [
            ("BOP", "Businessowners Policy"),
            ("GL", "General Liability"),
            ("WC", "Workers Compensation"),
            ("PROP", "Property"),
            ("CP", "Commercial Property"),
            ("UMB", "Umbrella"),
            ("UM", "Umbrella"),
            ("GAR", "Garage Liability"),
            ("DOL", "Dealers Open Lot"),
            ("AUTO", "Commercial Auto"),
            ("CA", "Commercial Auto"),
            ("PL", "Professional Liability"),
            ("EPLI", "Employment Practices Liability"),
            ("CY", "Cyber Liability"),
            ("CARGO", "Motor Truck Cargo"),
            ("IM", "Inland Marine"),
            ("DO", "Directors & Officers"),
            ("DNO", "Directors & Officers"),
            ("LIQ", "Liquor Liability"),
        ]

        for prefix, label in prefix_map:
            if key.startswith(prefix):
                return label

        return ""

    if not isinstance(profile, dict):
        profile = {}

    raw_policies = profile.get("policies") or []

    if isinstance(raw_policies, str):
        try:
            raw_policies = json.loads(raw_policies)
        except Exception:
            raw_policies = []

    if not isinstance(raw_policies, list):
        raw_policies = []

    cleaned_policies = []
    seen = set()

    for policy in raw_policies:
        if not isinstance(policy, dict):
            continue

        policy_number = clean(
            policy.get("policy_number")
            or policy.get("policyNumber")
            or policy.get("number")
            or policy.get("policy")
        )

        if not policy_number:
            continue

        if looks_like_account_number(policy_number):
            continue

        policy_key = compact(policy_number)
        if not policy_key or policy_key in seen:
            continue

        seen.add(policy_key)

        lob = infer_lob(
            policy_number,
            policy.get("line_of_business")
            or policy.get("policy_type")
            or policy.get("policyType")
            or policy.get("coverage")
            or policy.get("coverage_line")
            or policy.get("policy_line")
            or policy.get("line")
            or policy.get("lob")
        )

        if not lob:
            lob = "Unknown"

        policy["policy_number"] = policy_number
        policy["policyNumber"] = policy_number
        policy["line_of_business"] = lob
        policy["policy_type"] = lob
        policy["policyType"] = lob
        policy["coverage"] = lob
        policy["coverage_line"] = lob
        policy["coverageType"] = lob
        policy["policy_line"] = lob
        policy["line"] = lob
        policy["lob"] = lob

        cleaned_policies.append(policy)

    policy_lob_map = {}
    for policy in cleaned_policies:
        pn = compact(policy.get("policy_number"))
        lob = clean(policy.get("line_of_business"))
        if pn and lob and not is_unknown(lob):
            policy_lob_map[pn] = lob

    cleaned_claims = []

    if isinstance(claims, list):
        for claim in claims:
            if not isinstance(claim, dict):
                cleaned_claims.append(claim)
                continue

            claim_policy_number = clean(
                claim.get("policy_number")
                or claim.get("policyNumber")
                or claim.get("policy")
            )

            if looks_like_account_number(claim_policy_number):
                claim_policy_number = ""

            current_lob = (
                claim.get("line_of_business")
                or claim.get("policy_type")
                or claim.get("policyType")
                or claim.get("coverage")
                or claim.get("coverage_line")
                or claim.get("policy_line")
                or claim.get("line")
                or claim.get("lob")
            )

            lob = clean(policy_lob_map.get(compact(claim_policy_number)) or infer_lob(claim_policy_number, current_lob))

            if lob and not is_unknown(lob):
                claim["line_of_business"] = lob
                claim["policy_type"] = lob
                claim["policyType"] = lob
                claim["coverage"] = lob
                claim["coverage_line"] = lob
                claim["coverageType"] = lob
                claim["policy_line"] = lob
                claim["line"] = lob
                claim["lob"] = lob

            if claim_policy_number:
                claim["policy_number"] = claim_policy_number
                claim["policyNumber"] = claim_policy_number

            cleaned_claims.append(claim)
    else:
        cleaned_claims = claims

    if cleaned_policies:
        profile["policies"] = cleaned_policies
        profile["policy_schedule"] = cleaned_policies

        first_policy_number = clean(cleaned_policies[0].get("policy_number"))
        if first_policy_number:
            profile["policy_number"] = first_policy_number
            profile["main_policy_number"] = first_policy_number

    print("LOSSQ_POLICY_SCHEDULE_DISPLAY_CLEANUP_V2:", {
        "policies": [
            {
                "policy_number": p.get("policy_number"),
                "line_of_business": p.get("line_of_business"),
            }
            for p in cleaned_policies
        ]
    })

    return cleaned_claims, profile




# LOSSQ_LABEL_PDF_POLICY_SCHEDULE_REBUILD_V4
def lossq_rebuild_label_pdf_policy_schedule_v4(raw_text, carrier="", effective_date="", expiration_date=""):
  """
  Rebuilds policy schedule directly from the POLICY SCHEDULE section of a
  text-readable PDF. This prevents Account Number values from becoming policy
  rows and fills missing coverage names from clear policy prefixes.
  """
  import re

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").strip())

  def money_text(value):
    raw = clean(value).replace("$", "").replace(",", "")
    if raw in {"", "-", "None", "none", "null"}:
      return ""
    return raw

  def compact(value):
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())

  def looks_like_account_number(value):
    raw = clean(value).upper()
    key = compact(value)
    if not key:
      return False
    if key.startswith(("ACCT", "ACCOUNT", "CUST", "CUSTOMER", "CLIENT")):
      return True
    if "ACCT" in key or "ACCOUNT" in key or "CUSTOMER" in key or "CLIENT" in key:
      return True
    if re.search(r"\b(ACCT|ACCOUNT|CUSTOMER|CLIENT)\b", raw):
      return True
    return False

  def infer_lob(policy_number, current_lob=""):
    current_lob = clean(current_lob)
    if current_lob and current_lob.upper() not in {"UNKNOWN", "N/A", "NA", "NONE", "NULL", "NOT SET", "POLICY", "COVERAGE"}:
      return current_lob

    key = compact(policy_number)

    prefix_map = [
      ("BOP", "Businessowners Policy"),
      ("GL", "General Liability"),
      ("WC", "Workers Compensation"),
      ("PROP", "Property"),
      ("CP", "Commercial Property"),
      ("UMB", "Umbrella"),
      ("UM", "Umbrella"),
      ("GAR", "Garage Liability"),
      ("DOL", "Dealers Open Lot"),
      ("AUTO", "Commercial Auto"),
      ("CA", "Commercial Auto"),
      ("PL", "Professional Liability"),
      ("EPLI", "Employment Practices Liability"),
      ("CY", "Cyber Liability"),
      ("CARGO", "Motor Truck Cargo"),
      ("IM", "Inland Marine"),
      ("DO", "Directors & Officers"),
      ("DNO", "Directors & Officers"),
      ("LIQ", "Liquor Liability"),
    ]

    for prefix, label in prefix_map:
      if key.startswith(prefix):
        return label

    return ""

  source = str(raw_text or "")
  match = re.search(
    r"\bPOLICY SCHEDULE\b([\s\S]*?)(?:\bLOSS SUMMARY\b|\bCLAIM DETAIL\b|\bBROKER RECOMMENDATION\b|$)",
    source,
    flags=re.I,
  )

  if not match:
    return []

  section = match.group(1)
  lines = [clean(line) for line in section.splitlines() if clean(line)]

  policies = []
  current = None

  field_map = {
    "policy number": "policy_number",
    "line of business": "line_of_business",
    "coverage": "line_of_business",
    "policy type": "line_of_business",
    "policy line": "line_of_business",
    "carrier": "carrier",
    "writing carrier": "carrier",
    "effective date": "effective_date",
    "effective": "effective_date",
    "expiration date": "expiration_date",
    "expiration": "expiration_date",
    "current premium": "current_premium",
    "expiring premium": "expiring_premium",
    "target renewal premium": "target_renewal_premium",
    "exposure basis": "exposure_basis",
    "exposure value": "exposure_value",
    "employee count": "employee_count",
    "vehicle count": "vehicle_count",
    "driver count": "driver_count",
    "property tiv": "property_tiv",
  }

  for line in lines:
    m = re.match(r"^([^:]{2,80})\s*:\s*(.*)$", line)
    if not m:
      continue

    label = clean(m.group(1)).lower()
    value = clean(m.group(2))

    if label == "policy number":
      if current and current.get("policy_number"):
        policies.append(current)
      current = {"policy_number": value}
      continue

    if current is not None and label in field_map:
      current[field_map[label]] = value

  if current and current.get("policy_number"):
    policies.append(current)

  rebuilt = []
  seen = set()

  for policy in policies:
    policy_number = clean(policy.get("policy_number"))

    if not policy_number or looks_like_account_number(policy_number):
      continue

    key = compact(policy_number)
    if not key or key in seen:
      continue

    seen.add(key)

    lob = infer_lob(policy_number, policy.get("line_of_business"))
    if not lob:
      lob = "Unknown"

    policy_carrier = clean(policy.get("carrier")) or clean(carrier)
    policy_effective = clean(policy.get("effective_date")) or clean(effective_date)
    policy_expiration = clean(policy.get("expiration_date")) or clean(expiration_date)

    rebuilt.append({
      "policy_number": policy_number,
      "policyNumber": policy_number,
      "line_of_business": lob,
      "policy_type": lob,
      "policyType": lob,
      "coverage": lob,
      "coverage_line": lob,
      "coverageType": lob,
      "policy_line": lob,
      "line": lob,
      "lob": lob,
      "carrier": policy_carrier,
      "carrier_name": policy_carrier,
      "writing_carrier": policy_carrier,
      "effective_date": policy_effective,
      "effective": policy_effective,
      "expiration_date": policy_expiration,
      "expiration": policy_expiration,
      "current_premium": money_text(policy.get("current_premium")),
      "premium": money_text(policy.get("current_premium")),
      "expiring_premium": money_text(policy.get("expiring_premium")),
      "target_renewal_premium": money_text(policy.get("target_renewal_premium")),
      "exposure_basis": clean(policy.get("exposure_basis")),
      "exposure_value": money_text(policy.get("exposure_value")),
      "employee_count": money_text(policy.get("employee_count")),
      "vehicle_count": money_text(policy.get("vehicle_count")),
      "driver_count": money_text(policy.get("driver_count")),
      "property_tiv": money_text(policy.get("property_tiv")),
    })

  if rebuilt:
    print("LOSSQ_LABEL_PDF_POLICY_SCHEDULE_REBUILD_V4:", {
      "policies": [
        {"policy_number": item.get("policy_number"), "line_of_business": item.get("line_of_business")}
        for item in rebuilt
      ]
    })

  return rebuilt


# LOSSQ_LABEL_BASED_PDF_LOSS_RUN_PARSER_V1
def lossq_parse_label_based_pdf_loss_run_v1(file_path: str):
  """
  Universal fallback for text-readable PDFs that use label/value lines:
  Named Insured:, Policy Number:, Claim Number:, Paid:, Reserve:, etc.
  This is intentionally PDF-only and does not affect CSV/XLSX parsing.
  """
  import re

  try:
    from pypdf import PdfReader
  except Exception:
    PdfReader = None

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def money_text(value):
    raw = clean(value).replace("$", "").replace(",", "")
    if raw in {"", "-", "None", "none", "null"}:
      return ""
    return raw

  def money_float(value):
    raw = money_text(value)
    if not raw:
      return 0.0
    try:
      return float(raw)
    except Exception:
      return 0.0

  raw_text = ""

  try:
    if PdfReader:
      reader = PdfReader(file_path)
      raw_text = "\n".join(page.extract_text() or "" for page in reader.pages)
  except Exception as exc:
    print("LOSSQ_LABEL_PDF_TEXT_EXTRACT_WARNING:", str(exc)[:500])

  if not clean(raw_text):
    return [], {}

  lines = [clean(line) for line in raw_text.splitlines() if clean(line)]

  def get_first(*labels):
    for line in lines:
      for label in labels:
        m = re.match(rf"^{re.escape(label)}\s*:\s*(.+)$", line, flags=re.I)
        if m:
          value = clean(m.group(1))
          if value:
            return value
    return ""

  profile = {}

  business_name = get_first("Named Insured", "Business Name", "Insured", "Account Name")
  account_number = get_first("Account Number", "Customer Number", "Client Number")
  agency_name = get_first("Producing Agency", "Agency", "Broker")
  carrier = get_first("Writing Carrier", "Carrier", "Insurer")
  policy_period = get_first("Policy Period", "Policy Term")
  effective_date = get_first("Effective Date", "Effective")
  expiration_date = get_first("Expiration Date", "Expiration")
  state = get_first("State", "Primary State")
  evaluation_date = get_first("Evaluation Date", "Valuation Date")

  if business_name:
    profile["business_name"] = business_name
    profile["insured"] = business_name
    profile["insured_name"] = business_name
    profile["named_insured"] = business_name

  if account_number:
    profile["account_number"] = account_number
    profile["customer_number"] = account_number

  if agency_name:
    profile["agency_name"] = agency_name
    profile["producing_agency"] = agency_name

  if carrier:
    profile["carrier_name"] = carrier
    profile["writing_carrier"] = carrier

  if policy_period:
    profile["policy_period"] = policy_period

  if effective_date:
    profile["effective_date"] = effective_date
    profile["effective"] = effective_date

  if expiration_date:
    profile["expiration_date"] = expiration_date
    profile["expiration"] = expiration_date

  if evaluation_date:
    profile["evaluation_date"] = evaluation_date

  if state:
    profile["state"] = state

  # Parse repeated policy blocks.
  policies = []
  current_policy = None

  policy_field_map = {
    "policy number": "policy_number",
    "line of business": "line_of_business",
    "coverage": "line_of_business",
    "carrier": "carrier",
    "writing carrier": "carrier",
    "effective date": "effective_date",
    "expiration date": "expiration_date",
    "current premium": "current_premium",
    "expiring premium": "expiring_premium",
    "target renewal premium": "target_renewal_premium",
    "exposure basis": "exposure_basis",
    "exposure value": "exposure_value",
    "employee count": "employee_count",
    "vehicle count": "vehicle_count",
    "driver count": "driver_count",
    "property tiv": "property_tiv",
  }

  for line in lines:
    m = re.match(r"^([^:]{2,80})\s*:\s*(.+)$", line)
    if not m:
      continue

    label = clean(m.group(1)).lower()
    value = clean(m.group(2))

    if label == "claim number":
      if current_policy and current_policy.get("policy_number"):
        policies.append(current_policy)
      current_policy = None
      break

    if label == "policy number":
      if current_policy and current_policy.get("policy_number"):
        policies.append(current_policy)
      current_policy = {"policy_number": value}
      continue

    if current_policy is not None and label in policy_field_map:
      key = policy_field_map[label]
      current_policy[key] = money_text(value) if "premium" in key or key in {"exposure_value", "employee_count", "vehicle_count", "driver_count", "property_tiv"} else value

  if current_policy and current_policy.get("policy_number"):
    policies.append(current_policy)

  # Normalize policy rows.
  clean_policies = []
  seen_policies = set()

  for policy in policies:
    policy_number = clean(policy.get("policy_number"))
    if not policy_number:
      continue

    if policy_number.upper() in seen_policies:
      continue
    seen_policies.add(policy_number.upper())

    lob = clean(policy.get("line_of_business"))
    pol_carrier = clean(policy.get("carrier")) or carrier

    clean_policies.append({
      "policy_number": policy_number,
      "line_of_business": lob,
      "policy_type": lob,
      "coverage": lob,
      "carrier": pol_carrier,
      "carrier_name": pol_carrier,
      "writing_carrier": pol_carrier,
      "effective_date": clean(policy.get("effective_date")) or effective_date,
      "effective": clean(policy.get("effective_date")) or effective_date,
      "expiration_date": clean(policy.get("expiration_date")) or expiration_date,
      "expiration": clean(policy.get("expiration_date")) or expiration_date,
      "current_premium": money_text(policy.get("current_premium")),
      "premium": money_text(policy.get("current_premium")),
      "expiring_premium": money_text(policy.get("expiring_premium")),
      "target_renewal_premium": money_text(policy.get("target_renewal_premium")),
      "exposure_basis": clean(policy.get("exposure_basis")),
      "exposure_value": money_text(policy.get("exposure_value")),
      "employee_count": money_text(policy.get("employee_count")),
      "vehicle_count": money_text(policy.get("vehicle_count")),
      "driver_count": money_text(policy.get("driver_count")),
      "property_tiv": money_text(policy.get("property_tiv")),
    })

  # LOSSQ_LABEL_PDF_POLICY_SCHEDULE_REBUILD_CALL_V4
  rebuilt_pdf_policies = lossq_rebuild_label_pdf_policy_schedule_v4(raw_text, carrier, effective_date, expiration_date)
  if rebuilt_pdf_policies:
    clean_policies = rebuilt_pdf_policies

  if clean_policies:
    profile["policies"] = clean_policies
    profile["policy_schedule"] = clean_policies
    profile["policy_number"] = clean_policies[0].get("policy_number", "")
    profile["main_policy_number"] = clean_policies[0].get("policy_number", "")

  # Exposure rollup.
  exposure_inputs = {}

  def set_first(label, value):
    value = money_text(value)
    if value and not exposure_inputs.get(label):
      exposure_inputs[label] = value

  current_total = sum(money_float(p.get("current_premium")) for p in clean_policies)
  expiring_total = sum(money_float(p.get("expiring_premium")) for p in clean_policies)
  target_total = sum(money_float(p.get("target_renewal_premium")) for p in clean_policies)

  if current_total:
    exposure_inputs["Current Premium"] = str(int(current_total)) if current_total.is_integer() else f"{current_total:.2f}"
  if expiring_total:
    exposure_inputs["Expiring Premium"] = str(int(expiring_total)) if expiring_total.is_integer() else f"{expiring_total:.2f}"
  if target_total:
    exposure_inputs["Target Renewal Premium"] = str(int(target_total)) if target_total.is_integer() else f"{target_total:.2f}"

  for policy in clean_policies:
    basis = clean(policy.get("exposure_basis")).lower()
    value = money_text(policy.get("exposure_value"))

    if "revenue" in basis or "sales" in basis or "receipt" in basis:
      set_first("Revenue / Sales", value)
    elif "payroll" in basis:
      set_first("Payroll", value)
    elif "property" in basis or "tiv" in basis or "insured value" in basis:
      set_first("Property TIV", value)

    set_first("Employee Count", policy.get("employee_count"))
    set_first("Vehicle Count", policy.get("vehicle_count"))
    set_first("Driver Count", policy.get("driver_count"))
    set_first("Property TIV", policy.get("property_tiv"))

  if exposure_inputs:
    profile["exposure_inputs"] = exposure_inputs
    profile["exposures"] = exposure_inputs
    profile["current_premium"] = exposure_inputs.get("Current Premium", "")
    profile["expiring_premium"] = exposure_inputs.get("Expiring Premium", "")
    profile["target_renewal_premium"] = exposure_inputs.get("Target Renewal Premium", "")
    profile["payroll"] = exposure_inputs.get("Payroll", "")
    profile["revenue"] = exposure_inputs.get("Revenue / Sales", "")
    profile["sales"] = exposure_inputs.get("Revenue / Sales", "")
    profile["employee_count"] = exposure_inputs.get("Employee Count", "")
    profile["vehicle_count"] = exposure_inputs.get("Vehicle Count", "")
    profile["driver_count"] = exposure_inputs.get("Driver Count", "")
    profile["property_tiv"] = exposure_inputs.get("Property TIV", "")

  # Parse repeated claim blocks.
  claims = []
  current_claim = None

  claim_field_map = {
    "claim number": "claim_number",
    "policy number": "policy_number",
    "line of business": "line_of_business",
    "coverage": "line_of_business",
    "date of loss": "date_of_loss",
    "loss date": "date_of_loss",
    "date reported": "date_reported",
    "reported date": "date_reported",
    "date closed": "date_closed",
    "closed date": "date_closed",
    "status": "status",
    "paid": "paid",
    "paid amount": "paid",
    "reserve": "reserve",
    "total incurred": "total_incurred",
    "incurred": "total_incurred",
    "cause of loss": "cause_of_loss",
    "cause": "cause_of_loss",
    "description": "description",
    "claim notes": "description",
    "claim note": "description",
    "loss notes": "description",
    "loss note": "description",
    "notes": "description",
    "narrative": "description",
    "claim narrative": "description",
    "loss narrative": "description",
    "litigation": "litigation",
  }

  for line in lines:
    m = re.match(r"^([^:]{2,80})\s*:\s*(.+)$", line)
    if not m:
      continue

    label = clean(m.group(1)).lower()
    value = clean(m.group(2))

    if label == "claim number":
      if current_claim and current_claim.get("claim_number"):
        claims.append(current_claim)
      current_claim = {"claim_number": value, "claim_no": value}
      continue

    if current_claim is not None and label in claim_field_map:
      key = claim_field_map[label]
      current_claim[key] = money_text(value) if key in {"paid", "reserve", "total_incurred"} else value

  if current_claim and current_claim.get("claim_number"):
    claims.append(current_claim)

  normalized_claims = []
  for claim in claims:
    paid = money_text(claim.get("paid"))
    reserve = money_text(claim.get("reserve"))
    incurred = money_text(claim.get("total_incurred"))

    if not incurred:
      total = money_float(paid) + money_float(reserve)
      incurred = str(int(total)) if total.is_integer() else f"{total:.2f}"

    claim_number = clean(claim.get("claim_number"))
    if not claim_number:
      continue

    normalized_claims.append({
      "claim_number": claim_number,
      "claim_no": claim_number,
      "policy_number": clean(claim.get("policy_number")) or profile.get("policy_number", ""),
      "line_of_business": clean(claim.get("line_of_business")),
      "coverage": clean(claim.get("line_of_business")),
      "date_of_loss": clean(claim.get("date_of_loss")),
      "loss_date": clean(claim.get("date_of_loss")),
      "date_reported": clean(claim.get("date_reported")),
      "date_closed": clean(claim.get("date_closed")),
      "status": clean(claim.get("status")) or ("Closed" if clean(claim.get("date_closed")) else "Open"),
      "paid": paid,
      "paid_amount": paid,
      "reserve": reserve,
      "total_incurred": incurred,
      "incurred": incurred,
      "cause_of_loss": clean(claim.get("cause_of_loss")),
      "description": clean(claim.get("description")),
      "litigation": clean(claim.get("litigation")),
    })

  # LOSSQ_LABEL_PDF_NO_CLAIMS_PROFILE_UPLOAD_V1
  if not normalized_claims:
    no_claims_signal = bool(re.search(
      r"(?i)\\b(no\\s+claims|no\\s+losses|zero\\s+claims|claim\\s+count\\s*[:#-]?\\s*0|no\\s+claim\\s+activity|none\\s+reported|no\\s+claims\\s+reported)\\b",
      raw_text or "",
    ))

    if clean_policies and (profile or no_claims_signal):
      profile["claims"] = []
      profile["parsed_claims"] = []
      profile["lossq_no_claims_pdf_detected"] = True
      profile["loss_run_status"] = "No claims reported"
      profile["extraction_status"] = "passed"
      profile["extraction_score"] = 94
      profile["requires_review"] = False
      profile["claim_count"] = 0
      profile["total_claims"] = 0
      profile["open_claims"] = 0
      profile["closed_claims"] = 0
      profile["total_incurred"] = 0
      profile["total_paid"] = 0
      profile["total_reserve"] = 0
      profile["policies"] = clean_policies
      profile["policy_schedule"] = clean_policies

      normalized_claims, profile = lossq_clean_policy_schedule_display_names_v2([], profile)

      print("LOSSQ_LABEL_PDF_NO_CLAIMS_PROFILE_UPLOAD_V1:", {
        "business_name": profile.get("business_name"),
        "account_number": profile.get("account_number"),
        "policies": [
          {
            "policy_number": item.get("policy_number"),
            "line_of_business": item.get("line_of_business"),
          }
          for item in profile.get("policies", [])
          if isinstance(item, dict)
        ],
      })

      return [], profile

    return [], {}

  profile["claims"] = normalized_claims
  profile["parsed_claims"] = normalized_claims
  profile["lossq_label_based_pdf_detected"] = True
  profile["extraction_status"] = "passed"
  profile["extraction_score"] = 92
  profile["requires_review"] = False

  print("LOSSQ_LABEL_BASED_PDF_PARSED:", {
    "claims": len(normalized_claims),
    "policies": len(clean_policies),
    "business_name": profile.get("business_name"),
    "account_number": profile.get("account_number"),
  })

  # LOSSQ_LABEL_BASED_PDF_POLICY_DISPLAY_CLEANUP_V2
  normalized_claims, profile = lossq_clean_policy_schedule_display_names_v2(normalized_claims, profile)
  return normalized_claims, profile










# LOSSQ_BUSINESS_NAME_LABEL_REPORT_CLEAN_V1
def lossq_clean_business_name_label_report_v1(value):
  import re

  raw = re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  # Remove label prefixes accidentally captured from PDFs.
  raw = re.sub(
    r"(?i)^\s*(?:business\s+name|named\s+insured|insured|account\s+name|applicant|entity)\s*[:#\-\/]*\s*",
    "",
    raw,
  )

  # Remove trailing report/document words accidentally captured from headers.
  raw = re.sub(
    r"(?i)\s+(?:loss\s+run\s+report|loss\s+run|report|pdf|document)\s*$",
    "",
    raw,
  )

  raw = raw.strip(" :-|/")

  return raw

# LOSSQ_PDF_SAVE_TIME_BUSINESS_NAME_REPAIR_V2
def lossq_pdf_save_time_business_name_repair_v2(file_path, parsed_profile=None, parsed_claims=None, direct_profile=None):
  """
  Final save-time PDF business-name repair.

  This runs immediately before the upload save flow extends/saves claims.
  It updates BOTH parsed_profile and direct_profile so Account Snapshot receives
  the real insured/business name instead of labels like "/ Business Name".

  Universal only: no customer, carrier, or demo-file hardcoding.
  """
  import os
  import re

  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def compact(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def is_bad_name(value):
    k = compact(value)
    return k in {
      "",
      "businessnamenotset",
      "businessname",
      "namedinsured",
      "insured",
      "insuredname",
      "accountname",
      "applicant",
      "entity",
      "companyname",
      "namedinsuredbusinessname",
      "insuredbusinessname",
      "accountnamebusinessname",
      "unknown",
      "notset",
      "na",
      "none",
      "null",
    }

  def good_name(value):
    value = clean(value)
    if is_bad_name(value):
      return ""

    low = value.lower()
    blocked = [
      "loss run",
      "lossrun",
      "policy schedule",
      "claim detail",
      "claim block",
      "claim number",
      "policy number",
      "line of business",
      "coverage",
      "loss summary",
      "paid amount",
      "reserve amount",
      "total incurred",
      "date of loss",
      "effective date",
      "expiration date",
      "evaluation date",
      "valuation date",
      "account number",
      "carrier",
      "writing carrier",
      "producer",
      "producing agency",
      "report date",
    ]

    if any(part in low for part in blocked):
      return ""

    if len(value) < 4 or len(value) > 140:
      return ""

    return value

  def read_pdf_text():
    parts = [
      parsed_profile.get("raw_text"),
      parsed_profile.get("raw_text_preview"),
      parsed_profile.get("ocr_text"),
      parsed_profile.get("document_text"),
      direct_profile.get("raw_text"),
      direct_profile.get("raw_text_preview"),
    ]

    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      parts.append("\n".join((page.extract_text() or "") for page in reader.pages))
    except Exception as exc:
      print("LOSSQ_PDF_SAVE_TIME_BUSINESS_NAME_REPAIR_READ_ERROR_V2:", str(exc)[:200])

    return "\n".join(str(part or "") for part in parts if part)

  raw_text = read_pdf_text()
  normalized_text = re.sub(r"\s+", " ", raw_text or " ").strip()
  lines = [clean(line) for line in (raw_text or "").splitlines()]
  lines = [line for line in lines if line]

  candidates = []

  # 1) Strong same-line or near-line label extraction.
  label_patterns = [
    r"(?is)(?:named\s+insured|insured\s*/\s*business\s*name|insured\s+business\s+name|business\s+name|account\s+name|applicant|entity|company\s+name)\s*[:#\-]?\s+(.{4,140}?)(?=\s+(?:writing\s+carrier|carrier|account\s+number|policy\s+schedule|effective\s+date|expiration\s+date|valuation\s+date|claim\s+detail|loss\s+summary|producer|producing\s+agency)\b|$)",
    r"(?is)(?:named\s+insured|insured|business\s+name|account\s+name)\s*[:#\-]\s*(.{4,140}?)(?=\s+(?:writing\s+carrier|carrier|account\s+number|policy\s+schedule|effective\s+date|expiration\s+date|valuation\s+date|claim\s+detail|loss\s+summary)\b|$)",
  ]

  for pattern in label_patterns:
    for match in re.finditer(pattern, normalized_text):
      candidate = good_name(match.group(1))
      if candidate:
        candidates.append(candidate)

  # 2) Label on one line with value on following lines.
  label_keys = {
    "namedinsured",
    "insured",
    "businessname",
    "accountname",
    "applicant",
    "entity",
    "companyname",
    "insuredbusinessname",
    "namedinsuredbusinessname",
  }

  for idx, line in enumerate(lines):
    if compact(line) in label_keys:
      for j in range(idx + 1, min(idx + 10, len(lines))):
        candidate = good_name(lines[j])
        if candidate:
          candidates.append(candidate)
          break

  # 3) Whole-document entity suffix fallback.
  entity_pattern = re.compile(
    r"(?is)\b([A-Z][A-Za-z0-9&.,'’\- ]{2,110}?\s+"
    r"(?:LLC|L\.L\.C\.|Inc\.?|Incorporated|Corp\.?|Corporation|Co\.?|Company|PLLC|LP|LLP|Group|Services|Service|Agency|Associates|Partners|Enterprises|Holdings))\b"
  )

  for match in entity_pattern.finditer(normalized_text):
    candidate = good_name(match.group(1))
    if candidate:
      # Remove leading generic document words if extraction captured them.
      candidate = re.sub(r"(?i)^(loss\s+run|account\s+snapshot|submission|ready\s+for\s+submission)\s+", "", candidate).strip()
      candidate = good_name(candidate)
      if candidate:
        candidates.append(candidate)

  # 4) Claim-level fallback if claim rows have insured fields.
  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue
    for key in ["business_name", "named_insured", "insured_name", "account_name", "insured"]:
      candidate = good_name(claim.get(key))
      if candidate:
        candidates.append(candidate)

  # 5) Filename fallback only if there is no document-level candidate.
  if not candidates:
    try:
      base = os.path.basename(str(file_path or ""))
      base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
      base = re.sub(r"(?i)\b(lossq|loss|run|messy|clean|ready|submission|pdf|csv|xlsx|xls|test|v\d+|parser|friendly)\b", " ", base)
      base = re.sub(r"[_\-]+", " ", base)
      base = re.sub(r"\s+", " ", base).strip()
      candidate = good_name(base.title())
      if candidate:
        candidates.append(candidate)
    except Exception:
      pass

  business_name = ""
  for candidate in candidates:
    candidate = good_name(candidate)
    if candidate:
      business_name = lossq_clean_business_name_label_report_v1(candidate)
      break

  current = (
    parsed_profile.get("business_name")
    or parsed_profile.get("named_insured")
    or parsed_profile.get("insured_name")
    or parsed_profile.get("account_name")
    or direct_profile.get("business_name")
    or direct_profile.get("named_insured")
    or direct_profile.get("insured_name")
    or direct_profile.get("account_name")
  )

  business_name = lossq_clean_business_name_label_report_v1(business_name)
  if business_name and (is_bad_name(current) or business_name != clean(current)):
    for target in [parsed_profile, direct_profile]:
      target["business_name"] = business_name
      target["named_insured"] = business_name
      target["insured_name"] = business_name
      target["account_name"] = business_name

  print("LOSSQ_PDF_SAVE_TIME_BUSINESS_NAME_REPAIR_V2:", {
    "business_name": parsed_profile.get("business_name") or direct_profile.get("business_name"),
    "candidate_count": len(candidates),
    "current_before": current,
  })

  return parsed_profile, direct_profile

# LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_V1
def lossq_pdf_final_business_name_repair_v1(file_path, parsed_profile=None):
  """
  Universal final PDF business-name repair.

  Fixes messy PDFs where the account snapshot shows:
  - Business Name Not Set
  - / Business Name
  - Named Insured / Business Name

  No customer/file hardcoding.
  """
  import os
  import re

  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def bad_name(value):
    k = key(value)
    return k in {
      "",
      "businessnamenotset",
      "businessname",
      "namedinsured",
      "insured",
      "accountname",
      "applicant",
      "namedinsuredbusinessname",
      "insuredbusinessname",
      "unknown",
      "notset",
      "na",
      "none",
      "null",
    }

  def good_candidate(value):
    value = clean(value)
    if bad_name(value):
      return ""

    lowered = value.lower()
    blocked_parts = [
      "loss run",
      "policy schedule",
      "claim detail",
      "claim block",
      "claim number",
      "policy number",
      "line of business",
      "coverage",
      "loss summary",
      "paid amount",
      "reserve amount",
      "total incurred",
      "date of loss",
      "effective date",
      "expiration date",
      "valuation date",
      "account number",
      "carrier",
      "writing carrier",
      "producer",
      "producing agency",
      "report",
    ]

    if any(part in lowered for part in blocked_parts):
      return ""

    if len(value) < 4 or len(value) > 120:
      return ""

    return value

  def read_pdf_text():
    parts = [
      parsed_profile.get("raw_text"),
      parsed_profile.get("raw_text_preview"),
      parsed_profile.get("ocr_text"),
      parsed_profile.get("document_text"),
    ]

    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      parts.append("\n".join((page.extract_text() or "") for page in reader.pages))
    except Exception:
      pass

    return "\n".join(str(part or "") for part in parts if part)

  raw_text = read_pdf_text()
  lines = [clean(line) for line in raw_text.splitlines()]
  lines = [line for line in lines if line]

  label_keys = {
    "namedinsured",
    "businessname",
    "insured",
    "accountname",
    "applicant",
    "entity",
    "companyname",
  }

  # 1) Same-line labels: Business Name: ABC LLC
  for line in lines:
    m = re.match(
      r"(?i)^\s*(named\s+insured|business\s+name|insured|account\s+name|applicant|entity|company\s+name)\s*[:#\-]\s*(.+?)\s*$",
      line,
    )
    if m:
      candidate = good_candidate(m.group(2))
      if candidate:
        parsed_profile["business_name"] = candidate
        parsed_profile["named_insured"] = candidate
        parsed_profile["insured_name"] = candidate
        parsed_profile["account_name"] = candidate
        print("LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_V1:", {"business_name": candidate, "source": "same_line_label"})
        return parsed_profile

  # 2) Label on one line, value on following lines.
  for idx, line in enumerate(lines):
    line_key = key(line)

    is_label_line = (
      line_key in label_keys
      or line_key in {"namedinsuredbusinessname", "insuredbusinessname", "accountnamebusinessname"}
      or bool(re.match(r"(?i)^(named\s+insured|insured|account\s+name)\s*/\s*business\s+name$", line))
    )

    if not is_label_line:
      continue

    for j in range(idx + 1, min(idx + 8, len(lines))):
      candidate = good_candidate(lines[j])
      if candidate:
        parsed_profile["business_name"] = candidate
        parsed_profile["named_insured"] = candidate
        parsed_profile["insured_name"] = candidate
        parsed_profile["account_name"] = candidate
        print("LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_V1:", {"business_name": candidate, "source": "next_line_label"})
        return parsed_profile

  # 3) Entity suffix fallback: find the first clean company-like line.
  entity_suffix_pattern = re.compile(
    r"(?i)\b([A-Z][A-Za-z0-9&.,'’\- ]{2,90}\s+"
    r"(?:LLC|L\.L\.C\.|Inc\.?|Incorporated|Corp\.?|Corporation|Co\.?|Company|PLLC|LP|LLP|Group|Services|Service|Agency|Associates|Partners|Enterprises|Holdings))\b"
  )

  for line in lines:
    candidate = good_candidate(line)
    if not candidate:
      continue

    m = entity_suffix_pattern.search(candidate)
    if m:
      business_name = clean(m.group(1))
      if good_candidate(business_name):
        parsed_profile["business_name"] = business_name
        parsed_profile["named_insured"] = business_name
        parsed_profile["insured_name"] = business_name
        parsed_profile["account_name"] = business_name
        print("LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_V1:", {"business_name": business_name, "source": "entity_suffix"})
        return parsed_profile

  # 4) Filename fallback only when nothing usable exists.
  current = (
    parsed_profile.get("business_name")
    or parsed_profile.get("named_insured")
    or parsed_profile.get("insured_name")
    or parsed_profile.get("account_name")
  )

  if bad_name(current):
    try:
      base = os.path.basename(str(file_path or ""))
      base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
      base = re.sub(r"(?i)\b(lossq|loss|run|messy|clean|ready|submission|pdf|csv|xlsx|xls|test|v\d+|parser|friendly)\b", " ", base)
      base = re.sub(r"[_\-]+", " ", base)
      base = re.sub(r"\s+", " ", base).strip()
      candidate = good_candidate(base.title())
      if candidate:
        parsed_profile["business_name"] = candidate
        parsed_profile["named_insured"] = candidate
        parsed_profile["insured_name"] = candidate
        parsed_profile["account_name"] = candidate
        print("LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_V1:", {"business_name": candidate, "source": "filename_fallback"})
    except Exception:
      pass

  return parsed_profile

# LOSSQ_PDF_CLEAN_TABLE_CLAIM_REPAIR_V1
def lossq_pdf_clean_table_claim_repair_v1(file_path, parsed_claims=None, parsed_profile=None):
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      amount = float(raw or 0)
      return -amount if neg else amount
    except Exception:
      return 0.0

  def policy_like(value):
    raw = clean(value).upper()
    if not raw or not re.search(r"\d", raw):
      return False
    if any(token in raw for token in ["ACCT", "ACCOUNT", "CUSTOMER", "CUST", "CLIENT"]):
      return False
    return bool(re.search(r"^[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,9}$", raw))

  def claim_like(value):
    raw = clean(value).upper()
    if not raw or not re.search(r"\d", raw):
      return False
    if raw in {"CLAIM NUMBER", "CLAIM NO", "POLICY NUMBER", "LOSS SUMMARY"}:
      return False
    return bool(re.search(r"^[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){1,8}$", raw))

  def infer_line(policy_number, fallback=""):
    value = clean(fallback)
    if value:
      return value

    key = clean(policy_number).upper()
    mapping = [
      (("CARGO", "MTC"), "Motor Truck Cargo"),
      (("BOP",), "Businessowners Policy"),
      (("WC",), "Workers Compensation"),
      (("GL", "CGL"), "General Liability"),
      (("UMB", "UM", "EXCESS"), "Umbrella"),
      (("LIQ", "LIQUOR"), "Liquor Liability"),
      (("AUTO", "CA"), "Commercial Auto"),
      (("GAR",), "Garage Liability"),
      (("DOL",), "Dealers Open Lot"),
      (("CP", "PROP"), "Commercial Property"),
      (("CY", "CYBER"), "Cyber Liability"),
      (("EPLI",), "Employment Practices Liability"),
      (("PL",), "Professional Liability"),
      (("DO", "DNO"), "Directors & Officers"),
      (("IM",), "Inland Marine"),
    ]

    for prefixes, label in mapping:
      if any(key.startswith(prefix + "-") or key.startswith(prefix + "_") or key == prefix for prefix in prefixes):
        return label

    return ""

  def read_pdf_text():
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
      print("LOSSQ_PDF_CLEAN_TABLE_CLAIM_REPAIR_READ_ERROR_V1:", str(exc)[:200])
      return ""

  raw_text = read_pdf_text()
  if not raw_text:
    return parsed_claims, parsed_profile

  upper = raw_text.upper()

  def section_between(start_labels, end_labels):
    starts = [upper.find(label) for label in start_labels if upper.find(label) >= 0]
    if not starts:
      return ""
    start = min(starts)
    tail = upper[start:]
    ends = [tail.find(label) for label in end_labels if tail.find(label) > 0]
    end = start + min(ends) if ends else len(raw_text)
    return raw_text[start:end]

  date_pattern = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
  money_pattern = r"\$?\(?-?\d[\d,]*(?:\.\d{1,2})?\)?"
  status_pattern = r"Open|Closed|Reopened|Pending|Denied|Reported"
  claim_pattern = r"[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){1,8}"
  policy_pattern = r"[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,9}"

  claim_section = section_between(
    ["CLAIM DETAIL", "CLAIMS DETAIL", "CLAIM DETAILS", "CLAIMS DETAILS"],
    ["LOSS SUMMARY", "UNDERWRITING NOTES", "BROKER RECOMMENDATION"],
  )

  extracted_claims = []

  if claim_section:
    lines = [clean(line) for line in claim_section.splitlines() if clean(line)]
    lines = [
      line for line in lines
      if not re.search(r"(?i)^claim\s+number\s+policy\s+number\b", line)
      and not re.search(r"(?i)^claim\s+detail$", line)
      and not re.search(r"(?i)^claims\s+detail$", line)
    ]

    claim_text = "\n".join(lines)
    row_start = re.compile(rf"(?m)^\s*(?P<claim>{claim_pattern})\s+(?P<policy>{policy_pattern})\s+")
    starts = list(row_start.finditer(claim_text))
    seen = set()

    for index, match in enumerate(starts):
      claim_number = clean(match.group("claim")).upper()
      policy_number = clean(match.group("policy")).upper()

      if not claim_like(claim_number) or not policy_like(policy_number):
        continue

      next_start = starts[index + 1].start() if index + 1 < len(starts) else len(claim_text)
      rest = clean(claim_text[match.end():next_start])

      row_match = re.search(
        rf"^(?P<lob>.*?)(?P<loss>{date_pattern})\s+(?P<status>{status_pattern})\s+"
        rf"(?P<paid>{money_pattern})\s+(?P<reserve>{money_pattern})\s+(?P<total>{money_pattern})\s*(?P<description>.*)$",
        rest,
        flags=re.I | re.S,
      )

      if not row_match:
        row_match = re.search(
          rf"^(?P<lob>.*?)(?P<status>{status_pattern})\s+(?P<loss>{date_pattern})"
          rf"(?:\s+(?P<reported>{date_pattern}))?(?:\s+(?P<closed>{date_pattern}))?\s+"
          rf"(?P<paid>{money_pattern})\s+(?P<reserve>{money_pattern})\s+(?P<total>{money_pattern})\s*(?P<description>.*)$",
          rest,
          flags=re.I | re.S,
        )

      if not row_match:
        continue

      paid = money(row_match.group("paid"))
      reserve = money(row_match.group("reserve"))
      total = money(row_match.group("total"))

      if total <= 0 and (paid or reserve):
        total = paid + reserve

      if total <= 0 and paid <= 0 and reserve <= 0:
        continue

      dedupe_key = f"{claim_number}|{policy_number}"
      if dedupe_key in seen:
        continue

      seen.add(dedupe_key)

      line_of_business = infer_line(policy_number, row_match.group("lob"))
      status = clean(row_match.group("status")).title()

      extracted_claims.append({
        "claim_number": claim_number,
        "Claim Number": claim_number,
        "claim_id": claim_number,
        "policy_number": policy_number,
        "Policy Number": policy_number,
        "line_of_business": line_of_business,
        "claim_type": line_of_business,
        "coverage": line_of_business,
        "policy_type": line_of_business,
        "date_of_loss": clean(row_match.group("loss")),
        "loss_date": clean(row_match.group("loss")),
        "date_reported": clean(row_match.groupdict().get("reported")),
        "date_closed": clean(row_match.groupdict().get("closed")),
        "status": status,
        "claim_status": status,
        "paid_amount": paid,
        "total_paid": paid,
        "reserve_amount": reserve,
        "total_reserve": reserve,
        "total_incurred": total,
        "description": clean(row_match.group("description")),
      })

  if extracted_claims:
    existing_total = 0.0
    for claim in parsed_claims:
      if isinstance(claim, dict):
        existing_total += money(claim.get("total_incurred") or claim.get("incurred"))

    extracted_total = sum(money(claim.get("total_incurred")) for claim in extracted_claims)

    if len(extracted_claims) >= len(parsed_claims or []) or extracted_total > existing_total:
      parsed_claims = extracted_claims
      parsed_profile["claims"] = extracted_claims
      parsed_profile["parsed_claims"] = extracted_claims
      parsed_profile["claim_count"] = len(extracted_claims)
      parsed_profile["total_claims"] = len(extracted_claims)
      parsed_profile["open_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "open")
      parsed_profile["closed_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "closed")
      parsed_profile["total_paid"] = sum(money(c.get("paid_amount")) for c in extracted_claims)
      parsed_profile["total_reserve"] = sum(money(c.get("reserve_amount")) for c in extracted_claims)
      parsed_profile["total_incurred"] = sum(money(c.get("total_incurred")) for c in extracted_claims)

  print("LOSSQ_PDF_CLEAN_TABLE_CLAIM_REPAIR_V1:", {
    "extracted_claims": len(extracted_claims),
    "final_claims": len(parsed_claims or []),
    "sample_claim_numbers": [c.get("claim_number") for c in (parsed_claims or [])[:5] if isinstance(c, dict)],
  })

  return parsed_claims, parsed_profile

# LOSSQ_CLAIM_NOTES_TO_DESCRIPTION_V1
# LOSSQ_PDF_FULL_CLAIM_BLOCK_EXTRACT_BEFORE_SAVE_V1
def lossq_pdf_full_claim_block_extract_before_save_v1(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal final PDF claim-block extractor.

  Purpose:
  - PDF uploads should not fall short when the loss run uses Claim Block sections.
  - Repairs claim numbers that absorb labels like POLICY-NUMBER.
  - Pulls insured/business name from labeled PDF text.
  - Replaces parsed_claims only when the PDF extraction finds more complete claim rows.
  """
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      val = float(raw or 0)
      return -val if neg else val
    except Exception:
      return 0.0

  def clean_claim_number(value):
    raw = clean(value)
    if not raw:
      return ""

    raw = re.sub(
      r"(?i)(?:\s|_|-)*(?:"
      r"POLICY\s*(?:NUMBER|NO|#)|"
      r"POLICY[-_]*(?:NUMBER|NO|#)|"
      r"LINE\s*OF\s*BUSINESS|"
      r"COVERAGE|"
      r"CLAIM\s*STATUS|"
      r"STATUS|"
      r"PAID\s*AMOUNT|"
      r"PAID|"
      r"RESERVE\s*AMOUNT|"
      r"RESERVE|"
      r"GROSS\s*INCURRED|"
      r"NET\s*INCURRED|"
      r"TOTAL\s*INCURRED|"
      r"INCURRED|"
      r"DATE\s*OF\s*LOSS|"
      r"LOSS\s*DATE|"
      r"DESCRIPTION"
      r").*$",
      "",
      raw,
    ).strip(" -_:/|")

    raw = re.split(r"(?i)POLICY", raw)[0].strip(" -_:/|")
    return raw

  def read_pdf_text():
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
      print("LOSSQ_PDF_FULL_CLAIM_BLOCK_EXTRACT_READ_ERROR_V1:", str(exc)[:200])
      return ""

  raw_text = read_pdf_text()
  if not raw_text:
    return parsed_claims, parsed_profile

  lines = [clean(line) for line in raw_text.splitlines()]
  lines = [line for line in lines if line]

  def label_from_lines(label_names):
    wanted = {key(item) for item in label_names}
    for idx, line in enumerate(lines):
      # same-line label
      m = re.match(r"^\s*([^:]{2,60})\s*:\s*(.+?)\s*$", line)
      if m and key(m.group(1)) in wanted:
        val = clean(m.group(2))
        if val:
          return val

      # label on one line, value on next line
      if key(line.rstrip(":")) in wanted:
        for j in range(idx + 1, min(idx + 5, len(lines))):
          candidate = clean(lines[j])
          if candidate and key(candidate.rstrip(":")) not in wanted:
            return candidate
    return ""

  # LOSSQ_PDF_SAME_LINE_CLAIM_BLOCK_VALUES_V1
  def values_from_block(block_lines):
    result = {}

    def set_result(field_key, value):
      value = clean(value)
      if value and not result.get(field_key):
        result[field_key] = value

    # Universal support for PDFs rendered from table blocks:
    # Claim Number X Policy Number Y
    # Line of Business X Claim Status Y
    # Date of Loss X Paid Amount Y
    # Reserve Amount X Total Incurred Y
    joined = clean(" ".join(block_lines))

    label_patterns = [
      ("claimnumber", r"Claim\s*(?:Number|No\.?|#|ID)"),
      ("policynumber", r"Policy\s*(?:Number|No\.?|#)"),
      ("lineofbusiness", r"Line\s+of\s+Business|Coverage|Policy\s+Type|LOB"),
      ("claimstatus", r"Claim\s+Status|Status"),
      ("dateofloss", r"Date\s+of\s+Loss|Loss\s+Date"),
      ("datereported", r"Date\s+Reported|Reported\s+Date"),
      ("dateclosed", r"Date\s+Closed|Closed\s+Date"),
      ("paidamount", r"Paid\s+Amount|Total\s+Paid|Paid"),
      ("reserveamount", r"Reserve\s+Amount|Total\s+Reserve|Reserve"),
      ("totalincurred", r"Total\s+Incurred|Gross\s+Incurred|Net\s+Incurred|Incurred"),
      ("causeofloss", r"Cause\s+of\s+Loss|Cause"),
      ("description", r"Claim\s+Notes|Loss\s+Notes|Claim\s+Description|Loss\s+Description|Description|Narrative|Notes"),
      ("litigation", r"Litigation|Litigated"),
    ]

    all_labels = "|".join(f"(?:{pattern})" for _, pattern in label_patterns)

    for field_key, label_pattern in label_patterns:
      pattern = re.compile(
        rf"(?i)\b(?:{label_pattern})\b\s*[:#-]?\s*(?P<value>.*?)(?=\s+(?:{all_labels})\b|$)"
      )
      match = pattern.search(joined)
      if match:
        set_result(field_key, match.group("value"))

    # LOSSQ_CLAIM_NOTES_FULL_SENTENCE_CAPTURE_V1
    # The generic label scanner can stop early when a note contains words like
    # "status", "reserve", or "policy". Claim notes should be captured as a
    # full narrative sentence, stopping only at the next claim/summary boundary.
    notes_match = re.search(
      r"(?is)\b(?:Claim\s+Notes|Loss\s+Notes|Claim\s+Narrative|Loss\s+Narrative|Underwriting\s+Notes|Notes|Narrative)\b\s*[:#-]?\s*(?P<value>.*?)(?=\s+(?:Loss\s+Summary|Total\s+Claims|Open\s+Claims|Closed\s+Claims|Total\s+Paid|Total\s+Reserve|Total\s+Incurred|Umbrella\s+Loss\s+Activity|Claim\s+Block\s+\d+)\b|$)",
      joined,
    )
    if notes_match:
      note_value = clean(notes_match.group("value"))
      if note_value:
        set_result("description", note_value)
        result["claim_notes"] = note_value
        result["loss_description"] = note_value

    # Preserve existing colon and next-line behavior.
    idx = 0
    known_label_keys = {field_key for field_key, _ in label_patterns}

    while idx < len(block_lines):
      line = clean(block_lines[idx])
      if not line:
        idx += 1
        continue

      same = re.match(r"^\s*([^:]{2,80})\s*:\s*(.+?)\s*$", line)
      if same:
        label_key = key(same.group(1))
        set_result(label_key, same.group(2))
        idx += 1
        continue

      label_key = key(line.rstrip(":"))
      if label_key in known_label_keys:
        for j in range(idx + 1, min(idx + 4, len(block_lines))):
          candidate = clean(block_lines[j])
          if not candidate:
            continue
          if key(candidate.rstrip(":")) in known_label_keys:
            break
          set_result(label_key, candidate)
          break

      idx += 1

    return result
  # Build claim blocks. Prefer explicit Claim Block sections.
  block_starts = []
  for i, line in enumerate(lines):
    if re.match(r"(?i)^claim\s+block\s+\d+\b", line):
      block_starts.append(i)

  blocks = []
  if block_starts:
    for pos, start in enumerate(block_starts):
      end = block_starts[pos + 1] if pos + 1 < len(block_starts) else len(lines)
      blocks.append(lines[start:end])
  else:
    # Fallback for PDFs without Claim Block headings: split by Claim Number labels.
    claim_number_starts = []
    for i, line in enumerate(lines):
      if key(line.rstrip(":")) == "claimnumber" or re.match(r"(?i)^claim\s*number\s*:", line):
        claim_number_starts.append(i)
    for pos, start in enumerate(claim_number_starts):
      end = claim_number_starts[pos + 1] if pos + 1 < len(claim_number_starts) else len(lines)
      blocks.append(lines[start:end])

  extracted_claims = []
  seen = set()

  for block in blocks:
    data = values_from_block(block)

    claim_number = clean_claim_number(
      data.get("claimnumber")
      or data.get("claimno")
      or data.get("claim")
      or data.get("claimid")
    )

    policy_number = clean(
      data.get("policynumber")
      or data.get("policyno")
      or data.get("policy")
    )

    if not claim_number or not policy_number:
      continue

    line_of_business = clean(
      data.get("lineofbusiness")
      or data.get("coverage")
      or data.get("policytype")
      or data.get("lob")
    )

    paid = money(data.get("paid") or data.get("paidamount"))
    reserve = money(data.get("reserve") or data.get("reserveamount"))
    total = money(
      data.get("totalincurred")
      or data.get("grossincurred")
      or data.get("netincurred")
      or data.get("incurred")
    )
    if not total and (paid or reserve):
      total = paid + reserve

    status = clean(data.get("status") or data.get("claimstatus"))
    if status:
      if status.lower() in {"open", "opened"}:
        status = "Open"
      elif status.lower() in {"closed", "close"}:
        status = "Closed"

    dedupe_key = f"{claim_number.upper()}|{policy_number.upper()}"
    if dedupe_key in seen:
      continue
    seen.add(dedupe_key)

    claim = {
      "claim_number": claim_number,
      "Claim Number": claim_number,
      "policy_number": policy_number,
      "Policy Number": policy_number,
      "line_of_business": line_of_business,
      "claim_type": line_of_business,
      "coverage": line_of_business,
      "status": status,
      "claim_status": status,
      "date_of_loss": clean(data.get("dateofloss") or data.get("lossdate")),
      "date_reported": clean(data.get("datereported") or data.get("reporteddate")),
      "date_closed": clean(data.get("dateclosed") or data.get("closeddate")),
      "paid_amount": paid,
      "total_paid": paid,
      "reserve_amount": reserve,
      "total_reserve": reserve,
      "total_incurred": total,
      "cause_of_loss": clean(data.get("causeofloss") or data.get("cause")),
      "description": clean(data.get("description") or data.get("claimdescription")),
      "litigation": clean(data.get("litigation")),
    }

    extracted_claims.append(claim)

  # Pull profile identity from PDF labels.
  business_name = label_from_lines(["Named Insured", "Business Name", "Insured", "Account Name", "Applicant"])
  if business_name and key(business_name) not in {"businessnamenotset", "notset", "unknown", "na", "none"}:
    parsed_profile["business_name"] = business_name
    parsed_profile["named_insured"] = business_name
    parsed_profile["insured_name"] = business_name
    parsed_profile["account_name"] = business_name

  # If the PDF extraction found more rows, trust it. This fixes missed claim blocks.
  if extracted_claims and len(extracted_claims) >= len(parsed_claims or []):
    parsed_claims = extracted_claims
    parsed_profile["claims"] = extracted_claims
    parsed_profile["parsed_claims"] = extracted_claims
    parsed_profile["claim_count"] = len(extracted_claims)
    parsed_profile["total_claims"] = len(extracted_claims)
    parsed_profile["open_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "open")
    parsed_profile["closed_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "closed")
    parsed_profile["total_paid"] = sum(money(c.get("paid_amount")) for c in extracted_claims)
    parsed_profile["total_reserve"] = sum(money(c.get("reserve_amount")) for c in extracted_claims)
    parsed_profile["total_incurred"] = sum(money(c.get("total_incurred")) for c in extracted_claims)

  # LOSSQ_PDF_SAME_LINE_CLAIM_BLOCK_PROFILE_POLICY_REPAIR_V1
  def pdf_label_value(label_names):
    labels = "|".join(re.escape(item) for item in label_names)
    patterns = [
      rf"(?im)^\s*(?:{labels})\s*(?:[:#-])\s*(.+?)\s*$",
      rf"(?is)\b(?:{labels})\b\s*(?:[:#-])?\s*\n\s*(.+?)\s*(?:\n|$)",
    ]

    for pattern in patterns:
      match = re.search(pattern, raw_text)
      if match:
        value = clean(match.group(1))
        if value:
          return value

    return ""

  carrier_name = pdf_label_value(["Writing Carrier", "Carrier", "Carrier Name", "Insurance Carrier"])
  if carrier_name and key(carrier_name) not in {"carrier", "writingcarrier", "notset", "unknown", "na", "none"}:
    parsed_profile["carrier_name"] = carrier_name
    parsed_profile["writing_carrier"] = carrier_name
    parsed_profile["carrier"] = carrier_name

  def normalize_policy_date(value):
    raw = clean(value)
    match = re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", raw)
    if not match:
      return raw
    parts = re.split(r"[/-]", match.group(0))
    if len(parts) != 3:
      return raw
    month, day, year = parts
    if len(year) == 2:
      year = "20" + year
    try:
      return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    except Exception:
      return raw

  def line_from_policy_prefix(policy_number, fallback=""):
    fallback = clean(fallback)
    if fallback:
      return fallback

    raw = clean(policy_number).upper()
    prefix = raw.split("-")[0] if "-" in raw else raw.split("_")[0] if "_" in raw else raw

    mapping = {
      "GL": "General Liability",
      "CGL": "General Liability",
      "WC": "Workers Compensation",
      "BOP": "Businessowners Policy",
      "UMB": "Umbrella",
      "UM": "Umbrella",
      "EXCESS": "Umbrella",
      "LIQ": "Liquor Liability",
      "LIQUOR": "Liquor Liability",
      "AUTO": "Commercial Auto",
      "CA": "Commercial Auto",
      "GAR": "Garage Liability",
      "DOL": "Dealers Open Lot",
      "CP": "Commercial Property",
      "PROP": "Property",
      "CY": "Cyber Liability",
      "CYBER": "Cyber Liability",
      "PL": "Professional Liability",
      "EPLI": "Employment Practices Liability",
      "DO": "Directors & Officers",
      "DNO": "Directors & Officers",
      "IM": "Inland Marine",
      "CARGO": "Motor Truck Cargo",
    }
    return mapping.get(prefix, "")

  policy_id_pattern = r"[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,9}"
  date_pattern = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
  money_pattern = r"\$?\(?-?\d[\d,]*(?:\.\d{1,2})?\)?"

  def policy_like(value):
    raw = clean(value).upper()
    if not raw or not re.search(r"\d", raw):
      return False
    if any(token in raw for token in ["ACCT", "ACCOUNT", "CUSTOMER", "CUST", "CLIENT"]):
      return False
    return bool(re.search(r"^[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,9}$", raw))

  schedule_rows = []
  schedule_match = re.search(
    r"(?is)\bPOLICY\s+SCHEDULE\b[\s\S]*?(?=\bEXPOSURE\b|\bCLAIM\s+DETAIL\b|\bCLAIMS\s+DETAIL\b|\bLOSS\s+SUMMARY\b|$)",
    raw_text,
  )

  if schedule_match:
    for schedule_line in [clean(item) for item in schedule_match.group(0).splitlines() if clean(item)]:
      if re.search(r"(?i)^policy\s*(?:#|number|no\.)", schedule_line):
        continue

      row = re.match(rf"^(?P<policy>{policy_id_pattern})\s+(?P<rest>.+)$", schedule_line)
      if not row:
        continue

      policy_number = clean(row.group("policy")).upper()
      if not policy_like(policy_number):
        continue

      rest = clean(row.group("rest"))
      date_match = re.search(
        rf"(?P<eff>{date_pattern})\s+(?P<exp>{date_pattern})(?:\s+(?P<premium>{money_pattern}))?",
        rest,
      )

      if date_match:
        line_of_business = line_from_policy_prefix(policy_number, rest[:date_match.start()])
        effective_date = normalize_policy_date(date_match.group("eff"))
        expiration_date = normalize_policy_date(date_match.group("exp"))
        current_premium = money(date_match.group("premium"))
      else:
        line_of_business = line_from_policy_prefix(policy_number, rest)
        effective_date = ""
        expiration_date = ""
        current_premium = 0.0

      schedule_rows.append({
        "policy_number": policy_number,
        "policy_type": line_of_business,
        "line_of_business": line_of_business,
        "coverage": line_of_business,
        "carrier": carrier_name,
        "carrier_name": carrier_name,
        "writing_carrier": carrier_name,
        "effective_date": effective_date,
        "expiration_date": expiration_date,
        "current_premium": current_premium,
        "claim_count": 0,
        "total_incurred": 0.0,
      })

  final_claims = parsed_claims if isinstance(parsed_claims, list) else []
  claim_counts = {}
  claim_totals = {}

  for claim in final_claims:
    if not isinstance(claim, dict):
      continue

    policy_number = clean(claim.get("policy_number") or claim.get("Policy Number")).upper()
    if not policy_like(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
    claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + money(claim.get("total_incurred"))

  if schedule_rows:
    existing = parsed_profile.get("policy_schedule") or parsed_profile.get("policies") or []
    merged = {}

    if isinstance(existing, list):
      for item in existing:
        if not isinstance(item, dict):
          continue

        policy_number = clean(item.get("policy_number") or item.get("Policy Number")).upper()
        if policy_like(policy_number):
          merged[policy_number] = dict(item)

    for row in schedule_rows:
      policy_number = row["policy_number"]
      target = merged.get(policy_number, {})
      target.update({k: v for k, v in row.items() if v not in ("", None)})
      target["claim_count"] = claim_counts.get(policy_number, 0)
      target["total_incurred"] = claim_totals.get(policy_number, 0.0)

      if carrier_name:
        target["carrier"] = carrier_name
        target["carrier_name"] = carrier_name
        target["writing_carrier"] = carrier_name

      merged[policy_number] = target

    policies = list(merged.values())
    parsed_profile["policy_schedule"] = policies
    parsed_profile["policies"] = policies
    parsed_profile["policy_numbers"] = [item.get("policy_number") for item in policies if item.get("policy_number")]
  print("LOSSQ_PDF_FULL_CLAIM_BLOCK_EXTRACT_BEFORE_SAVE_V1:", {
    "existing_claims": len(parsed_profile.get("claims") or []),
    "extracted_claims": len(extracted_claims),
    "final_claims": len(parsed_claims or []),
    "business_name": parsed_profile.get("business_name"),
    "sample_claim_numbers": [c.get("claim_number") for c in (parsed_claims or [])[:5] if isinstance(c, dict)],
  })

  return parsed_claims, parsed_profile


# LOSSQ_PDF_VERTICAL_POLICY_SECTION_REPAIR_V1
def lossq_pdf_vertical_policy_section_repair_v1(file_path, parsed_claims=None, parsed_profile=None, direct_profile=None):
  """PDF-only repair for vertical carrier policy-section PDFs."""
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", " ").strip())

  def money(value):
    raw = clean(value).replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
      return 0.0
    try:
      return float(match.group(0))
    except Exception:
      return 0.0

  def norm_date(value):
    raw = clean(value)
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", raw)
    if not match:
      return raw
    month, day, year = match.groups()
    if len(year) == 2:
      year = "20" + year
    return f"{int(month):02d}/{int(day):02d}/{year}"

  def money_text(value):
    amount = money(value)
    if amount <= 0:
      return "0"
    return str(int(amount)) if float(amount).is_integer() else f"{amount:.2f}"

  try:
    from pypdf import PdfReader
    raw_text = "\n".join((page.extract_text() or "") for page in PdfReader(file_path).pages)
  except Exception as exc:
    print("LOSSQ_PDF_VERTICAL_POLICY_SECTION_REPAIR_READ_ERROR_V1:", str(exc)[:200])
    return parsed_claims, parsed_profile, direct_profile

  if "POLICY SECTION" not in raw_text or "Policy Number" not in raw_text:
    return parsed_claims, parsed_profile, direct_profile

  def put_profile(key, value, aliases=()):
    value = clean(value)
    if not value:
      return
    parsed_profile[key] = value
    direct_profile[key] = value
    for alias in aliases:
      parsed_profile[alias] = value
      direct_profile[alias] = value

  insured = re.search(r"(?is)\bInsured\s+Name\s*:\s*(.+?)(?=\n|\s+Account\s+Number)", raw_text)
  if insured:
    put_profile("business_name", insured.group(1), ("insured_name", "named_insured", "account_name"))

  account = re.search(r"(?is)\bAccount\s+Number\s*:\s*([A-Z0-9._/-]+)", raw_text)
  if account:
    put_profile("account_number", account.group(1), ("account_no", "account"))

  customer = re.search(r"(?is)\bCustomer\s+Number\s*:\s*([A-Z0-9._/-]+)", raw_text)
  if customer:
    put_profile("customer_number", customer.group(1))

  carrier = re.search(r"(?is)\bCarrier\s*:\s*(.+?)(?=\n|\s+Producer)", raw_text)
  if carrier:
    put_profile("carrier_name", carrier.group(1), ("writing_carrier", "carrier"))

  evaluation = re.search(r"(?is)\bEvaluation\s+Date\s*:\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", raw_text)
  if evaluation:
    put_profile("evaluation_date", norm_date(evaluation.group(1)), ("valuation_date",))

  current_premium = re.search(r"(?is)\bCurrent\s+Premium\s*:\s*(\$?\s*\d[\d,]*(?:\.\d{2})?)", raw_text)
  if current_premium:
    put_profile("current_premium", money_text(current_premium.group(1)), ("expiring_premium",))

  target_premium = re.search(r"(?is)\bTarget\s+Renewal\s+Premium\s*:\s*(\$?\s*\d[\d,]*(?:\.\d{2})?)", raw_text)
  if target_premium:
    put_profile("target_renewal_premium", money_text(target_premium.group(1)))

  exposure_notes = re.search(r"(?is)\bExposure\s+Notes\s*:\s*(.+?)(?=\n\s*POLICY\s+SECTION|$)", raw_text)
  if exposure_notes:
    exposure_text = clean(exposure_notes.group(1))
    beds = re.search(r"(?i)\b(\d{1,6})\s+beds?\b", exposure_text)
    payroll = re.search(r"(?i)\bpayroll\s*\$?\s*([0-9][0-9,]*(?:\.\d{2})?)", exposure_text)
    revenue = re.search(r"(?i)\bannual\s+revenue\s*\$?\s*([0-9][0-9,]*(?:\.\d{2})?)", exposure_text)
    employees = re.search(r"(?i)\bemployee\s+count\s*(\d{1,6})", exposure_text)
    locations = re.search(r"(?i)\blocation\s+count\s*(\d{1,6})", exposure_text)
    if beds:
      put_profile("bed_count", beds.group(1), ("beds",))
    if payroll:
      put_profile("payroll", payroll.group(1).replace(',', ''))
    if revenue:
      put_profile("revenue", revenue.group(1).replace(',', ''), ("annual_revenue", "sales"))
    if employees:
      put_profile("employee_count", employees.group(1))
    if locations:
      put_profile("location_count", locations.group(1), ("locations", "locationCount"))

  policy_re = re.compile(r"(?is)\bPOLICY\s+SECTION\s+\d+\s*:\s*(.+?)(?=\n\s*POLICY\s+SECTION\s+\d+\s*:|\Z)")
  claim_re = re.compile(
    r"(?is)\bClaim\s+([A-Z0-9][A-Z0-9./-]{2,45})\s+Status\s+([A-Za-z]+)\s+"
    r"Claimant\s+(.+?)\s+Jurisdiction\s+([A-Z]{2})\s+"
    r"Loss\s+Date\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+Reported\s+Date\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+"
    r"Closed\s+Date\s+(OPEN|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+Adjuster\s+(.+?)\s+"
    r"Paid\s+(\$?\s*\d[\d,]*(?:\.\d{2})?)\s+Reserve\s+(\$?\s*\d[\d,]*(?:\.\d{2})?)\s+"
    r"Total\s+Incurred\s+(\$?\s*\d[\d,]*(?:\.\d{2})?)\s+Description\s+(.+?)(?=\s+Claim\s+[A-Z0-9][A-Z0-9./-]{2,45}\s+Status\b|\s+POLICY\s+SECTION\s+\d+\s*:|\s+NO\s+CLAIMS\s+REPORTED|\Z)"
  )

  policies = []
  extracted_claims = []

  for section_match in policy_re.finditer(raw_text):
    section = section_match.group(1)
    section_title = clean(section.splitlines()[0] if section.splitlines() else "")

    policy_number_match = re.search(r"(?is)\bPolicy\s+Number\s+([A-Z0-9][A-Z0-9./-]{2,60})", section)
    line_match = re.search(r"(?is)\bLine\s+of\s+Business\s+(.+?)(?=\s+Policy\s+Term|\n)", section)
    term_match = re.search(r"(?is)\bPolicy\s+Term\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*-\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", section)
    premium_match = re.search(r"(?is)\bPremium\s+(\$?\s*\d[\d,]*(?:\.\d{2})?)", section)
    section_total_match = re.search(r"(?is)\bTotal\s+Incurred\s+(\$?\s*\d[\d,]*(?:\.\d{2})?)", section)

    policy_number = clean(policy_number_match.group(1)).upper() if policy_number_match else ""
    line_name = clean(line_match.group(1)) if line_match else section_title

    if policy_number:
      policies.append({
        "policy_number": policy_number,
        "policy_type": line_name,
        "line_of_business": line_name,
        "coverage": line_name,
        "carrier": parsed_profile.get("carrier_name", ""),
        "carrier_name": parsed_profile.get("carrier_name", ""),
        "writing_carrier": parsed_profile.get("writing_carrier", parsed_profile.get("carrier_name", "")),
        "effective_date": norm_date(term_match.group(1)) if term_match else parsed_profile.get("effective_date", ""),
        "expiration_date": norm_date(term_match.group(2)) if term_match else parsed_profile.get("expiration_date", ""),
        "premium": money_text(premium_match.group(1)) if premium_match else "",
        "current_premium": money_text(premium_match.group(1)) if premium_match else "",
        "claim_count": 0,
        "total_incurred": money(section_total_match.group(1)) if section_total_match else 0.0,
      })

    for claim_match in claim_re.finditer(section):
      claim_number = clean(claim_match.group(1)).upper()
      status = clean(claim_match.group(2))
      claimant = clean(claim_match.group(3))
      state = clean(claim_match.group(4)).upper()
      date_of_loss = norm_date(claim_match.group(5))
      date_reported = norm_date(claim_match.group(6))
      closed_raw = clean(claim_match.group(7))
      date_closed = "" if closed_raw.upper() == "OPEN" else norm_date(closed_raw)
      adjuster = clean(claim_match.group(8))
      paid = money(claim_match.group(9))
      reserve = money(claim_match.group(10))
      incurred = money(claim_match.group(11)) or paid + reserve
      description = clean(claim_match.group(12))

      extracted_claims.append({
        "business_name": parsed_profile.get("business_name", ""),
        "named_insured": parsed_profile.get("named_insured", ""),
        "insured_name": parsed_profile.get("insured_name", ""),
        "carrier_name": parsed_profile.get("carrier_name", ""),
        "writing_carrier": parsed_profile.get("writing_carrier", parsed_profile.get("carrier_name", "")),
        "policy_number": policy_number,
        "policy": policy_number,
        "policy_type": line_name,
        "line_of_business": line_name,
        "claim_type": line_name,
        "coverage": line_name,
        "claim_number": claim_number,
        "claim_no": claim_number,
        "status": status,
        "claim_status": status,
        "claimant": claimant,
        "claimant_name": claimant,
        "jurisdiction": state,
        "state": state,
        "date_of_loss": date_of_loss,
        "loss_date": date_of_loss,
        "date_reported": date_reported,
        "reported_date": date_reported,
        "date_closed": date_closed,
        "closed_date": date_closed,
        "paid_amount": paid,
        "paid": paid,
        "reserve_amount": reserve,
        "reserve": reserve,
        "total_incurred": incurred,
        "incurred": incurred,
        "total": incurred,
        "adjuster": adjuster,
        "examiner": adjuster,
        "description": description,
        "loss_description": description,
      })

  if policies:
    stats = {}
    for claim in extracted_claims:
      policy_key = clean(claim.get("policy_number")).upper()
      stats.setdefault(policy_key, {"claim_count": 0, "total_incurred": 0.0})
      stats[policy_key]["claim_count"] += 1
      stats[policy_key]["total_incurred"] += money(claim.get("total_incurred"))

    for policy in policies:
      policy_key = clean(policy.get("policy_number")).upper()
      if policy_key in stats:
        policy["claim_count"] = stats[policy_key]["claim_count"]
        policy["total_incurred"] = stats[policy_key]["total_incurred"]

    parsed_profile["policies"] = policies
    parsed_profile["policy_schedule"] = policies
    parsed_profile["policy_numbers"] = [p.get("policy_number") for p in policies if p.get("policy_number")]
    if parsed_profile.get("policy_numbers") and not parsed_profile.get("policy_number"):
      parsed_profile["policy_number"] = parsed_profile["policy_numbers"][0]
      parsed_profile["main_policy"] = parsed_profile["policy_numbers"][0]

    direct_profile["policies"] = policies
    direct_profile["policy_schedule"] = policies
    direct_profile["policy_numbers"] = parsed_profile.get("policy_numbers")

  existing_total = sum(money(c.get("total_incurred") or c.get("incurred")) for c in parsed_claims if isinstance(c, dict))
  extracted_total = sum(money(c.get("total_incurred") or c.get("incurred")) for c in extracted_claims)

  if extracted_claims and (len(extracted_claims) >= len(parsed_claims or []) or extracted_total > existing_total):
    parsed_claims = extracted_claims
    parsed_profile["claims"] = extracted_claims
    parsed_profile["parsed_claims"] = extracted_claims
    parsed_profile["claim_count"] = len(extracted_claims)
    parsed_profile["total_claims"] = len(extracted_claims)
    parsed_profile["open_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "open")
    parsed_profile["closed_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "closed")
    parsed_profile["total_paid"] = sum(money(c.get("paid_amount")) for c in extracted_claims)
    parsed_profile["total_reserve"] = sum(money(c.get("reserve_amount")) for c in extracted_claims)
    parsed_profile["total_incurred"] = extracted_total
    direct_profile["claims"] = extracted_claims
    direct_profile["parsed_claims"] = extracted_claims
    direct_profile["claim_count"] = len(extracted_claims)
    direct_profile["total_claims"] = len(extracted_claims)
    direct_profile["open_claims"] = parsed_profile["open_claims"]
    direct_profile["closed_claims"] = parsed_profile["closed_claims"]
    direct_profile["total_paid"] = parsed_profile["total_paid"]
    direct_profile["total_reserve"] = parsed_profile["total_reserve"]
    direct_profile["total_incurred"] = extracted_total

  print("LOSSQ_PDF_VERTICAL_POLICY_SECTION_REPAIR_V1:", {
    "policies": len(policies),
    "claims": len(extracted_claims),
    "final_claims": len(parsed_claims or []),
    "total_incurred": parsed_profile.get("total_incurred"),
    "claim_numbers": [c.get("claim_number") for c in (parsed_claims or [])[:10] if isinstance(c, dict)],
  })

  return parsed_claims, parsed_profile, direct_profile


# LOSSQ_PDF_MESSY_BLOCK_MINI_REPAIR_V1
def lossq_pdf_messy_block_mini_repair_v1(file_path, parsed_claims=None, parsed_profile=None, direct_profile=None):
  """
  PDF-only universal overlay for messy claim-block loss runs.

  It reads labeled evidence from the uploaded PDF and only replaces claims/profile
  when it extracts stronger claim-block results than the current parser.
  """
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", " ").strip())

  def money(value):
    raw = clean(value).replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
      return 0.0
    try:
      return float(match.group(0))
    except Exception:
      return 0.0


  def money_text(value):
    amount = money(value)
    if amount <= 0:
      return ""
    if float(amount).is_integer():
      return str(int(amount))
    return f"{amount:.2f}"

  def plausible_premium(value):
    amount = money(value)
    return amount > 0 and amount < 100000000

  def norm_date(value):
    raw = clean(value)
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", raw)
    if not match:
      return raw
    month, day, year = match.groups()
    if len(year) == 2:
      year = "20" + year
    return f"{int(month):02d}/{int(day):02d}/{year}"

  try:
    from pypdf import PdfReader
    raw_text = "\n".join((page.extract_text() or "") for page in PdfReader(file_path).pages)
  except Exception as exc:
    print("LOSSQ_PDF_MESSY_BLOCK_MINI_REPAIR_READ_ERROR_V1:", str(exc)[:200])
    return parsed_claims, parsed_profile, direct_profile

  if not clean(raw_text):
    return parsed_claims, parsed_profile, direct_profile

  if "Claim Number" not in raw_text or "Line / Coverage" not in raw_text:
    return parsed_claims, parsed_profile, direct_profile

  def put_profile(key, value, aliases=()):
    value = clean(value)
    if not value:
      return
    parsed_profile[key] = value
    direct_profile[key] = value
    for alias in aliases:
      parsed_profile[alias] = value
      direct_profile[alias] = value

  def title_carrier(value):
    value = clean(value).strip(" :-|")
    if not value:
      return ""
    value = value.title()
    for old, new in {
      " Llc": " LLC",
      " Inc": " Inc.",
      " Co": " Co.",
      " Ltd": " Ltd.",
      " Usa": " USA",
    }.items():
      value = value.replace(old, new)
    return value

  heading = re.search(
    r"(?im)^\s*([A-Z][A-Z0-9&.,' -]{4,120}?)\s*-\s*(?:LOSS|CLAIM|CLAIMS)\s+(?:EXPERIENCE|RUN|SUMMARY)\b",
    raw_text,
  )
  if heading:
    put_profile("carrier_name", title_carrier(heading.group(1)), ("writing_carrier", "carrier"))

  insured = re.search(r"(?is)\bNamed\s+Insured\s*:\s*(.+?)(?=\s+Account\s*#|\n)", raw_text)
  if insured:
    put_profile("business_name", clean(insured.group(1)), ("insured_name", "named_insured", "account_name"))

  account = re.search(r"(?is)\bAccount\s*#\s*:\s*([A-Z0-9./_-]+)", raw_text)
  if account:
    put_profile("account_number", clean(account.group(1)), ("account_no", "account"))

  customer = re.search(r"(?is)\bCustomer\s+No\s*:\s*([A-Z0-9./_-]+)", raw_text)
  if customer:
    put_profile("customer_number", clean(customer.group(1)))

  agency = re.search(r"(?is)\bProducing\s+Agency\s*:\s*(.+?)(?=\s+Evaluation\s+Date|\n)", raw_text)
  if agency:
    put_profile("producing_agency", clean(agency.group(1)), ("agency_name", "producer"))

  evaluation_date = re.search(r"(?is)\bEvaluation\s+Date\s*:\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", raw_text)
  if evaluation_date:
    put_profile("evaluation_date", norm_date(evaluation_date.group(1)), ("valuation_date",))

  term = re.search(
    r"(?is)\bPolicy\s+Term\s*:\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+to\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    raw_text,
  )
  if term:
    put_profile("effective_date", norm_date(term.group(1)), ("policy_effective_date",))
    put_profile("expiration_date", norm_date(term.group(2)), ("policy_expiration_date",))

  current_premium = re.search(
    r"(?is)\bCurrent\s+Premium\s*:\s*(\$?\s*\d[\d,]*(?:\.\d{2})?)",
    raw_text,
  )
  if current_premium:
    put_profile("current_premium", money_text(current_premium.group(1)), ("expiring_premium",))

  target_renewal_premium = re.search(
    r"(?is)\bTarget\s+Renewal\s+Premium\s*:\s*(\$?\s*\d[\d,]*(?:\.\d{2})?)",
    raw_text,
  )
  if target_renewal_premium:
    put_profile("target_renewal_premium", money_text(target_renewal_premium.group(1)))

  carrier_name = parsed_profile.get("carrier_name") or direct_profile.get("carrier_name") or ""

  schedule_match = re.search(
    r"(?is)\bPolicy\s+Schedule\b(.+?)(?=\n\s*(?:Carrier\s+comment|CLAIM\s+DETAIL|Claim\s+Number)\b)",
    raw_text,
  )
  schedule_text = schedule_match.group(1) if schedule_match else ""

  policy_re = re.compile(
    r"(?im)^\s*([A-Za-z][A-Za-z &'/-]{2,70}?)\s+([A-Z]{1,12}-[A-Z0-9-]{3,})\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*-\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(\$?[\d,]+(?:\.\d{2})?)"
  )

  policies = []
  seen_policies = set()

  for match in policy_re.finditer(schedule_text):
    line_name = clean(match.group(1))
    policy_number = clean(match.group(2)).upper()

    if policy_number in seen_policies:
      continue

    seen_policies.add(policy_number)

    policies.append({
      "policy_number": policy_number,
      "policy_type": line_name,
      "line_of_business": line_name,
      "coverage": line_name,
      "carrier": carrier_name,
      "carrier_name": carrier_name,
      "writing_carrier": carrier_name,
      "effective_date": norm_date(match.group(3)),
      "expiration_date": norm_date(match.group(4)),
      "premium": clean(match.group(5)),
      "current_premium": clean(match.group(5)),
      "claim_count": 0,
      "total_incurred": 0.0,
    })

  policy_premium_total = sum(money(policy.get("premium")) for policy in policies if isinstance(policy, dict))
  if policy_premium_total and not plausible_premium(parsed_profile.get("current_premium")):
    put_profile("current_premium", money_text(policy_premium_total), ("expiring_premium",))

  claim_re = re.compile(
    r"(?is)\bClaim\s+Number\s+([A-Z0-9][A-Z0-9./-]{2,45})\s+Policy\s+([A-Z0-9][A-Z0-9./-]{2,45})"
  )

  starts = list(claim_re.finditer(raw_text))
  extracted_claims = []

  for index, match in enumerate(starts):
    claim_number = clean(match.group(1)).upper()
    policy_number = clean(match.group(2)).upper()

    if claim_number == policy_number:
      continue
    if not re.search(r"\d", claim_number):
      continue
    if re.search(r"\b(POLICY|CLAIM|NUMBER|LINE|COVERAGE|STATUS|SUMMARY|TOTAL)\b", claim_number):
      continue

    end = starts[index + 1].start() if index + 1 < len(starts) else len(raw_text)
    body = raw_text[match.end():end]

    line_status = re.search(
      r"(?is)\bLine\s*/\s*Coverage\s+(.+?)\s+Status\s+(.+?)(?=\s+Claimant\b|\n)",
      body,
    )
    line_name = clean(line_status.group(1)) if line_status else ""
    status = clean(line_status.group(2)) if line_status else ""

    claimant_state = re.search(r"(?is)\bClaimant\s+(.+?)\s+Jurisdiction\s+([A-Z]{2})\b", body)
    claimant = clean(claimant_state.group(1)) if claimant_state else ""
    state = clean(claimant_state.group(2)).upper() if claimant_state else ""

    dates = re.search(
      r"(?is)\bDate\s+of\s+Loss\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+Reported\s*/\s*Closed\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*/\s*(OPEN|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
      body,
    )
    date_of_loss = norm_date(dates.group(1)) if dates else ""
    date_reported = norm_date(dates.group(2)) if dates else ""
    closed_raw = clean(dates.group(3)) if dates else ""
    date_closed = "" if closed_raw.upper() == "OPEN" else norm_date(closed_raw)

    amounts = re.search(
      r"(?is)\bPaid\s+(\$?[\d,]+(?:\.\d{2})?)\s+Reserve\s*/\s*Incurred\s+(\$?[\d,]+(?:\.\d{2})?)\s*/\s*(\$?[\d,]+(?:\.\d{2})?)",
      body,
    )
    paid = money(amounts.group(1)) if amounts else 0.0
    reserve = money(amounts.group(2)) if amounts else 0.0
    incurred = money(amounts.group(3)) if amounts else paid + reserve

    examiner = ""
    examiner_match = re.search(r"(?is)\bExaminer\s+(.+?)\s+Description\b", body)
    if examiner_match:
      examiner = clean(examiner_match.group(1))

    description = ""
    description_match = re.search(
      r"(?is)\bDescription\s+(.+?)(?=\n\s*(?:CLAIM\s+DETAIL\s+BLOCK|LOSS\s+SUMMARY)\b|$)",
      body,
    )
    if description_match:
      description = clean(description_match.group(1))

    extracted_claims.append({
      "business_name": parsed_profile.get("business_name", ""),
      "named_insured": parsed_profile.get("named_insured", ""),
      "insured_name": parsed_profile.get("insured_name", ""),
      "carrier_name": carrier_name,
      "writing_carrier": carrier_name,
      "producing_agency": parsed_profile.get("producing_agency", ""),
      "policy_number": policy_number,
      "policy": policy_number,
      "policy_type": line_name,
      "line_of_business": line_name,
      "claim_type": line_name,
      "coverage": line_name,
      "claim_number": claim_number,
      "claim_no": claim_number,
      "status": status,
      "claim_status": status,
      "claimant": claimant,
      "claimant_name": claimant,
      "jurisdiction": state,
      "state": state,
      "date_of_loss": date_of_loss,
      "loss_date": date_of_loss,
      "date_reported": date_reported,
      "reported_date": date_reported,
      "date_closed": date_closed,
      "closed_date": date_closed,
      "paid_amount": paid,
      "paid": paid,
      "reserve_amount": reserve,
      "reserve": reserve,
      "total_incurred": incurred,
      "incurred": incurred,
      "total": incurred,
      "adjuster": examiner,
      "examiner": examiner,
      "description": description,
      "loss_description": description,
    })

  if policies:
    stats = {}
    for claim in extracted_claims:
      policy_key = clean(claim.get("policy_number")).upper()
      stats.setdefault(policy_key, {"claim_count": 0, "total_incurred": 0.0})
      stats[policy_key]["claim_count"] += 1
      stats[policy_key]["total_incurred"] += money(claim.get("total_incurred"))

    for policy in policies:
      policy_key = clean(policy.get("policy_number")).upper()
      if policy_key in stats:
        policy["claim_count"] = stats[policy_key]["claim_count"]
        policy["total_incurred"] = stats[policy_key]["total_incurred"]

    parsed_profile["policies"] = policies
    parsed_profile["policy_schedule"] = policies
    parsed_profile["policy_numbers"] = [p.get("policy_number") for p in policies if p.get("policy_number")]
    direct_profile["policies"] = policies
    direct_profile["policy_schedule"] = policies
    direct_profile["policy_numbers"] = parsed_profile["policy_numbers"]

  existing_total = sum(money(c.get("total_incurred") or c.get("incurred")) for c in parsed_claims if isinstance(c, dict))
  extracted_total = sum(money(c.get("total_incurred") or c.get("incurred")) for c in extracted_claims)

  if extracted_claims and (len(extracted_claims) >= len(parsed_claims or []) or extracted_total > existing_total):
    parsed_claims = extracted_claims
    parsed_profile["claims"] = extracted_claims
    parsed_profile["parsed_claims"] = extracted_claims
    parsed_profile["claim_count"] = len(extracted_claims)
    parsed_profile["total_claims"] = len(extracted_claims)
    parsed_profile["open_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "open")
    parsed_profile["closed_claims"] = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "closed")
    parsed_profile["total_paid"] = sum(money(c.get("paid_amount")) for c in extracted_claims)
    parsed_profile["total_reserve"] = sum(money(c.get("reserve_amount")) for c in extracted_claims)
    parsed_profile["total_incurred"] = extracted_total

    direct_profile["claims"] = extracted_claims
    direct_profile["parsed_claims"] = extracted_claims
    direct_profile["claim_count"] = len(extracted_claims)
    direct_profile["total_claims"] = len(extracted_claims)
    direct_profile["open_claims"] = parsed_profile["open_claims"]
    direct_profile["closed_claims"] = parsed_profile["closed_claims"]
    direct_profile["total_paid"] = parsed_profile["total_paid"]
    direct_profile["total_reserve"] = parsed_profile["total_reserve"]
    direct_profile["total_incurred"] = extracted_total

  print("LOSSQ_PDF_MESSY_BLOCK_MINI_REPAIR_V1:", {
    "business_name": parsed_profile.get("business_name"),
    "carrier_name": parsed_profile.get("carrier_name"),
    "producing_agency": parsed_profile.get("producing_agency"),
    "policies": len(policies),
    "claims": len(parsed_claims or []),
    "claim_numbers": [c.get("claim_number") for c in (parsed_claims or [])[:10] if isinstance(c, dict)],
  })

  return parsed_claims, parsed_profile, direct_profile



# LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_V2
def lossq_pdf_wide_claims_table_final_save_rescue_v2(file_path, parsed_claims=None, parsed_profile=None, direct_profile=None):
  """
  Final universal PDF claims-table rescue immediately before DB save.

  Adds value without replacing the existing parser:
  - only activates for PDFs with a real CLAIMS DETAIL table
  - extracts row-based claim data from text-readable wide tables
  - does not hardcode company names, carriers, provinces, claim numbers, or demo files
  - only replaces parsed_claims when it finds at least as many real claims as current parser
  """
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def money_float(value):
    raw = re.sub(r"[^0-9.\-]", "", str(value or ""))
    try:
      return float(raw or 0)
    except Exception:
      return 0.0

  def money_text(value):
    amount = money_float(value)
    if abs(amount - round(amount)) < 0.005:
      return str(int(round(amount)))
    return f"{amount:.2f}"

  def read_pdf_text():
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
      print("LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_READ_ERROR_V2:", str(exc)[:200])
      return ""

  raw_text = read_pdf_text()
  if not raw_text or not re.search(r"(?i)\bCLAIMS\s+DETAIL\b", raw_text):
    return parsed_claims, parsed_profile, direct_profile

  normalized = re.sub(r"\s+", " ", raw_text).strip()

  start = re.search(r"(?i)\bCLAIMS\s+DETAIL\b", normalized)
  if not start:
    return parsed_claims, parsed_profile, direct_profile

  section = normalized[start.end():]
  end = re.search(r"(?i)\bLOSS\s+SUMMARY\b|\bLOSS\s+SUMMARY\s+BY\s+COVERAGE\b|\bEarned\s+Premium\b|\bOpen\s+Claims\b|\bClosed\s+Claims\b", section)
  if end:
    section = section[:end.start()]

  claim_matches = list(re.finditer(r"\b[A-Z]{1,10}-\d{4}-\d{3,8}\b", section))
  if not claim_matches:
    print("LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_V2:", {
      "rescued_claims": 0,
      "reason": "no_claim_row_ids_found",
      "existing_claims": len(parsed_claims),
    })
    return parsed_claims, parsed_profile, direct_profile

  policy_number = clean(
    parsed_profile.get("policy_number")
    or parsed_profile.get("main_policy_number")
    or direct_profile.get("policy_number")
    or direct_profile.get("main_policy_number")
  )

  if not policy_number:
    m_policy = re.search(r"(?i)\bPolicy\s+Number\s*[:#-]\s*([A-Z0-9][A-Z0-9 ./_-]{3,80}?)(?=\s+IBC\s+Line\s+of\s+Business|\s+Policy\s+Period|\s+Retroactive\s+Date|\s+Occurrence\s+Limit|$)", normalized)
    if m_policy:
      policy_number = clean(m_policy.group(1))

  carrier = clean(
    parsed_profile.get("carrier_name")
    or parsed_profile.get("writing_carrier")
    or direct_profile.get("carrier_name")
    or direct_profile.get("writing_carrier")
  )

  business_name = clean(
    parsed_profile.get("business_name")
    or parsed_profile.get("insured_name")
    or parsed_profile.get("named_insured")
    or parsed_profile.get("insured")
    or direct_profile.get("business_name")
    or direct_profile.get("insured_name")
    or direct_profile.get("named_insured")
  )

  lob = clean(
    parsed_profile.get("line_of_business")
    or parsed_profile.get("policy_type")
    or parsed_profile.get("coverage")
    or direct_profile.get("line_of_business")
    or direct_profile.get("policy_type")
    or direct_profile.get("coverage")
  )

  if not lob:
    m_lob = re.search(r"(?i)\bIBC\s+Line\s+of\s+Business\s*[:#-]\s*(.{3,160}?)(?=\s+Policy\s+Period|\s+Retroactive\s+Date|\s+Occurrence\s+Limit|\s+Aggregate\s+Limit|\s+Deductible|\s+Report\s+Date|\s+Currency|$)", normalized)
    if m_lob:
      lob = clean(m_lob.group(1))

  coverage_terms = [
    "Products & Completed Operations",
    "Products and Completed Operations",
    "Completed Operations",
    "Commercial General Liability",
    "Commercial Property",
    "General Liability",
    "Employer's Liability",
    "Employers Liability",
    "Professional Liability",
    "Commercial Auto",
    "Bodily Injury",
    "Property Damage",
    "Completed Ops",
    "Liquor Liability",
    "Business Interruption",
    "Cyber Liability",
    "Umbrella",
    "Excess",
  ]

  incident_start = re.compile(
    r"(?i)\b(excavation|slip|fall|faulty|underground|repetitive|rented|fire|water|theft|collision|injury|damage|alleged|disputed|claim|leak|wind|hail|property|bodily|remediation|foundation|surgery|utility|strain)\b"
  )

  rescued = []
  seen = set()

  for idx, match in enumerate(claim_matches):
    claim_number = clean(match.group(0))
    if not claim_number or claim_number.upper() in seen:
      continue

    next_start = claim_matches[idx + 1].start() if idx + 1 < len(claim_matches) else len(section)
    chunk = clean(section[match.start():next_start])

    dates = re.findall(r"\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4})\b", chunk)
    if len(dates) < 2:
      continue

    amount_matches = re.findall(r"\$\s*\d[\d,]*(?:\.\d{2})?", chunk)
    if len(amount_matches) < 4:
      continue

    status_match = re.search(r"(?i)\b(Open|Closed|Reopened|Pending|Ouvert|Ouverte|Fermé|Fermée|Clos|Clôturé)\b(?=\s+\$\s*\d)", chunk)
    if not status_match:
      status_match = re.search(r"(?i)\b(Open|Closed|Reopened|Pending|Ouvert|Ouverte|Fermé|Fermée|Clos|Clôturé)\b", chunk)

    if not status_match:
      continue

    seen.add(claim_number.upper())

    loss_date = dates[0]
    reported_date = dates[1]
    status = clean(status_match.group(1))

    if status.lower() in {"ouvert", "ouverte"}:
      status = "Open"
    elif status.lower() in {"fermé", "fermée", "clos", "clôturé"}:
      status = "Closed"

    paid_indemnity = money_text(amount_matches[-4])
    paid_expense = money_text(amount_matches[-3])
    reserve = money_text(amount_matches[-2])
    total_incurred = money_text(amount_matches[-1])
    paid_total = money_text(money_float(paid_indemnity) + money_float(paid_expense))

    body_start = chunk.find(reported_date) + len(reported_date)
    body_end = status_match.start()
    body = clean(chunk[body_start:body_end])

    coverage = ""
    coverage_index = -1
    for term in sorted(coverage_terms, key=len, reverse=True):
      pos = body.lower().rfind(term.lower())
      if pos >= 0 and pos > coverage_index:
        coverage = term
        coverage_index = pos

    pre_coverage = body[:coverage_index].strip(" -–") if coverage_index >= 0 else body

    claimant = ""
    description = pre_coverage

    incident_match = incident_start.search(pre_coverage)
    if incident_match and incident_match.start() > 0:
      claimant = clean(pre_coverage[:incident_match.start()])
      description = clean(pre_coverage[incident_match.start():])
    else:
      words = pre_coverage.split()
      if len(words) <= 6:
        claimant = pre_coverage
      else:
        claimant = clean(" ".join(words[:4]))
        description = clean(" ".join(words[4:]))

    note_text = ""
    note_match = re.search(r"(?i)\bClaim\s+" + re.escape(claim_number) + r"\b(.{0,320})", normalized)
    if note_match:
      note_text = clean(note_match.group(1))

    litigation = ""
    attorney = ""
    if re.search(r"(?i)\b(active\s+litigation|litigation|counsel\s+retained|attorney|lawyer)\b", note_text):
      litigation = "Yes"
      attorney = "Yes"

    rescued.append({
      "claim_number": claim_number,
      "claim_no": claim_number,
      "policy_number": policy_number,
      "line_of_business": lob or coverage,
      "coverage": coverage or lob,
      "claim_type": coverage or lob,
      "date_of_loss": loss_date,
      "loss_date": loss_date,
      "date_reported": reported_date,
      "reported_date": reported_date,
      "status": status,
      "claimant": claimant,
      "description": note_text or description or body,
      "cause_of_loss": coverage,
      "paid_indemnity": paid_indemnity,
      "paid_expense": paid_expense,
      "paid": paid_total,
      "paid_amount": paid_total,
      "reserve": reserve,
      "reserve_amount": reserve,
      "total_incurred": total_incurred,
      "incurred": total_incurred,
      "carrier_name": carrier,
      "writing_carrier": carrier,
      "business_name": business_name,
      "named_insured": business_name,
      "litigation": litigation,
      "attorney_involved": attorney,
      "represented": attorney,
    })

  if not rescued:
    print("LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_V2:", {
      "rescued_claims": 0,
      "reason": "rows_failed_validation",
      "existing_claims": len(parsed_claims),
    })
    return parsed_claims, parsed_profile, direct_profile

  if parsed_claims and len(rescued) < len(parsed_claims):
    print("LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_V2_SKIPPED:", {
      "rescued_claims": len(rescued),
      "existing_claims": len(parsed_claims),
    })
    return parsed_claims, parsed_profile, direct_profile

  total_paid = sum(money_float(c.get("paid_amount")) for c in rescued)
  total_reserve = sum(money_float(c.get("reserve_amount")) for c in rescued)
  total_incurred = sum(money_float(c.get("total_incurred")) for c in rescued)
  open_claims = sum(1 for c in rescued if clean(c.get("status")).lower() == "open")
  closed_claims = sum(1 for c in rescued if clean(c.get("status")).lower() == "closed")

  for target in (parsed_profile, direct_profile):
    target["claims"] = rescued
    target["parsed_claims"] = rescued
    target["claim_count"] = len(rescued)
    target["total_claims"] = len(rescued)
    target["open_claims"] = open_claims
    target["closed_claims"] = closed_claims
    target["total_paid"] = total_paid
    target["total_reserve"] = total_reserve
    target["total_incurred"] = total_incurred

  print("LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_V2:", {
    "rescued_claims": len(rescued),
    "open_claims": open_claims,
    "closed_claims": closed_claims,
    "total_incurred": total_incurred,
    "claim_numbers": [c.get("claim_number") for c in rescued],
  })

  return rescued, parsed_profile, direct_profile


# LOSSQ_PDF_ACCOUNT_PROFILE_GRID_REPAIR_V1
def lossq_pdf_account_profile_grid_repair_v1(file_path, parsed_profile=None, direct_profile=None):
  """
  Final universal PDF profile-grid repair.

  Fixes text-readable PDF account headers where:
  - Insured Name is beside Province in a grid/table.
  - Carrier address province/city appears later in the footer.
  - Footer confidentiality language is accidentally captured as insured name.

  No customer, carrier, filename, or demo-file hardcoding.
  """
  import re

  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def compact(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def read_pdf_text():
    parts = []
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      parts.append("\n".join((page.extract_text() or "") for page in reader.pages))
    except Exception as exc:
      print("LOSSQ_PDF_ACCOUNT_PROFILE_GRID_REPAIR_READ_ERROR_V1:", str(exc)[:200])
    return "\n".join(part for part in parts if part)

  raw_text = read_pdf_text()
  if not raw_text:
    return parsed_profile, direct_profile

  normalized_text = re.sub(r"\s+", " ", raw_text or " ").strip()

  def first_match(patterns):
    for pattern in patterns:
      match = re.search(pattern, normalized_text, flags=re.I | re.S)
      if match:
        value = clean(match.group(1))
        if value:
          return value
    return ""

  blocked_name_bits = (
    "authorized broker",
    "unauthorized reproduction",
    "this document is confidential",
    "ibc codes referenced",
    "report generated",
    "commercial lines underwriting",
    "loss run report",
  )

  def good_business_name(value):
    value = clean(value)
    low = value.lower()
    if not value or len(value) < 4 or len(value) > 140:
      return ""
    if any(bit in low for bit in blocked_name_bits):
      return ""
    if compact(value) in {"insured", "insuredname", "businessname", "namedinsured", "accountname", "unknown", "na", "none"}:
      return ""
    return value

  def clean_carrier(value):
    value = clean(value)
    value = re.sub(r"(?i)\s+[–-]\s+(?:commercial|personal|claims|underwriting).*$", "", value).strip()
    value = re.sub(r"(?i)\s+commercial\s+lines\s+underwriting.*$", "", value).strip()
    value = re.sub(r"\s*\|\s*.*$", "", value).strip()
    value = re.sub(r"(?i),\s*[A-Za-z .'-]+,\s*(?:[A-Z]{2}|[A-Za-z]+).*$", "", value).strip()
    value = clean(value)
    if len(value) < 3 or len(value) > 120:
      return ""
    return value

  province_to_code = {
    "alberta": "AB",
    "britishcolumbia": "BC",
    "manitoba": "MB",
    "newbrunswick": "NB",
    "newfoundlandandlabrador": "NL",
    "novascotia": "NS",
    "northwestterritories": "NT",
    "nunavut": "NU",
    "ontario": "ON",
    "princeedwardisland": "PE",
    "quebec": "QC",
    "québec": "QC",
    "saskatchewan": "SK",
    "yukon": "YT",
  }

  insured_name = good_business_name(first_match([
    r"\bInsured\s+Name\s*[:#-]\s*(.{3,140}?)(?=\s+Province\s*[:#-]|\s+State\s*[:#-]|\s+Policy\s+Number\s*[:#-]|\s+IBC\s+Line\s+of\s+Business\s*[:#-]|\s+Policy\s+Period\s*[:#-]|\s+Currency\s*[:#-]|\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|\s+CLAIMS\s+DETAIL\b|$)",
    r"\b(?:Named\s+Insured|Business\s+Name|Account\s+Name|Applicant|Company\s+Name)\s*[:#-]\s*(.{3,140}?)(?=\s+Province\s*[:#-]|\s+State\s*[:#-]|\s+Policy\s+Number\s*[:#-]|\s+Carrier\s*[:#-]|\s+Writing\s+Carrier\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|\s+CLAIMS\s+DETAIL\b|$)",
  ]))

  province_name = first_match([
    r"\bProvince\s*[:#-]\s*([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'-]{1,45}?)(?=\s+Policy\s+Number\s*[:#-]|\s+IBC\s+Line\s+of\s+Business\s*[:#-]|\s+Policy\s+Period\s*[:#-]|\s+Retroactive\s+Date\s*[:#-]|\s+Occurrence\s+Limit\s*[:#-]|\s+Aggregate\s+Limit\s*[:#-]|\s+Deductible\s*[:#-]|\s+Report\s+Date\s*[:#-]|\s+Currency\s*[:#-]|\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|\s+CLAIMS\s+DETAIL\b|$)",
    r"\bState\s*/\s*Province\s*[:#-]\s*([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'-]{1,45}?)(?=\s+Policy\s+Number\s*[:#-]|\s+Carrier\s*[:#-]|\s+Currency\s*[:#-]|$)",
  ])
  province_key = compact(province_name)
  province_code = province_to_code.get(province_key, "")

  currency = first_match([
    r"\bCurrency\s*[:#-]\s*([A-Z]{3})(?=\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|\s+CLAIMS\s+DETAIL\b|$)",
  ])

  policy_number = first_match([
    r"\bPolicy\s+Number\s*[:#-]\s*([A-Z0-9][A-Z0-9 ./_-]{3,80}?)(?=\s+IBC\s+Line\s+of\s+Business\s*[:#-]|\s+Policy\s+Period\s*[:#-]|\s+Retroactive\s+Date\s*[:#-]|\s+Occurrence\s+Limit\s*[:#-]|\s+Aggregate\s+Limit\s*[:#-]|\s+Deductible\s*[:#-]|\s+Report\s+Date\s*[:#-]|\s+Currency\s*[:#-]|\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|$)",
  ])

  lob = first_match([
    r"\bIBC\s+Line\s+of\s+Business\s*[:#-]\s*(.{3,160}?)(?=\s+Policy\s+Period\s*[:#-]|\s+Retroactive\s+Date\s*[:#-]|\s+Occurrence\s+Limit\s*[:#-]|\s+Aggregate\s+Limit\s*[:#-]|\s+Deductible\s*[:#-]|\s+Report\s+Date\s*[:#-]|\s+Currency\s*[:#-]|\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|$)",
    r"\bLine\s+of\s+Business\s*[:#-]\s*(.{3,120}?)(?=\s+Policy\s+Period\s*[:#-]|\s+Effective\s+Date\s*[:#-]|\s+Expiration\s+Date\s*[:#-]|$)",
  ])

  policy_period = first_match([
    r"\bPolicy\s+Period\s*[:#-]\s*([0-9]{4}[-/][0-9]{2}[-/][0-9]{2}\s+(?:to|through|-|–)\s+[0-9]{4}[-/][0-9]{2}[-/][0-9]{2})",
    r"\bPolicy\s+Period\s*[:#-]\s*([0-9]{1,2}[/][0-9]{1,2}[/][0-9]{2,4}\s+(?:to|through|-|–)\s+[0-9]{1,2}[/][0-9]{1,2}[/][0-9]{2,4})",
  ])

  effective_date = ""
  expiration_date = ""
  if policy_period:
    m_period = re.search(r"([0-9]{4}[-/][0-9]{2}[-/][0-9]{2}|[0-9]{1,2}[/][0-9]{1,2}[/][0-9]{2,4})\s+(?:to|through|-|–)\s+([0-9]{4}[-/][0-9]{2}[-/][0-9]{2}|[0-9]{1,2}[/][0-9]{1,2}[/][0-9]{2,4})", policy_period, flags=re.I)
    if m_period:
      effective_date = clean(m_period.group(1))
      expiration_date = clean(m_period.group(2))

  report_date = first_match([
    r"\bReport\s+Date\s*[:#-]\s*([A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{4}[-/]\d{2}[-/]\d{2}|\d{1,2}/\d{1,2}/\d{2,4})(?=\s+Currency\s*[:#-]|\s+Prepared\s+By\s*[:#-]|\s+EXPOSURE\s+SUMMARY\b|$)",
    r"\bEvaluation\s+Date\s*[:#-]\s*([A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{4}[-/]\d{2}[-/]\d{2}|\d{1,2}/\d{1,2}/\d{2,4})",
    r"\bValuation\s+Date\s*[:#-]\s*([A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{4}[-/]\d{2}[-/]\d{2}|\d{1,2}/\d{1,2}/\d{2,4})",
  ])

  carrier = clean_carrier(first_match([
    r"\bPrepared\s+By\s*[:#-]\s*([A-Z][A-Za-z0-9&.'’\- ]{3,120}?)(?=\s+[–-]\s+(?:Commercial|Personal|Claims|Underwriting)\b|\s+EXPOSURE\s+SUMMARY\b|\s+CLAIMS\s+DETAIL\b|$)",
    r"\bWriting\s+Carrier\s*[:#-]\s*([A-Z][A-Za-z0-9&.'’\- ]{3,120}?)(?=\s+Policy\s+Number\s*[:#-]|\s+Effective\s+Date\s*[:#-]|\s+Expiration\s+Date\s*[:#-]|$)",
    r"\bCarrier\s*[:#-]\s*([A-Z][A-Za-z0-9&.'’\- ]{3,120}?)(?=,\s*[A-Za-z .'-]+,\s*(?:[A-Z]{2}|[A-Za-z]+)\b|\s+\|\s+|$)",
  ]))

  if not carrier:
    for line in [clean(x) for x in raw_text.splitlines()[:15]]:
      if re.search(r"(?i)\binsurance\b", line) and not re.search(r"(?i)loss\s+run|commercial\s+lines|report", line):
        carrier = clean_carrier(line.title() if line.isupper() else line)
        break

  is_canada = bool(province_code or currency.upper() == "CAD" or re.search(r"(?i)\bIBC\s+\d+\b|\bWSIB\b|\bCanada\b|\bOntario\b|\bManitoba\b|\bQuebec\b|\bQuébec\b", normalized_text))

  def apply(target):
    if insured_name:
      target["business_name"] = insured_name
      target["insured"] = insured_name
      target["insured_name"] = insured_name
      target["named_insured"] = insured_name
      target["account_name"] = insured_name

    if carrier:
      target["carrier_name"] = carrier
      target["writing_carrier"] = carrier
      target["carrier"] = carrier

    if policy_number:
      target["policy_number"] = policy_number
      target["main_policy_number"] = policy_number

    if effective_date:
      target["effective_date"] = effective_date
      target["effective"] = effective_date

    if expiration_date:
      target["expiration_date"] = expiration_date
      target["expiration"] = expiration_date

    if report_date:
      target["evaluation_date"] = report_date
      target["valuation_date"] = report_date
      target["report_date"] = report_date

    if currency:
      target["currency"] = currency.upper()
      target["default_currency"] = currency.upper()

    if province_name:
      target["province"] = province_name
      target["province_name"] = province_name
      target["state_name"] = province_name
      if province_code:
        target["state"] = province_code
        target["province_code"] = province_code
        target["state_province"] = province_code
      else:
        target["state"] = province_name
        target["state_province"] = province_name

    if is_canada:
      target["country"] = "Canada"
      target["market_country"] = "Canada"
      target["country_market"] = "Canada"
      target["market"] = "Canada"

    current_schedule = target.get("policy_schedule") or target.get("policies")
    schedule_missing = not isinstance(current_schedule, list) or not any(isinstance(p, dict) and clean(p.get("policy_number")) for p in current_schedule)

    if policy_number and schedule_missing:
      row = {
        "policy_number": policy_number,
        "policyNumber": policy_number,
        "line_of_business": lob,
        "policy_type": lob,
        "coverage": lob,
        "carrier": carrier,
        "carrier_name": carrier,
        "writing_carrier": carrier,
        "policy_period": policy_period,
        "effective_date": effective_date,
        "effective": effective_date,
        "expiration_date": expiration_date,
        "expiration": expiration_date,
        "state": province_code or province_name,
        "province": province_name,
        "currency": currency.upper() if currency else "",
      }
      target["policies"] = [row]
      target["policy_schedule"] = [row]
    elif policy_number and isinstance(current_schedule, list):
      for row in current_schedule:
        if not isinstance(row, dict):
          continue
        if clean(row.get("policy_number")) == policy_number:
          if lob and not clean(row.get("line_of_business")):
            row["line_of_business"] = lob
            row["policy_type"] = lob
            row["coverage"] = lob
          if carrier:
            row["carrier"] = row.get("carrier") or carrier
            row["carrier_name"] = row.get("carrier_name") or carrier
            row["writing_carrier"] = row.get("writing_carrier") or carrier
          if effective_date:
            row["effective_date"] = row.get("effective_date") or effective_date
            row["effective"] = row.get("effective") or effective_date
          if expiration_date:
            row["expiration_date"] = row.get("expiration_date") or expiration_date
            row["expiration"] = row.get("expiration") or expiration_date
      target["policies"] = current_schedule
      target["policy_schedule"] = current_schedule

  apply(parsed_profile)
  apply(direct_profile)

  print("LOSSQ_PDF_ACCOUNT_PROFILE_GRID_REPAIR_V1:", {
    "business_name": insured_name,
    "province": province_name,
    "province_code": province_code,
    "country": "Canada" if is_canada else "",
    "currency": currency,
    "carrier": carrier,
    "policy_number": policy_number,
  })

  return parsed_profile, direct_profile


# LOSSQ_PDF_WIDE_CLAIMS_DETAIL_TABLE_REPAIR_V1
def lossq_pdf_wide_claims_detail_table_repair_v1(file_path, parsed_claims=None, parsed_profile=None, direct_profile=None):
  """
  Universal repair for text-readable PDF loss runs with a wide CLAIMS DETAIL table.

  It only replaces parsed claims when it finds at least as many real claim rows as
  the existing parser found. No company, carrier, customer, or demo-file hardcoding.
  """
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_claims, parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def money_float(value):
    raw = re.sub(r"[^0-9.\-]", "", str(value or ""))
    try:
      return float(raw or 0)
    except Exception:
      return 0.0

  def money_text(value):
    amount = money_float(value)
    if abs(amount - round(amount)) < 0.005:
      return str(int(round(amount)))
    return f"{amount:.2f}"

  def fmt_amount(amount):
    amount = float(amount or 0)
    if abs(amount - round(amount)) < 0.005:
      return str(int(round(amount)))
    return f"{amount:.2f}"

  def read_pdf_text():
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
      print("LOSSQ_PDF_WIDE_CLAIMS_DETAIL_READ_ERROR_V1:", str(exc)[:200])
      return ""

  raw_text = read_pdf_text()
  if not raw_text or not re.search(r"(?i)\bCLAIMS\s+DETAIL\b", raw_text):
    return parsed_claims, parsed_profile, direct_profile

  normalized_text = re.sub(r"\s+", " ", raw_text or " ").strip()
  section = normalized_text
  m_start = re.search(r"(?i)\bCLAIMS\s+DETAIL\b", section)
  if m_start:
    section = section[m_start.end():]
  m_end = re.search(r"(?i)\bLOSS\s+SUMMARY\b|\bEarned\s+Premium\b|\bThis\s+report\s+reflects\b", section)
  if m_end:
    section = section[:m_end.start()]

  policy_number = clean(parsed_profile.get("policy_number") or parsed_profile.get("main_policy_number"))
  if not policy_number:
    m_policy = re.search(r"(?i)\bPolicy\s+Number\s*[:#-]\s*([A-Z0-9][A-Z0-9 ./_-]{3,80}?)(?=\s+IBC\s+Line\s+of\s+Business\s*[:#-]|\s+Policy\s+Period\s*[:#-]|\s+Retroactive\s+Date\s*[:#-]|$)", normalized_text)
    if m_policy:
      policy_number = clean(m_policy.group(1))

  lob = clean(parsed_profile.get("line_of_business") or parsed_profile.get("policy_type") or parsed_profile.get("coverage"))
  if not lob:
    m_lob = re.search(r"(?i)\bIBC\s+Line\s+of\s+Business\s*[:#-]\s*(.{3,160}?)(?=\s+Policy\s+Period\s*[:#-]|\s+Retroactive\s+Date\s*[:#-]|\s+Occurrence\s+Limit\s*[:#-]|$)", normalized_text)
    if m_lob:
      lob = clean(m_lob.group(1))

  carrier = clean(parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier") or direct_profile.get("carrier_name") or direct_profile.get("writing_carrier"))
  business_name = clean(parsed_profile.get("business_name") or parsed_profile.get("insured_name") or parsed_profile.get("named_insured") or direct_profile.get("business_name"))

  existing_by_number = {}
  for item in parsed_claims:
    if isinstance(item, dict) and clean(item.get("claim_number")):
      existing_by_number[clean(item.get("claim_number")).upper()] = item

  row_pattern = re.compile(
    r"(?P<claim>[A-Z0-9]{1,12}-\d{2,4}-\d{3,8})\s+"
    r"(?P<loss>\d{4}[-/]\d{2}[-/]\d{2}|\d{1,2}/\d{1,2}/\d{2,4})\s+"
    r"(?P<reported>\d{4}[-/]\d{2}[-/]\d{2}|\d{1,2}/\d{1,2}/\d{2,4})\s+"
    r"(?P<details>.+?)\s+"
    r"(?P<status>Open|Closed|Reopened|Pending|Ouvert|Ouverte|Fermé|Fermée|Clos|Clôturé)\s+"
    r"\$?\s*(?P<paid_indemnity>[\d,]+(?:\.\d{2})?)\s+"
    r"\$?\s*(?P<paid_expense>[\d,]+(?:\.\d{2})?)\s+"
    r"\$?\s*(?P<reserve>[\d,]+(?:\.\d{2})?)\s+"
    r"\$?\s*(?P<total>[\d,]+(?:\.\d{2})?)",
    flags=re.I | re.S,
  )

  coverage_terms = [
    "Products and Completed Operations",
    "Products & Completed Operations",
    "Completed Operations",
    "Employer's Liability",
    "Employers Liability",
    "Professional Liability",
    "Commercial Auto",
    "General Liability",
    "Commercial Property",
    "Business Interruption",
    "Liquor Liability",
    "Cyber Liability",
    "Property Damage",
    "Bodily Injury",
    "Completed Ops",
    "Non-Owned Auto",
    "Umbrella",
    "Excess",
  ]

  incident_start = re.compile(
    r"(?i)\b(?:slip|fall|fire|water|wind|hail|theft|damage|injur|collision|alleged|property|bodily|medical|vehicle|equipment|employee|customer|contractor|faulty|failed|repetitive|underground|excavation|rented|lost|leak|flood|smoke)\b"
  )

  extracted_claims = []
  seen = set()

  for match in row_pattern.finditer(section):
    claim_number = clean(match.group("claim"))
    if not claim_number or claim_number.upper() in seen:
      continue
    seen.add(claim_number.upper())

    details = clean(match.group("details"))
    coverage = ""
    coverage_start = -1
    for term in sorted(coverage_terms, key=len, reverse=True):
      idx = details.lower().rfind(term.lower())
      if idx >= 0 and idx >= coverage_start:
        coverage = term
        coverage_start = idx

    pre_coverage = details[:coverage_start].strip(" -–") if coverage_start >= 0 else details
    description = pre_coverage
    claimant = ""

    m_incident = incident_start.search(pre_coverage)
    if m_incident and m_incident.start() >= 3:
      claimant = clean(pre_coverage[:m_incident.start()])
      description = clean(pre_coverage[m_incident.start():])
    else:
      words = pre_coverage.split()
      if len(words) <= 6:
        claimant = pre_coverage
        description = ""
      else:
        claimant = clean(" ".join(words[:4]))
        description = clean(" ".join(words[4:]))

    paid_indemnity = money_text(match.group("paid_indemnity"))
    paid_expense = money_text(match.group("paid_expense"))
    paid_total = money_float(paid_indemnity) + money_float(paid_expense)
    reserve = money_text(match.group("reserve"))
    total = money_text(match.group("total"))

    status = clean(match.group("status"))
    if status.lower() in {"ouvert", "ouverte"}:
      status = "Open"
    elif status.lower() in {"fermé", "fermée", "clos", "clôturé"}:
      status = "Closed"

    existing = existing_by_number.get(claim_number.upper(), {})

    claim = {
      "claim_number": claim_number,
      "claim_no": claim_number,
      "policy_number": clean(existing.get("policy_number")) or policy_number,
      "line_of_business": clean(existing.get("line_of_business")) or lob,
      "coverage": clean(existing.get("coverage")) or coverage or lob,
      "claim_type": clean(existing.get("claim_type")) or coverage or lob,
      "date_of_loss": clean(match.group("loss")),
      "loss_date": clean(match.group("loss")),
      "date_reported": clean(match.group("reported")),
      "reported_date": clean(match.group("reported")),
      "status": status,
      "claimant": clean(existing.get("claimant")) or claimant,
      "description": clean(existing.get("description")) or description or details,
      "cause_of_loss": clean(existing.get("cause_of_loss")) or coverage,
      "paid_indemnity": paid_indemnity,
      "paid_expense": paid_expense,
      "paid": fmt_amount(paid_total),
      "paid_amount": fmt_amount(paid_total),
      "reserve": reserve,
      "reserve_amount": reserve,
      "total_incurred": total,
      "incurred": total,
      "carrier_name": clean(existing.get("carrier_name")) or carrier,
      "writing_carrier": clean(existing.get("writing_carrier")) or carrier,
      "business_name": clean(existing.get("business_name")) or business_name,
      "named_insured": clean(existing.get("named_insured")) or business_name,
      "litigation": clean(existing.get("litigation")),
      "attorney_involved": clean(existing.get("attorney_involved")),
    }

    note_pattern = re.compile(r"(?i)\bClaim\s+" + re.escape(claim_number) + r"\b(.{0,260})")
    note_match = note_pattern.search(normalized_text)
    if note_match:
      note_text = clean(note_match.group(1))
      if note_text and not claim.get("description"):
        claim["description"] = note_text
      if re.search(r"(?i)\b(active\s+litigation|litigation|counsel\s+retained|attorney|lawyer)\b", note_text):
        claim["litigation"] = claim.get("litigation") or "Yes"
        claim["attorney_involved"] = claim.get("attorney_involved") or "Yes"
        claim["represented"] = claim.get("represented") or "Yes"

    extracted_claims.append(claim)

  if not extracted_claims or len(extracted_claims) < len(parsed_claims):
    return parsed_claims, parsed_profile, direct_profile

  total_paid = sum(money_float(c.get("paid_amount")) for c in extracted_claims)
  total_reserve = sum(money_float(c.get("reserve_amount")) for c in extracted_claims)
  total_incurred = sum(money_float(c.get("total_incurred")) for c in extracted_claims)
  open_claims = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "open")
  closed_claims = sum(1 for c in extracted_claims if clean(c.get("status")).lower() == "closed")

  for target in (parsed_profile, direct_profile):
    target["claims"] = extracted_claims
    target["parsed_claims"] = extracted_claims
    target["claim_count"] = len(extracted_claims)
    target["total_claims"] = len(extracted_claims)
    target["open_claims"] = open_claims
    target["closed_claims"] = closed_claims
    target["total_paid"] = total_paid
    target["total_reserve"] = total_reserve
    target["total_incurred"] = total_incurred

  print("LOSSQ_PDF_WIDE_CLAIMS_DETAIL_TABLE_REPAIR_V1:", {
    "claims": len(extracted_claims),
    "open_claims": open_claims,
    "closed_claims": closed_claims,
    "total_incurred": total_incurred,
    "claim_numbers": [c.get("claim_number") for c in extracted_claims],
  })

  return extracted_claims, parsed_profile, direct_profile


# LOSSQ_PDF_FINAL_CARRIER_LABEL_REPAIR_V1
def lossq_pdf_final_carrier_label_repair_v1(file_path, parsed_profile=None, direct_profile=None):
  """
  Final universal PDF carrier repair.

  Purpose:
  - Runs late in the upload flow so later profile cleanup cannot overwrite the true PDF carrier.
  - Reads carrier from generic labels only: Carrier, Writing Carrier, Insurance Carrier, Carrier Name.
  - Does not hardcode customer, carrier, policy number, claim number, or sample file.
  """
  import re

  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".pdf"):
    return parsed_profile, direct_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|")

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def read_pdf_text():
    try:
      from pypdf import PdfReader
      reader = PdfReader(file_path)
      return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
      print("LOSSQ_PDF_FINAL_CARRIER_LABEL_REPAIR_READ_ERROR_V1:", str(exc)[:200])
      return ""

  def looks_like_policy_or_account(value):
    raw = clean(value).upper()
    if not raw:
      return True
    if re.search(r"\b(ACCT|ACCOUNT|CUSTOMER|CLIENT)\b", raw):
      return True
    if re.search(r"^[A-Z0-9]{2,}(?:[-_][A-Z0-9]{2,}){2,9}$", raw) and re.search(r"\d", raw):
      return True
    return False

  def good_carrier(value):
    candidate = clean(value)
    if not candidate:
      return False

    low = key(candidate)
    bad_exact = {
      "carrier",
      "carriername",
      "writingcarrier",
      "insurancecarrier",
      "notset",
      "unknown",
      "na",
      "none",
      "generalliability",
      "workerscompensation",
      "businessownerspolicy",
      "businessownerpolicy",
      "umbrella",
      "liquorliability",
      "commercialauto",
      "garageliability",
      "dealersopenlot",
      "commercialproperty",
      "property",
      "cyberliability",
      "professionalliability",
      "employmentpracticesliability",
      "directorsofficers",
      "inlandmarine",
      "motortruckcargo",
    }

    if low in bad_exact:
      return False

    if looks_like_policy_or_account(candidate):
      return False

    if re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", candidate):
      return False

    if re.search(r"\$\s*\d", candidate):
      return False

    if re.search(
      r"(?i)\b(claim\s+number|policy\s+number|line\s+of\s+business|claim\s+status|paid\s+amount|reserve\s+amount|total\s+incurred|date\s+of\s+loss|loss\s+summary|policy\s+schedule)\b",
      candidate,
    ):
      return False

    return bool(re.search(r"[A-Za-z]", candidate)) and len(candidate) >= 3

  raw_text = read_pdf_text()
  if not raw_text:
    return parsed_profile, direct_profile

  lines = [clean(line) for line in raw_text.splitlines() if clean(line)]
  carrier_value = ""

  # Same-line labels: Carrier - ABC, Carrier: ABC, Writing Carrier # ABC
  for line in lines:
    match = re.match(
      r"(?i)^\s*(?:writing\s+carrier|insurance\s+carrier|carrier\s+name|carrier)\s*(?:[:#\-\u2013\u2014])\s*(.+?)\s*$",
      line,
    )
    if match and good_carrier(match.group(1)):
      carrier_value = clean(match.group(1))
      break

  # Label on one line, value on a following line.
  if not carrier_value:
    carrier_labels = {"writingcarrier", "insurancecarrier", "carriername", "carrier"}
    for idx, line in enumerate(lines):
      if key(line.rstrip(":#-")) in carrier_labels:
        for j in range(idx + 1, min(idx + 4, len(lines))):
          candidate = clean(lines[j])
          if good_carrier(candidate):
            carrier_value = candidate
            break
        if carrier_value:
          break

  if carrier_value:
    for target in (parsed_profile, direct_profile):
      target["carrier_name"] = carrier_value
      target["writing_carrier"] = carrier_value
      target["carrier"] = carrier_value

      for list_key in ("policy_schedule", "policies"):
        rows = target.get(list_key)
        if isinstance(rows, list):
          for row in rows:
            if isinstance(row, dict):
              row["carrier"] = carrier_value
              row["carrier_name"] = carrier_value
              row["writing_carrier"] = carrier_value

  print("LOSSQ_PDF_FINAL_CARRIER_LABEL_REPAIR_V1:", {
    "carrier_found": carrier_value,
    "parsed_profile_carrier": parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier") or parsed_profile.get("carrier"),
  })

  return parsed_profile, direct_profile


# LOSSQ_EXCEL_ACCOUNT_CUSTOMER_PRECEDENCE_V1
def lossq_excel_account_customer_precedence_v1(file_path, parsed_profile=None, direct_profile=None):
  """
  Excel-only guardrail: keep Account Number and Customer Number separate.
  If both exist in a workbook, Account Number wins for account_number.
  """
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  if not str(file_path or "").lower().endswith((".xlsx", ".xlsm")):
    return parsed_profile, direct_profile

  try:
    import re
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    def clean(value):
      return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

    def key(value):
      return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

    account_labels = {
      "accountnumber",
      "accountno",
      "account",
      "acctnumber",
      "acctno",
      "insuredaccountnumber",
    }

    customer_labels = {
      "customernumber",
      "customerno",
      "customerid",
      "clientnumber",
      "clientno",
    }

    def usable_identifier(value):
      value = clean(value)
      if not value or value.lower() in {"field", "value", "none", "null", "na", "n/a"}:
        return ""
      if len(value) > 80:
        return ""
      if not re.search(r"[A-Za-z0-9]", value):
        return ""
      if re.search(r"(?i)\b(policy|claim|premium|carrier|insured|agency|effective|expiration|date)\b", value):
        return ""
      return value

    def next_value(row_values, index):
      for value in row_values[index + 1:index + 6]:
        cleaned = usable_identifier(value)
        if cleaned:
          return cleaned
      return ""

    account_number = ""
    customer_number = ""

    for sheet in workbook.worksheets:
      for row in sheet.iter_rows(values_only=True):
        row_values = list(row or [])
        for idx, cell in enumerate(row_values):
          label_key = key(cell)
          cell_text = clean(cell)

          if not account_number and label_key in account_labels:
            account_number = next_value(row_values, idx)

          if not customer_number and label_key in customer_labels:
            customer_number = next_value(row_values, idx)

          if not account_number:
            same_cell_account = re.search(r"(?i)\baccount\s*(?:number|no\.?|#)\s*[:#-]\s*([A-Z0-9][A-Z0-9._/-]{2,60})", cell_text)
            if same_cell_account:
              account_number = usable_identifier(same_cell_account.group(1))

          if not customer_number:
            same_cell_customer = re.search(r"(?i)\bcustomer\s*(?:number|no\.?|#|id)\s*[:#-]\s*([A-Z0-9][A-Z0-9._/-]{2,60})", cell_text)
            if same_cell_customer:
              customer_number = usable_identifier(same_cell_customer.group(1))

        if account_number and customer_number:
          break
      if account_number and customer_number:
        break

    if account_number:
      for target in (parsed_profile, direct_profile):
        target["account_number"] = account_number
        target["account_no"] = account_number
        target["account"] = account_number

    if customer_number:
      for target in (parsed_profile, direct_profile):
        target["customer_number"] = customer_number

    print("LOSSQ_EXCEL_ACCOUNT_CUSTOMER_PRECEDENCE_V1:", {
      "account_number": account_number,
      "customer_number": customer_number,
    })

    return parsed_profile, direct_profile
  except Exception as exc:
    print("LOSSQ_EXCEL_ACCOUNT_CUSTOMER_PRECEDENCE_ERROR_V1:", str(exc)[:250])
    return parsed_profile, direct_profile


# LOSSQ_FINAL_UPLOAD_CLAIM_PROFILE_CLEANUP_V1
def lossq_final_upload_claim_profile_cleanup_v1(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal final cleanup before saving uploaded loss run data.

  Fixes messy parser artifacts across accepted upload types:
  - Claim numbers that accidentally absorb nearby labels like POLICY-NUMBER.
  - Missing insured/business name when the uploaded document contains a labeled insured.
  - Applies clean values back to profile claims and parsed claims.
  """
  import os
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def bad_business_value(value):
    key = re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())
    return key in {
      "",
      "businessnamenotset",
      "notset",
      "unknown",
      "na",
      "none",
      "null",
    }

  def clean_claim_number(value):
    raw = clean_text(value)
    if not raw:
      return raw

    # Remove labels that can get glued to claim numbers during messy PDF extraction.
    cleaned = re.sub(
      r"(?i)(?:\s|_|-)*(?:"
      r"POLICY\s*(?:NUMBER|NO|#)|"
      r"POLICY[-_]*(?:NUMBER|NO|#)|"
      r"LINE\s*OF\s*BUSINESS|"
      r"CLAIM\s*STATUS|"
      r"STATUS|"
      r"PAID|"
      r"RESERVE|"
      r"TOTAL\s*INCURRED|"
      r"INCURRED|"
      r"DATE\s*OF\s*LOSS|"
      r"DESCRIPTION"
      r").*$",
      "",
      raw,
    ).strip(" -_:/|")

    # If the value still contains a glued POLICY token, cut there.
    cleaned = re.split(r"(?i)POLICY", cleaned)[0].strip(" -_:/|")

    return cleaned or raw

  def read_upload_text():
    text_parts = [
      parsed_profile.get("raw_text"),
      parsed_profile.get("raw_text_preview"),
      parsed_profile.get("ocr_text"),
      parsed_profile.get("document_text"),
    ]

    lower_path = str(file_path or "").lower()

    if lower_path.endswith(".pdf"):
      try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pdf_text = "\n".join((page.extract_text() or "") for page in reader.pages)
        text_parts.append(pdf_text)
      except Exception:
        pass

    return "\n".join(str(part or "") for part in text_parts if part)

  def good_business_candidate(value):
    value = clean_text(value).strip(" :|-")
    if not value:
      return ""

    lower = value.lower()
    blocked = [
      "business name not set",
      "policy number",
      "claim number",
      "loss summary",
      "claim detail",
      "policy schedule",
      "account number",
      "effective date",
      "expiration date",
      "valuation date",
      "writing carrier",
      "carrier",
    ]
    if any(token == lower or lower.startswith(token + " ") for token in blocked):
      return ""

    if len(value) < 4 or len(value) > 120:
      return ""

    return value

  def extract_business_name_from_upload():
    raw_text = read_upload_text()

    label_patterns = [
      r"(?im)^\s*(?:Named\s+Insured|Insured|Business\s+Name|Account\s+Name|Applicant|Entity)\s*[:#-]?\s*(.+?)\s*$",
      r"(?im)\b(?:Named\s+Insured|Insured|Business\s+Name|Account\s+Name|Applicant|Entity)\b\s*[:#-]\s*([^\n\r]+)",
    ]

    for pattern in label_patterns:
      for match in re.finditer(pattern, raw_text or ""):
        candidate = good_business_candidate(match.group(1))
        if candidate:
          return candidate

    # Fallback: derive a clean display name from filename only when no document label is available.
    try:
      base = os.path.basename(str(file_path or ""))
      base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
      base = re.sub(r"(?i)\b(lossq|loss|run|messy|clean|ready|submission|pdf|csv|xlsx|xls|test|v\d+)\b", " ", base)
      base = re.sub(r"[_\-]+", " ", base)
      base = re.sub(r"\s+", " ", base).strip()
      if base:
        return base.title()
    except Exception:
      pass

    return ""

  # Clean claim numbers in parsed claims.
  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue

    original_claim_number = claim.get("claim_number") or claim.get("Claim Number") or claim.get("claim_no") or claim.get("Claim #")
    fixed_claim_number = clean_claim_number(original_claim_number)

    if fixed_claim_number and fixed_claim_number != original_claim_number:
      claim["claim_number"] = fixed_claim_number
      claim["Claim Number"] = fixed_claim_number

  # Keep profile claim arrays aligned.
  for claim_key in ["claims", "parsed_claims"]:
    profile_claims = parsed_profile.get(claim_key)
    if isinstance(profile_claims, list):
      for claim in profile_claims:
        if not isinstance(claim, dict):
          continue
        original_claim_number = claim.get("claim_number") or claim.get("Claim Number") or claim.get("claim_no") or claim.get("Claim #")
        fixed_claim_number = clean_claim_number(original_claim_number)
        if fixed_claim_number and fixed_claim_number != original_claim_number:
          claim["claim_number"] = fixed_claim_number
          claim["Claim Number"] = fixed_claim_number

  # Fill missing insured/business name from document text or filename-derived fallback.
  current_business = (
    parsed_profile.get("business_name")
    or parsed_profile.get("named_insured")
    or parsed_profile.get("insured_name")
    or parsed_profile.get("account_name")
  )

  if bad_business_value(current_business):
    business_name = extract_business_name_from_upload()
    if business_name:
      parsed_profile["business_name"] = business_name
      parsed_profile["named_insured"] = business_name
      parsed_profile["insured_name"] = business_name
      parsed_profile["account_name"] = business_name

  print("LOSSQ_FINAL_UPLOAD_CLAIM_PROFILE_CLEANUP_V1:", {
    "claims": len(parsed_claims),
    "business_name": parsed_profile.get("business_name"),
    "sample_claim_numbers": [
      item.get("claim_number") for item in parsed_claims[:3] if isinstance(item, dict)
    ],
  })

  return parsed_claims, parsed_profile

def parse_file(file_path: str, filename: str):
  # LOSSQ_PARSE_FILE_SAFE_DEFAULTS_V1
  parsed_claims = []
  parsed_profile = {}
  claims = []
  profile = {}
  policies = []

  lower_name = str(filename or "").lower()

  if lower_name.endswith(".pdf"):
    # LOSSQ_LABEL_BASED_PDF_PRIORITY_V1
    try:
      label_pdf_claims, label_pdf_profile = lossq_parse_label_based_pdf_loss_run_v1(file_path)

      # LOSSQ_LABEL_PDF_RETURN_NO_CLAIMS_PROFILE_V2
      label_pdf_no_claims_profile = (
        isinstance(label_pdf_profile, dict)
        and (
          label_pdf_profile.get("lossq_no_claims_pdf_detected")
          or label_pdf_profile.get("loss_run_status") == "No claims reported"
        )
        and isinstance(label_pdf_profile.get("policies") or label_pdf_profile.get("policy_schedule"), list)
      )

      if label_pdf_claims or label_pdf_no_claims_profile:
        # LOSSQ_LABEL_BASED_PDF_PRIORITY_POLICY_DISPLAY_CLEANUP_V2
        label_pdf_claims, label_pdf_profile = lossq_clean_policy_schedule_display_names_v2(label_pdf_claims, label_pdf_profile)
        print("LOSSQ_LABEL_PDF_PRIORITY_RETURN_V2:", {
          "claims": len(label_pdf_claims or []),
          "no_claims_profile": bool(label_pdf_no_claims_profile),
          "business_name": label_pdf_profile.get("business_name") if isinstance(label_pdf_profile, dict) else "",
          "policy_count": len((label_pdf_profile.get("policies") or label_pdf_profile.get("policy_schedule") or [])) if isinstance(label_pdf_profile, dict) else 0,
        })
        return label_pdf_claims, label_pdf_profile
    except Exception as label_pdf_exc:
      print("LOSSQ_LABEL_BASED_PDF_PARSE_ERROR:", str(label_pdf_exc)[:500])

    result = parse_loss_run_file(file_path, filename)

    profile = result.get("profile") or {}
    policies = result.get("policies") or []
    claims = result.get("claims") or []
    validation = result.get("validation") or {}

    raw_text_preview = result.get("raw_text_preview", "")[:50000]

    # LOSSQ_APPLY_EXPOSURE_INPUTS_TO_UPLOAD_PROFILE_V1
    exposure_inputs = {}
    exposure_inputs.update(extract_exposure_inputs_from_raw_text(raw_text_preview) or {})
    exposure_inputs.update(extract_exposure_inputs_from_parsed_rows(parsed_claims) or {})

    if exposure_inputs:
      profile.update({k: v for k, v in exposure_inputs.items() if v not in ("", None, [], {})})
      validation["exposure_inputs"] = exposure_inputs
      validation["exposures"] = exposure_inputs
    claims = lossq_repair_pdf_claims_from_raw_text(raw_text_preview, claims)
    result["claims"] = claims
    profile = extract_universal_profile_from_text(
      raw_text=raw_text_preview,
      existing_profile=profile,
      claims=claims,
      filename=filename,
    )

    raw_exposure_inputs = extract_exposure_inputs_from_raw_text(raw_text_preview)
    for exposure_field, exposure_value in raw_exposure_inputs.items():
      if exposure_value not in ("", None, [], {}):
        profile[exposure_field] = profile.get(exposure_field) or exposure_value

    profile["policies"] = merge_policy_lists_for_upload(
      profile.get("policies"),
      policies,
    )
    profile["validation"] = validation
    profile["raw_text_preview"] = raw_text_preview

    # LOSSQ_DISABLE_AUTO_EXPOSURE_PARSE_MERGE_V1
    # Exposure Inputs are now manual only. Do not auto-merge premium/exposure fields from uploads.
    return claims, profile

  # LOSSQ_DO_NOT_PARSE_XLSX_AS_CSV_V1
  # XLSX files are ZIP workbooks and must not be read by csv.reader.
  if lower_name.endswith(".csv"):
    # LOSSQ_FRENCH_CSV_SECTION_PARSER_CALL_V1
    try:
      french_claims, french_profile = lossq_parse_french_section_csv_v1(file_path)
      if french_claims or (isinstance(french_profile, dict) and (french_profile.get("business_name") or french_profile.get("account_number") or french_profile.get("policy_number"))):
        return french_claims, french_profile
    except Exception as exc:
      print("LOSSQ_FRENCH_CSV_SECTION_PARSE_ERROR_V1", str(exc)[:500])

    # LOSSQ_CSV_PARSE_ORDER_V3
    # Important: clean flat CSVs may still expose account/profile fields.
    # Do not return a profile-only section result before trying the clean flat parser.
    section_claims, section_profile = _lossq_live_extract_section_based_csv(file_path)
    if section_claims:
      return section_claims, section_profile

    try:
      clean_flat_claims, clean_flat_profile = lossq_parse_clean_flat_csv_v1(file_path)
      if clean_flat_claims:
        return clean_flat_claims, clean_flat_profile
    except Exception as exc:
      print("LOSSQ_CLEAN_FLAT_CSV_PARSE_ERROR:", str(exc)[:500])

    if section_profile.get("account_number") or section_profile.get("business_name"):
      return section_claims, section_profile

    if parse_claims_from_excel:
      claims = parse_claims_from_excel(file_path)
      return claims, {}
    return [], {}


  if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
    # LOSSQ_XLSX_EXCEL_PARSER_BRANCH_V1
    # Do not parse XLSX as CSV. Use the Excel parser only.
    if parse_claims_from_excel:
      claims = parse_claims_from_excel(file_path)
      return claims, {}
    return [], {}


def parse_date(value: Any):
  if not value:
    return None

  if isinstance(value, datetime):
    return value.date().isoformat()

  raw = str(value).strip()

  if not raw or raw.lower() in ["needs review", "not set", "none", "nan"]:
    return None

  formats = ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"]

  for fmt in formats:
    try:
      return datetime.strptime(raw, fmt).date().isoformat()
    except Exception:
      pass

  return raw


def days_between(start_value: Any, end_value: Any):
  start = parse_date(start_value)
  end = parse_date(end_value)

  if not start:
    return None

  try:
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) if end else datetime.now()
    return max((end_dt - start_dt).days, 0)
  except Exception:
    return None


def pick(data: dict, keys: list[str], default=None):
  for key in keys:
    if key in data and data[key] not in [None, "", "Needs Review", "Not Set"]:
      return data[key]

  return default


def clean_profile_value(value):
  if value is None:
    return ""

  cleaned = str(value).strip()

  if cleaned.lower() in ["", "none", "nan", "needs review", "not set"]:
    return ""

  return cleaned


def is_bad_policy_key_for_upload(value: Any):
  cleaned = clean_profile_value(value).upper().replace(" ", "").strip()

  if not cleaned:
    return True

  bad_values = {
    "LINE-COVERAGE",
    "LINECOVERAGE",
    "POLICY",
    "POLICYNUMBER",
    "POLICY-NUMBER",
    "ACCOUNTNUMBER",
    "ACCOUNT-NUMBER",
    "EXPOSUREBASIS",
    "EXPOSURE-BASIS",
    "CURRENT-PREMIUM",
    "EXPIRING-PREMIUM",
    "TARGET-RENEWAL",
    "TARGETRENEWAL",
  }

  if cleaned in bad_values:
    return True

  if "COVERAGE" in cleaned and not any(ch.isdigit() for ch in cleaned):
    return True

  return False


def choose_upload_account_key(profile_data: dict, direct_profile: dict | None = None):
  direct_profile = direct_profile or {}
  candidates = [
    profile_data.get("account_number"),
    profile_data.get("customer_number"),
    direct_profile.get("account_number"),
    direct_profile.get("customer_number"),
    profile_data.get("policy_number"),
    direct_profile.get("policy_number"),
  ]

  for candidate in candidates:
    cleaned = clean_profile_value(candidate)
    if cleaned and not is_bad_policy_key_for_upload(cleaned):
      return cleaned

  # LOSSQ_DO_NOT_USE_POLICY_SCHEDULE_AS_ACCOUNT_KEY_V1
  # Policies identify coverage. They are not account/customer numbers.
  # Never use policy schedule rows to populate account_number.
  return ""


# LOSSQ_TABULAR_UPLOAD_POLICY_SCHEDULE_FROM_CLAIMS_V1
def build_policy_schedule_from_claims_for_upload(claims):
  """
  CSV/XLSX files often do not include a profile-level policy schedule.
  Build one from claim rows so the dashboard keeps all account claims after reload/back navigation.
  """
  schedule = {}

  for claim in claims or []:
    if not isinstance(claim, dict):
      continue

    policy_number = clean_profile_value(
      claim.get("policy_number")
      or claim.get("policyNumber")
      or claim.get("policy_no")
      or claim.get("policy")
    )

    if not policy_number or is_bad_policy_key_for_upload(policy_number):
      continue

    key = policy_number.strip().upper()

    line_of_business = clean_profile_value(
      claim.get("line_of_business")
      or claim.get("coverage")
      or claim.get("coverage_line")
      or claim.get("lob")
      or claim.get("policy_type")
    )

    if key not in schedule:
      schedule[key] = {
        "policy_number": policy_number,
        "line_of_business": line_of_business or "Unknown",
        "claim_count": 0,
        "total_incurred": 0,
      }

    schedule[key]["claim_count"] += 1

    try:
      incurred_raw = (
        claim.get("total_incurred")
        or claim.get("incurred")
        or claim.get("loss_amount")
        or claim.get("amount")
        or 0
      )
      incurred = float(str(incurred_raw).replace("$", "").replace(",", "").strip() or 0)
    except Exception:
      incurred = 0

    schedule[key]["total_incurred"] += incurred

    if line_of_business and schedule[key].get("line_of_business") in ("", "Unknown", None):
      schedule[key]["line_of_business"] = line_of_business

  return list(schedule.values())


def merge_policy_lists_for_upload(*policy_lists):
  merged = {}

  for policy_list in policy_lists:
    if not isinstance(policy_list, list):
      continue

    for item in policy_list:
      if not isinstance(item, dict):
        continue

      key = clean_profile_value(
        item.get("policy_number") or item.get("policy") or item.get("number")
      ).upper()

      if not key or is_bad_policy_key_for_upload(key):
        key = clean_profile_value(
          item.get("line_of_business") or item.get("coverage") or item.get("policy_type")
        ).upper()

      if not key:
        continue

      existing = merged.get(key, {})
      combined = dict(existing)

      for field, value in item.items():
        if value not in ("", None) and not combined.get(field):
          combined[field] = value

      merged[key] = combined

  return list(merged.values())


def ensure_claim_timeline_columns(db: Session):
  required_columns = {
    "date_reported": "VARCHAR",
    "date_closed": "VARCHAR",
    "open_days": "INTEGER",
    "claim_age": "INTEGER",
  }

  try:
    inspector = inspect(db.bind)
    existing_columns = [column["name"] for column in inspector.get_columns("claims")]

    for column_name, column_type in required_columns.items():
      if column_name not in existing_columns:
        db.execute(text(f"ALTER TABLE claims ADD COLUMN {column_name} {column_type}"))

    db.commit()
  except Exception as e:
    # LOSSQ_UPLOAD_ERROR_TRACE_V1
    print("LOSSQ_UPLOAD_ERROR_TRACE:", traceback.format_exc())
    db.rollback()
    print(f"Claim timeline column check failed: {e}")

def ensure_account_profile_columns(db: Session):
  required_columns = {
    "writing_carrier": "VARCHAR",
    "account_number": "VARCHAR",
    "customer_number": "VARCHAR",
    "producer_number": "VARCHAR",
    "policies": "TEXT",
    "validation": "TEXT",
    "raw_text_preview": "TEXT",

    # LOSSQ_EXPOSURE_INPUT_FIELDS_V1
    "current_premium": "VARCHAR",
    "expiring_premium": "VARCHAR",
    "target_renewal_premium": "VARCHAR",
    "line_of_business": "VARCHAR",
    "state": "VARCHAR",
    "class_code": "VARCHAR",
    "class_codes": "VARCHAR",
    "limits": "VARCHAR",
    "coverage_limit": "VARCHAR",
    "deductible": "VARCHAR",
    "retention": "VARCHAR",
    "payroll": "VARCHAR",
    "revenue": "VARCHAR",
    "sales": "VARCHAR",
    "receipts": "VARCHAR",
    "employee_count": "VARCHAR",
    "vehicle_count": "VARCHAR",
    "driver_count": "VARCHAR",
    "property_tiv": "VARCHAR",
    "tiv": "VARCHAR",
    "building_value": "VARCHAR",
    "contents_value": "VARCHAR",
    "square_footage": "VARCHAR",
    "location_count": "VARCHAR",
    "unit_count": "VARCHAR",
    "cargo_limit": "VARCHAR",
    "umbrella_limit": "VARCHAR",
    "experience_mod": "VARCHAR",
    "mod": "VARCHAR",
    "exposure_change_percent": "VARCHAR",
    "cyber_revenue": "VARCHAR",
    "professional_revenue": "VARCHAR",
    "exposure_basis": "VARCHAR",
    "underwriter_notes": "TEXT",
  }

  try:
    inspector = inspect(db.bind)
    existing_columns = [
      column["name"] for column in inspector.get_columns("account_profiles")
    ]

    for column_name, column_type in required_columns.items():
      if column_name not in existing_columns:
        db.execute(
          text(
            f"ALTER TABLE account_profiles ADD COLUMN {column_name} {column_type}"
          )
        )

    db.commit()
  except Exception as e:
    print("LOSSQ_UPLOAD_ERROR_TRACE:", traceback.format_exc())
    db.rollback()
    print(f"Account profile column check failed: {e}")


def clean_cause_of_loss(value: Any):
  # LOSSQ_CLEAN_CAUSE_OF_LOSS_V1
  # Prevent parser/table headers from leaking into the Cause of Loss field.

  text_value = clean_profile_value(value)

  if not text_value:
    return ""

  stop_phrases = [
    "Total Claims",
    "Claims Total",
    "Claim Count",
    "Total Paid",
    "Paid Total",
    "Total Reserve",
    "Reserve Total",
    "Total Incurred",
    "Incurred Total",
    "Loss Summary",
    "Policy Schedule",
    "Claim #",
    "Claim Number",
    "Date of Loss",
    "Loss Date",
    "Reported Date",
    "Status Paid",
    "Status Reserve",
  ]

  for phrase in stop_phrases:
    index = text_value.lower().find(phrase.lower())
    if index > 0:
      text_value = text_value[:index].strip()

  text_value = re.sub(r"\s+", " ", text_value).strip(".,-;:")

  # Keep it readable for claim detail cards.
  if len(text_value) > 140:
    text_value = text_value[:140].rsplit(" ", 1)[0].strip(".,-;:")

  return text_value


# LOSSQ_CLAIMANT_FROM_UPLOAD_ROW_V1
def lossq_clean_claimant_value(value):
  clean = re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())
  if not clean:
    return ""

  key = re.sub(r"[^a-z0-9]+", "", clean.lower())
  blocked = {
    "",
    "claimant",
    "claimantname",
    "name",
    "na",
    "n/a",
    "none",
    "null",
    "unknown",
    "unavailable",
    "claim",
    "claimnumber",
    "policynumber",
    "policy",
    "status",
    "description",
    "totalincurred",
    "paid",
    "reserve",
    "dateofloss",
    "datereported",
  }

  if key in blocked:
    return ""

  if re.fullmatch(r"[$,\d.\s]+", clean):
    return ""

  if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", clean):
    return ""

  return clean


def lossq_extract_claimant_from_raw_claim(raw_claim):
  if not isinstance(raw_claim, dict):
    return ""

  keys = [
    "claimant",
    "Claimant",
    "claimant_name",
    "Claimant Name",
    "claimantName",
    "injured_worker",
    "Injured Worker",
    "injured_party",
    "Injured Party",
    "injured_person",
    "Injured Person",
    "employee_name",
    "Employee Name",
    "worker_name",
    "Worker Name",
    "plaintiff",
    "Plaintiff",
    "party_name",
    "Party Name",
    "driver_name",
    "Driver Name",
    "customer_name",
    "Customer Name",
    "third_party_name",
    "Third Party Name",
  ]

  for key in keys:
    value = lossq_clean_claimant_value(raw_claim.get(key))
    if value:
      return value

  for key, value in raw_claim.items():
    key_clean = re.sub(r"[^a-z0-9]+", "", str(key or "").lower())
    if any(token in key_clean for token in [
      "claimant",
      "injuredworker",
      "injuredparty",
      "injuredperson",
      "employee",
      "plaintiff",
      "thirdparty",
    ]):
      cleaned = lossq_clean_claimant_value(value)
      if cleaned:
        return cleaned

  return ""


def lossq_apply_claimant_to_normalized_claim(normalized_claim, raw_claim):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  current = lossq_clean_claimant_value(normalized_claim.get("claimant"))
  if current:
    return normalized_claim

  claimant = lossq_extract_claimant_from_raw_claim(raw_claim)
  if claimant:
    normalized_claim["claimant"] = claimant
    normalized_claim["claimant_name"] = claimant
    print("LOSSQ_CLAIMANT_EXTRACTED_FROM_UPLOAD", {
      "claim_number": str(normalized_claim.get("claim_number") or raw_claim.get("claim_number") or raw_claim.get("Claim #") or "")[:80],
      "claimant": claimant[:120],
    })

  return normalized_claim


# LOSSQ_CLAIMANT_COLUMN_ENSURE_V1
def ensure_claimant_column(db):
  try:
    rows = db.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name = 'claims'")).fetchall()
    existing = {str(row[0]).lower() for row in rows}
    if "claimant" not in existing:
      db.execute(text("ALTER TABLE claims ADD COLUMN claimant VARCHAR"))
      db.commit()
  except Exception:
    try:
      db.rollback()
    except Exception:
      pass
    try:
      db.execute(text("ALTER TABLE claims ADD COLUMN claimant VARCHAR"))
      db.commit()
    except Exception:
      try:
        db.rollback()
      except Exception:
        pass


# LOSSQ_CLAIM_DETAIL_COLUMNS_ENSURE_V1
def ensure_claim_detail_columns(db):
  columns = {
    "claimant": "VARCHAR",
    "jurisdiction_state": "VARCHAR",
    "adjuster": "VARCHAR",
    "examiner": "VARCHAR",
  }

  for column_name, column_type in columns.items():
    try:
      db.execute(text(f"ALTER TABLE claims ADD COLUMN IF NOT EXISTS {column_name} {column_type}"))
      db.commit()
    except Exception:
      try:
        db.rollback()
      except Exception:
        pass


# LOSSQ_FINAL_SAVE_CSV_FIELD_REPAIR_V3
def lossq_final_clean_v3(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def lossq_final_key_v3(value):
  return re.sub(r"[^a-z0-9]+", "", lossq_final_clean_v3(value).lower())


def lossq_final_good_v3(value):
  raw = lossq_final_clean_v3(value)
  return bool(raw and raw.lower() not in {"-", "na", "n/a", "none", "null", "unknown"})


def lossq_final_account_like_v3(value):
  raw = lossq_final_clean_v3(value).upper()
  return bool("ACCT" in raw or "ACCOUNT" in raw or "CUSTOMER" in raw or "CLIENT" in raw or "CUST" in raw)


def lossq_final_policy_like_v3(value):
  raw = lossq_final_clean_v3(value).upper()
  if not raw or lossq_final_account_like_v3(raw):
    return False
  if not re.search(r"\d", raw):
    return False
  if re.search(r"\b(GL|BOP|WC|AUTO|CA|AL|LIQ|LIQUOR|PROP|CP|UMB|UM|IM|CARGO|GAR|DOL|CY|EPL|DO|PL)\b", raw):
    return True
  return bool("-" in raw and len(raw) >= 6)


def lossq_final_first_v3(source, *labels):
  if not isinstance(source, dict):
    return ""

  wanted = {lossq_final_key_v3(label) for label in labels}

  for key, value in source.items():
    if lossq_final_key_v3(key) in wanted:
      clean = lossq_final_clean_v3(value)
      if lossq_final_good_v3(clean):
        return clean

  return ""


def lossq_final_fix_claim_detail_v3(normalized_claim, raw_claim):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  claimant = (
    lossq_final_first_v3(normalized_claim, "claimant", "claimant name")
    or lossq_final_first_v3(raw_claim, "claimant", "claimant name", "injured worker", "injured party", "employee name", "plaintiff", "customer name", "third party name")
  )

  jurisdiction_state = (
    lossq_final_first_v3(normalized_claim, "jurisdiction_state", "jurisdiction/state", "jurisdiction", "state", "venue_state", "venue state", "loss state")
    or lossq_final_first_v3(raw_claim, "jurisdiction_state", "jurisdiction/state", "jurisdiction", "state", "venue_state", "venue state", "loss state")
  )

  adjuster = (
    lossq_final_first_v3(normalized_claim, "adjuster", "examiner", "adjuster/examiner", "claim adjuster", "claim examiner", "file handler")
    or lossq_final_first_v3(raw_claim, "adjuster", "examiner", "adjuster/examiner", "claim adjuster", "claim examiner", "file handler")
  )

  if claimant:
    normalized_claim["claimant"] = claimant

  if jurisdiction_state:
    normalized_claim["jurisdiction_state"] = jurisdiction_state
    normalized_claim["venue_state"] = normalized_claim.get("venue_state") or jurisdiction_state

  if adjuster:
    normalized_claim["adjuster"] = adjuster
    normalized_claim["examiner"] = normalized_claim.get("examiner") or adjuster

  # Remove frontend/API aliases that are not DB model columns.
  normalized_claim.pop("claimant_name", None)
  normalized_claim.pop("jurisdiction", None)
  normalized_claim.pop("state", None)
  normalized_claim.pop("adjuster_examiner", None)

  return normalized_claim


def lossq_final_repair_profile_account_and_exposures_v3(parsed_profile):
  if not isinstance(parsed_profile, dict):
    return parsed_profile

  account_number = (
    parsed_profile.get("account_number")
    or parsed_profile.get("customer_number")
    or parsed_profile.get("accountNumber")
    or parsed_profile.get("customerNumber")
  )

  if account_number and lossq_final_account_like_v3(account_number):
    parsed_profile["account_number"] = lossq_final_clean_v3(account_number)
    parsed_profile["customer_number"] = parsed_profile.get("customer_number") or parsed_profile["account_number"]

  # Never allow account number to become main policy.
  for field in ["policy_number", "main_policy"]:
    if parsed_profile.get(field) and lossq_final_account_like_v3(parsed_profile.get(field)):
      parsed_profile[field] = ""

  policies = parsed_profile.get("policies") if isinstance(parsed_profile.get("policies"), list) else []
  first_policy = ""
  for policy in policies:
    if isinstance(policy, dict):
      candidate = policy.get("policy_number") or policy.get("Policy Number")
      if lossq_final_policy_like_v3(candidate):
        first_policy = candidate
        break

  if first_policy:
    parsed_profile["policy_number"] = parsed_profile.get("policy_number") or first_policy
    parsed_profile["main_policy"] = parsed_profile.get("main_policy") or first_policy

  # Build exposure inputs from exposure rows / policies when direct fields are blank.
  exposure_rows = parsed_profile.get("exposures") if isinstance(parsed_profile.get("exposures"), list) else []
  if not exposure_rows:
    exposure_rows = [p for p in policies if isinstance(p, dict)]

  exposure_inputs = parsed_profile.get("exposure_inputs") if isinstance(parsed_profile.get("exposure_inputs"), dict) else {}
  if exposure_rows:
    exposure_inputs["exposure_rows"] = exposure_rows

    def nums(*keys):
      values = []
      for row in exposure_rows:
        if not isinstance(row, dict):
          continue
        for key in keys:
          value = row.get(key)
          if value in ("", None):
            continue
          try:
            values.append(float(str(value).replace("$", "").replace(",", "")))
            break
          except Exception:
            pass
      return values

    sum_fields = [
      ("current_premium", ["current_premium", "Current Premium"]),
      ("expiring_premium", ["expiring_premium", "Expiring Premium"]),
      ("target_renewal_premium", ["target_renewal_premium", "Target Renewal Premium"]),
    ]

    max_fields = [
      ("payroll", ["payroll", "Payroll"]),
      ("revenue", ["revenue", "Revenue", "Sales", "Gross Sales"]),
      ("employee_count", ["employee_count", "Employee Count", "employees"]),
      ("vehicle_count", ["vehicle_count", "Vehicle Count", "vehicles"]),
      ("driver_count", ["driver_count", "Driver Count", "drivers"]),
      ("property_tiv", ["property_tiv", "Property TIV", "TIV"]),
    ]

    for target, keys in sum_fields:
      values = nums(*keys)
      if values:
        exposure_inputs[target] = sum(values)
        parsed_profile[target] = parsed_profile.get(target) or sum(values)

    for target, keys in max_fields:
      values = nums(*keys)
      if values:
        exposure_inputs[target] = max(values)
        parsed_profile[target] = parsed_profile.get(target) or max(values)

    parsed_profile["exposure_inputs"] = exposure_inputs
    parsed_profile["exposures"] = exposure_rows
    parsed_profile["exposure_inputs_used"] = True

  return parsed_profile


# LOSSQ_FINAL_CSV_ACCOUNT_AND_MISSING_CLAIMS_V4
def lossq_v4_clean(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def lossq_v4_key(value):
  return re.sub(r"[^a-z0-9]+", "", lossq_v4_clean(value).lower())


def lossq_v4_good(value):
  raw = lossq_v4_clean(value)
  return bool(raw and raw.lower() not in {"-", "na", "n/a", "none", "null", "unknown"})


def lossq_v4_money(value):
  raw = lossq_v4_clean(value)
  if not raw:
    return ""
  cleaned = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
  try:
    return float(cleaned)
  except Exception:
    return raw


def lossq_v4_bool(value):
  raw = lossq_v4_clean(value).lower()
  if raw in {"yes", "y", "true", "1", "litigated", "suit", "attorney"}:
    return True
  if raw in {"no", "n", "false", "0", "none", "-", "na", "n/a", ""}:
    return False
  return bool(raw)


def lossq_v4_account_like(value):
  raw = lossq_v4_clean(value).upper()
  return bool("ACCT" in raw or "ACCOUNT" in raw or "CUSTOMER" in raw or "CLIENT" in raw or "CUST" in raw)


def lossq_v4_policy_like(value):
  raw = lossq_v4_clean(value).upper()
  if not raw or lossq_v4_account_like(raw):
    return False
  if not re.search(r"\d", raw):
    return False
  if re.search(r"\b(GL|BOP|WC|AUTO|CA|AL|LIQ|LIQUOR|PROP|CP|UMB|UM|IM|CARGO|GAR|DOL|CY|EPL|DO|PL)\b", raw):
    return True
  return bool("-" in raw and len(raw) >= 6)


def lossq_v4_first(row_map, *labels):
  for label in labels:
    value = row_map.get(lossq_v4_key(label), "")
    if lossq_v4_good(value):
      return lossq_v4_clean(value)
  return ""


def lossq_v4_row_map(headers, row):
  mapped = {}
  for idx, header in enumerate(headers):
    header_key = lossq_v4_key(header)
    if header_key:
      mapped[header_key] = lossq_v4_clean(row[idx]) if idx < len(row) else ""
  return mapped


def lossq_v4_parse_csv_sections(file_path):
  import csv

  result = {
    "account_number": "",
    "customer_number": "",
    "exposure_rows": [],
    "claims": [],
  }

  if not str(file_path or "").lower().endswith(".csv"):
    return result

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return result

  # True account/customer number from label-value rows.
  account_keys = {
    "accountnumber",
    "accountno",
    "accountid",
    "customernumber",
    "customerno",
    "customerid",
    "clientnumber",
    "clientno",
    "clientid",
  }

  for row in rows[:200]:
    if len(row) < 2:
      continue

    label_key = lossq_v4_key(row[0])
    value = lossq_v4_clean(row[1])

    if label_key in account_keys and lossq_v4_good(value):
      result["account_number"] = value
      result["customer_number"] = value
      break

  def find_header(required_groups, section_words):
    section_seen = not section_words

    for idx, row in enumerate(rows):
      row_text = " ".join(lossq_v4_clean(cell).lower() for cell in row if lossq_v4_clean(cell))
      row_keys = {lossq_v4_key(cell) for cell in row if lossq_v4_clean(cell)}

      if section_words and any(word in row_text for word in section_words):
        section_seen = True
        continue

      if not section_seen:
        continue

      if all(any(option in row_keys for option in group) for group in required_groups):
        return idx, row

    return None, []

  # Exposure / policy table.
  exposure_idx, exposure_headers = find_header(
    [
      {"policynumber", "policyno", "policy"},
      {"lineofbusiness", "coverage", "policytype", "lob", "currentpremium", "exposurebasis"},
    ],
    ["exposure", "policy information", "policy schedule"],
  )

  if exposure_idx is not None:
    for row in rows[exposure_idx + 1:]:
      row_text = " ".join(lossq_v4_clean(cell).lower() for cell in row if lossq_v4_clean(cell))

      if not any(lossq_v4_clean(cell) for cell in row):
        break

      if any(stop in row_text for stop in ["claims detail", "claim detail", "loss summary", "underwriting notes"]):
        break

      row_map = lossq_v4_row_map(exposure_headers, row)
      policy_number = lossq_v4_first(row_map, "Policy Number", "Policy No", "Policy")
      line = lossq_v4_first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB")

      if not lossq_v4_policy_like(policy_number) and not line:
        continue

      result["exposure_rows"].append({
        "policy_number": policy_number,
        "policy_type": line,
        "line_of_business": line,
        "carrier": lossq_v4_first(row_map, "Carrier", "Writing Carrier"),
        "effective_date": lossq_v4_first(row_map, "Effective Date", "Policy Effective Date"),
        "expiration_date": lossq_v4_first(row_map, "Expiration Date", "Policy Expiration Date"),
        "exposure_basis": lossq_v4_first(row_map, "Exposure Basis", "Basis"),
        "exposure_value": lossq_v4_first(row_map, "Exposure Value", "Exposure"),
        "payroll": lossq_v4_money(lossq_v4_first(row_map, "Payroll")),
        "revenue": lossq_v4_money(lossq_v4_first(row_map, "Revenue", "Sales", "Gross Sales")),
        "employee_count": lossq_v4_money(lossq_v4_first(row_map, "Employee Count", "Employees")),
        "vehicle_count": lossq_v4_money(lossq_v4_first(row_map, "Vehicle Count", "Vehicles", "Autos")),
        "driver_count": lossq_v4_money(lossq_v4_first(row_map, "Driver Count", "Drivers")),
        "property_tiv": lossq_v4_money(lossq_v4_first(row_map, "Property TIV", "TIV", "Total Insured Value")),
        "current_premium": lossq_v4_money(lossq_v4_first(row_map, "Current Premium")),
        "expiring_premium": lossq_v4_money(lossq_v4_first(row_map, "Expiring Premium")),
        "target_renewal_premium": lossq_v4_money(lossq_v4_first(row_map, "Target Renewal Premium")),
      })

  # Claims detail table. This is intentionally broad so LIQ, DOL, CYBER, IM, etc. are not dropped.
  claim_idx, claim_headers = find_header(
    [
      {"claimnumber", "claimno", "claim", "claimid"},
      {"policynumber", "policyno", "policy", "paid", "reserve", "totalincurred"},
    ],
    ["claims detail", "claim detail", "claims"],
  )

  if claim_idx is not None:
    for row in rows[claim_idx + 1:]:
      row_text = " ".join(lossq_v4_clean(cell).lower() for cell in row if lossq_v4_clean(cell))

      if not any(lossq_v4_clean(cell) for cell in row):
        break

      if any(stop in row_text for stop in ["underwriting notes", "loss summary", "exposure / policy", "account information"]):
        break

      row_map = lossq_v4_row_map(claim_headers, row)

      claim_number = lossq_v4_first(row_map, "Claim Number", "Claim #", "Claim No", "Claim ID", "Claim")
      policy_number = lossq_v4_first(row_map, "Policy Number", "Policy No", "Policy")

      if not lossq_v4_good(claim_number):
        continue

      if lossq_v4_key(claim_number) in {"claimnumber", "claimno", "claimid", "claim"}:
        continue

      result["claims"].append({
        "claim_number": claim_number,
        "policy_number": policy_number,
        "line_of_business": lossq_v4_first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claim_type": lossq_v4_first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claimant": lossq_v4_first(row_map, "Claimant", "Claimant Name", "Injured Worker", "Injured Party", "Employee Name", "Plaintiff", "Customer Name", "Third Party Name"),
        "jurisdiction_state": lossq_v4_first(row_map, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "venue_state": lossq_v4_first(row_map, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "adjuster": lossq_v4_first(row_map, "Adjuster", "Adjuster/Examiner", "Examiner", "Claim Adjuster", "Claim Examiner", "File Handler"),
        "examiner": lossq_v4_first(row_map, "Examiner", "Adjuster/Examiner", "Adjuster", "Claim Examiner", "File Handler"),
        "date_of_loss": lossq_v4_first(row_map, "Date of Loss", "Loss Date"),
        "date_reported": lossq_v4_first(row_map, "Date Reported", "Reported Date"),
        "date_closed": lossq_v4_first(row_map, "Date Closed", "Closed Date"),
        "status": lossq_v4_first(row_map, "Status", "Claim Status"),
        "cause_of_loss": lossq_v4_first(row_map, "Cause of Loss", "Loss Cause", "Cause"),
        "description": lossq_v4_first(row_map, "Claim Notes", "Loss Notes", "Notes", "Narrative", "Claim Description", "Description", "Loss Description"),
        "paid_amount": lossq_v4_money(lossq_v4_first(row_map, "Paid", "Paid Amount", "Total Paid")),
        "reserve_amount": lossq_v4_money(lossq_v4_first(row_map, "Reserve", "Reserve Amount", "Outstanding Reserve")),
        "total_incurred": lossq_v4_money(lossq_v4_first(row_map, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred")),
        "litigation": lossq_v4_bool(lossq_v4_first(row_map, "Litigation", "Litigated", "Suit Filed", "Lawsuit", "Legal Status")),
        "litigation_status": lossq_v4_first(row_map, "Litigation Status", "Legal Status", "Suit Status", "Lawsuit Status"),
        "attorney_assigned": lossq_v4_first(row_map, "Attorney Assigned", "Attorney", "Attorney Name", "Attorney Involvement", "Counsel", "Claimant Counsel", "Plaintiff Attorney", "Defense Counsel", "Represented", "Claimant Represented"),
        "suit_filed": lossq_v4_bool(lossq_v4_first(row_map, "Suit Filed", "Lawsuit Filed", "Complaint Filed")),
        "venue_state": lossq_v4_first(row_map, "Venue State", "Venue", "Jurisdiction", "Jurisdiction/State", "Loss State", "State"),
        "flag": lossq_v4_first(row_map, "Flag", "Flags", "Red Flag", "Red Flags", "Claim Flag", "Alert", "Concern"),
      })

  return result


def lossq_v4_merge_csv_sections_before_save(file_path, parsed_claims, parsed_profile):
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  context = lossq_v4_parse_csv_sections(file_path)

  account_number = lossq_v4_clean(context.get("account_number"))
  if account_number:
    parsed_profile["account_number"] = account_number
    parsed_profile["customer_number"] = parsed_profile.get("customer_number") or account_number

  exposure_rows = context.get("exposure_rows") or []
  if exposure_rows:
    parsed_profile["exposures"] = exposure_rows

    exposure_inputs = parsed_profile.get("exposure_inputs") if isinstance(parsed_profile.get("exposure_inputs"), dict) else {}
    exposure_inputs["exposure_rows"] = exposure_rows

    def numbers(field):
      values = []
      for row in exposure_rows:
        try:
          value = row.get(field)
          if value not in ("", None):
            values.append(float(value))
        except Exception:
          pass
      return values

    for field in ["current_premium", "expiring_premium", "target_renewal_premium"]:
      values = numbers(field)
      if values:
        parsed_profile[field] = sum(values)
        exposure_inputs[field] = sum(values)

    for field in ["payroll", "revenue", "employee_count", "vehicle_count", "driver_count", "property_tiv"]:
      values = numbers(field)
      if values:
        parsed_profile[field] = max(values)
        exposure_inputs[field] = max(values)

    parsed_profile["exposure_inputs"] = exposure_inputs
    parsed_profile["exposure_inputs_used"] = True

  merged = []
  by_key = {}

  def claim_merge_key(claim):
    claim_number = lossq_v4_clean(claim.get("claim_number") or claim.get("Claim Number")).upper()
    policy_number = lossq_v4_clean(claim.get("policy_number") or claim.get("Policy Number")).upper()
    return f"{claim_number}|{policy_number}"

  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue

    copy_claim = dict(claim)
    merged.append(copy_claim)
    by_key[claim_merge_key(copy_claim)] = copy_claim

  overlay_fields = [
    "policy_number",
    "line_of_business",
    "claim_type",
    "claimant",
    "jurisdiction_state",
    "venue_state",
    "adjuster",
    "examiner",
    "date_of_loss",
    "date_reported",
    "date_closed",
    "status",
    "cause_of_loss",
    "description",
    "paid_amount",
    "reserve_amount",
    "total_incurred",
    "litigation",
    "attorney_assigned",
  ]

  added_claims = 0
  updated_claims = 0

  for csv_claim in context.get("claims") or []:
    mk = claim_merge_key(csv_claim)

    if mk in by_key:
      target = by_key[mk]
      for field in overlay_fields:
        value = csv_claim.get(field)
        if value not in ("", None):
          if field in {"claimant", "jurisdiction_state", "venue_state", "adjuster", "examiner"} or not target.get(field):
            target[field] = value
      updated_claims += 1
    else:
      merged.append(dict(csv_claim))
      by_key[mk] = merged[-1]
      added_claims += 1

  parsed_claims = merged

  # Rebuild policy schedule from exposures and all claims.
  claim_counts = {}
  claim_totals = {}
  claim_lines = {}

  for claim in parsed_claims:
    policy_number = lossq_v4_clean(claim.get("policy_number")).upper()
    if not lossq_v4_policy_like(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1

    try:
      claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + float(claim.get("total_incurred") or 0)
    except Exception:
      pass

    line = lossq_v4_clean(claim.get("line_of_business") or claim.get("claim_type"))
    if line:
      claim_lines[policy_number] = line

  policies_by_number = {}

  for exposure in exposure_rows:
    policy_number = lossq_v4_clean(exposure.get("policy_number")).upper()
    if not lossq_v4_policy_like(policy_number):
      continue

    policies_by_number[policy_number] = {
      "policy_number": exposure.get("policy_number"),
      "policy_type": exposure.get("policy_type") or exposure.get("line_of_business") or claim_lines.get(policy_number),
      "line_of_business": exposure.get("line_of_business") or exposure.get("policy_type") or claim_lines.get(policy_number),
      "carrier": exposure.get("carrier") or parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier"),
      "effective_date": exposure.get("effective_date") or parsed_profile.get("effective_date"),
      "expiration_date": exposure.get("expiration_date") or parsed_profile.get("expiration_date"),
      "claim_count": claim_counts.get(policy_number, 0),
      "total_incurred": claim_totals.get(policy_number, 0),
      "current_premium": exposure.get("current_premium"),
      "expiring_premium": exposure.get("expiring_premium"),
      "target_renewal_premium": exposure.get("target_renewal_premium"),
    }

  for policy_number, count in claim_counts.items():
    if policy_number not in policies_by_number:
      policies_by_number[policy_number] = {
        "policy_number": policy_number,
        "policy_type": claim_lines.get(policy_number, ""),
        "line_of_business": claim_lines.get(policy_number, ""),
        "carrier": parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier"),
        "effective_date": parsed_profile.get("effective_date"),
        "expiration_date": parsed_profile.get("expiration_date"),
        "claim_count": count,
        "total_incurred": claim_totals.get(policy_number, 0),
      }

  if policies_by_number:
    parsed_profile["policies"] = list(policies_by_number.values())
    parsed_profile["policy_schedule"] = parsed_profile["policies"]

    current_policy = parsed_profile.get("policy_number") or parsed_profile.get("main_policy")
    if not lossq_v4_policy_like(current_policy):
      first_policy = parsed_profile["policies"][0].get("policy_number")
      parsed_profile["policy_number"] = first_policy
      parsed_profile["main_policy"] = first_policy

  print("LOSSQ_FINAL_CSV_ACCOUNT_AND_MISSING_CLAIMS_V4", {
    "account_number": str(parsed_profile.get("account_number") or "")[:80],
    "csv_claims": len(context.get("claims") or []),
    "added_claims": added_claims,
    "updated_claims": updated_claims,
    "final_claims": len(parsed_claims),
    "exposure_rows": len(exposure_rows),
  })

  return parsed_claims, parsed_profile


# LOSSQ_TRUE_ACCOUNT_NUMBER_FROM_UPLOAD_CSV_V1
def lossq_true_account_number_value(value):
  raw = re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())
  if not raw:
    return ""

  compact = re.sub(r"[^A-Z0-9]+", "", raw.upper())

  true_account_signal = any(token in compact for token in [
    "ACCT",
    "ACCOUNT",
    "CUSTOMER",
    "CLIENT",
    "CUST",
  ])

  if true_account_signal:
    return raw

  return ""


def lossq_extract_true_account_number_from_upload_csv(file_path):
  import csv

  if not str(file_path or "").lower().endswith(".csv"):
    return ""

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return ""

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

  account_labels = {
    "accountnumber",
    "accountno",
    "accountid",
    "customernumber",
    "customerno",
    "customerid",
    "clientnumber",
    "clientno",
    "clientid",
  }

  for row in rows[:200]:
    if len(row) < 2:
      continue

    if key(row[0]) in account_labels:
      account_number = lossq_true_account_number_value(row[1])
      if account_number:
        return account_number

  return ""


# LOSSQ_CANADA_UPLOAD_SUPPORT_HOOKS_V3
def lossq_canada_context_text_v3(*values):
  parts = []

  def add(value):
    if value in (None, "", [], {}):
      return
    if isinstance(value, dict):
      for key, item in value.items():
        add(key)
        add(item)
      return
    if isinstance(value, (list, tuple, set)):
      for item in value:
        add(item)
      return
    parts.append(str(value))

  for value in values:
    add(value)

  return " ".join(parts)[:50000]



# LOSSQ_CANADA_UPLOAD_CLEANUP_V1_1
def lossq_canada_clean_text_v11(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

def lossq_canada_key_v11(value):
  return re.sub(r"[^a-z0-9]+", "", lossq_canada_clean_text_v11(value).lower())

def lossq_canada_is_context_v11(profile_data=None, claims=None):
  text = lossq_canada_clean_text_v11(profile_data)
  if isinstance(profile_data, dict):
    text += " " + " ".join(lossq_canada_clean_text_v11(v) for v in profile_data.values())
  if isinstance(claims, list):
    for claim in claims[:20]:
      if isinstance(claim, dict):
        text += " " + " ".join(lossq_canada_clean_text_v11(v) for v in claim.values())
  lower = text.lower()
  return any(token in lower for token in ["canada", "cad", "ca$", "c$", "ontario", "quebec", "québec", "alberta", "british columbia", "wsib", "worksafebc", "wcb"])

def lossq_canada_line_v11(value):
  raw_value = lossq_canada_clean_text_v11(value)
  key = lossq_canada_key_v11(raw_value)
  mapping = {
    "cgl": "General Liability",
    "commercialgeneralliability": "General Liability",
    "fleetautomobile": "Commercial Auto",
    "commercialautomobile": "Commercial Auto",
    "automobile": "Commercial Auto",
    "auto": "Commercial Auto",
    "wcbwsib": "Workers Compensation",
    "wsib": "Workers Compensation",
    "wcb": "Workers Compensation",
    "worksafebc": "Workers Compensation",
    "workerscompensation": "Workers Compensation",
    "errorsandomissions": "Professional Liability",
    "eo": "Professional Liability",
    "professional liability": "Professional Liability",
    "professionalliability": "Professional Liability",
    "cyberliability": "Cyber",
    "cyber": "Cyber",
    "excessliability": "Umbrella",
    "umbrella": "Umbrella",
    "commercialproperty": "Commercial Property",
    "property": "Commercial Property",
  }
  return mapping.get(key) or raw_value

def lossq_canada_date_v11(value):
  raw_value = lossq_canada_clean_text_v11(value)
  if not raw_value:
    return value
  month_map = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10, "october": 10,
    "nov": 11, "november": 11, "dec": 12, "december": 12,
  }
  m = re.match(r"^(\d{1,2})[-\s]([A-Za-zéûÉÛ]{3,})[-\s](\d{2,4})$", raw_value)
  if m:
    day = int(m.group(1))
    mon = month_map.get(m.group(2).lower().replace("é", "e"))
    year = int(m.group(3))
    if year < 100:
      year += 2000
    if mon:
      return f"{year:04d}-{mon:02d}-{day:02d}"
  m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", raw_value)
  if m:
    day = int(m.group(1))
    month = int(m.group(2))
    year = int(m.group(3))
    if year < 100:
      year += 2000
    if 1 <= day <= 31 and 1 <= month <= 12:
      return f"{year:04d}-{month:02d}-{day:02d}"
  m = re.match(r"^(20\d{2}|19\d{2})[/-](\d{1,2})[/-](\d{1,2})$", raw_value)
  if m:
    year = int(m.group(1))
    middle = int(m.group(2))
    last = int(m.group(3))
    if middle > 12 and 1 <= last <= 12:
      return f"{year:04d}-{last:02d}-{middle:02d}"
    return f"{year:04d}-{middle:02d}-{last:02d}"
  return value

def lossq_canada_policy_number_v11(value):
  policy_number = lossq_canada_clean_text_v11(value).upper().replace(" ", "")
  if not policy_number:
    return ""
  blocked = {"CLAIMSDETAIL", "CLAIMNUMBER", "UNDERWRITINGNOTES", "NOTE", "EXPECTEDTOTALCLAIMS", "EXPECTEDOPENCLAIMS", "EXPECTEDTOTALINCURREDCAD"}
  if lossq_canada_key_v11(policy_number).upper() in blocked:
    return ""
  if re.search(r"-(?:19|20)\d{2}-", policy_number):
    return policy_number
  return ""

def lossq_canada_read_csv_preamble_v11(file_path):
  result = {}
  try:
    import csv
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
      reader = csv.reader(handle)
      for row in reader:
        cells = [lossq_canada_clean_text_v11(cell) for cell in row]
        cells = [cell for cell in cells if cell]
        if len(cells) < 2:
          continue
        key = lossq_canada_key_v11(cells[0])
        value = cells[1]
        if key in {"exposurepolicyinformation", "claimsdetail", "underwritingnotes"}:
          break
        if key in {"insurer", "carrier", "insurancecompany", "underwriter"}:
          result["carrier_name"] = result.get("carrier_name") or value
          result["writing_carrier"] = result.get("writing_carrier") or value
        elif key in {"brokerage", "producingagency", "agency"}:
          result["producing_agency"] = result.get("producing_agency") or value
        elif key in {"broker", "producer"}:
          result["agency_name"] = result.get("agency_name") or value
        elif key in {"province", "provincecode"}:
          result["province"] = result.get("province") or value
        elif key in {"postalcode", "postcode"}:
          result["postal_code"] = result.get("postal_code") or value
        elif key == "country":
          result["country"] = value
        elif key in {"currency", "losscurrency"}:
          result["currency"] = value
  except Exception as exc:
    print("LOSSQ_CANADA_PREAMBLE_READ_SKIPPED_V11:", str(exc)[:200])
  return result

def lossq_canada_upload_cleanup_v11(profile_data, claims=None, file_path=None):
  if not isinstance(profile_data, dict):
    return profile_data
  if not lossq_canada_is_context_v11(profile_data, claims):
    return profile_data

  preamble = lossq_canada_read_csv_preamble_v11(file_path) if file_path else {}

  carrier = preamble.get("carrier_name") or profile_data.get("carrier_name") or profile_data.get("writing_carrier")
  if carrier:
    profile_data["carrier_name"] = carrier
    profile_data["writing_carrier"] = carrier

  if preamble.get("producing_agency"):
    profile_data["producing_agency"] = preamble.get("producing_agency")
  if preamble.get("agency_name"):
    profile_data["agency_name"] = preamble.get("agency_name")
  if preamble.get("postal_code"):
    profile_data["postal_code"] = preamble.get("postal_code")
  if preamble.get("province"):
    profile_data["province"] = preamble.get("province")
  profile_data["country"] = "Canada"
  profile_data["currency"] = "CAD"

  claim_counts = {}
  claim_totals = {}
  if isinstance(claims, list):
    for claim in claims:
      if not isinstance(claim, dict):
        continue
      policy_number = lossq_canada_clean_text_v11(claim.get("policy_number") or claim.get("policy"))
      if not policy_number:
        continue
      claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
      try:
        claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + float(claim.get("total_incurred") or 0)
      except Exception:
        pass

  cleaned_policies = []
  seen = set()
  for source_key in ["policies", "policy_schedule"]:
    rows = profile_data.get(source_key)
    if not isinstance(rows, list):
      continue
    for row in rows:
      if not isinstance(row, dict):
        continue
      policy_number = lossq_canada_policy_number_v11(row.get("policy_number") or row.get("policy") or row.get("policy_no"))
      if not policy_number or policy_number in seen:
        continue
      seen.add(policy_number)
      line = lossq_canada_line_v11(row.get("line_of_business") or row.get("coverage") or row.get("policy_type"))
      item = dict(row)
      item["policy_number"] = policy_number
      item["line_of_business"] = line
      item["coverage"] = line
      item["policy_type"] = line
      item["carrier"] = carrier or item.get("carrier") or item.get("carrier_name") or item.get("writing_carrier") or ""
      item["carrier_name"] = carrier or item.get("carrier_name") or item.get("carrier") or ""
      item["writing_carrier"] = carrier or item.get("writing_carrier") or item.get("carrier") or ""
      item["effective_date"] = lossq_canada_date_v11(item.get("effective_date"))
      item["expiration_date"] = lossq_canada_date_v11(item.get("expiration_date"))
      item["claim_count"] = claim_counts.get(policy_number, item.get("claim_count") or item.get("claims") or 0)
      item["claims"] = item["claim_count"]
      item["total_incurred"] = claim_totals.get(policy_number, item.get("total_incurred") or 0.0)
      cleaned_policies.append(item)

  if cleaned_policies:
    profile_data["policies"] = cleaned_policies
    profile_data["policy_schedule"] = cleaned_policies
    profile_data["policy_numbers"] = [p.get("policy_number") for p in cleaned_policies]

  for key in ["effective_date", "expiration_date", "policy_effective_date", "policy_expiration_date"]:
    if profile_data.get(key):
      profile_data[key] = lossq_canada_date_v11(profile_data.get(key))

  print("LOSSQ_CANADA_UPLOAD_CLEANUP_V1_1", {
    "carrier": profile_data.get("carrier_name"),
    "producing_agency": profile_data.get("producing_agency"),
    "policy_count": len(profile_data.get("policies") or []),
    "country": profile_data.get("country"),
    "currency": profile_data.get("currency"),
  })

  return profile_data

def lossq_canada_profile_hook_v3(profile_data, parsed_claims=None):
  if lossq_canada_enhance_profile_for_canada is None:
    return profile_data
  if not isinstance(profile_data, dict):
    return profile_data

  try:
    context = lossq_canada_context_text_v3(profile_data, parsed_claims)
    enhanced = lossq_canada_enhance_profile_for_canada(dict(profile_data), context)
    if not isinstance(enhanced, dict):
      return profile_data

    for key in [
      "country",
      "currency",
      "postal_code",
      "province",
      "province_code",
      "state",
      "carrier_name",
      "writing_carrier",
      "producing_agency",
      "effective_date",
      "expiration_date",
      "evaluation_date",
      "valuation_date",
    ]:
      value = enhanced.get(key)
      if value not in ("", None, [], {}):
        profile_data[key] = value

    print("LOSSQ_CANADA_PROFILE_HOOK_V3", {
      "country": profile_data.get("country"),
      "currency": profile_data.get("currency"),
      "province": profile_data.get("province") or profile_data.get("province_code"),
      "postal_code": profile_data.get("postal_code"),
    })
  except Exception as exc:
    print("LOSSQ_CANADA_PROFILE_HOOK_SKIPPED_V3:", str(exc)[:200])

  return profile_data


def lossq_canada_claim_hook_v3(normalized_claim, raw_claim=None):
  if lossq_canada_enhance_claim_for_canada is None:
    return normalized_claim
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  raw_claim = raw_claim if isinstance(raw_claim, dict) else {}

  try:
    combined = {}
    combined.update(raw_claim)
    combined.update(normalized_claim)

    context = lossq_canada_context_text_v3(raw_claim, normalized_claim)
    enhanced = lossq_canada_enhance_claim_for_canada(combined, context)
    if not isinstance(enhanced, dict):
      return normalized_claim

    for key in [
      "jurisdiction_state",
      "venue_state",
      "line_of_business",
      "paid_amount",
      "reserve_amount",
      "total_incurred",
      "date_of_loss",
      "date_reported",
      "date_closed",
    ]:
      value = enhanced.get(key)
      if value not in ("", None, [], {}):
        normalized_claim[key] = value

  except Exception as exc:
    print("LOSSQ_CANADA_CLAIM_HOOK_SKIPPED_V3:", str(exc)[:200])

  return normalized_claim

# LOSSQ_SAFE_MONEY_FLOAT_CURRENCY_V4
def lossq_safe_money_float_currency_v4(value, default=0.0):
  if value in (None, "", [], {}):
    return default

  if isinstance(value, (int, float)):
    try:
      return float(value)
    except Exception:
      return default

  raw_value = str(value or "").replace("\u00a0", " ").strip()
  if not raw_value:
    return default

  if raw_value.lower() in {"n/a", "na", "none", "null", "-", "--", "—"}:
    return default

  negative = raw_value.startswith("-") or ("(" in raw_value and ")" in raw_value)

  cleaned = raw_value
  cleaned = re.sub(r"(?i)\b(?:cad|usd|cdn|cnd)\b", "", cleaned)
  cleaned = cleaned.replace("CA$", "")
  cleaned = cleaned.replace("US$", "")
  cleaned = cleaned.replace("C$", "")
  cleaned = cleaned.replace("$", "")
  cleaned = cleaned.replace("(", "")
  cleaned = cleaned.replace(")", "")
  cleaned = re.sub(r"[^0-9,\.\- ]+", "", cleaned).strip()

  compact = cleaned.replace(" ", "")
  if not compact:
    return default

  if "," in compact and "." not in compact:
    if compact.count(",") == 1 and re.search(r",\d{1,2}$", compact):
      compact = compact.replace(",", ".")
    else:
      compact = compact.replace(",", "")
  else:
    compact = compact.replace(",", "")

  match = re.search(r"-?\d+(?:\.\d+)?", compact)
  if not match:
    return default

  try:
    amount = float(match.group(0))
  except Exception:
    return default

  if negative and amount > 0:
    amount = -amount

  return amount


# LOSSQ_CLAIM_LEGAL_SIGNAL_STRICT_SAVE_V1
def lossq_claim_legal_signal_strict_save_v1(normalized_claim, raw_claim=None):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  raw_claim = raw_claim if isinstance(raw_claim, dict) else {}

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def get_any(source, labels):
    if not isinstance(source, dict):
      return None
    wanted = {key(label) for label in labels}
    for label in labels:
      if label in source and source.get(label) not in (None, ""):
        return source.get(label)
    for source_key, source_value in source.items():
      if key(source_key) in wanted and source_value not in (None, ""):
        return source_value
    return None

  def yes_no_signal(value):
    if value is True:
      return True
    if value is False:
      return False
    text = clean(value).lower()
    if not text:
      return None
    yes_values = {"yes", "y", "true", "1"}
    no_values = {
      "no", "n", "false", "0", "none", "n/a", "na", "none present",
      "not present", "no attorney", "no attorney present", "no litigation",
      "no litigation present", "no suit", "no suit filed", "not represented",
      "unrepresented"
    }
    if text in yes_values:
      return True
    if text in no_values:
      return False
    if any(phrase in text for phrase in [
      "no attorney", "no litigation", "no suit", "none present",
      "not represented", "unrepresented"
    ]):
      return False
    return None

  attorney_labels = [
    "Attorney Assigned", "Attorney Involved", "attorney_assigned",
    "attorneyAssigned", "attorney_involved", "attorneyInvolved",
    "Counsel Assigned", "Defense Counsel"
  ]

  litigation_labels = [
    "Litigation", "Litigation Flag", "litigation", "litigation_flag",
    "litigationFlag", "Suit Filed", "suit_filed", "suitFiled",
    "Lawsuit", "Represented", "represented"
  ]

  attorney_value = get_any(raw_claim, attorney_labels)
  if attorney_value is None:
    attorney_value = get_any(normalized_claim, attorney_labels)

  litigation_value = get_any(raw_claim, litigation_labels)
  if litigation_value is None:
    litigation_value = get_any(normalized_claim, litigation_labels)

  attorney_signal = yes_no_signal(attorney_value)
  litigation_signal = yes_no_signal(litigation_value)

  legal_flag_values = {"attorney", "litigation", "suit", "open reserve"}
  existing_flag = clean(normalized_claim.get("flag") or normalized_claim.get("claim_flag") or normalized_claim.get("risk_flag"))

  if attorney_signal is True:
    normalized_claim["attorney_assigned"] = True
    normalized_claim["attorney_involved"] = True
    normalized_claim["flag"] = "Attorney"
    normalized_claim["claim_flag"] = "Attorney"
    return normalized_claim

  if litigation_signal is True:
    normalized_claim["litigation"] = True
    normalized_claim["litigation_flag"] = True
    normalized_claim["flag"] = "Litigation"
    normalized_claim["claim_flag"] = "Litigation"
    return normalized_claim

  if attorney_signal is False or litigation_signal is False:
    normalized_claim["attorney_assigned"] = False
    normalized_claim["attorney_involved"] = False
    normalized_claim["suit_filed"] = False
    normalized_claim["litigation"] = False
    normalized_claim["litigation_flag"] = False
    normalized_claim["represented"] = False
    if existing_flag.lower() in legal_flag_values:
      normalized_claim["flag"] = ""
      normalized_claim["claim_flag"] = ""
      normalized_claim["risk_flag"] = ""
    return normalized_claim

  if existing_flag.lower() in legal_flag_values:
    normalized_claim["flag"] = ""
    normalized_claim["claim_flag"] = ""
    normalized_claim["risk_flag"] = ""

  return normalized_claim


# LOSSQ_FINAL_CLAIM_AMOUNT_DATE_COERCE_BEFORE_SAVE_V1
def lossq_final_claim_amount_date_coerce_before_save_v1(normalized_claim):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  def money(value):
    try:
      return lossq_safe_money_float_currency_v4(value, 0.0)
    except Exception:
      try:
        raw_value = str(value or "").replace("\\u00a0", " ").strip()
        raw_value = re.sub(r"(?i)\\b(?:cad|usd|cdn|cnd)\\b", "", raw_value)
        raw_value = raw_value.replace("CA$", "").replace("US$", "").replace("C$", "").replace("$", "")
        raw_value = raw_value.replace(",", "").replace("(", "-").replace(")", "")
        match = re.search(r"-?\\d+(?:\\.\\d+)?", raw_value)
        return float(match.group(0)) if match else 0.0
      except Exception:
        return 0.0

  for field in ["paid_amount", "reserve_amount", "total_incurred"]:
    normalized_claim[field] = money(normalized_claim.get(field))

  if normalized_claim.get("total_incurred", 0.0) == 0.0:
    paid = normalized_claim.get("paid_amount", 0.0) or 0.0
    reserve = normalized_claim.get("reserve_amount", 0.0) or 0.0
    if paid or reserve:
      normalized_claim["total_incurred"] = paid + reserve

  def repair_date(value):
    raw_date = str(value or "").strip()
    if not raw_date:
      return value
    match = re.match(r"^(20\\d{2}|19\\d{2})-(\\d{1,2})-(\\d{1,2})$", raw_date)
    if not match:
      return value
    year = int(match.group(1))
    middle = int(match.group(2))
    last = int(match.group(3))
    # Repair accidental YYYY-DD-MM from Canadian DD/MM/YYYY parsing.
    if middle > 12 and 1 <= last <= 12:
      return f"{year:04d}-{last:02d}-{middle:02d}"
    return raw_date

  for field in ["date_of_loss", "date_reported", "date_closed"]:
    normalized_claim[field] = repair_date(normalized_claim.get(field))

  return normalized_claim

def normalize_claim_data(raw: dict, fallback_policy_number: str, current_user: dict):
  extracted_policy_number = clean_profile_value(
    pick(raw, ["policy_number", "policy_no", "policy"], "")
  )

  final_policy_number = extracted_policy_number or clean_profile_value(fallback_policy_number)

  date_of_loss = parse_date(
    pick(raw, ["date_of_loss", "loss_date", "date_of_accident", "accident_date"])
  )

  date_reported = parse_date(
    pick(raw, ["date_reported", "reported_date", "report_date"])
  )

  date_closed = parse_date(
    pick(raw, ["date_closed", "closed_date", "closure_date"])
  )

  status = pick(raw, ["status", "claim_status"], "Open")

  open_days = days_between(date_reported or date_of_loss, date_closed)
  claim_age = days_between(date_of_loss, None)

  normalized = {
    "claim_number": pick(raw, ["claim_number", "claim_no", "claim_id"], "Unknown"),
    "policy_id": raw.get("policy_id"),
    "policy_number": final_policy_number,
    "line_of_business": pick(raw, ["line_of_business", "lob", "coverage_line"]),
    "claim_type": pick(raw, ["claim_type", "type"]),
    "cause_of_loss": clean_cause_of_loss(pick(raw, ["cause_of_loss", "cause"])),
    "claimant_type": pick(raw, ["claimant_type"]),
    "date_of_loss": date_of_loss,
    "date_reported": date_reported,
    "date_closed": date_closed,
    "open_days": open_days,
    "claim_age": claim_age,
    "status": status,
    "description": pick(raw, ["description", "loss_description", "claim_description", "claim_notes", "loss_notes", "notes", "narrative", "Claim Notes", "Loss Notes", "Notes", "Narrative", "Claim Description", "Description", "Loss Description"]),
    "paid_amount": lossq_safe_money_float_currency_v4(pick(raw, ["paid_amount", "paid", "total_paid"], 0), 0.0),
    "reserve_amount": lossq_safe_money_float_currency_v4(pick(raw, ["reserve_amount", "reserve", "outstanding_reserve"], 0), 0.0),
    "total_incurred": lossq_safe_money_float_currency_v4(pick(raw, ["total_incurred", "incurred", "total"], 0), 0.0),
    "litigation": bool(pick(raw, ["litigation", "is_litigated"], False)),
    "litigation_status": pick(raw, ["litigation_status"]),
    "attorney_assigned": bool(pick(raw, ["attorney_assigned"], False)),
    "suit_filed": bool(pick(raw, ["suit_filed"], False)),
    "venue_state": pick(raw, ["venue_state"]),
    "injury_type": pick(raw, ["injury_type"]),
    "flag": pick(raw, ["flag"]),
    "organization_id": current_user["organization_id"],
    "uploaded_by_user_id": current_user["user_id"],
    "uploaded_at": datetime.now().isoformat(),
  }

  # LOSSQ_NORMALIZE_ROW_POLICY_PRESERVATION_V1
  normalized_claim = preserve_row_policy_fields(
    raw=raw,
    normalized=normalized,
    fallback_policy_number=fallback_policy_number,
  )

  # LOSSQ_CANADA_NORMALIZE_CLAIM_CALL_V3
  normalized_claim = lossq_canada_claim_hook_v3(normalized_claim, raw)
  return normalized_claim



# LOSSQ_ATTORNEY_FLAGS_BEFORE_SAVE_V1
def lossq_claim_text_clean_v1(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def lossq_claim_pick_v1(data, *keys):
  if not isinstance(data, dict):
    return ""

  lowered = {str(k).strip().lower(): v for k, v in data.items()}

  for key in keys:
    if key in data and lossq_claim_text_clean_v1(data.get(key)):
      return lossq_claim_text_clean_v1(data.get(key))

    lookup = str(key).strip().lower()
    if lookup in lowered and lossq_claim_text_clean_v1(lowered.get(lookup)):
      return lossq_claim_text_clean_v1(lowered.get(lookup))

  return ""


def lossq_claim_truthy_v1(value):
  clean = lossq_claim_text_clean_v1(value)
  low = clean.lower()

  if not clean:
    return False

  if low in {"yes", "y", "true", "1", "attorney", "represented", "litigated", "suit filed", "open"}:
    return True

  if low in {"no", "n", "false", "0", "none", "n/a", "na", "not represented", "not litigated", "unknown", "-"}:
    return False

  return any(token in low for token in [
    "attorney",
    "counsel",
    "law firm",
    "law office",
    "esq",
    "plaintiff",
    "defense",
    "litigation",
    "lawsuit",
    "suit filed",
    "docket",
    "represented",
  ])


def lossq_claim_money_float_v1(value):
  raw_value = lossq_claim_text_clean_v1(value)
  if not raw_value:
    return 0.0

  raw_value = raw_value.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
  raw_value = re.sub(r"[^0-9.\-]+", "", raw_value)

  try:
    return float(raw_value or 0)
  except Exception:
    return 0.0


def lossq_apply_attorney_flags_before_save_v1(normalized_claim, raw_claim=None):
  normalized_claim = normalized_claim if isinstance(normalized_claim, dict) else {}
  raw_claim = raw_claim if isinstance(raw_claim, dict) else {}

  combined_values = []
  for source in (raw_claim, normalized_claim):
    if isinstance(source, dict):
      for value in source.values():
        clean = lossq_claim_text_clean_v1(value)
        if clean:
          combined_values.append(clean)

  combined_text = " | ".join(combined_values)
  combined_low = combined_text.lower()

  attorney_value = lossq_claim_pick_v1(
    raw_claim,
    "Attorney Assigned",
    "Attorney",
    "Attorney Name",
    "Attorney Involvement",
    "Counsel",
    "Claimant Counsel",
    "Plaintiff Counsel",
    "Plaintiff Attorney",
    "Defense Counsel",
    "Law Firm",
    "Attorney Firm",
    "Represented",
    "Representation",
    "Claimant Represented",
  )

  litigation_value = lossq_claim_pick_v1(
    raw_claim,
    "Litigation",
    "Litigated",
    "Litigation Status",
    "Suit Filed",
    "Lawsuit",
    "Legal Status",
    "Court",
    "Docket",
  )

  suit_value = lossq_claim_pick_v1(
    raw_claim,
    "Suit Filed",
    "Lawsuit Filed",
    "Complaint Filed",
    "Docket",
  )

  flag_value = lossq_claim_pick_v1(
    raw_claim,
    "Flag",
    "Flags",
    "Red Flag",
    "Red Flags",
    "Claim Flag",
    "Alert",
    "Concern",
  )

  venue_state = lossq_claim_pick_v1(
    raw_claim,
    "Venue State",
    "Venue",
    "Jurisdiction",
    "Jurisdiction/State",
    "Loss State",
    "State",
  )

  attorney_signal = (
    lossq_claim_truthy_v1(attorney_value)
    or any(token in combined_low for token in [
      "attorney",
      "claimant counsel",
      "plaintiff counsel",
      "defense counsel",
      "law firm",
      "represented by",
    ])
  )

  litigation_signal = (
    lossq_claim_truthy_v1(litigation_value)
    or lossq_claim_truthy_v1(suit_value)
    or any(token in combined_low for token in [
      "litigation",
      "litigated",
      "lawsuit",
      "suit filed",
      "complaint filed",
      "docket",
      "court filing",
    ])
  )

  suit_signal = (
    lossq_claim_truthy_v1(suit_value)
    or any(token in combined_low for token in [
      "suit filed",
      "lawsuit filed",
      "complaint filed",
      "docket",
    ])
  )

  reserve = lossq_claim_money_float_v1(
    normalized_claim.get("reserve_amount")
    or raw_claim.get("Reserve")
    or raw_claim.get("Reserve Amount")
    or raw_claim.get("Outstanding Reserve")
  )

  incurred = lossq_claim_money_float_v1(
    normalized_claim.get("total_incurred")
    or raw_claim.get("Total Incurred")
    or raw_claim.get("Incurred")
  )

  status = lossq_claim_text_clean_v1(
    normalized_claim.get("status")
    or raw_claim.get("Status")
    or raw_claim.get("Claim Status")
  ).lower()

  flags = []

  existing_flag = lossq_claim_text_clean_v1(normalized_claim.get("flag") or flag_value)
  if existing_flag and existing_flag.lower() not in {"none", "n/a", "na", "-", "no"}:
    flags.append(existing_flag)

  if attorney_signal:
    flags.append("Attorney Involved")

  if litigation_signal:
    flags.append("Litigation")

  if suit_signal:
    flags.append("Suit Filed")

  if reserve > 0:
    flags.append("Outstanding Reserve")

  if status in {"open", "reopened", "reopen", "pending", "active"}:
    flags.append("Open Claim")

  if "reopen" in combined_low:
    flags.append("Reopened Claim")

  if incurred >= 100000:
    flags.append("Large Loss")

  if any(token in combined_low for token in ["subrogation", "subro"]):
    flags.append("Subrogation")

  if any(token in combined_low for token in ["fraud", "siu", "special investigation"]):
    flags.append("SIU / Fraud Review")

  if any(token in combined_low for token in ["disputed", "coverage dispute", "denied", "denial", "reservation of rights"]):
    flags.append("Coverage / Liability Dispute")

  if any(token in combined_low for token in ["late reported", "late report", "late notice"]):
    flags.append("Late Reported")

  if any(token in combined_low for token in ["fatality", "death", "deceased"]):
    flags.append("Fatality")

  if any(token in combined_low for token in ["surgery", "amputation", "fracture", "hospitalized", "serious injury"]):
    flags.append("Severe Injury")

  unique_flags = []
  seen = set()
  for item in flags:
    clean_item = lossq_claim_text_clean_v1(item)
    key = clean_item.lower()
    if clean_item and key not in seen:
      unique_flags.append(clean_item)
      seen.add(key)

  if unique_flags:
    normalized_claim["flag"] = "; ".join(unique_flags[:8])

  if attorney_signal:
    normalized_claim["attorney_assigned"] = True

  if litigation_signal:
    normalized_claim["litigation"] = True
    normalized_claim["litigation_status"] = lossq_claim_text_clean_v1(litigation_value) or ("Suit Filed" if suit_signal else "Litigation indicated")

  if suit_signal:
    normalized_claim["suit_filed"] = True

  if venue_state and not lossq_claim_text_clean_v1(normalized_claim.get("venue_state")):
    normalized_claim["venue_state"] = venue_state

  return normalized_claim


# LOSSQ_CLEAN_EXPOSURE_LIMITS_FIELD_V1
def lossq_clean_exposure_limits_field(profile_data: dict):
  """
  Keep exposure_basis separate from policy limits.
  If limits contains a full exposure sentence, replace it with coverage_limit when available.
  """
  profile_data = dict(profile_data or {})

  raw_limits = str(profile_data.get("limits") or "").strip()
  lower_limits = raw_limits.lower()

  looks_like_exposure_basis = any(token in lower_limits for token in [
    "payroll",
    "revenue",
    "employees",
    "vehicles",
    "drivers",
    "umbrella",
    "gl limit",
    "exposure basis",
  ])

  if looks_like_exposure_basis:
    coverage_limit = profile_data.get("coverage_limit") or profile_data.get("policy_limit") or ""
    profile_data["limits"] = coverage_limit or ""

  return profile_data




# LOSSQ_FINAL_PROFILE_DATA_BUSINESS_NAME_REPAIR_V3
def lossq_final_profile_data_business_name_repair_v3(file_path, profile_data=None, direct_profile=None):
  """
  Final account-profile business-name repair after extract_profile_data().

  This fixes the Account Snapshot when the parser captured label text such as:
  - / Business Name
  - Business Name
  - Named Insured / Business Name

  Universal only. No customer, carrier, or demo-file hardcoding.
  """
  import os
  import re

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def compact(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def clean_business(value):
    value = clean(value)

    value = re.sub(
      r"(?i)^\s*(?:business\s+name|named\s+insured|insured|insured\s*/\s*business\s*name|account\s+name|applicant|entity|company\s+name)\s*[:#\-\/]*\s*",
      "",
      value,
    )

    value = re.sub(
      r"(?i)\s+(?:loss\s+run\s+report|loss\s+run|report|pdf|document)\s*$",
      "",
      value,
    )

    return clean(value)

  def is_bad(value):
    k = compact(value)
    return k in {
      "",
      "businessname",
      "businessnamenotset",
      "namedinsured",
      "insured",
      "insuredname",
      "insuredbusinessname",
      "namedinsuredbusinessname",
      "accountname",
      "accountnamebusinessname",
      "applicant",
      "entity",
      "companyname",
      "unknown",
      "notset",
      "na",
      "none",
      "null",
    }

  def good(value):
    value = clean_business(value)
    if is_bad(value):
      return ""

    low = value.lower()
    blocked = [
      "loss run",
      "policy schedule",
      "claim detail",
      "claim block",
      "claim number",
      "policy number",
      "line of business",
      "coverage",
      "loss summary",
      "paid amount",
      "reserve amount",
      "total incurred",
      "date of loss",
      "effective date",
      "expiration date",
      "evaluation date",
      "valuation date",
      "account number",
      "customer number",
      "carrier",
      "writing carrier",
      "producer",
      "producing agency",
    ]

    if any(part in low for part in blocked):
      return ""

    if len(value) < 4 or len(value) > 140:
      return ""

    return value

  def read_pdf_text():
    parts = [
      profile_data.get("raw_text"),
      profile_data.get("raw_text_preview"),
      profile_data.get("ocr_text"),
      profile_data.get("document_text"),
      direct_profile.get("raw_text"),
      direct_profile.get("raw_text_preview"),
      direct_profile.get("ocr_text"),
      direct_profile.get("document_text"),
    ]

    if str(file_path or "").lower().endswith(".pdf"):
      try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        parts.append("\n".join((page.extract_text() or "") for page in reader.pages))
      except Exception as exc:
        print("LOSSQ_FINAL_PROFILE_DATA_BUSINESS_NAME_REPAIR_READ_ERROR_V3:", str(exc)[:200])

    return "\n".join(str(part or "") for part in parts if part)

  raw_text = read_pdf_text()
  lines = [clean(line) for line in raw_text.splitlines()]
  lines = [line for line in lines if line]

  candidates = []

  # Same-line labels.
  same_line = re.compile(
    r"(?i)^\s*(?:named\s+insured|insured\s*/\s*business\s*name|insured|business\s+name|account\s+name|applicant|entity|company\s+name)\s*[:#\-]\s*(.+?)\s*$"
  )

  for line in lines:
    m = same_line.match(line)
    if m:
      candidate = good(m.group(1))
      if candidate:
        candidates.append(candidate)

  # Label on one line, value on next lines.
  label_keys = {
    "namedinsured",
    "insured",
    "businessname",
    "accountname",
    "applicant",
    "entity",
    "companyname",
    "insuredbusinessname",
    "namedinsuredbusinessname",
  }

  for idx, line in enumerate(lines):
    if compact(line) in label_keys:
      for j in range(idx + 1, min(idx + 10, len(lines))):
        candidate = good(lines[j])
        if candidate:
          candidates.append(candidate)
          break

  # Entity suffix fallback from document text.
  normalized_text = re.sub(r"\s+", " ", raw_text or "").strip()
  entity_pattern = re.compile(
    r"(?is)\b([A-Z][A-Za-z0-9&.,'’\- ]{2,110}?\s+"
    r"(?:LLC|L\.L\.C\.|Inc\.?|Incorporated|Corp\.?|Corporation|Co\.?|Company|PLLC|LP|LLP|Group|Services|Service|Agency|Associates|Partners|Enterprises|Holdings))\b"
  )

  for match in entity_pattern.finditer(normalized_text):
    candidate = good(match.group(1))
    if candidate:
      candidates.append(candidate)

  # Filename fallback only if no document candidate exists.
  if not candidates:
    try:
      base = os.path.basename(str(file_path or ""))
      base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
      base = re.sub(r"(?i)\b(lossq|loss|run|messy|clean|ready|submission|pdf|csv|xlsx|xls|test|v\d+|parser|friendly)\b", " ", base)
      base = re.sub(r"[_\-]+", " ", base)
      base = re.sub(r"\s+", " ", base).strip()
      candidate = good(base.title())
      if candidate:
        candidates.append(candidate)
    except Exception:
      pass

  current = (
    profile_data.get("business_name")
    or profile_data.get("named_insured")
    or profile_data.get("insured_name")
    or profile_data.get("account_name")
  )

  business_name = ""
  for candidate in candidates:
    candidate = good(candidate)
    if candidate:
      business_name = candidate
      break

  # If current value is only a label, replace it.
  if business_name and (is_bad(current) or clean_business(current) != business_name):
    for target in [profile_data, direct_profile]:
      target["business_name"] = business_name
      target["named_insured"] = business_name
      target["insured_name"] = business_name
      target["account_name"] = business_name

  # If no candidate was found, at least remove the bad label so UI does not show it.
  if not business_name and is_bad(current):
    for field in ["business_name", "named_insured", "insured_name", "account_name"]:
      profile_data[field] = ""

  print("LOSSQ_FINAL_PROFILE_DATA_BUSINESS_NAME_REPAIR_V3:", {
    "business_name": profile_data.get("business_name"),
    "candidate_count": len(candidates),
    "current_before": current,
  })

  return profile_data, direct_profile



# LOSSQ_EXCEL_ONLY_PROFILE_REPAIR_V2
def lossq_excel_only_profile_repair_v2(file_path, profile_data=None, direct_profile=None, parsed_claims=None):
  """
  Excel-only final profile repair.

  This does not run for PDF or CSV.
  Purpose:
  - Fix Excel uploads where claims parse but Account Snapshot uses filename fallback.
  - Pull insured/business name and dates from parsed claim rows first.
  - Fallback to workbook cells if claim rows do not contain profile fields.
  """
  import re
  import os
  import datetime as _dt

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  lower_path = str(file_path or "").lower()
  if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xls")):
    return profile_data, direct_profile

  def clean(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%m/%d/%Y")
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|/")

  def compact(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def normalize_date(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%m/%d/%Y")

    raw = clean(value)
    if not raw:
      return ""

    # Excel serial date fallback.
    if re.fullmatch(r"\d{5}", raw):
      try:
        serial = int(raw)
        dt = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=serial)
        if 1990 <= dt.year <= 2100:
          return dt.strftime("%m/%d/%Y")
      except Exception:
        pass

    formats = [
      "%m/%d/%Y",
      "%m/%d/%y",
      "%Y-%m-%d",
      "%m-%d-%Y",
      "%m-%d-%y",
      "%B %d, %Y",
      "%b %d, %Y",
    ]

    for fmt in formats:
      try:
        return _dt.datetime.strptime(raw, fmt).strftime("%m/%d/%Y")
      except Exception:
        pass

    match = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", raw)
    if match:
      mm, dd, yy = match.group(1), match.group(2), match.group(3)
      if len(yy) == 2:
        yy = "20" + yy
      try:
        return _dt.datetime(int(yy), int(mm), int(dd)).strftime("%m/%d/%Y")
      except Exception:
        pass

    return ""

  def clean_business(value):
    value = clean(value)

    # Remove timestamp/filename fallback noise.
    value = re.sub(r"^\d{8,20}\s+", "", value)
    value = re.sub(r"(?i)^\s*(?:lossq\s+)?(?:excel\s+)?(?:loss\s+run\s+)?", "", value)

    # Remove captured labels.
    value = re.sub(
      r"(?i)^\s*(?:business\s+name|named\s+insured|insured|insured\s*/\s*business\s*name|account\s+name|applicant|entity|company\s+name)\s*[:#\-\/]*\s*",
      "",
      value,
    )

    value = re.sub(
      r"(?i)\s+(?:loss\s+run\s+report|loss\s+run|report|excel|xlsx|xls|document)\s*$",
      "",
      value,
    )

    return clean(value)

  def is_bad_business(value):
    k = compact(value)
    return (
      not k
      or k in {
        "businessname",
        "businessnamenotset",
        "namedinsured",
        "insured",
        "insuredname",
        "accountname",
        "applicant",
        "entity",
        "companyname",
        "unknown",
        "notset",
        "na",
        "none",
        "null",
      }
      or k.startswith("lossqexcellossrun")
      or k.startswith("excellossrun")
      or bool(re.match(r"^\d{8,20}", clean(value)))
    )

  def good_business(value):
    value = clean_business(value)
    if is_bad_business(value):
      return ""

    low = value.lower()
    blocked = [
      "policy schedule",
      "claim detail",
      "claims detail",
      "claim number",
      "policy number",
      "line of business",
      "coverage",
      "loss summary",
      "paid amount",
      "reserve amount",
      "total incurred",
      "effective date",
      "expiration date",
      "evaluation date",
      "valuation date",
      "account number",
      "customer number",
      "carrier",
      "writing carrier",
      "producer",
      "producing agency",
    ]

    if any(part in low for part in blocked):
      return ""

    if len(value) < 4 or len(value) > 140:
      return ""

    return value

  def first_from_claims(keys, is_date=False):
    for claim in parsed_claims:
      if not isinstance(claim, dict):
        continue
      for key in keys:
        value = claim.get(key)
        if is_date:
          normalized = normalize_date(value)
          if normalized:
            return normalized
        else:
          candidate = good_business(value)
          if candidate:
            return candidate
    return ""

  # 1) First trust values already parsed into claim rows.
  business_name = first_from_claims([
    "business_name",
    "named_insured",
    "insured_name",
    "account_name",
    "insured",
  ])

  effective_date = first_from_claims([
    "effective_date",
    "policy_effective_date",
    "policy_effective",
    "eff_date",
    "effective",
  ], is_date=True)

  expiration_date = first_from_claims([
    "expiration_date",
    "policy_expiration_date",
    "policy_expiration",
    "exp_date",
    "expiration",
  ], is_date=True)

  evaluation_date = first_from_claims([
    "evaluation_date",
    "valuation_date",
    "loss_run_date",
    "report_date",
    "as_of_date",
    "as-of_date",
  ], is_date=True)

  # 2) Workbook cell fallback.
  rows = []
  try:
    if lower_path.endswith(".xlsx"):
      from openpyxl import load_workbook
      wb = load_workbook(file_path, data_only=True, read_only=True)
      for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
          values = [clean(cell) for cell in row]
          if any(values):
            rows.append(values)
    elif lower_path.endswith(".xls"):
      try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
          for r in range(sheet.nrows):
            values = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(values):
              rows.append(values)
      except Exception as exc:
        print("LOSSQ_EXCEL_ONLY_PROFILE_REPAIR_XLS_SKIPPED_V2:", str(exc)[:200])
  except Exception as exc:
    print("LOSSQ_EXCEL_ONLY_PROFILE_REPAIR_WORKBOOK_READ_ERROR_V2:", str(exc)[:200])

  flat = []
  for row in rows:
    for cell in row:
      if clean(cell):
        flat.append(clean(cell))

  def label_value(labels):
    wanted = {compact(label) for label in labels}

    # Same row label -> value to right.
    for row in rows:
      for idx, cell in enumerate(row):
        if compact(cell) in wanted:
          for j in range(idx + 1, min(idx + 8, len(row))):
            value = clean(row[j])
            if value and compact(value) not in wanted:
              return value

    # Same cell "Label: Value".
    for cell in flat:
      match = re.match(r"^\s*([^:]{2,80})\s*:\s*(.+?)\s*$", cell)
      if match and compact(match.group(1)) in wanted:
        return clean(match.group(2))

    # Vertical label -> next non-empty cell.
    for idx, cell in enumerate(flat):
      if compact(cell) in wanted:
        for j in range(idx + 1, min(idx + 8, len(flat))):
          value = clean(flat[j])
          if value and compact(value) not in wanted:
            return value

    return ""

  if not business_name:
    business_name = good_business(label_value([
      "Business Name",
      "Named Insured",
      "Insured",
      "Insured Name",
      "Account Name",
      "Applicant",
      "Entity",
      "Company Name",
    ]))

  if not effective_date:
    effective_date = normalize_date(label_value([
      "Effective Date",
      "Policy Effective Date",
      "Policy Effective",
      "Eff Date",
      "Effective",
      "Policy Period Start",
    ]))

  if not expiration_date:
    expiration_date = normalize_date(label_value([
      "Expiration Date",
      "Policy Expiration Date",
      "Policy Expiration",
      "Exp Date",
      "Expiration",
      "Policy Period End",
    ]))

  if not evaluation_date:
    evaluation_date = normalize_date(label_value([
      "Evaluation Date",
      "Valuation Date",
      "Loss Run Date",
      "Report Date",
      "As Of Date",
      "As-Of Date",
    ]))

  # 3) Clean current filename fallback if it is already present.
  current_business = (
    profile_data.get("business_name")
    or profile_data.get("named_insured")
    or profile_data.get("insured_name")
    or profile_data.get("account_name")
  )

  if not business_name:
    candidate = good_business(current_business)
    if candidate:
      business_name = candidate

  # 4) Filename fallback only if there is no workbook/claim value.
  if not business_name:
    try:
      base = os.path.basename(str(file_path or ""))
      base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
      base = re.sub(r"^\d{8,20}\s*", "", base)
      base = re.sub(r"(?i)\b(lossq|loss|run|excel|messy|clean|ready|submission|pdf|csv|xlsx|xls|test|v\d+|parser|friendly)\b", " ", base)
      base = re.sub(r"[_\-]+", " ", base)
      base = re.sub(r"\s+", " ", base).strip()
      business_name = good_business(base.title())
    except Exception:
      business_name = ""

  if business_name:
    for target in [profile_data, direct_profile]:
      target["business_name"] = business_name
      target["named_insured"] = business_name
      target["insured_name"] = business_name
      target["account_name"] = business_name

  if effective_date:
    profile_data["effective_date"] = effective_date
    profile_data["policy_effective_date"] = effective_date
    direct_profile["effective_date"] = effective_date

  if expiration_date:
    profile_data["expiration_date"] = expiration_date
    profile_data["policy_expiration_date"] = expiration_date
    direct_profile["expiration_date"] = expiration_date

  if evaluation_date:
    profile_data["evaluation_date"] = evaluation_date
    profile_data["valuation_date"] = evaluation_date
    direct_profile["evaluation_date"] = evaluation_date

  print("LOSSQ_EXCEL_ONLY_PROFILE_REPAIR_V2:", {
    "business_name": profile_data.get("business_name"),
    "effective_date": profile_data.get("effective_date"),
    "expiration_date": profile_data.get("expiration_date"),
    "evaluation_date": profile_data.get("evaluation_date"),
    "claims": len(parsed_claims),
    "rows": len(rows),
  })

  return profile_data, direct_profile



# LOSSQ_EXCEL_ONLY_POLICY_SCHEDULE_REPAIR_V1
def lossq_excel_only_policy_schedule_repair_v1(file_path, profile_data=None, direct_profile=None, parsed_claims=None):
  """
  Excel-only policy schedule repair.

  Purpose:
  - Preserve Excel policy schedule rows even when a policy has zero claims.
  - Keeps Umbrella, Cyber, Liquor, BOP, WC, GL, Auto, etc. on the account.
  - Does not run for PDF or CSV.
  - No customer/file hardcoding.
  """
  import re
  import datetime as _dt

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  direct_profile = direct_profile if isinstance(direct_profile, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  lower_path = str(file_path or "").lower()
  if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xls")):
    return profile_data, direct_profile

  def clean(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm_key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def normalize_policy_number(value):
    raw = clean(value).upper()
    raw = re.sub(r"\s+", "", raw)
    raw = raw.strip(" :;,.|/\\")
    return raw

  def looks_like_account(value):
    text = normalize_policy_number(value)
    return bool(re.search(r"(ACCT|ACCOUNT|CUSTOMER|CUST|CLIENT)", text))

  def looks_like_policy(value):
    text = normalize_policy_number(value)
    if not text or looks_like_account(text):
      return False

    # Generic commercial policy number pattern:
    # GL-2026-MBG200, UMB-2026-MBG203, WC-2026-APR101, etc.
    if re.match(r"^[A-Z]{1,10}-\d{2,6}-[A-Z0-9]{2,20}$", text):
      return True

    # Also allow compact carrier formats like GL2026MBG200 if needed.
    if re.match(r"^[A-Z]{2,10}\d{4}[A-Z0-9]{2,20}$", text):
      return True

    return False

  def line_from_policy_prefix(policy_number):
    p = normalize_policy_number(policy_number)
    prefix = p.split("-")[0] if "-" in p else re.match(r"^[A-Z]+", p).group(0) if re.match(r"^[A-Z]+", p) else ""

    mapping = {
      "GL": "General Liability",
      "CGL": "General Liability",
      "WC": "Workers Compensation",
      "BOP": "Businessowners Policy",
      "CP": "Commercial Property",
      "PROP": "Property",
      "PROPERTY": "Property",
      "UMB": "Umbrella",
      "UM": "Umbrella",
      "EXCESS": "Umbrella",
      "LIQ": "Liquor Liability",
      "LIQUOR": "Liquor Liability",
      "AUTO": "Commercial Auto",
      "CA": "Commercial Auto",
      "GAR": "Garage Liability",
      "DOL": "Dealers Open Lot",
      "CY": "Cyber Liability",
      "CYBER": "Cyber Liability",
      "PL": "Professional Liability",
      "EPLI": "Employment Practices Liability",
      "DO": "Directors & Officers",
      "DNO": "Directors & Officers",
      "IM": "Inland Marine",
      "CARGO": "Motor Truck Cargo",
    }

    return mapping.get(prefix, "")

  def normalize_date(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")

    raw = clean(value)
    if not raw:
      return ""

    if re.fullmatch(r"\d{5}", raw):
      try:
        dt = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(raw))
        if 1990 <= dt.year <= 2100:
          return dt.strftime("%Y-%m-%d")
      except Exception:
        pass

    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"]:
      try:
        return _dt.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
      except Exception:
        pass

    m = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", raw)
    if m:
      mm, dd, yy = m.group(1), m.group(2), m.group(3)
      if len(yy) == 2:
        yy = "20" + yy
      try:
        return _dt.datetime(int(yy), int(mm), int(dd)).strftime("%Y-%m-%d")
      except Exception:
        pass

    return ""

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      val = float(raw or 0)
      return -val if neg else val
    except Exception:
      return 0.0

  rows = []

  try:
    if lower_path.endswith(".xlsx"):
      from openpyxl import load_workbook
      wb = load_workbook(file_path, data_only=True, read_only=True)
      for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
          values = [clean(cell) for cell in row]
          if any(values):
            rows.append(values)
    elif lower_path.endswith(".xls"):
      try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
          for r in range(sheet.nrows):
            values = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(values):
              rows.append(values)
      except Exception as exc:
        print("LOSSQ_EXCEL_ONLY_POLICY_SCHEDULE_REPAIR_XLS_SKIPPED_V1:", str(exc)[:200])
  except Exception as exc:
    print("LOSSQ_EXCEL_ONLY_POLICY_SCHEDULE_REPAIR_READ_ERROR_V1:", str(exc)[:200])
    return profile_data, direct_profile

  # Existing policies from profile.
  policy_map = {}

  def add_policy(policy_number, line_of_business="", carrier="", effective_date="", expiration_date="", premium=0.0):
    policy_number = normalize_policy_number(policy_number)
    if not looks_like_policy(policy_number):
      return

    existing = policy_map.get(policy_number, {})
    line = clean(line_of_business) or clean(existing.get("line_of_business")) or line_from_policy_prefix(policy_number)
    carrier_value = clean(carrier) or clean(existing.get("carrier")) or clean(profile_data.get("carrier_name") or profile_data.get("writing_carrier"))
    eff = normalize_date(effective_date) or normalize_date(existing.get("effective_date")) or normalize_date(profile_data.get("effective_date"))
    exp = normalize_date(expiration_date) or normalize_date(existing.get("expiration_date")) or normalize_date(profile_data.get("expiration_date"))
    prem = money(premium) or money(existing.get("premium")) or money(existing.get("current_premium"))

    policy_map[policy_number] = {
      "policy_number": policy_number,
      "line_of_business": line,
      "coverage": line,
      "policy_type": line,
      "carrier": carrier_value,
      "effective_date": eff,
      "expiration_date": exp,
      "premium": prem,
      "current_premium": prem,
    }

  for source in [
    profile_data.get("policies"),
    profile_data.get("policy_schedule"),
    direct_profile.get("policies"),
    direct_profile.get("policy_schedule"),
  ]:
    if isinstance(source, list):
      for item in source:
        if isinstance(item, dict):
          add_policy(
            item.get("policy_number") or item.get("policy") or item.get("policy_no"),
            item.get("line_of_business") or item.get("coverage") or item.get("policy_type"),
            item.get("carrier") or item.get("carrier_name") or item.get("writing_carrier"),
            item.get("effective_date") or item.get("policy_effective_date"),
            item.get("expiration_date") or item.get("policy_expiration_date"),
            item.get("premium") or item.get("current_premium"),
          )

  # Claim counts/totals from parsed claims.
  claim_counts = {}
  claim_totals = {}
  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue
    policy_number = normalize_policy_number(claim.get("policy_number") or claim.get("Policy Number"))
    if looks_like_policy(policy_number):
      claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
      claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + money(
        claim.get("total_incurred") or claim.get("incurred") or claim.get("total")
      )
      add_policy(
        policy_number,
        claim.get("line_of_business") or claim.get("claim_type") or claim.get("coverage"),
        claim.get("carrier_name") or claim.get("writing_carrier"),
        claim.get("effective_date"),
        claim.get("expiration_date"),
        0,
      )

  # Workbook policy extraction.
  header_aliases = {
    "policy": ["policy", "policy number", "policy no", "policy #", "policy_number"],
    "line": ["line", "line of business", "coverage", "policy type", "lob"],
    "carrier": ["carrier", "writing carrier", "insurance carrier"],
    "effective": ["effective", "effective date", "policy effective date", "eff date"],
    "expiration": ["expiration", "expiration date", "policy expiration date", "exp date"],
    "premium": ["premium", "current premium", "annual premium", "written premium"],
  }

  def find_header_index(row, aliases):
    alias_keys = {norm_key(a) for a in aliases}
    for i, cell in enumerate(row):
      if norm_key(cell) in alias_keys:
        return i
    return None

  for r_idx, row in enumerate(rows):
    # Header row table extraction.
    policy_idx = find_header_index(row, header_aliases["policy"])
    if policy_idx is not None:
      line_idx = find_header_index(row, header_aliases["line"])
      carrier_idx = find_header_index(row, header_aliases["carrier"])
      eff_idx = find_header_index(row, header_aliases["effective"])
      exp_idx = find_header_index(row, header_aliases["expiration"])
      prem_idx = find_header_index(row, header_aliases["premium"])

      for data_row in rows[r_idx + 1 : min(r_idx + 25, len(rows))]:
        if policy_idx >= len(data_row):
          continue
        policy_number = normalize_policy_number(data_row[policy_idx])

        if not looks_like_policy(policy_number):
          continue

        add_policy(
          policy_number,
          data_row[line_idx] if line_idx is not None and line_idx < len(data_row) else "",
          data_row[carrier_idx] if carrier_idx is not None and carrier_idx < len(data_row) else "",
          data_row[eff_idx] if eff_idx is not None and eff_idx < len(data_row) else "",
          data_row[exp_idx] if exp_idx is not None and exp_idx < len(data_row) else "",
          data_row[prem_idx] if prem_idx is not None and prem_idx < len(data_row) else 0,
        )

    # Pair/near-row extraction for loose worksheets.
    for c_idx, cell in enumerate(row):
      policy_number = normalize_policy_number(cell)
      if not looks_like_policy(policy_number):
        continue

      nearby = []
      for rr in rows[max(0, r_idx - 2): min(len(rows), r_idx + 3)]:
        nearby.extend(rr)

      nearby_text = " ".join(clean(x) for x in nearby)
      line = ""
      for value in nearby:
        maybe_line = clean(value)
        if maybe_line and norm_key(maybe_line) in {
          "generalliability", "workerscompensation", "businessownerspolicy",
          "umbrella", "liquorliability", "commercialauto", "garage liability",
          "garagekeepers", "cyberliability", "professionalliability"
        }:
          line = maybe_line
          break

      line = line or line_from_policy_prefix(policy_number)

      add_policy(
        policy_number,
        line,
        profile_data.get("carrier_name") or profile_data.get("writing_carrier"),
        profile_data.get("effective_date"),
        profile_data.get("expiration_date"),
        0,
      )

  policies = []
  for policy_number, item in sorted(policy_map.items()):
    item["claim_count"] = claim_counts.get(policy_number, 0)
    item["claims"] = claim_counts.get(policy_number, 0)
    item["total_incurred"] = claim_totals.get(policy_number, 0.0)
    policies.append(item)

  if policies:
    profile_data["policies"] = policies
    profile_data["policy_schedule"] = policies
    profile_data["policy_numbers"] = [p["policy_number"] for p in policies]
    direct_profile["policies"] = policies
    direct_profile["policy_schedule"] = policies
    direct_profile["policy_numbers"] = [p["policy_number"] for p in policies]

  print("LOSSQ_EXCEL_ONLY_POLICY_SCHEDULE_REPAIR_V1:", {
    "policies": [p.get("policy_number") for p in policies],
    "count": len(policies),
  })

  return profile_data, direct_profile



# LOSSQ_EXCEL_POLICY_SCHEDULE_BEFORE_PROFILE_SAVE_V2
def lossq_excel_policy_schedule_before_profile_save_v2(file_path, profile_data=None, parsed_claims=None):
  """
  Excel-only final policy schedule repair immediately before account profile save.

  Purpose:
  - Preserve policy schedule rows that have zero claims, such as Umbrella.
  - Does not run for PDF or CSV.
  - Does not hardcode carrier, insured, policy number, or sample file.
  """
  import re
  import datetime as _dt

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  lower_path = str(file_path or "").lower()
  if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xls")):
    return profile_data

  def clean(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def norm_policy(value):
    return clean(value).upper().replace(" ", "").strip(" :;,.|/\\")

  def is_account(value):
    text = norm_policy(value)
    return bool(re.search(r"(ACCT|ACCOUNT|CUSTOMER|CUST|CLIENT)", text))

  def is_policy(value):
    text = norm_policy(value)
    if not text or is_account(text):
      return False
    return bool(
      re.match(r"^[A-Z]{1,10}-\d{2,6}-[A-Z0-9]{2,20}$", text)
      or re.match(r"^[A-Z]{2,10}\d{4}[A-Z0-9]{2,20}$", text)
    )

  def line_from_prefix(policy_number):
    p = norm_policy(policy_number)
    prefix = p.split("-")[0] if "-" in p else ""
    mapping = {
      "GL": "General Liability",
      "CGL": "General Liability",
      "WC": "Workers Compensation",
      "BOP": "Businessowners Policy",
      "UMB": "Umbrella",
      "UM": "Umbrella",
      "EXCESS": "Umbrella",
      "LIQ": "Liquor Liability",
      "AUTO": "Commercial Auto",
      "CA": "Commercial Auto",
      "GAR": "Garage Liability",
      "DOL": "Dealers Open Lot",
      "CP": "Commercial Property",
      "PROP": "Property",
      "CY": "Cyber Liability",
      "CYBER": "Cyber Liability",
      "PL": "Professional Liability",
      "EPLI": "Employment Practices Liability",
      "DO": "Directors & Officers",
      "DNO": "Directors & Officers",
      "IM": "Inland Marine",
      "CARGO": "Motor Truck Cargo",
    }
    return mapping.get(prefix, "")

  def normalize_date(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")
    raw = clean(value)
    if not raw:
      return ""

    if re.fullmatch(r"\d{5}", raw):
      try:
        dt = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(raw))
        if 1990 <= dt.year <= 2100:
          return dt.strftime("%Y-%m-%d")
      except Exception:
        pass

    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"]:
      try:
        return _dt.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
      except Exception:
        pass

    return ""

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)
    try:
      val = float(raw or 0)
      return -val if neg else val
    except Exception:
      return 0.0

  def better_line(value):
    text = clean(value)
    if not text:
      return ""

    k = key(text)
    mapping = {
      "generalliability": "General Liability",
      "cgl": "General Liability",
      "workerscompensation": "Workers Compensation",
      "workerscomp": "Workers Compensation",
      "businessownerspolicy": "Businessowners Policy",
      "bop": "Businessowners Policy",
      "umbrella": "Umbrella",
      "umbrellaliability": "Umbrella",
      "excessliability": "Umbrella",
      "liquorliability": "Liquor Liability",
      "commercialauto": "Commercial Auto",
      "garage liability": "Garage Liability",
      "garageliability": "Garage Liability",
      "property": "Property",
      "commercialproperty": "Commercial Property",
      "cyberliability": "Cyber Liability",
    }
    return mapping.get(k, text if any(word in k for word in ["liability", "compensation", "umbrella", "policy", "property", "auto"]) else "")

  rows = []
  try:
    if lower_path.endswith(".xlsx"):
      from openpyxl import load_workbook
      wb = load_workbook(file_path, data_only=True, read_only=True)
      for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
          values = [clean(cell) for cell in row]
          if any(values):
            rows.append(values)
    elif lower_path.endswith(".xls"):
      try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
          for r in range(sheet.nrows):
            values = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(values):
              rows.append(values)
      except Exception as exc:
        print("LOSSQ_EXCEL_POLICY_SCHEDULE_BEFORE_PROFILE_SAVE_XLS_SKIPPED_V2:", str(exc)[:200])
  except Exception as exc:
    print("LOSSQ_EXCEL_POLICY_SCHEDULE_BEFORE_PROFILE_SAVE_READ_ERROR_V2:", str(exc)[:300])
    return profile_data

  policy_map = {}

  def add_policy(policy_number, line="", carrier="", effective="", expiration="", premium=0):
    policy_number = norm_policy(policy_number)
    if not is_policy(policy_number):
      return

    existing = policy_map.get(policy_number, {})
    lob = clean(line) or existing.get("line_of_business") or line_from_prefix(policy_number)
    carrier_value = clean(carrier) or existing.get("carrier") or clean(profile_data.get("carrier_name") or profile_data.get("writing_carrier"))
    eff = normalize_date(effective) or existing.get("effective_date") or normalize_date(profile_data.get("effective_date"))
    exp = normalize_date(expiration) or existing.get("expiration_date") or normalize_date(profile_data.get("expiration_date"))
    prem = money(premium) or money(existing.get("premium"))

    policy_map[policy_number] = {
      "policy_number": policy_number,
      "line_of_business": lob,
      "coverage": lob,
      "policy_type": lob,
      "carrier": carrier_value,
      "effective_date": eff,
      "expiration_date": exp,
      "premium": prem,
      "current_premium": prem,
    }

  # Start with existing profile policy rows.
  for source in [profile_data.get("policies"), profile_data.get("policy_schedule")]:
    if isinstance(source, list):
      for item in source:
        if isinstance(item, dict):
          add_policy(
            item.get("policy_number") or item.get("policy") or item.get("policy_no"),
            item.get("line_of_business") or item.get("coverage") or item.get("policy_type"),
            item.get("carrier") or item.get("carrier_name") or item.get("writing_carrier"),
            item.get("effective_date") or item.get("policy_effective_date"),
            item.get("expiration_date") or item.get("policy_expiration_date"),
            item.get("premium") or item.get("current_premium"),
          )

  # Add policies from parsed claim rows.
  claim_counts = {}
  claim_totals = {}
  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue
    policy_number = norm_policy(claim.get("policy_number") or claim.get("Policy Number"))
    if not is_policy(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
    claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + money(
      claim.get("total_incurred") or claim.get("incurred") or claim.get("total")
    )

    add_policy(
      policy_number,
      claim.get("line_of_business") or claim.get("claim_type") or claim.get("coverage"),
      claim.get("carrier_name") or claim.get("writing_carrier"),
      claim.get("effective_date"),
      claim.get("expiration_date"),
      0,
    )

  # Detect header table rows in workbook.
  aliases = {
    "policy": {"policynumber", "policyno", "policy", "policy#", "policynumber"},
    "line": {"lineofbusiness", "coverage", "policytype", "lob", "line"},
    "carrier": {"carrier", "writingcarrier", "insurancecarrier"},
    "effective": {"effectivedate", "policyeffectivedate", "effective", "effdate"},
    "expiration": {"expirationdate", "policyexpirationdate", "expiration", "expdate"},
    "premium": {"premium", "currentpremium", "annualpremium", "writtenpremium"},
  }

  def find_col(row, names):
    for i, cell in enumerate(row):
      if key(cell) in names:
        return i
    return None

  for r_idx, row in enumerate(rows):
    policy_col = find_col(row, aliases["policy"])
    if policy_col is None:
      continue

    line_col = find_col(row, aliases["line"])
    carrier_col = find_col(row, aliases["carrier"])
    eff_col = find_col(row, aliases["effective"])
    exp_col = find_col(row, aliases["expiration"])
    prem_col = find_col(row, aliases["premium"])

    for data_row in rows[r_idx + 1 : min(len(rows), r_idx + 40)]:
      if policy_col >= len(data_row):
        continue

      policy_number = norm_policy(data_row[policy_col])
      if not is_policy(policy_number):
        continue

      add_policy(
        policy_number,
        data_row[line_col] if line_col is not None and line_col < len(data_row) else "",
        data_row[carrier_col] if carrier_col is not None and carrier_col < len(data_row) else "",
        data_row[eff_col] if eff_col is not None and eff_col < len(data_row) else "",
        data_row[exp_col] if exp_col is not None and exp_col < len(data_row) else "",
        data_row[prem_col] if prem_col is not None and prem_col < len(data_row) else 0,
      )

  # Loose scan: any policy-looking cell is preserved even if no claims exist.
  for r_idx, row in enumerate(rows):
    for c_idx, cell in enumerate(row):
      policy_number = norm_policy(cell)
      if not is_policy(policy_number):
        continue

      nearby_cells = []
      for rr in rows[max(0, r_idx - 2): min(len(rows), r_idx + 3)]:
        nearby_cells.extend(rr)

      line = ""
      for value in nearby_cells:
        maybe = better_line(value)
        if maybe:
          line = maybe
          break

      add_policy(
        policy_number,
        line or line_from_prefix(policy_number),
        profile_data.get("carrier_name") or profile_data.get("writing_carrier"),
        profile_data.get("effective_date"),
        profile_data.get("expiration_date"),
        0,
      )

  policies = []
  for policy_number in sorted(policy_map.keys()):
    item = policy_map[policy_number]
    item["claim_count"] = claim_counts.get(policy_number, 0)
    item["claims"] = claim_counts.get(policy_number, 0)
    item["total_incurred"] = claim_totals.get(policy_number, 0.0)
    policies.append(item)

  if policies:
    profile_data["policies"] = policies
    profile_data["policy_schedule"] = policies
    profile_data["policy_numbers"] = [p["policy_number"] for p in policies]

  print("LOSSQ_EXCEL_POLICY_SCHEDULE_BEFORE_PROFILE_SAVE_V2:", {
    "count": len(policies),
    "policies": [p.get("policy_number") for p in policies],
  })

  return profile_data



# LOSSQ_EXCEL_ZERO_CLAIM_POLICY_REPAIR_V3
def lossq_excel_zero_claim_policy_repair_v3(file_path, profile_data=None, parsed_claims=None):
  """
  Excel-only final policy schedule repair.

  Runs only for .xlsx/.xls.
  Preserves policies from the Excel policy schedule even when they have zero claims,
  such as Umbrella / Excess Liability.
  Does not run for PDF or CSV.
  """
  import re
  import datetime as _dt

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  lower_path = str(file_path or "").lower()
  if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xls")):
    return profile_data

  def clean(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def norm_policy(value):
    return clean(value).upper().replace(" ", "").strip(" :;,.|/\\")

  def is_account_number(value):
    text = norm_policy(value)
    return bool(re.search(r"(ACCT|ACCOUNT|CUSTOMER|CUST|CLIENT)", text))

  def is_policy_number(value):
    text = norm_policy(value)
    if not text or is_account_number(text):
      return False

    return bool(
      re.match(r"^[A-Z]{1,10}-\d{2,6}-[A-Z0-9]{2,20}$", text)
      or re.match(r"^[A-Z]{2,10}\d{4}[A-Z0-9]{2,20}$", text)
    )

  def line_from_policy(policy_number):
    text = norm_policy(policy_number)
    prefix = text.split("-")[0] if "-" in text else ""

    mapping = {
      "GL": "General Liability",
      "CGL": "General Liability",
      "WC": "Workers Compensation",
      "BOP": "Businessowners Policy",
      "UMB": "Umbrella",
      "UM": "Umbrella",
      "EXCESS": "Umbrella",
      "LIQ": "Liquor Liability",
      "LIQUOR": "Liquor Liability",
      "AUTO": "Commercial Auto",
      "CA": "Commercial Auto",
      "GAR": "Garage Liability",
      "DOL": "Dealers Open Lot",
      "CP": "Commercial Property",
      "PROP": "Property",
      "CY": "Cyber Liability",
      "CYBER": "Cyber Liability",
      "PL": "Professional Liability",
      "EPLI": "Employment Practices Liability",
      "DO": "Directors & Officers",
      "DNO": "Directors & Officers",
      "IM": "Inland Marine",
      "CARGO": "Motor Truck Cargo",
    }

    return mapping.get(prefix, "")

  def normalize_line(value, policy_number=""):
    raw = clean(value)
    k = key(raw)

    mapping = {
      "generalliability": "General Liability",
      "cgl": "General Liability",
      "workerscompensation": "Workers Compensation",
      "workerscomp": "Workers Compensation",
      "wc": "Workers Compensation",
      "businessownerspolicy": "Businessowners Policy",
      "bop": "Businessowners Policy",
      "umbrella": "Umbrella",
      "umbrellaliability": "Umbrella",
      "excessliability": "Umbrella",
      "excess": "Umbrella",
      "liquorliability": "Liquor Liability",
      "liquor": "Liquor Liability",
      "commercialauto": "Commercial Auto",
      "garage liability": "Garage Liability",
      "garageliability": "Garage Liability",
      "property": "Property",
      "commercialproperty": "Commercial Property",
      "cyberliability": "Cyber Liability",
    }

    if k in mapping:
      return mapping[k]

    if raw and any(token in k for token in ["liability", "compensation", "umbrella", "policy", "property", "auto", "cyber"]):
      return raw

    return line_from_policy(policy_number)

  def normalize_date(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")

    raw = clean(value)
    if not raw:
      return ""

    if re.fullmatch(r"\d{5}", raw):
      try:
        dt = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(raw))
        if 1990 <= dt.year <= 2100:
          return dt.strftime("%Y-%m-%d")
      except Exception:
        pass

    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"]:
      try:
        return _dt.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
      except Exception:
        pass

    return ""

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0

    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]+", "", raw)

    try:
      val = float(raw or 0)
      return -val if neg else val
    except Exception:
      return 0.0

  rows = []

  try:
    if lower_path.endswith(".xlsx"):
      from openpyxl import load_workbook
      wb = load_workbook(file_path, data_only=True, read_only=True)
      for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
          values = [clean(cell) for cell in row]
          if any(values):
            rows.append(values)

    elif lower_path.endswith(".xls"):
      try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
          for r in range(sheet.nrows):
            values = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(values):
              rows.append(values)
      except Exception as exc:
        print("LOSSQ_EXCEL_ZERO_CLAIM_POLICY_REPAIR_XLS_SKIPPED_V3:", str(exc)[:200])

  except Exception as exc:
    print("LOSSQ_EXCEL_ZERO_CLAIM_POLICY_REPAIR_READ_ERROR_V3:", str(exc)[:300])
    return profile_data

  policy_map = {}

  def add_policy(policy_number, line="", carrier="", effective="", expiration="", premium=0):
    policy_number = norm_policy(policy_number)
    if not is_policy_number(policy_number):
      return

    existing = policy_map.get(policy_number, {})
    line_value = normalize_line(line, policy_number) or existing.get("line_of_business") or line_from_policy(policy_number)

    carrier_value = (
      clean(carrier)
      or existing.get("carrier")
      or clean(profile_data.get("carrier_name") or profile_data.get("writing_carrier"))
    )

    effective_value = (
      normalize_date(effective)
      or existing.get("effective_date")
      or normalize_date(profile_data.get("effective_date"))
    )

    expiration_value = (
      normalize_date(expiration)
      or existing.get("expiration_date")
      or normalize_date(profile_data.get("expiration_date"))
    )

    premium_value = money(premium) or money(existing.get("premium")) or money(existing.get("current_premium"))

    policy_map[policy_number] = {
      "policy_number": policy_number,
      "line_of_business": line_value,
      "coverage": line_value,
      "policy_type": line_value,
      "carrier": carrier_value,
      "effective_date": effective_value,
      "expiration_date": expiration_value,
      "premium": premium_value,
      "current_premium": premium_value,
    }

  # 1. Preserve any policies already on the profile.
  for source in [profile_data.get("policies"), profile_data.get("policy_schedule")]:
    if isinstance(source, list):
      for item in source:
        if not isinstance(item, dict):
          continue

        add_policy(
          item.get("policy_number") or item.get("policy") or item.get("policy_no"),
          item.get("line_of_business") or item.get("coverage") or item.get("policy_type"),
          item.get("carrier") or item.get("carrier_name") or item.get("writing_carrier"),
          item.get("effective_date") or item.get("policy_effective_date"),
          item.get("expiration_date") or item.get("policy_expiration_date"),
          item.get("premium") or item.get("current_premium"),
        )

  # 2. Add all claim policies and count them.
  claim_counts = {}
  claim_totals = {}

  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue

    policy_number = norm_policy(claim.get("policy_number") or claim.get("Policy Number"))
    if not is_policy_number(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
    claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + money(
      claim.get("total_incurred") or claim.get("incurred") or claim.get("total")
    )

    add_policy(
      policy_number,
      claim.get("line_of_business") or claim.get("claim_type") or claim.get("coverage"),
      claim.get("carrier_name") or claim.get("writing_carrier"),
      claim.get("effective_date"),
      claim.get("expiration_date"),
      0,
    )

  # 3. Read policy schedule table from Excel if present.
  aliases = {
    "policy": {"policynumber", "policyno", "policy", "policy#", "policynum"},
    "line": {"lineofbusiness", "coverage", "policytype", "lob", "line"},
    "carrier": {"carrier", "writingcarrier", "insurancecarrier"},
    "effective": {"effectivedate", "policyeffectivedate", "effective", "effdate"},
    "expiration": {"expirationdate", "policyexpirationdate", "expiration", "expdate"},
    "premium": {"premium", "currentpremium", "annualpremium", "writtenpremium"},
  }

  def find_col(row, options):
    for idx, cell in enumerate(row):
      if key(cell) in options:
        return idx
    return None

  for r_idx, row in enumerate(rows):
    policy_col = find_col(row, aliases["policy"])
    if policy_col is None:
      continue

    line_col = find_col(row, aliases["line"])
    carrier_col = find_col(row, aliases["carrier"])
    eff_col = find_col(row, aliases["effective"])
    exp_col = find_col(row, aliases["expiration"])
    prem_col = find_col(row, aliases["premium"])

    for data_row in rows[r_idx + 1 : min(len(rows), r_idx + 60)]:
      if policy_col >= len(data_row):
        continue

      policy_number = norm_policy(data_row[policy_col])
      if not is_policy_number(policy_number):
        continue

      add_policy(
        policy_number,
        data_row[line_col] if line_col is not None and line_col < len(data_row) else "",
        data_row[carrier_col] if carrier_col is not None and carrier_col < len(data_row) else "",
        data_row[eff_col] if eff_col is not None and eff_col < len(data_row) else "",
        data_row[exp_col] if exp_col is not None and exp_col < len(data_row) else "",
        data_row[prem_col] if prem_col is not None and prem_col < len(data_row) else 0,
      )

  # 4. Loose scan fallback: preserve any standalone policy-looking cell.
  for r_idx, row in enumerate(rows):
    for c_idx, cell in enumerate(row):
      policy_number = norm_policy(cell)
      if not is_policy_number(policy_number):
        continue

      nearby = []
      for nearby_row in rows[max(0, r_idx - 2): min(len(rows), r_idx + 3)]:
        nearby.extend(nearby_row)

      line_value = ""
      for nearby_cell in nearby:
        maybe_line = normalize_line(nearby_cell, policy_number)
        if maybe_line:
          line_value = maybe_line
          break

      add_policy(
        policy_number,
        line_value or line_from_policy(policy_number),
        profile_data.get("carrier_name") or profile_data.get("writing_carrier"),
        profile_data.get("effective_date"),
        profile_data.get("expiration_date"),
        0,
      )

  policies = []
  for policy_number in sorted(policy_map.keys()):
    item = policy_map[policy_number]
    item["claim_count"] = claim_counts.get(policy_number, 0)
    item["claims"] = claim_counts.get(policy_number, 0)
    item["total_incurred"] = claim_totals.get(policy_number, 0.0)
    policies.append(item)

  if policies:
    profile_data["policies"] = policies
    profile_data["policy_schedule"] = policies
    profile_data["policy_numbers"] = [item["policy_number"] for item in policies]

  print("LOSSQ_EXCEL_ZERO_CLAIM_POLICY_REPAIR_V3:", {
    "policy_count": len(policies),
    "policies": [item.get("policy_number") for item in policies],
  })

  return profile_data

# LOSSQ_EXCEL_MULTISHEET_POLICY_SCHEDULE_AUTHORITY_V1
def lossq_excel_multisheet_policy_schedule_authority_v1(file_path, profile_data=None, parsed_claims=None):
  """
  Excel-only final authority layer for multi-sheet loss runs.

  Purpose:
  - Trust workbook Account/Profile and Policy Schedule sheets over filename fallback.
  - Preserve Policy Schedule rows as the authoritative account policy list.
  - Prevent claim numbers from becoming policy type / coverage rows.
  - Prevent claim-detail rows from overwriting clean policy schedule rows.
  - Keep zero-claim policies such as Umbrella visible.
  - Does not run for PDF or CSV.
  - No customer, carrier, policy, claim, or demo-file hardcoding.
  """
  import re
  import datetime as _dt

  profile_data = profile_data if isinstance(profile_data, dict) else {}
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []

  lower_path = str(file_path or "").lower()
  if not (lower_path.endswith(".xlsx") or lower_path.endswith(".xls")):
    return profile_data

  def clean(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
      return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|")

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    raw = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
      return float(raw)
    except Exception:
      return 0.0

  def normalize_date(value):
    value = clean(value)
    if not value:
      return ""
    match = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", value)
    if match:
      month, day, year = match.groups()
      if len(year) == 2:
        year = "20" + year
      return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    match = re.search(r"\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b", value)
    if match:
      year, month, day = match.groups()
      return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return value

  def split_policy_period(value):
    value = clean(value)
    dates = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b", value)
    if len(dates) >= 2:
      return normalize_date(dates[0]), normalize_date(dates[1])
    return "", ""

  policy_prefixes = {
    "GL": "General Liability",
    "CGL": "General Liability",
    "WC": "Workers Compensation",
    "AUTO": "Commercial Auto",
    "CA": "Commercial Auto",
    "CARGO": "Motor Truck Cargo",
    "MTC": "Motor Truck Cargo",
    "BOP": "Businessowners Policy",
    "CP": "Commercial Property",
    "PROP": "Property",
    "UMB": "Umbrella",
    "UM": "Umbrella",
    "EXCESS": "Umbrella / Excess",
    "XS": "Umbrella / Excess",
    "LIQ": "Liquor Liability",
    "LQ": "Liquor Liability",
    "CY": "Cyber Liability",
    "CYBER": "Cyber Liability",
    "PL": "Professional Liability",
    "EPLI": "Employment Practices Liability",
    "DO": "Directors & Officers",
    "DNO": "Directors & Officers",
    "IM": "Inland Marine",
    "GAR": "Garage Liability",
    "DOL": "Dealers Open Lot",
  }

  def first_segment(value):
    return clean(value).upper().split("-")[0].strip()

  def line_from_policy(policy_number):
    upper = clean(policy_number).upper()
    first = first_segment(upper)
    if first in policy_prefixes:
      return policy_prefixes[first]
    for prefix, label in policy_prefixes.items():
      if upper.startswith(prefix):
        return label
    return ""

  def looks_like_claim_number(value):
    upper = clean(value).upper()
    if not upper:
      return False
    if re.search(r"-\d{5,}$", upper) and not re.search(r"-\d{4}-", upper):
      return True
    return False

  def is_policy_number(value):
    upper = clean(value).upper()
    if not upper:
      return False
    if any(token in upper for token in ["ACCT", "ACCOUNT", "CUSTOMER", "CLIENT", "CUST"]):
      return False
    if looks_like_claim_number(upper):
      return False
    first = first_segment(upper)
    if first in policy_prefixes:
      return True
    if re.search(r"-\d{4}-", upper):
      return True
    return False

  def normalize_line(value, policy_number=""):
    policy_line = line_from_policy(policy_number)
    if policy_line:
      return policy_line

    raw = clean(value)
    if not raw:
      return ""

    low = raw.lower()
    if looks_like_claim_number(raw) or is_policy_number(raw):
      return ""

    mapping = {
      "general liability": "General Liability",
      "commercial general liability": "General Liability",
      "workers compensation": "Workers Compensation",
      "workers comp": "Workers Compensation",
      "commercial auto": "Commercial Auto",
      "businessowners policy": "Businessowners Policy",
      "business owners policy": "Businessowners Policy",
      "bop": "Businessowners Policy",
      "umbrella": "Umbrella",
      "umbrella / excess": "Umbrella / Excess",
      "excess liability": "Umbrella / Excess",
      "motor truck cargo": "Motor Truck Cargo",
      "cargo": "Motor Truck Cargo",
      "liquor liability": "Liquor Liability",
      "property": "Property",
      "commercial property": "Commercial Property",
      "cyber liability": "Cyber Liability",
      "professional liability": "Professional Liability",
      "employment practices liability": "Employment Practices Liability",
      "directors and officers": "Directors & Officers",
      "directors & officers": "Directors & Officers",
      "inland marine": "Inland Marine",
      "garage liability": "Garage Liability",
      "dealers open lot": "Dealers Open Lot",
    }

    compact = re.sub(r"[^a-z0-9]+", " ", low).strip()
    if compact in mapping:
      return mapping[compact]

    if any(word in low for word in ["liability", "compensation", "umbrella", "property", "auto", "cargo", "cyber", "marine"]):
      return raw

    return ""

  def good_business(value):
    value = clean(value)
    if not value:
      return ""
    low = value.lower()
    bad_fragments = [
      "lossq",
      "loss run",
      "policy schedule",
      "claim detail",
      "claim number",
      "policy number",
      "account number",
      "business name",
      "named insured",
      "insured name",
      "carrier",
      "effective date",
      "expiration date",
      "valuation date",
      "evaluation date",
      ".xlsx",
      ".xls",
    ]
    if any(fragment in low for fragment in bad_fragments):
      return ""
    if len(value) < 3:
      return ""
    return value

  sheets = []
  try:
    if lower_path.endswith(".xlsx"):
      from openpyxl import load_workbook
      wb = load_workbook(file_path, data_only=True, read_only=True)
      for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
          values = [clean(cell) for cell in row]
          if any(values):
            rows.append(values)
        sheets.append((clean(ws.title), rows))
    else:
      try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
          rows = []
          for r in range(sheet.nrows):
            values = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(values):
              rows.append(values)
          sheets.append((clean(sheet.name), rows))
      except Exception:
        return profile_data
  except Exception as workbook_exc:
    print("LOSSQ_EXCEL_MULTISHEET_POLICY_SCHEDULE_AUTHORITY_V1_ERROR:", str(workbook_exc)[:500])
    return profile_data

  label_values = {}

  label_aliases = {
    "business_name": {"namedinsured", "insured", "insuredname", "accountname", "businessname", "companyname", "applicant"},
    "account_number": {"accountnumber", "accountno", "account", "customernumber", "clientnumber"},
    "carrier_name": {"writingcarrier", "carrier", "carriername", "insurancecarrier"},
    "policy_period": {"policyperiod", "policyterm", "coverageperiod"},
    "effective_date": {"effectivedate", "policyeffectivedate", "policyeffective", "effdate"},
    "expiration_date": {"expirationdate", "policyexpirationdate", "policyexpiration", "expdate"},
    "evaluation_date": {"evaluationdate", "valuationdate", "asofdate", "lossrundate"},
  }

  reverse_labels = {}
  for field, aliases in label_aliases.items():
    for alias in aliases:
      reverse_labels[alias] = field

  for sheet_name, rows in sheets:
    for row in rows:
      for idx, cell in enumerate(row):
        cell_clean = clean(cell)
        cell_key = key(cell_clean)

        if cell_key in reverse_labels and idx + 1 < len(row):
          value = clean(row[idx + 1])
          if value:
            label_values.setdefault(reverse_labels[cell_key], value)

        if ":" in cell_clean:
          left, right = cell_clean.split(":", 1)
          left_key = key(left)
          if left_key in reverse_labels and clean(right):
            label_values.setdefault(reverse_labels[left_key], clean(right))


  # LOSSQ_EXCEL_PROFILE_HEADER_ROW_VALUES_V1
  # Some Excel account profile sheets place labels across one row and values
  # directly underneath. Example:
  # Named Insured | Writing Carrier | Carrier | Account Number
  # Metro Courier | Keystone ...   | Keystone | MCI-...
  # The older same-row next-cell scan can accidentally read adjacent labels
  # such as Expiration Date or Current Premium as values. This block prefers
  # below-row values when a row contains multiple profile labels.
  def lossq_excel_label_value_usable_v1(field, value):
    raw = clean(value)
    if not raw:
      return False

    raw_key = key(raw)
    if raw_key in reverse_labels:
      return False

    if field == "business_name":
      return bool(good_business(raw))

    if field == "account_number":
      return not is_policy_number(raw) and not looks_like_claim_number(raw) and "account" not in raw_key

    if field == "carrier_name":
      low = raw.lower()
      return not any(fragment in low for fragment in ["policy schedule", "claim detail", "claim number", "policy number"])

    if field in {"effective_date", "expiration_date", "evaluation_date"}:
      normalized = normalize_date(raw)
      if not normalized or raw_key in {"effectivedate", "expirationdate", "currentpremium", "policynumber"}:
        return False
      return any(char.isdigit() for char in normalized)

    if field == "policy_period":
      start, end = split_policy_period(raw)
      return bool(start and end)

    return True

  for sheet_name, rows in sheets:
    for row_index, row in enumerate(rows):
      label_positions = []

      for idx, cell in enumerate(row):
        field = reverse_labels.get(key(cell))
        if field:
          label_positions.append((field, idx))

      if not label_positions:
        continue

      row_has_multiple_labels = len(label_positions) >= 2
      next_rows = rows[row_index + 1: row_index + 4]

      for field, idx in label_positions:
        candidates = []

        # Only trust same-row next-cell when the row is not a label header row.
        if not row_has_multiple_labels and idx + 1 < len(row):
          candidates.append(row[idx + 1])

        # Prefer values below the label in the same column.
        for next_row in next_rows:
          if idx < len(next_row):
            candidates.append(next_row[idx])
          if idx + 1 < len(next_row):
            candidates.append(next_row[idx + 1])

        for candidate in candidates:
          candidate = clean(candidate)
          if lossq_excel_label_value_usable_v1(field, candidate):
            existing = clean(label_values.get(field))
            if not lossq_excel_label_value_usable_v1(field, existing):
              label_values[field] = candidate
            else:
              # For multi-label header rows, below-row values are more authoritative.
              label_values[field] = candidate
            break

  business_name = good_business(label_values.get("business_name"))
  if business_name:
    for field in ["business_name", "insured_name", "named_insured", "account_name", "insured"]:
      profile_data[field] = business_name

  account_number = clean(label_values.get("account_number"))
  if account_number and not is_policy_number(account_number):
    profile_data["account_number"] = account_number
    profile_data["customer_number"] = account_number

  carrier_name = clean(label_values.get("carrier_name"))
  if carrier_name:
    profile_data["carrier_name"] = carrier_name
    profile_data["writing_carrier"] = carrier_name
    profile_data["carrier"] = carrier_name

  effective_date = normalize_date(label_values.get("effective_date"))
  expiration_date = normalize_date(label_values.get("expiration_date"))

  period_start, period_end = split_policy_period(label_values.get("policy_period"))
  if not effective_date and period_start:
    effective_date = period_start
  if not expiration_date and period_end:
    expiration_date = period_end

  if effective_date:
    profile_data["effective_date"] = effective_date
  if expiration_date:
    profile_data["expiration_date"] = expiration_date

  evaluation_date = normalize_date(label_values.get("evaluation_date"))
  if evaluation_date:
    profile_data["evaluation_date"] = evaluation_date
    profile_data["valuation_date"] = evaluation_date

  claim_counts = {}
  claim_totals = {}
  claim_carriers = {}

  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue
    policy_number = clean(claim.get("policy_number") or claim.get("Policy Number")).upper()
    if not is_policy_number(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
    claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + money(
      claim.get("total_incurred") or claim.get("incurred") or claim.get("total")
    )

    claim_carrier = clean(claim.get("carrier_name") or claim.get("writing_carrier"))
    if claim_carrier:
      claim_carriers[policy_number] = claim_carrier

  header_aliases = {
    "policy_number": {"policynumber", "policyno", "policy", "policy#", "policynum"},
    "line": {"policytypecoverage", "policytype", "linecoverage", "lineofbusiness", "coverage", "lob", "line"},
    "carrier": {"carrier", "writingcarrier", "carriername", "insurancecarrier"},
    "effective_date": {"effective", "effectivedate", "policyeffective", "policyeffectivedate", "effdate"},
    "expiration_date": {"expiration", "expirationdate", "policyexpiration", "policyexpirationdate", "expdate"},
    "premium": {"premium", "currentpremium", "annualpremium", "writtenpremium", "policypremium"},
  }

  def header_field_for(cell):
    cell_key = key(cell)
    for field, aliases in header_aliases.items():
      if cell_key in aliases:
        return field
    return ""

  policy_map = {}

  def add_policy(policy_number, line_value="", carrier_value="", eff_value="", exp_value="", premium_value=""):
    policy_number = clean(policy_number).upper()
    if not is_policy_number(policy_number):
      return

    final_line = normalize_line(line_value, policy_number) or line_from_policy(policy_number) or "Other"
    final_carrier = clean(carrier_value) or claim_carriers.get(policy_number) or profile_data.get("carrier_name") or profile_data.get("writing_carrier") or ""
    final_eff = normalize_date(eff_value) or profile_data.get("effective_date") or effective_date or ""
    final_exp = normalize_date(exp_value) or profile_data.get("expiration_date") or expiration_date or ""
    final_premium = clean(premium_value)

    existing = policy_map.get(policy_number, {})
    policy_map[policy_number] = {
      "policy_number": policy_number,
      "policy_type": final_line or existing.get("policy_type") or "",
      "line_of_business": final_line or existing.get("line_of_business") or "",
      "coverage": final_line or existing.get("coverage") or "",
      "line": final_line or existing.get("line") or "",
      "carrier": final_carrier or existing.get("carrier") or "",
      "carrier_name": final_carrier or existing.get("carrier_name") or "",
      "writing_carrier": final_carrier or existing.get("writing_carrier") or "",
      "effective_date": final_eff or existing.get("effective_date") or "",
      "policy_effective_date": final_eff or existing.get("policy_effective_date") or "",
      "expiration_date": final_exp or existing.get("expiration_date") or "",
      "policy_expiration_date": final_exp or existing.get("policy_expiration_date") or "",
      "premium": final_premium or existing.get("premium") or "",
      "current_premium": final_premium.replace("$", "").replace(",", "") if final_premium else existing.get("current_premium", ""),
      "claim_count": claim_counts.get(policy_number, 0),
      "claims": claim_counts.get(policy_number, 0),
      "total_incurred": claim_totals.get(policy_number, 0.0),
    }

  for sheet_name, rows in sheets:
    if "claim" in sheet_name.lower():
      continue

    active_header = None

    for row in rows:
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))
      row_keys = {key(cell) for cell in row}

      if "claimnumber" in row_keys or "claim detail" in row_text or "claims detail" in row_text:
        active_header = None
        continue

      maybe_header = {}
      for idx, cell in enumerate(row):
        field = header_field_for(cell)
        if field:
          maybe_header[field] = idx

      if "policy_number" in maybe_header and ("line" in maybe_header or "effective_date" in maybe_header or "expiration_date" in maybe_header):
        active_header = maybe_header
        continue

      if not active_header:
        continue

      policy_i = active_header.get("policy_number")
      if policy_i is None or policy_i >= len(row):
        continue

      policy_number = clean(row[policy_i]).upper()
      if not is_policy_number(policy_number):
        continue

      def row_value(field):
        idx = active_header.get(field)
        if idx is None or idx >= len(row):
          return ""
        return clean(row[idx])

      add_policy(
        policy_number=policy_number,
        line_value=row_value("line"),
        carrier_value=row_value("carrier"),
        eff_value=row_value("effective_date"),
        exp_value=row_value("expiration_date"),
        premium_value=row_value("premium"),
      )

  if not policy_map:
    return profile_data

  policies = list(policy_map.values())
  policy_numbers = [item["policy_number"] for item in policies if item.get("policy_number")]

  profile_data["policies"] = policies
  profile_data["policy_schedule"] = policies
  profile_data["policy_numbers"] = policy_numbers

  if policy_numbers and not is_policy_number(profile_data.get("policy_number")):
    profile_data["policy_number"] = policy_numbers[0]
  if policy_numbers and not is_policy_number(profile_data.get("main_policy")):
    profile_data["main_policy"] = policy_numbers[0]


  # LOSSQ_EXCEL_TOP_LEVEL_PROFILE_FINALIZE_V1
  # Final top-level cleanup after policy schedule authority. The policy table can
  # be correct while the account snapshot still holds filename/header fallback values.
  def lossq_excel_invalid_account_date_v1(value):
    raw = clean(value)
    if not raw:
      return True

    raw_key = key(raw)
    if raw_key in reverse_labels:
      return True

    if raw_key in {"currentpremium", "expiringpremium", "targetrenewalpremium", "policynumber", "claimnumber"}:
      return True

    normalized = normalize_date(raw)
    if not normalized:
      return True

    return not any(char.isdigit() for char in normalized)

  corrected_business_name = good_business(label_values.get("business_name"))
  current_business_name = clean(profile_data.get("business_name"))
  current_business_key = current_business_name.lower()

  if corrected_business_name and (
    not good_business(current_business_name)
    or "lossq" in current_business_key
    or ".xlsx" in current_business_key
    or ".xls" in current_business_key
    or current_business_key.startswith("20")
  ):
    for business_field in ["business_name", "insured_name", "named_insured", "account_name", "insured"]:
      profile_data[business_field] = corrected_business_name

  first_policy_effective = ""
  first_policy_expiration = ""

  for policy in policies:
    if not isinstance(policy, dict):
      continue

    if not first_policy_effective:
      first_policy_effective = normalize_date(
        policy.get("effective_date")
        or policy.get("policy_effective_date")
        or policy.get("effective")
      )

    if not first_policy_expiration:
      first_policy_expiration = normalize_date(
        policy.get("expiration_date")
        or policy.get("policy_expiration_date")
        or policy.get("expiration")
      )

  if first_policy_effective and lossq_excel_invalid_account_date_v1(profile_data.get("effective_date")):
    profile_data["effective_date"] = first_policy_effective

  if first_policy_expiration and lossq_excel_invalid_account_date_v1(profile_data.get("expiration_date")):
    profile_data["expiration_date"] = first_policy_expiration


  # LOSSQ_EXCEL_ACCOUNT_PROFILE_TWO_COLUMN_FINAL_V1
  # Final Excel account snapshot repair for simple Account Profile sheets:
  # Field | Value
  # Named Insured | ...
  # Account Number | ...
  # Writing Carrier | ...
  # Policy Period | ...
  # This prevents filename fallback and adjacent labels/carriers from being saved
  # as the account snapshot values.
  excel_profile_fields = {}

  excel_two_column_aliases = {
    "namedinsured": "business_name",
    "insured": "business_name",
    "insuredname": "business_name",
    "accountname": "business_name",
    "businessname": "business_name",
    "companyname": "business_name",
    "accountnumber": "account_number",
    "accountno": "account_number",
    "account": "account_number",
    "customernumber": "account_number",
    "clientnumber": "account_number",
    "writingcarrier": "carrier_name",
    "carrier": "carrier_name",
    "carriername": "carrier_name",
    "insurancecarrier": "carrier_name",
    "policyperiod": "policy_period",
    "policyterm": "policy_period",
    "coverageperiod": "policy_period",
    "effectivedate": "effective_date",
    "policyeffectivedate": "effective_date",
    "expirationdate": "expiration_date",
    "policyexpirationdate": "expiration_date",
    "evaluationdate": "evaluation_date",
    "valuationdate": "evaluation_date",
    "asofdate": "evaluation_date",
    "lossrundate": "evaluation_date",
  }

  def lossq_excel_valid_account_number_v1(value):
    raw = clean(value)
    if not raw:
      return False

    raw_key = key(raw)
    low = raw.lower()

    if any(fragment in low for fragment in ["insurance", "carrier", "company", "co.", "mutual", "policy schedule", "claim detail"]):
      return False

    if raw_key in reverse_labels:
      return False

    if is_policy_number(raw) or looks_like_claim_number(raw):
      return False

    if not any(char.isdigit() for char in raw):
      return False

    return True

  for sheet_name, rows in sheets:
    sheet_key = key(sheet_name)

    # Prefer account/profile-style sheets, but still allow universal Field/Value rows.
    for row in rows:
      if len(row) < 2:
        continue

      left = clean(row[0])
      right = clean(row[1])
      field = excel_two_column_aliases.get(key(left))

      if not field or not right:
        continue

      if field == "business_name":
        if good_business(right):
          excel_profile_fields[field] = right

      elif field == "account_number":
        if lossq_excel_valid_account_number_v1(right):
          excel_profile_fields[field] = right

      elif field == "carrier_name":
        if right and not is_policy_number(right) and not looks_like_claim_number(right):
          excel_profile_fields[field] = right

      elif field in {"effective_date", "expiration_date", "evaluation_date"}:
        normalized_value = normalize_date(right)
        if normalized_value and any(char.isdigit() for char in normalized_value):
          excel_profile_fields[field] = normalized_value

      elif field == "policy_period":
        start_date, end_date = split_policy_period(right)
        if start_date and end_date:
          excel_profile_fields["effective_date"] = start_date
          excel_profile_fields["expiration_date"] = end_date
          excel_profile_fields["policy_period"] = right

  final_business = good_business(excel_profile_fields.get("business_name"))
  if final_business:
    for business_field in ["business_name", "insured_name", "named_insured", "account_name", "insured"]:
      profile_data[business_field] = final_business

  final_account_number = excel_profile_fields.get("account_number")
  if lossq_excel_valid_account_number_v1(final_account_number):
    profile_data["account_number"] = final_account_number
    profile_data["customer_number"] = final_account_number

  final_carrier = clean(excel_profile_fields.get("carrier_name"))
  if final_carrier:
    profile_data["carrier_name"] = final_carrier
    profile_data["writing_carrier"] = final_carrier
    profile_data["carrier"] = final_carrier

  if excel_profile_fields.get("effective_date"):
    profile_data["effective_date"] = excel_profile_fields["effective_date"]

  if excel_profile_fields.get("expiration_date"):
    profile_data["expiration_date"] = excel_profile_fields["expiration_date"]

  if excel_profile_fields.get("evaluation_date"):
    profile_data["evaluation_date"] = excel_profile_fields["evaluation_date"]
    profile_data["valuation_date"] = excel_profile_fields["evaluation_date"]

  print("LOSSQ_EXCEL_ACCOUNT_PROFILE_TWO_COLUMN_FINAL_V1:", {
    "business_name": profile_data.get("business_name"),
    "account_number": profile_data.get("account_number"),
    "carrier_name": profile_data.get("carrier_name"),
    "effective_date": profile_data.get("effective_date"),
    "expiration_date": profile_data.get("expiration_date"),
    "evaluation_date": profile_data.get("evaluation_date"),
  })

  print("LOSSQ_EXCEL_MULTISHEET_POLICY_SCHEDULE_AUTHORITY_V1:", {
    "business_name": profile_data.get("business_name"),
    "effective_date": profile_data.get("effective_date"),
    "expiration_date": profile_data.get("expiration_date"),
    "policy_count": len(policies),
    "policy_numbers": policy_numbers,
    "lines": [p.get("line_of_business") for p in policies],
  })

  return profile_data



def extract_profile_data(
  parsed_claims: list[dict],
  fallback_policy_number: str,
  direct_profile: dict | None = None,
):
  direct_profile = direct_profile or {}

  profile = {
    "business_name": clean_profile_value(direct_profile.get("business_name")),
    "carrier_name": clean_profile_value(direct_profile.get("carrier_name")),
    "writing_carrier": clean_profile_value(
      direct_profile.get("writing_carrier") or direct_profile.get("carrier_name")
    ),
    "agency_name": clean_profile_value(direct_profile.get("agency_name")),
    "account_number": clean_profile_value(
      direct_profile.get("account_number") or direct_profile.get("customer_number")
    ),
    "customer_number": clean_profile_value(
      direct_profile.get("customer_number") or direct_profile.get("account_number")
    ),
    "producer_number": clean_profile_value(direct_profile.get("producer_number")),
    "policy_number": clean_profile_value(
      direct_profile.get("policy_number") or direct_profile.get("account_number")
    ),
    "effective_date": parse_date(direct_profile.get("effective_date")) or "",
    "expiration_date": parse_date(direct_profile.get("expiration_date")) or "",
    "evaluation_date": parse_date(direct_profile.get("evaluation_date")) or "",
    "policies": direct_profile.get("policies") or [],
    "validation": direct_profile.get("validation") or {},
    "raw_text_preview": direct_profile.get("raw_text_preview") or "",
  }


  exposure_input_fields = [
    "current_premium",
    "expiring_premium",
    "target_renewal_premium",
    "line_of_business",
    "state",
    "class_code",
    "class_codes",
    "limits",
    "coverage_limit",
    "deductible",
    "retention",
    "payroll",
    "revenue",
    "sales",
    "receipts",
    "employee_count",
    "vehicle_count",
    "driver_count",
    "property_tiv",
    "tiv",
    "building_value",
    "contents_value",
    "square_footage",
    "location_count",
    "unit_count",
    "cargo_limit",
    "umbrella_limit",
    "experience_mod",
    "mod",
    "exposure_change_percent",
    "cyber_revenue",
    "professional_revenue",
    "exposure_basis",
    "underwriter_notes",
  ]

  for field in exposure_input_fields:
    value = direct_profile.get(field)
    if value not in ("", None, [], {}):
      profile[field] = value


  for item in parsed_claims:
    if not profile["business_name"]:
      profile["business_name"] = clean_profile_value(
        pick(item, ["business_name", "insured_name", "named_insured", "account_name"], "")
      )

    if not profile["carrier_name"]:
      profile["carrier_name"] = clean_profile_value(
        pick(item, ["carrier_name", "insurance_carrier", "carrier"], "")
      )

    if not profile["writing_carrier"]:
      profile["writing_carrier"] = clean_profile_value(
        pick(item, ["writing_carrier", "carrier_name", "insurance_carrier", "carrier"], "")
      )

    if not profile["agency_name"]:
      profile["agency_name"] = clean_profile_value(
        pick(item, ["agency_name", "broker_name", "agency", "producer_name"], "")
      )

    if not profile["account_number"]:
      profile["account_number"] = clean_profile_value(
        pick(item, ["account_number", "customer_number", "account_no", "customer_no"], "")
      )

    if not profile["customer_number"]:
      profile["customer_number"] = profile["account_number"]

    if not profile["policy_number"]:
      profile["policy_number"] = clean_profile_value(
        pick(item, ["policy_number", "policy_no", "policy"], "")
      )

    if not profile["effective_date"]:
      profile["effective_date"] = parse_date(
        pick(item, ["effective_date", "policy_effective_date"])
      ) or ""

    if not profile["expiration_date"]:
      profile["expiration_date"] = parse_date(
        pick(item, ["expiration_date", "policy_expiration_date", "expiry_date"])
      ) or ""

  if not profile["policy_number"]:
    profile["policy_number"] = clean_profile_value(
      profile.get("account_number") or fallback_policy_number
    )

  if not profile["writing_carrier"]:
    profile["writing_carrier"] = profile["carrier_name"]

  return profile


def serialize_json(value, fallback):
  try:
    if value is None:
      return json.dumps(fallback)
    if isinstance(value, str):
      return value
    return json.dumps(value)
  except Exception:
    return json.dumps(fallback)


def serialize_json(value, fallback):
  try:
    if value is None:
      return json.dumps(fallback)

    if isinstance(value, str):
      return value

    return json.dumps(value)
  except Exception:
    return json.dumps(fallback)


def strict_money_value_for_exposure(value):
  # LOSSQ_STRICT_EXPOSURE_MONEY_VALUES_V1
  # Exposure money fields must look like actual dollars.
  # This prevents policy/account numbers like PV-ACCT-572914 from becoming Property TIV.
  text_value = str(value or "")

  money_match = re.search(r"\$\s*[0-9][0-9,]*(?:\.\d{2})?", text_value)
  if money_match:
    return money_match.group(0).replace(" ", "")

  return ""


def derive_exposure_inputs_from_policy_schedule(profile_data: dict):
  # LOSSQ_POLICY_SCHEDULE_TO_EXPOSURE_INPUTS_V1
  # Copies exposure/premium values that were detected inside policy schedule rows
  # into top-level Exposure Inputs fields.

  profile_data = profile_data or {}
  policies = profile_data.get("policies") or []

  if not isinstance(policies, list):
    return profile_data

  money_values_for_premium = []

  def first_money(value):
    return strict_money_value_for_exposure(value)

  def first_number(value):
    match = re.search(r"\b[0-9][0-9,]*\b", str(value or ""))
    return match.group(0) if match else ""

  def set_if_blank(field, value):
    value = str(value or "").strip()
    if value and not profile_data.get(field):
      profile_data[field] = value

  def scan_text(value):
    text_value = str(value or "")
    lower = text_value.lower()

    if "payroll" in lower:
      set_if_blank("payroll", first_money(text_value))

    if "revenue" in lower or "sales" in lower:
      money = first_money(text_value)
      set_if_blank("revenue", money)
      set_if_blank("sales", money)

    if "receipt" in lower:
      set_if_blank("receipts", first_money(text_value))

    if "vehicle" in lower:
      vehicle_match = re.search(r"vehicles?\s*[:\-]?\s*([0-9,]+)", text_value, re.I)
      set_if_blank("vehicle_count", vehicle_match.group(1) if vehicle_match else first_number(text_value))

    if "driver" in lower:
      driver_match = re.search(r"drivers?\s*[:\-]?\s*([0-9,]+)", text_value, re.I)
      set_if_blank("driver_count", driver_match.group(1) if driver_match else first_number(text_value))

    if "employee" in lower:
      employee_match = re.search(r"employees?\s*[:\-]?\s*([0-9,]+)", text_value, re.I)
      set_if_blank("employee_count", employee_match.group(1) if employee_match else first_number(text_value))

    if "tiv" in lower or "total insured value" in lower:
      money = first_money(text_value)
      set_if_blank("property_tiv", money)
      set_if_blank("tiv", money)

    if "limit" in lower:
      money = first_money(text_value)
      set_if_blank("coverage_limit", money)
      set_if_blank("limits", text_value.strip())

    if "deductible" in lower:
      set_if_blank("deductible", first_money(text_value))

    if "retention" in lower or "sir" in lower:
      set_if_blank("retention", first_money(text_value))

    if "class" in lower and "code" in lower:
      class_match = re.search(r"class(?:ification)?\s*codes?\s*[:\-]?\s*([A-Za-z0-9,\- ]+)", text_value, re.I)
      if class_match:
        set_if_blank("class_codes", class_match.group(1).strip())
        set_if_blank("class_code", class_match.group(1).strip())

  for policy in policies:
    if not isinstance(policy, dict):
      continue

    line = (
      policy.get("line_of_business")
      or policy.get("policy_type")
      or policy.get("coverage")
      or policy.get("line")
      or ""
    )

    if line and not profile_data.get("line_of_business"):
      profile_data["line_of_business"] = str(line).strip()

    for field_name, value in policy.items():
      value_text = str(value or "").strip()
      if not value_text:
        continue

      scan_text(value_text)

      field_lower = str(field_name or "").lower()

      if "premium" in field_lower:
        money = first_money(value_text)
        if money:
          money_values_for_premium.append(money)

      # Some clean policy tables put exposure basis in one column and premium in the next.
      # If a row has exposure text and another field is only a money value, treat that money as premium.
      if re.fullmatch(r"\$?\s*[0-9][0-9,]*(?:\.\d{2})?", value_text):
        row_text = " ".join(str(v or "") for v in policy.values()).lower()
        if any(word in row_text for word in ["payroll", "vehicles", "drivers", "revenue", "limit", "tiv"]):
          money_values_for_premium.append(value_text.replace(" ", ""))

  if money_values_for_premium and not profile_data.get("current_premium"):
    total = 0.0
    for item in money_values_for_premium:
      try:
        total += float(str(item).replace("$", "").replace(",", "").strip())
      except Exception:
        pass

    if total > 0:
      profile_data["current_premium"] = f"${total:,.0f}"

  if not profile_data.get("exposure_basis"):
    basis_parts = []
    for field in ["payroll", "revenue", "vehicle_count", "driver_count", "property_tiv", "coverage_limit"]:
      if profile_data.get(field):
        basis_parts.append(f"{field.replace('_', ' ').title()}: {profile_data.get(field)}")
    if basis_parts:
      profile_data["exposure_basis"] = "; ".join(basis_parts)

  return profile_data




# LOSSQ_UPSERT_POLICY_SCHEDULE_SANITIZE_V6
def lossq_sanitize_profile_policies_before_upsert_v6(profile_data: dict):
  """
  Final account-profile save cleaner.
  Removes account/customer identifiers from policy schedules and restores missing
  policy line names from policy prefixes before AccountProfile.policies is saved.
  """
  import json
  import re

  if not isinstance(profile_data, dict):
    return profile_data

  profile_data = dict(profile_data)

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").strip())

  def compact(value):
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())

  def is_account_identifier(value):
    raw = clean(value).upper()
    key = compact(value)
    if not key:
      return False
    if key.startswith(("ACCT", "ACCOUNT", "CUST", "CUSTOMER", "CLIENT")):
      return True
    if "ACCT" in key or "ACCOUNT" in key or "CUSTOMER" in key or "CLIENT" in key:
      return True
    if re.search(r"\b(ACCT|ACCOUNT|CUSTOMER|CLIENT)\b", raw):
      return True
    return False

  def is_generic_line(value):
    value = clean(value).upper()
    return value in {"", "UNKNOWN", "N/A", "NA", "NONE", "NULL", "NOT SET", "POLICY", "COVERAGE", "-"}

  def infer_lob(policy_number, current_lob=""):
    current_lob = clean(current_lob)
    if current_lob and not is_generic_line(current_lob):
      return current_lob

    key = compact(policy_number)
    prefix_map = [
      ("BOP", "Businessowners Policy"),
      ("GL", "General Liability"),
      ("WC", "Workers Compensation"),
      ("PROP", "Property"),
      ("CP", "Commercial Property"),
      ("UMB", "Umbrella"),
      ("UM", "Umbrella"),
      ("GAR", "Garage Liability"),
      ("DOL", "Dealers Open Lot"),
      ("AUTO", "Commercial Auto"),
      ("CA", "Commercial Auto"),
      ("PL", "Professional Liability"),
      ("EPLI", "Employment Practices Liability"),
      ("CY", "Cyber Liability"),
      ("CARGO", "Motor Truck Cargo"),
      ("IM", "Inland Marine"),
      ("DO", "Directors & Officers"),
      ("DNO", "Directors & Officers"),
      ("LIQ", "Liquor Liability"),
    ]

    for prefix, label in prefix_map:
      if key.startswith(prefix):
        return label

    return ""

  def rows_from(value):
    if isinstance(value, list):
      return value

    if isinstance(value, str):
      try:
        loaded = json.loads(value)
        return loaded if isinstance(loaded, list) else []
      except Exception:
        return []

    return []

  raw_rows = []

  for source_key in ["policies", "policy_schedule"]:
    raw_rows.extend(rows_from(profile_data.get(source_key)))

  policy_numbers = profile_data.get("policy_numbers")
  if isinstance(policy_numbers, str):
    policy_numbers = [item.strip() for item in policy_numbers.split(",") if item.strip()]

  if isinstance(policy_numbers, list):
    for policy_number in policy_numbers:
      raw_rows.append({"policy_number": policy_number})

  by_key = {}

  for row in raw_rows:
    if isinstance(row, str):
      row = {"policy_number": row}

    if not isinstance(row, dict):
      continue

    policy_number = clean(
      row.get("policy_number")
      or row.get("policyNumber")
      or row.get("policy_no")
      or row.get("policy")
      or row.get("number")
    )

    if not policy_number or is_account_identifier(policy_number):
      continue

    key = compact(policy_number)
    if not key:
      continue

    existing = by_key.get(key, {})
    merged = dict(existing)

    for k, v in row.items():
      if v not in ("", None, [], {}):
        merged[k] = v

    current_line = (
      merged.get("line_of_business")
      or merged.get("policy_type")
      or merged.get("policyType")
      or merged.get("coverage")
      or merged.get("coverage_line")
      or merged.get("coverageType")
      or merged.get("policy_line")
      or merged.get("lob")
      or merged.get("line")
    )

    lob = infer_lob(policy_number, current_line)
    if not lob:
      lob = "Unknown"

    carrier = clean(
      merged.get("carrier")
      or merged.get("carrier_name")
      or merged.get("writing_carrier")
      or profile_data.get("carrier_name")
      or profile_data.get("writing_carrier")
    )

    effective = clean(
      merged.get("effective_date")
      or merged.get("effective")
      or profile_data.get("effective_date")
      or profile_data.get("effective")
    )

    expiration = clean(
      merged.get("expiration_date")
      or merged.get("expiration")
      or profile_data.get("expiration_date")
      or profile_data.get("expiration")
    )

    merged["policy_number"] = policy_number
    merged["policyNumber"] = policy_number
    merged["line_of_business"] = lob
    merged["policy_type"] = lob
    merged["policyType"] = lob
    merged["coverage"] = lob
    merged["coverage_line"] = lob
    merged["coverageType"] = lob
    merged["policy_line"] = lob
    merged["line"] = lob
    merged["lob"] = lob

    if carrier:
      merged["carrier"] = carrier
      merged["carrier_name"] = carrier
      merged["writing_carrier"] = carrier

    if effective:
      merged["effective_date"] = effective
      merged["effective"] = effective

    if expiration:
      merged["expiration_date"] = expiration
      merged["expiration"] = expiration

    by_key[key] = merged

  cleaned_policies = list(by_key.values())

  if cleaned_policies:
    profile_data["policies"] = cleaned_policies
    profile_data["policy_schedule"] = cleaned_policies
    profile_data["policy_numbers"] = [
      item.get("policy_number")
      for item in cleaned_policies
      if item.get("policy_number")
    ]

    current_policy = clean(
      profile_data.get("policy_number")
      or profile_data.get("main_policy")
      or profile_data.get("main_policy_number")
    )

    if not current_policy or is_account_identifier(current_policy):
      first_policy = cleaned_policies[0].get("policy_number", "")
      profile_data["policy_number"] = first_policy
      profile_data["main_policy"] = first_policy
      profile_data["main_policy_number"] = first_policy

  print("LOSSQ_UPSERT_POLICY_SCHEDULE_SANITIZE_V6:", {
    "policy_numbers": profile_data.get("policy_numbers"),
    "policies": [
      {
        "policy_number": item.get("policy_number"),
        "line_of_business": item.get("line_of_business"),
      }
      for item in cleaned_policies
    ],
  })

  return profile_data



# LOSSQ_EXPOSURE_INPUT_UNIVERSAL_TERMINOLOGY_V1
def lossq_exposure_input_universal_terminology_v1(profile_data):
  if not isinstance(profile_data, dict):
    return profile_data

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm_key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def set_if_blank(field, value):
    cleaned = clean(value)
    if cleaned and not clean(profile_data.get(field)):
      profile_data[field] = cleaned

  def money_clean(value):
    text = clean(value)
    if not text:
      return ""
    text = re.sub(r"(?i)\b(?:cad|cdn|cnd|usd)\b", "", text)
    text = text.replace("CA$", "").replace("C$", "").replace("US$", "")
    text = text.replace("$", "").replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else ""

  def count_clean(value):
    text = clean(value).replace(",", "")
    match = re.search(r"\d+", text)
    return match.group(0) if match else ""

  aliases = {
    # Premium
    "currentpremium": "current_premium",
    "annualpremium": "current_premium",
    "writtenpremium": "current_premium",
    "policypremium": "current_premium",
    "termpremium": "current_premium",
    "inforcepremium": "current_premium",
    "expiringpremium": "expiring_premium",
    "priorpremium": "expiring_premium",
    "previoustermpremium": "expiring_premium",
    "targetrenewalpremium": "target_renewal_premium",
    "renewaltargetpremium": "target_renewal_premium",

    # Line / classification
    "primarylineofbusiness": "line_of_business",
    "lineofbusiness": "line_of_business",
    "lob": "line_of_business",
    "coverage": "line_of_business",
    "policytype": "line_of_business",
    "classcode": "class_code",
    "classcodes": "class_codes",
    "naics": "naics",
    "sic": "sic",
    "industrycode": "industry_code",

    # Geography: U.S. and Canada
    "state": "state",
    "primarystate": "state",
    "jurisdiction": "state",
    "province": "state",
    "provincecode": "state",
    "territory": "state",
    "riskprovince": "state",
    "postalcode": "postal_code",
    "postcode": "postal_code",
    "zip": "zip_code",
    "zipcode": "zip_code",

    # Payroll / revenue
    "payroll": "payroll",
    "annualpayroll": "payroll",
    "grosspayroll": "payroll",
    "estimatedpayroll": "payroll",
    "insurableearnings": "payroll",
    "remuneration": "payroll",
    "wages": "payroll",
    "annualrevenue": "revenue",
    "revenue": "revenue",
    "grosssales": "revenue",
    "sales": "sales",
    "turnover": "revenue",
    "receipts": "receipts",
    "grossreceipts": "receipts",
    "professionalrevenue": "professional_revenue",
    "cyberrevenue": "cyber_revenue",

    # Counts
    "employeecount": "employee_count",
    "employees": "employee_count",
    "numberofemployees": "employee_count",
    "fte": "employee_count",
    "fulltimeequivalents": "employee_count",
    "vehiclecount": "vehicle_count",
    "vehicles": "vehicle_count",
    "scheduledautos": "vehicle_count",
    "powerunits": "vehicle_count",
    "fleetunits": "vehicle_count",
    "ownedautos": "vehicle_count",
    "drivercount": "driver_count",
    "drivers": "driver_count",
    "listeddrivers": "driver_count",
    "operators": "driver_count",

    # Property / limits
    "propertytiv": "property_tiv",
    "tiv": "property_tiv",
    "totalinsuredvalue": "property_tiv",
    "statementofvalues": "property_tiv",
    "sov": "property_tiv",
    "locationtiv": "property_tiv",
    "buildingandcontents": "property_tiv",
    "buildingvalue": "building_value",
    "buildinglimit": "building_value",
    "buildingsuminsured": "building_value",
    "contentsvalue": "contents_value",
    "contentslimit": "contents_value",
    "businesspersonalproperty": "contents_value",
    "bpp": "contents_value",
    "stockandequipment": "contents_value",
    "policylimits": "limits",
    "limitofliability": "limits",
    "coverageLimit": "coverage_limit",
    "coveragelimit": "coverage_limit",
    "eachoccurrence": "limits",
    "aggregate": "limits",
    "deductible": "deductible",
    "retention": "retention",
    "sir": "retention",
    "selfinsuredretention": "retention",
    "umbrellalimit": "umbrella_limit",
    "excesslimit": "umbrella_limit",
    "cargolimit": "cargo_limit",
  }

  def normalize_line(value):
    text = clean(value).lower()
    compact = norm_key(value)
    if compact in {"cgl", "commercialgeneralliability", "generalliability"} or "general liability" in text:
      return "General Liability"
    if any(x in text for x in ["fleet", "auto", "automobile", "scheduled autos", "business auto"]):
      return "Commercial Auto"
    if any(x in text for x in ["wcb", "wsib", "worksafebc", "cnesst", "workers", "compensation"]):
      return "Workers Compensation"
    if any(x in text for x in ["errors", "omissions", "e&o", "professional"]):
      return "Professional Liability"
    if "cyber" in text:
      return "Cyber"
    if any(x in text for x in ["umbrella", "excess"]):
      return "Umbrella / Excess"
    if any(x in text for x in ["property", "bop", "package"]):
      return "Property / Package"
    if any(x in text for x in ["cargo", "transit", "inland marine"]):
      return "Cargo / Inland Marine"
    return clean(value)

  # Map any raw/Canadian/U.S. terminology keys already captured by parser.
  for source in [profile_data, profile_data.get("exposure_inputs"), profile_data.get("exposures")]:
    if isinstance(source, dict):
      for source_key, source_value in list(source.items()):
        mapped = aliases.get(norm_key(source_key))
        if mapped:
          set_if_blank(mapped, source_value)

  # Clean money/count fields.
  for field in [
    "current_premium", "expiring_premium", "target_renewal_premium",
    "payroll", "revenue", "sales", "receipts", "professional_revenue",
    "cyber_revenue", "property_tiv", "tiv", "building_value",
    "contents_value", "limits", "coverage_limit", "deductible",
    "retention", "umbrella_limit", "cargo_limit",
  ]:
    if profile_data.get(field) not in (None, ""):
      fixed = money_clean(profile_data.get(field))
      if fixed:
        profile_data[field] = fixed

  for field in ["employee_count", "vehicle_count", "driver_count", "location_count", "unit_count", "square_footage"]:
    if profile_data.get(field) not in (None, ""):
      fixed = count_clean(profile_data.get(field))
      if fixed:
        profile_data[field] = fixed

  # Split combined fleet wording: '12 vehicles / 18 drivers'.
  combined = clean(profile_data.get("vehicle_count") or profile_data.get("fleet_size") or profile_data.get("fleet"))
  if combined:
    vehicle_match = re.search(r"(\d+)\s*(?:vehicles?|autos?|power units?|fleet units?)", combined, re.I)
    driver_match = re.search(r"(\d+)\s*(?:drivers?|operators?)", combined, re.I)
    if vehicle_match:
      profile_data["vehicle_count"] = vehicle_match.group(1)
    if driver_match:
      profile_data["driver_count"] = driver_match.group(1)

  # Build multi-line exposure from policy schedule when multiple policies exist.
  lines = []
  for row in profile_data.get("policies") or profile_data.get("policy_schedule") or []:
    if isinstance(row, dict):
      line = normalize_line(row.get("line_of_business") or row.get("policy_type") or row.get("coverage") or row.get("line"))
      if line and line not in lines:
        lines.append(line)

  if len(lines) > 1:
    profile_data["line_of_business"] = "Multi-line: " + ", ".join(lines)
    profile_data["primary_line_of_business"] = profile_data["line_of_business"]
  elif len(lines) == 1:
    profile_data["line_of_business"] = lines[0]
    profile_data["primary_line_of_business"] = lines[0]
  elif profile_data.get("line_of_business"):
    profile_data["line_of_business"] = normalize_line(profile_data.get("line_of_business"))
    profile_data["primary_line_of_business"] = profile_data["line_of_business"]

  # Province/state normalization, including Canadian postal code first-letter logic.
  province_map = {
    "ontario": "ON", "on": "ON", "alberta": "AB", "ab": "AB",
    "british columbia": "BC", "bc": "BC", "quebec": "QC", "québec": "QC", "qc": "QC",
    "manitoba": "MB", "mb": "MB", "saskatchewan": "SK", "sk": "SK",
    "nova scotia": "NS", "ns": "NS", "new brunswick": "NB", "nb": "NB",
    "newfoundland": "NL", "newfoundland and labrador": "NL", "nl": "NL",
    "prince edward island": "PE", "pei": "PE", "pe": "PE",
    "yukon": "YT", "yt": "YT", "northwest territories": "NT", "nt": "NT",
    "nunavut": "NU", "nu": "NU",
  }
  postal_prefix_map = {
    "A": "NL", "B": "NS", "C": "PE", "E": "NB", "G": "QC", "H": "QC", "J": "QC",
    "K": "ON", "L": "ON", "M": "ON", "N": "ON", "P": "ON", "R": "MB", "S": "SK",
    "T": "AB", "V": "BC", "X": "NT", "Y": "YT",
  }
  postal = clean(profile_data.get("postal_code") or profile_data.get("postcode"))
  postal_state = postal_prefix_map.get(postal.upper().replace(" ", "")[:1], "")
  explicit_state = province_map.get(clean(profile_data.get("province") or profile_data.get("province_code") or profile_data.get("state")).lower(), "")
  final_state = postal_state or explicit_state
  if final_state:
    profile_data["state"] = final_state
    profile_data["province"] = final_state
    profile_data["province_code"] = final_state

  print("LOSSQ_EXPOSURE_INPUT_UNIVERSAL_TERMINOLOGY_V1", {
    "line_of_business": profile_data.get("line_of_business"),
    "state": profile_data.get("state"),
    "vehicle_count": profile_data.get("vehicle_count"),
    "driver_count": profile_data.get("driver_count"),
    "property_tiv": profile_data.get("property_tiv"),
  })

  return profile_data


# LOSSQ_MARKET_INTELLIGENCE_PROFILE_APPLY_V1
def lossq_apply_market_intelligence_to_profile_v1(profile_data, raw_text=None):
  if not isinstance(profile_data, dict):
    return profile_data

  if lossq_normalize_market_profile is None:
    return profile_data

  try:
    raw_parts = []
    if raw_text:
      raw_parts.append(str(raw_text))

    for key in [
      "raw_text",
      "document_text",
      "upload_text",
      "ocr_text",
      "extracted_text",
      "loss_run_text",
      "business_name",
      "carrier_name",
      "writing_carrier",
      "insurer",
      "state",
      "province",
      "province_code",
      "postal_code",
      "postcode",
      "line_of_business",
      "exposure_basis",
    ]:
      value = profile_data.get(key)
      if value not in (None, ""):
        raw_parts.append(str(value))

    market_raw_text = "\n".join(raw_parts)
    normalized = lossq_normalize_market_profile(profile_data, market_raw_text)

    if not isinstance(normalized, dict):
      return profile_data

    # Keep this safe for the existing database schema.
    # These flattened fields only persist if AccountProfile/upsert already supports them.
    context = normalized.get("market_context") or {}
    region_context = context.get("region_context") or {}

    if context.get("country"):
      normalized.setdefault("market_country", context.get("country"))
    if context.get("currency"):
      normalized.setdefault("market_currency", context.get("currency"))
    if context.get("language"):
      normalized.setdefault("market_language", context.get("language"))
    if context.get("region_code"):
      normalized.setdefault("market_region_code", context.get("region_code"))
    if region_context.get("regulator"):
      normalized.setdefault("market_regulator", region_context.get("regulator"))
    if region_context.get("date_format"):
      normalized.setdefault("market_date_format", region_context.get("date_format"))

    print("LOSSQ_MARKET_INTELLIGENCE_PROFILE_APPLIED_V1", {
      "country": normalized.get("market_country"),
      "region": normalized.get("market_region_code") or normalized.get("state"),
      "currency": normalized.get("market_currency"),
      "regulator": normalized.get("market_regulator"),
      "line_of_business": normalized.get("line_of_business"),
      "carrier_name": normalized.get("carrier_name"),
      "vehicle_count": normalized.get("vehicle_count"),
      "driver_count": normalized.get("driver_count"),
      "property_tiv": normalized.get("property_tiv"),
    })

    return normalized
  except Exception as exc:
    print("LOSSQ_MARKET_INTELLIGENCE_PROFILE_APPLY_FAILED_V1", str(exc))
    return profile_data

def upsert_account_profile(db: Session, profile_data: dict, current_user: dict):
  # LOSSQ_UPSERT_POLICY_SCHEDULE_SANITIZE_CALL_V6
  profile_data = lossq_sanitize_profile_policies_before_upsert_v6(profile_data)
  # LOSSQ_MARKET_INTELLIGENCE_PROFILE_APPLY_CALL_V1
  profile_data = lossq_apply_market_intelligence_to_profile_v1(profile_data)
  # LOSSQ_EXPOSURE_INPUT_UNIVERSAL_TERMINOLOGY_CALL_V1
  profile_data = lossq_exposure_input_universal_terminology_v1(profile_data)

  policy_number = clean_profile_value(
    profile_data.get("policy_number")
    or profile_data.get("main_policy")
    or profile_data.get("main_policy_number")
    or profile_data.get("account_number")
  )

  if not policy_number:
    return None

  existing = (
    db.query(AccountProfile)
   .filter(AccountProfile.organization_id == current_user["organization_id"])
   .filter(AccountProfile.policy_number == policy_number)
   .first()
  )

  fields_to_save = [
    "business_name",
    "carrier_name",
    "writing_carrier",
    "agency_name",
    "account_number",
    "customer_number",
    "producer_number",
    "policy_number",
    "effective_date",
    "expiration_date",
    "evaluation_date",
    "raw_text_preview",

    # LOSSQ_SAVE_EXPOSURE_FIELDS_TO_PROFILE_V1
    "current_premium",
    "expiring_premium",
    "target_renewal_premium",
    "line_of_business",
    "state",
    "class_code",
    "class_codes",
    "limits",
    "coverage_limit",
    "deductible",
    "retention",
    "payroll",
    "revenue",
    "sales",
    "receipts",
    "employee_count",
    "vehicle_count",
    "driver_count",
    "property_tiv",
    "tiv",
    "building_value",
    "contents_value",
    "square_footage",
    "location_count",
    "unit_count",
    "cargo_limit",
    "umbrella_limit",
    "experience_mod",
    "mod",
    "exposure_change_percent",
    "cyber_revenue",
    "professional_revenue",
    "exposure_basis",
    "underwriter_notes",
  ]

  policies_json = serialize_json(profile_data.get("policies") or [], [])
  validation_json = serialize_json(profile_data.get("validation") or {}, {})

  if existing:
    for field in fields_to_save:
      value = clean_profile_value(profile_data.get(field))

      if value and hasattr(existing, field):
        setattr(existing, field, value)

    if hasattr(existing, "policies"):
      existing.policies = policies_json

    if hasattr(existing, "validation"):
      existing.validation = validation_json

    return existing

  # LOSSQ_FINAL_CARRIER_BAD_VALUE_CLEANUP_V1
  bad_final_carrier_values = {"exposure value", "exposure basis", "premium", "annual premium", "policy number", "line of business"}
  for carrier_key in ["carrier_name", "writing_carrier", "carrier"]:
    if str(profile_data.get(carrier_key) or "").strip().lower() in bad_final_carrier_values:
      profile_data[carrier_key] = ""

  # Backfill carrier from first real policy carrier if profile carrier was cleared.
  if not profile_data.get("carrier_name") and isinstance(profile_data.get("policies"), list):
    for policy_item in profile_data.get("policies") or []:
      possible_carrier = str(
        policy_item.get("carrier_name") or policy_item.get("writing_carrier") or policy_item.get("carrier") or ""
      ).strip()
      if possible_carrier and possible_carrier.lower() not in bad_final_carrier_values:
        profile_data["carrier_name"] = possible_carrier
        profile_data["writing_carrier"] = possible_carrier
        break

  # LOSSQ_FINAL_RAW_TEXT_CARRIER_BACKFILL_V1
  # Universal fallback for PDFs/text exports where carrier is present in account text
  # but policy schedule carrier cells are blank.
  if not profile_data.get("carrier_name"):
    raw_text_for_carrier = str(profile_data.get("raw_text_preview") or "")
    carrier_patterns = [
      r"(?i)\bwriting\s+carrier\b\s*[:\-]?\s*([A-Z][A-Za-z0-9&.,'\- ]{3,80})",
      r"(?i)\binsurance\s+carrier\b\s*[:\-]?\s*([A-Z][A-Za-z0-9&.,'\- ]{3,80})",
      r"(?i)\bcarrier\b\s*[:\-]?\s*([A-Z][A-Za-z0-9&.,'\- ]{3,80})",
    ]

    for carrier_pattern in carrier_patterns:
      carrier_match = re.search(carrier_pattern, raw_text_for_carrier)
      if not carrier_match:
        continue

      possible_carrier = str(carrier_match.group(1) or "").strip()
      possible_carrier = re.split(
        r"\b(Named Insured|Account Number|Policy Number|Effective|Expiration|Evaluation|Producing Agency|Agency|Exposure Value|Exposure Basis|Policy Schedule|Claim Detail)\b",
        possible_carrier,
        flags=re.I,
      )[0].strip(" :-|")

      if possible_carrier and possible_carrier.lower() not in bad_final_carrier_values:
        profile_data["carrier_name"] = possible_carrier
        profile_data["writing_carrier"] = possible_carrier
        break

  new_profile = AccountProfile(
    business_name=profile_data.get("business_name") or "Business Name Not Set",
    carrier_name=profile_data.get("carrier_name") or "Carrier Not Set",
    writing_carrier=profile_data.get("writing_carrier")
    or profile_data.get("carrier_name")
    or "Carrier Not Set",
    agency_name=profile_data.get("agency_name") or "Agency Not Set",
    account_number=profile_data.get("account_number") or policy_number,
    customer_number=profile_data.get("customer_number")
    or profile_data.get("account_number")
    or policy_number,
    producer_number=profile_data.get("producer_number") or "",
    policy_number=policy_number,
    effective_date=profile_data.get("effective_date") or "Not Set",
    expiration_date=profile_data.get("expiration_date") or "Not Set",
    evaluation_date=profile_data.get("evaluation_date") or "",
    policies=policies_json,
    validation=validation_json,
    raw_text_preview=profile_data.get("raw_text_preview") or "",
    organization_id=current_user["organization_id"],
  )

  # LOSSQ_UPSERT_ACCOUNT_PROFILE_EXPOSURE_FIELDS_V1
  # Ensure captured CSV/PDF/XLSX exposure inputs are saved on newly-created profiles.
  for exposure_field in [
    "current_premium",
    "expiring_premium",
    "target_renewal_premium",
    "line_of_business",
    "state",
    "class_code",
    "class_codes",
    "limits",
    "coverage_limit",
    "deductible",
    "retention",
    "payroll",
    "revenue",
    "sales",
    "receipts",
    "employee_count",
    "vehicle_count",
    "driver_count",
    "property_tiv",
    "tiv",
    "building_value",
    "contents_value",
    "square_footage",
    "location_count",
    "unit_count",
    "cargo_limit",
    "umbrella_limit",
    "experience_mod",
    "mod",
    "exposure_change_percent",
    "cyber_revenue",
    "professional_revenue",
    "exposure_basis",
    "underwriter_notes",
  ]:
    exposure_value = profile_data.get(exposure_field)
    if exposure_value not in ("", None, [], {}) and hasattr(new_profile, exposure_field):
      setattr(new_profile, exposure_field, exposure_value)

  db.add(new_profile)
  db.flush()
  db.refresh(new_profile)
  return new_profile

@router.post("/loss-run")
async def upload_loss_run(
  file: UploadFile = File(...),
  policy_number: str = Form(default=""),
  db: Session = Depends(get_db),
  current_user: dict = Depends(require_permission("upload")),
):
  # LOSSQ_UPLOAD_ROUTE_SAFE_DEFAULTS_V1
  # Initialize shared upload parse variables for every accepted file type.
  # This prevents PDF/XLSX/XLS paths from falling into rescue handling
  # with undefined parsed_claims / parsed_profile variables.
  parsed_claims = []
  parsed_profile = {}
  rescue_claims = []
  rescue_profile = {}

  # LOSSQ_UPLOAD_ROUTE_ERROR_REDACTED_V1
  try:
    await validate_upload_file_security(file)
    return await save_uploaded_files(
      files=[file],
      policy_number=policy_number,
      db=db,
      current_user=current_user,
    )
  except HTTPException:
    raise
  except Exception as e:
    print("LOSSQ_UPLOAD_ERROR_TRACE:", traceback.format_exc())
    # LOSSQ_UPLOAD_PROCESSING_FAILED_ROOT_CAUSE_V1
    print("LOSSQ_UPLOAD_PROCESSING_FAILED_ROOT_CAUSE:", str(e)[:1000])
    raise HTTPException(
      status_code=500,
      detail={
        "message": "Internal server error",
        "error": "Upload processing failed.",
      },
    )


@router.post("/loss-runs")
async def upload_multiple_loss_runs(
  files: List[UploadFile] = File(...),
  policy_number: str = Form(default=""),
  db: Session = Depends(get_db),
  current_user: dict = Depends(require_permission("upload")),
):
  try:
    # LOSSQ_UPLOAD_LOSS_RUN_VALIDATE_MULTIPLE_V1
    for upload_file in files:
      await validate_upload_file_security(upload_file)
    return await save_uploaded_files(
      files=files,
      policy_number=policy_number,
      db=db,
      current_user=current_user,
    )
  except HTTPException:
    raise
  except Exception as e:
    print("LOSSQ_UPLOAD_ERROR_TRACE:", traceback.format_exc())
    raise HTTPException(
      status_code=500,
      detail={
        "message": "Internal server error",
        "error": "Upload processing failed.",
      },
    )


@router.post("/debug-loss-run")
async def debug_loss_run_parser(
  file: UploadFile = File(...),
  current_user: dict = Depends(require_permission("upload")),
):
  # LOSSQ_UPLOAD_DEBUG_VALIDATE_FILE_V1
  safe_upload_filename = await validate_upload_file_security(file)
  timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
  safe_filename = (safe_upload_filename or "debug_loss_run.pdf").replace(" ", "_")
  file_path = os.path.join(UPLOAD_DIR, f"DEBUG-{timestamp}_{safe_filename}")

  with open(file_path, "wb") as buffer:
    shutil.copyfileobj(file.file, buffer)

  result = parse_loss_run_file(file_path, safe_filename)

  profile = result.get("profile") or {}
  policies = result.get("policies") or []
  claims = result.get("claims") or []
  validation = result.get("validation") or {}

  return {
    "profile": profile,
    "policy_count": len(policies),
    "policies": policies,
    "claim_count": len(claims),
    "claims": claims,
    "validation": validation,
    "raw_text_preview": result.get("raw_text_preview", "")[:3000],
  }


# LOSSQ_BETA_UPLOAD_GUARDRAILS_V1
def lossq_beta_clean_text(value):
  return re.sub(r"\s+", " ", str(value or "").strip())

def lossq_beta_norm_key(value):
  return lossq_beta_clean_text(value).upper()

def lossq_beta_valid_policy_key(value):
  key = lossq_beta_norm_key(value)
  if not key:
    return False
  bad = {
    "POLICY NOT SET",
    "NOT SET",
    "UNKNOWN",
    "N/A",
    "NONE",
    "LOSS SUMMARY",
    "METRIC",
    "TOTAL CLAIMS",
    "NOTE",
    "NOTES",
  }
  if key in bad:
    return False
  # LOSSQ_UNIVERSAL_MULTI_SEGMENT_POLICY_ID_V1
  # Accept universal carrier/account-prefixed policy IDs:
  # ABC-GL-2025-1234, ACCT-WC-2025-0001, ORG-LIAB-2025-55, etc.
  if bool(re.search(r"[A-Z0-9]{2,}[-_][A-Z0-9]{2,}[-_](19|20)\d{2}[-_][A-Z0-9]{2,}", key)):
    return True

  return bool(re.search(r"[A-Z]{2,10}[-_ ]?\d{4}[-_ ][A-Z0-9]+", key)) or bool(re.search(r"[A-Z]{2,10}-\d+", key))

def lossq_beta_valid_claim_number(value):
  key = lossq_beta_norm_key(value)
  if not key:
    return False

  blocked_exact = {
    "NOTE",
    "NOTES",
    "METRIC",
    "VALUE",
    "FIELD",
    "LOSS SUMMARY",
    "UNDERWRITING NOTES",
    "TOTAL CLAIMS",
    "OPEN CLAIMS",
    "CLOSED CLAIMS",
    "TOTAL PAID",
    "TOTAL RESERVE",
    "TOTAL INCURRED",
    "LARGEST LOSS",
    "LOSS RATIO",
    "CURRENT PREMIUM",
    "EXPIRING PREMIUM",
    "TARGET RENEWAL PREMIUM",
    "PAYROLL",
    "REVENUE / SALES",
    "EMPLOYEE COUNT",
    "VEHICLE COUNT",
    "DRIVER COUNT",
    "PROPERTY TIV",
  }
  if key in blocked_exact:
    return False

  blocked_contains = [
    "FICTIONAL TEST",
    "DESIGNED TO TEST",
    "NOT AFFILIATED",
    "LOSS SUMMARY",
    "UNDERWRITING NOTES",
    "EXPOSURE INPUTS",
    "POLICY SCHEDULE",
    "ACCOUNT INFORMATION",
  ]
  if any(item in key for item in blocked_contains):
    return False

  if not re.search(r"\d", key):
    return False

  # LOSSQ_REJECT_POLICY_FRAGMENT_AS_CLAIM_V2
  # Reject policy schedule fragments that look like line + year + policy suffix.
  # Rejected: GL-2025, CY-2025, BOP-2025, GL-2025-3101-GENERAL.
  # Accepted: carrier/account-prefixed commercial claim IDs with line and numeric claim segments.
  line_tokens = (
    "GL", "WC", "AUTO", "AU", "PROP", "PR", "CP", "BOP", "CY", "CYBER",
    "UMB", "EXCESS", "EPLI", "EPL", "DO", "DNO", "EO", "PL", "IM",
    "CRIME", "FID", "FIDUCIARY", "CARGO", "MTC", "LIAB", "ABUSE",
    "MOLESTATION", "GAR", "GARAGE"
  )

  policy_fragment_pattern = r"^(" + "|".join(line_tokens) + r")[-_ ]?(19|20)\d{2}([-_ ][A-Z0-9]+){0,3}$"
  if re.match(policy_fragment_pattern, key):
    return False

  # Real claim numbers usually include an account/carrier prefix before the line token.
  real_prefixed_claim_pattern = r"^[A-Z0-9]{2,}[-_](" + "|".join(line_tokens) + r")[-_]\d{2,4}[-_]\d{3,8}$"
  if re.match(real_prefixed_claim_pattern, key):
    return True

  # Accept explicit claim IDs.
  if re.search(r"\b(CLM|CLAIM)[-_ ]?[A-Z0-9]{3,}", key):
    return True

  # Accept structured alphanumeric claim IDs with at least 3 meaningful segments,
  # but only if they are not policy-fragment shaped.
  if re.search(r"^[A-Z0-9]{2,}[-_][A-Z0-9]{2,}[-_][A-Z0-9]{2,}([-_][A-Z0-9]{2,})?$", key):
    return True

  compact = re.sub(r"[^A-Z0-9]", "", key)
  if len(compact) >= 8 and re.search(r"\d", compact) and re.search(r"[A-Z]", compact):
    return True

  return False

# LOSSQ_UNIVERSAL_REAL_CLAIM_ROW_EVIDENCE_V1
def lossq_beta_money_to_float(value):
  try:
    raw = str(value or "").strip()
    if not raw:
      return 0.0

    raw = raw.replace("$", "").replace(",", "").replace(" ", "")
    negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()")

    if raw in {"", "-", "--", "N/A", "NA", "NONE", "NULL"}:
      return 0.0

    number = float(raw)
    return -number if negative else number
  except Exception:
    return 0.0


def lossq_beta_extract_money_triplet_from_text(item):
  if not isinstance(item, dict):
    return {}

  text_parts = []
  for key in [
    "description",
    "loss_description",
    "claim_description",
    "cause_of_loss",
    "narrative",
    "notes",
    "raw_text",
  ]:
    value = item.get(key)
    if value not in ("", None):
      text_parts.append(str(value))

  text_value = " ".join(text_parts)
  if not text_value.strip():
    return {}

  # Remove dates so date numbers are not mistaken for claim dollars.
  scrubbed = re.sub(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", " ", text_value)
  scrubbed = re.sub(r"\b(19|20)\d{2}\b", " ", scrubbed)

  line_patterns = [
    r"general\s+liability",
    r"liquor\s+liability",
    r"workers?\s+comp(?:ensation)?",
    r"business\s*owners?\s+policy",
    r"\bbop\b",
    r"cyber\s+liability",
    r"commercial\s+auto",
    r"auto\s+liability",
    r"cargo",
    r"property",
    r"umbrella",
    r"excess",
    r"epli",
    r"employment\s+practices",
    r"directors?\s+and\s+officers?",
    r"\bd\s*&\s*o\b",
    r"professional\s+liability",
    r"errors?\s+and\s+omissions?",
    r"inland\s+marine",
    r"crime",
    r"abuse",
    r"molestation",
    r"garage",
  ]

  segments = []
  for pattern in line_patterns:
    match = re.search(pattern, scrubbed, re.IGNORECASE)
    if match:
      segments.append(scrubbed[match.end(): match.end() + 220])

  # Fallback to full text if no known commercial line label was found.
  if not segments:
    segments.append(scrubbed[:260])

  for segment in segments:
    tokens = re.findall(r"\$?\(?\d[\d,]*(?:\.\d+)?\)?", segment)
    numbers = [lossq_beta_money_to_float(token) for token in tokens]

    # Keep zeros because reserve can be 0. Require at least one positive value.
    clean_numbers = [n for n in numbers if n >= 0]
    if len(clean_numbers) >= 3 and any(n > 0 for n in clean_numbers[:3]):
      paid = clean_numbers[0]
      reserve = clean_numbers[1]
      total = clean_numbers[2]
      if total <= 0 and (paid > 0 or reserve > 0):
        total = paid + reserve
      return {
        "paid_amount": paid,
        "reserve_amount": reserve,
        "total_incurred": total,
      }

    if len(clean_numbers) >= 2 and any(n > 0 for n in clean_numbers[:2]):
      paid = clean_numbers[0]
      reserve = 0.0
      total = clean_numbers[1]
      if total <= 0 and paid > 0:
        total = paid
      return {
        "paid_amount": paid,
        "reserve_amount": reserve,
        "total_incurred": total,
      }

  return {}


def lossq_beta_get_claim_amounts(item):
  if not isinstance(item, dict):
    return {}

  paid = lossq_beta_money_to_float(
    item.get("paid_amount")
    or item.get("paid")
    or item.get("Paid")
    or item.get("Paid Amount")
    or item.get("Total Paid")
  )
  reserve = lossq_beta_money_to_float(
    item.get("reserve_amount")
    or item.get("reserve")
    or item.get("Reserve")
    or item.get("Reserve Amount")
    or item.get("Outstanding Reserve")
  )
  total = lossq_beta_money_to_float(
    item.get("total_incurred")
    or item.get("incurred")
    or item.get("Total Incurred")
    or item.get("Incurred")
    or item.get("Total")
    or item.get("total")
  )

  if total <= 0 and (paid > 0 or reserve > 0):
    total = paid + reserve

  amounts = {
    "paid_amount": paid,
    "reserve_amount": reserve,
    "total_incurred": total,
  }

  if not any(value > 0 for value in amounts.values()):
    recovered = lossq_beta_extract_money_triplet_from_text(item)
    if recovered:
      amounts.update(recovered)

  return amounts


def lossq_beta_apply_recovered_amounts(item):
  if not isinstance(item, dict):
    return item, {}

  def current_amount(key):
    return lossq_beta_money_to_float(item.get(key))

  current_paid = current_amount("paid_amount")
  current_reserve = current_amount("reserve_amount")
  current_total = current_amount("total_incurred")

  text_amounts = lossq_beta_extract_money_triplet_from_text(item)

  use_text_override = False
  override_reason = ""

  if text_amounts and any(lossq_beta_money_to_float(v) > 0 for v in text_amounts.values()):
    text_paid = lossq_beta_money_to_float(text_amounts.get("paid_amount"))
    text_reserve = lossq_beta_money_to_float(text_amounts.get("reserve_amount"))
    text_total = lossq_beta_money_to_float(text_amounts.get("total_incurred"))

    text_total_matches_parts = abs((text_paid + text_reserve) - text_total) <= max(2.0, text_total * 0.02)

    current_total_conflicts = (
      current_total > 0
      and text_total > 0
      and abs(current_total - text_total) > max(100.0, text_total * 0.10)
    )

    current_amounts_missing = (
      current_paid <= 0
      or ("reserve_amount" not in item and "reserve" not in item)
      or current_total <= 0
    )

    # LOSSQ_ROW_TEXT_AMOUNT_OVERRIDE_V1
    # Universal safeguard:
    # If the claim narrative/row text contains a clean paid + reserve = total triplet,
    # trust that row-level triplet over a conflicting summary/header amount.
    if text_total_matches_parts and (current_total_conflicts or current_amounts_missing):
      use_text_override = True
      override_reason = "row_text_triplet_conflict" if current_total_conflicts else "row_text_triplet_missing_amounts"

  if use_text_override:
    amounts = text_amounts
  else:
    amounts = lossq_beta_get_claim_amounts(item)

  if not amounts or not any(lossq_beta_money_to_float(value) > 0 for value in amounts.values()):
    return item, {}

  changed = {}

  for key in ["paid_amount", "reserve_amount", "total_incurred"]:
    current = lossq_beta_money_to_float(item.get(key))
    recovered = lossq_beta_money_to_float(amounts.get(key))

    should_replace = False

    if use_text_override:
      should_replace = True
    elif current <= 0 and recovered >= 0:
      should_replace = True

    if should_replace:
      before = item.get(key)
      item[key] = recovered
      if str(before) != str(recovered):
        changed[key] = recovered

  if changed:
    print("LOSSQ_BETA_AMOUNT_RECOVERY_APPLIED:", {
      "claim_number": lossq_beta_clean_text(
        item.get("claim_number") or item.get("claim_id") or item.get("Claim Number")
      ),
      "policy_number": lossq_beta_clean_text(
        item.get("policy_number") or item.get("Policy Number") or item.get("policy")
      ),
      "reason": override_reason or "missing_amount_recovery",
      **changed,
    })

  return item, changed


def lossq_beta_has_commercial_line_context(item):
  if not isinstance(item, dict):
    return False

  text_value = " ".join(
    str(value or "")
    for key, value in item.items()
    if key in {
      "line_of_business",
      "claim_type",
      "policy_type",
      "coverage",
      "Coverage",
      "Line of Business",
      "description",
      "loss_description",
      "claim_description",
      "cause_of_loss",
    }
  ).upper()

  line_terms = [
    "GENERAL LIABILITY",
    "LIQUOR LIABILITY",
    "WORKERS COMPENSATION",
    "WORKERS COMP",
    "BUSINESSOWNERS POLICY",
    "BUSINESS OWNERS POLICY",
    "BOP",
    "CYBER LIABILITY",
    "CYBER",
    "COMMERCIAL AUTO",
    "AUTO LIABILITY",
    "CARGO",
    "PROPERTY",
    "UMBRELLA",
    "EXCESS",
    "EPLI",
    "EMPLOYMENT PRACTICES",
    "DIRECTORS AND OFFICERS",
    "D&O",
    "PROFESSIONAL LIABILITY",
    "ERRORS AND OMISSIONS",
    "E&O",
    "INLAND MARINE",
    "CRIME",
    "ABUSE",
    "MOLESTATION",
    "GARAGE",
  ]

  return any(term in text_value for term in line_terms)


def lossq_beta_has_real_claim_row_evidence(item):
  if not isinstance(item, dict):
    return False

  description = lossq_beta_clean_text(
    item.get("description")
    or item.get("loss_description")
    or item.get("claim_description")
    or item.get("cause_of_loss")
    or ""
  )

  policy_number = (
    item.get("policy_number")
    or item.get("Policy Number")
    or item.get("policy_no")
    or item.get("policy")
    or ""
  )

  has_policy_context = bool(policy_number and lossq_beta_valid_policy_key(policy_number))
  has_line_context = lossq_beta_has_commercial_line_context(item)

  amounts = lossq_beta_get_claim_amounts(item)
  has_financial_context = bool(amounts and any(value > 0 for value in amounts.values()))

  has_loss_narrative = bool(len(description) >= 12 and re.search(r"[A-Za-z]", description))

  # Universal claim-row test:
  # A row can survive an imperfect/generated claim number only if the row still
  # looks like an actual claim: commercial line + financial values + policy or narrative context.
  return bool(
    has_line_context
    and has_financial_context
    and (has_policy_context or has_loss_narrative)
  )


def lossq_beta_filter_claim_rows(parsed_claims):
  clean_claims = []
  removed_rows = []

  for item in parsed_claims or []:
    if not isinstance(item, dict):
      removed_rows.append({"reason": "not_dict", "row": str(item)[:160]})
      continue

    item, recovered_amounts = lossq_beta_apply_recovered_amounts(item)

    claim_number = (
      item.get("claim_number")
      or item.get("Claim Number")
      or item.get("claim_id")
      or item.get("Claim ID")
      or item.get("claim_no")
      or ""
    )

    policy_number = (
      item.get("policy_number")
      or item.get("Policy Number")
      or item.get("policy_no")
      or item.get("Policy No")
      or item.get("policy")
      or ""
    )

    description = (
      item.get("description")
      or item.get("loss_description")
      or item.get("claim_description")
      or ""
    )

    valid_claim_number = lossq_beta_valid_claim_number(claim_number)
    real_claim_evidence = lossq_beta_has_real_claim_row_evidence(item)

    if not valid_claim_number and not real_claim_evidence:
      removed_rows.append({
        "reason": "invalid_claim_number",
        "claim_number": lossq_beta_clean_text(claim_number),
        "description": lossq_beta_clean_text(description)[:120],
      })
      continue

    # Keep claim if policy is valid or parser will fallback later.
    if policy_number and not lossq_beta_valid_policy_key(policy_number):
      removed_rows.append({
        "reason": "invalid_policy_number",
        "claim_number": lossq_beta_clean_text(claim_number),
        "policy_number": lossq_beta_clean_text(policy_number),
      })
      continue

    if real_claim_evidence and not valid_claim_number:
      print("LOSSQ_BETA_REAL_CLAIM_ROW_RESCUED:", {
        "claim_number": lossq_beta_clean_text(claim_number),
        "policy_number": lossq_beta_clean_text(policy_number),
        "line_of_business": lossq_beta_clean_text(
          item.get("line_of_business") or item.get("claim_type") or item.get("policy_type")
        ),
        "paid_amount": item.get("paid_amount"),
        "reserve_amount": item.get("reserve_amount"),
        "total_incurred": item.get("total_incurred"),
      })

    clean_claims.append(item)

  return clean_claims, removed_rows


def lossq_beta_collect_upload_policy_keys(parsed_profile, parsed_claims, fallback_policy_number=""):
  keys = set()

  def add(value):
    value = lossq_beta_norm_key(value)
    if lossq_beta_valid_policy_key(value):
      keys.add(value)

  if isinstance(parsed_profile, dict):
    for name in ["policy_number", "main_policy_number", "main_policy"]:
      add(parsed_profile.get(name))

    policies = parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []
    if isinstance(policies, list):
      for policy in policies:
        if isinstance(policy, dict):
          add(policy.get("policy_number") or policy.get("policy") or policy.get("policy_no"))

  for claim in parsed_claims or []:
    if isinstance(claim, dict):
      add(claim.get("policy_number") or claim.get("policy") or claim.get("policy_no"))

  add(fallback_policy_number)

  return sorted(keys)

def lossq_beta_purge_prior_upload_data(db, current_user, policy_keys):
  result = {
    "policy_keys": policy_keys or [],
    "deleted_claims": 0,
    "deleted_upload_history": 0,
  }

  if not db or not current_user or not policy_keys:
    return result

  org_id = current_user.get("organization_id") if isinstance(current_user, dict) else None
  if not org_id:
    return result

  upper_keys = [lossq_beta_norm_key(key) for key in policy_keys if lossq_beta_valid_policy_key(key)]
  if not upper_keys:
    return result

  try:
    deleted_claims = (
      db.query(Claim)
     .filter(Claim.organization_id == org_id)
     .filter(func.upper(func.trim(Claim.policy_number)).in_(upper_keys))
     .delete(synchronize_session=False)
    )
    result["deleted_claims"] = int(deleted_claims or 0)
  except Exception as exc:
    result["claim_purge_warning"] = str(exc)[:200]

  try:
    if "UploadHistory" in globals():
      deleted_uploads = (
        db.query(UploadHistory)
       .filter(UploadHistory.organization_id == org_id)
       .filter(func.upper(func.trim(UploadHistory.policy_number)).in_(upper_keys))
       .delete(synchronize_session=False)
      )
      result["deleted_upload_history"] = int(deleted_uploads or 0)
  except Exception as exc:
    result["upload_history_purge_warning"] = str(exc)[:200]

  return result


# LOSSQ_SECTION_CSV_PROFILE_DATE_REPAIR_V1
# LOSSQ_PRODUCING_AGENCY_EXTRACTION_V1
def lossq_section_csv_clean(value):
  return re.sub(r"\s+", " ", str(value or "").strip())

def lossq_section_csv_key(value):
  return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

def lossq_section_csv_date(value):
  raw = lossq_section_csv_clean(value)
  if not raw:
    return ""

  m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", raw)
  if m:
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if year < 100:
      year += 2000 if year < 50 else 1900
    if 1 <= month <= 12 and 1 <= day <= 31:
      return f"{month:02d}/{day:02d}/{year}"

  m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", raw)
  if m:
    year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if 1 <= month <= 12 and 1 <= day <= 31:
      return f"{month:02d}/{day:02d}/{year}"

  return raw

def lossq_section_csv_valid_carrier(value):
  text_value = lossq_section_csv_clean(value)
  low = text_value.lower()

  if not text_value:
    return False

  bad_exact = {
    "carrier",
    "writing carrier",
    "effective date",
    "expiration date",
    "valuation date",
    "evaluation date",
    "as of date",
    "policy number",
    "main policy",
    "account number",
    "producer",
    "named insured",
  }

  if low in bad_exact:
    return False

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", text_value):
    return False

  return True

def lossq_section_csv_apply_profile_date_repair(file_path, parsed_profile):
  """
  Universal repair for section-based CSV loss runs.

  It reads Account Information and Policy Schedule sections directly from the CSV,
  then merges dates/carrier/account/policies into parsed_profile before saving.
  """
  parsed_profile = parsed_profile or {}

  try:
    import csv as _csv

    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
      rows = [row for row in _csv.reader(f)]
  except Exception:
    return parsed_profile

  account_info = {}
  policies = []
  current_header = []

  for raw_row in rows:
    row = [lossq_section_csv_clean(cell) for cell in raw_row]
    if not any(row):
      continue

    first = row[0].strip()
    first_key = lossq_section_csv_key(first)

    # Three-column Account Information rows:
    # Section, Field, Value
    if first_key == "accountinformation" and len(row) >= 3:
      field = lossq_section_csv_key(row[1])
      value = row[2]

      if field in {"carrier", "carriercarriername"} and lossq_section_csv_valid_carrier(value):
        lossq_set_once(account_info, "carrier_name", value)
      elif field in {"writingcarrier"} and lossq_section_csv_valid_carrier(value):
        lossq_set_once(account_info, "writing_carrier", value)
      elif field in {"namedinsured", "insured", "businessname"}:
        account_info["business_name"] = value
        account_info["named_insured"] = value
      elif field in {"producer", "producingagency", "agency", "agencyname", "broker", "brokerage"}:
        account_info["agency_name"] = value
        account_info["producer"] = value
        account_info["producing_agency"] = value
      elif field in {"accountnumber", "customernumber"}:
        account_info["account_number"] = value
        account_info["customer_number"] = value
      elif field in {"mainpolicy", "mainpolicynumber", "policynumber"}:
        lossq_set_once(account_info, "policy_number", value)
      elif field in {"effectivedate", "policyeffectivedate", "effective"}:
        account_info["effective_date"] = lossq_section_csv_date(value)
        account_info["policy_effective_date"] = lossq_section_csv_date(value)
      elif field in {"expirationdate", "expirydate", "policyexpirationdate", "expiration"}:
        account_info["expiration_date"] = lossq_section_csv_date(value)
        account_info["policy_expiration_date"] = lossq_section_csv_date(value)
      elif field in {"valuationdate", "evaluationdate", "asofdate", "reportdate"}:
        fixed_date = lossq_section_csv_date(value)
        account_info["valuation_date"] = fixed_date
        account_info["evaluation_date"] = fixed_date
        account_info["loss_run_valuation_date"] = fixed_date

      continue

    # Header rows.
    if first_key == "section" and len(row) > 1:
      current_header = row
      continue

    # Policy Schedule rows that follow:
    # Section, Policy Number, Line of Business, Carrier, Effective Date, Expiration Date...
    if first_key == "policyschedule" and current_header:
      mapped = {}
      for idx, header_name in enumerate(current_header):
        if idx >= len(row):
          continue
        key = lossq_section_csv_key(header_name)
        value = row[idx]

        if key == "policynumber":
          mapped["policy_number"] = value
        elif key in {"lineofbusiness", "policytype", "coverage", "linecoverage"}:
          mapped["line_of_business"] = value
          mapped["policy_type"] = value
          mapped["coverage"] = value
        elif key in {"carrier", "carriername", "writingcarrier"} and lossq_section_csv_valid_carrier(value):
          mapped["carrier"] = value
          mapped["carrier_name"] = value
        elif key in {"effectivedate", "policyeffectivedate", "effective"}:
          mapped["effective_date"] = lossq_section_csv_date(value)
          mapped["policy_effective_date"] = lossq_section_csv_date(value)
        elif key in {"expirationdate", "expirydate", "policyexpirationdate", "expiration"}:
          mapped["expiration_date"] = lossq_section_csv_date(value)
          mapped["policy_expiration_date"] = lossq_section_csv_date(value)
        elif key in {"currentpremium", "premium"}:
          mapped["current_premium"] = value
        elif key in {"exposurebasis"}:
          mapped["exposure_basis"] = value
        elif key in {"exposurevalue"}:
          mapped["exposure_value"] = value
        elif key == "state":
          mapped["state"] = value

      if mapped.get("policy_number"):
        policies.append(mapped)

  for key, value in account_info.items():
    if value:
      parsed_profile[key] = value

  if policies:
    parsed_profile["policies"] = policies
    parsed_profile["policy_schedule"] = policies

    if not parsed_profile.get("policy_number"):
      parsed_profile["policy_number"] = policies[0].get("policy_number", "")

    if not parsed_profile.get("effective_date"):
      parsed_profile["effective_date"] = policies[0].get("effective_date", "")
      parsed_profile["policy_effective_date"] = policies[0].get("effective_date", "")

    if not parsed_profile.get("expiration_date"):
      parsed_profile["expiration_date"] = policies[0].get("expiration_date", "")
      parsed_profile["policy_expiration_date"] = policies[0].get("expiration_date", "")

    if not parsed_profile.get("carrier_name"):
      for policy in policies:
        if lossq_section_csv_valid_carrier(policy.get("carrier_name")):
          parsed_profile["carrier_name"] = policy.get("carrier_name")
          break

    if not parsed_profile.get("writing_carrier"):
      parsed_profile["writing_carrier"] = parsed_profile.get("carrier_name", "")

  return parsed_profile


# LOSSQ_PROFILE_FIRST_VALID_VALUE_WINS_V1
def lossq_set_once(target, key, value):
  value = lossq_section_csv_clean(value)
  if not value:
    return

  current = lossq_section_csv_clean(target.get(key))
  if not current:
    target[key] = value


# LOSSQ_MESSY_CSV_LABEL_VALUE_VALIDATION_V1
def lossq_csv_is_header_or_label_value(value):
  clean = lossq_section_csv_clean(value)
  key = lossq_section_csv_key(clean)

  if not clean:
    return True

  label_keys = {
    "section",
    "field",
    "value",
    "policy",
    "policynumber",
    "policytype",
    "policytypecoverage",
    "lineofbusiness",
    "coverage",
    "carrier",
    "carriername",
    "writingcarrier",
    "effective",
    "effectivedate",
    "policyeffectivedate",
    "expiration",
    "expirationdate",
    "policyexpirationdate",
    "expiry",
    "expirydate",
    "annualpremium",
    "currentpremium",
    "premium",
    "exposure",
    "exposurebasis",
    "exposurevalue",
    "claims",
    "claimcount",
    "totalincurred",
    "claimdetail",
    "losssummary",
    "accountprofile",
    "policy schedule",
    "policyschedule",
    "producer",
    "producingagency",
    "agency",
    "agencyname",
    "broker",
    "brokerage",
    "adjuster",
    "claimhandler",
    "examiner",
    "downloadedby",
    "createdby",
  }

  return key in {lossq_section_csv_key(item) for item in label_keys}


def lossq_csv_valid_profile_date_value(value):
  clean = lossq_section_csv_clean(value)
  if lossq_csv_is_header_or_label_value(clean):
    return False

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", clean):
    return True

  if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", clean):
    return True

  return False


def lossq_csv_valid_profile_text_value(value):
  clean = lossq_section_csv_clean(value)
  if lossq_csv_is_header_or_label_value(clean):
    return False

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", clean):
    return False

  if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", clean):
    return False

  return True


def lossq_profile_date_or_blank(value):
  clean = lossq_section_csv_clean(value)

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", clean):
    return lossq_section_csv_date(clean)

  if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", clean):
    return lossq_section_csv_date(clean)

  return ""


# LOSSQ_MESSY_CSV_LABEL_PAIR_PROFILE_REPAIR_V1

# LOSSQ_POLICY_PERIOD_RANGE_SPLIT_V1
def lossq_policy_period_range_dates(value):
  """
  Universal parser for combined policy period values like:
  03/01/2025 - 03/01/2026
  03/01/2025 to 03/01/2026
  Effective 03/01/2025 Expiration 03/01/2026
  """
  try:
    import re

    clean = lossq_section_csv_clean(value)
    if not clean:
      return "", ""

    date_matches = re.findall(
      r"\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\b",
      clean,
    )

    if len(date_matches) >= 2:
      effective = lossq_profile_date_or_blank(date_matches[0])
      expiration = lossq_profile_date_or_blank(date_matches[1])
      return effective, expiration

    return "", ""
  except Exception as exc:
    print("LOSSQ_POLICY_PERIOD_RANGE_SPLIT_ERROR:", str(exc)[:200])
    return "", ""


def lossq_csv_label_pair_profile_repair(file_path, parsed_profile):
  """
  Universal repair for messy CSV exports where account profile fields are stored
  as label/value pairs across rows instead of Section, Field, Value format.

  Example:
  Policy Effective Date, 01/01/2025, Policy Expiration Date, 01/01/2026
  Evaluation Date, 06/30/2025
  """
  parsed_profile = parsed_profile or {}

  filename = str(file_path or "").lower()
  if not filename.endswith(".csv"):
    return parsed_profile

  try:
    import csv as _csv

    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
      rows = [row for row in _csv.reader(f)]
  except Exception:
    return parsed_profile

  account_info = {}
  policies = []

  def put_profile(label, value):
    field = lossq_section_csv_key(label)
    value = lossq_section_csv_clean(value)

    if not value:
      return

    if field in {"namedinsured", "insured", "businessname", "accountname"} and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "business_name", value)
      lossq_set_once(account_info, "named_insured", value)
      lossq_set_once(account_info, "insured", value)

    elif field in {"dba"} and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "dba", value)

    elif field in {"carrier", "carriername", "insurancecarrier"} and lossq_section_csv_valid_carrier(value) and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "carrier_name", value)

    elif field in {"writingcarrier", "underwritingcarrier"} and lossq_section_csv_valid_carrier(value) and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "writing_carrier", value)

    elif field in {"producingagency", "producer", "agency", "agencyname", "broker", "brokerage"} and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "agency_name", value)
      lossq_set_once(account_info, "producer", value)
      lossq_set_once(account_info, "producing_agency", value)

    elif field in {"accountnumber", "customernumber", "accountid"} and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "account_number", value)
      lossq_set_once(account_info, "customer_number", value)

    elif field in {"mainpolicy", "mainpolicynumber", "policynumber"} and lossq_csv_valid_profile_text_value(value):
      lossq_set_once(account_info, "policy_number", value)

    elif field in {"policyeffectivedate", "effectivedate", "effective"} and lossq_csv_valid_profile_date_value(value):
      fixed = lossq_section_csv_date(value)
      account_info["effective_date"] = fixed
      account_info["policy_effective_date"] = fixed

    elif field in {"policyexpirationdate", "expirationdate", "expirydate", "expiration", "expiry"} and lossq_csv_valid_profile_date_value(value):
      fixed = lossq_section_csv_date(value)
      account_info["expiration_date"] = fixed
      account_info["policy_expiration_date"] = fixed

    elif field in {"policyperiod", "policyterm", "period", "coverageperiod", "policydates", "daterange"}:
      effective, expiration = lossq_policy_period_range_dates(value)
      if effective and not account_info.get("effective_date"):
        account_info["effective_date"] = effective
        account_info["policy_effective_date"] = effective
      if expiration and not account_info.get("expiration_date"):
        account_info["expiration_date"] = expiration
        account_info["policy_expiration_date"] = expiration
      if effective or expiration:
        print("LOSSQ_POLICY_PERIOD_RANGE_PROFILE_DATES:", {
          "effective_date": account_info.get("effective_date"),
          "expiration_date": account_info.get("expiration_date"),
        })

    elif field in {"evaluationdate", "valuationdate", "valuedasof", "asofdate", "reportdate", "lossrunvaluationdate"} and lossq_csv_valid_profile_date_value(value):
      fixed = lossq_section_csv_date(value)
      account_info["evaluation_date"] = fixed
      account_info["valuation_date"] = fixed
      account_info["loss_run_valuation_date"] = fixed

  # LOSSQ_MESSY_CSV_PROFILE_SCAN_STOPS_BEFORE_CLAIMS_V1
  # Read label/value pairs only from account/profile area. Claim detail rows may include
  # producers, adjusters, examiners, or claim handlers that are not the producing agency.
  for row in rows[:80]:
    clean_row = [lossq_section_csv_clean(cell) for cell in row]
    row_key_text = " ".join(lossq_section_csv_key(cell) for cell in clean_row)

    if any(stop_key in row_key_text for stop_key in [
      "claimdetail",
      "claimnumber",
      "dateofloss",
      "losssummary",
      "trailingexportnoise",
      "underwritingnotes",
    ]):
      break

    for idx in range(0, len(clean_row) - 1):
      label = clean_row[idx]
      value = clean_row[idx + 1]

      if not label or not value:
        continue

      put_profile(label, value)


  # LOSSQ_PROFILE_DATES_FROM_POLICY_SCHEDULE_V1
  # If the account-level dates were not captured from label/value rows,
  # use the first valid policy schedule effective/expiration dates.
  try:
    schedule_rows = account_info.get("policies") or account_info.get("policy_schedule") or []
    if isinstance(schedule_rows, list):
      for policy_row in schedule_rows:
        if not isinstance(policy_row, dict):
          continue

        effective = (
          policy_row.get("effective_date")
          or policy_row.get("policy_effective_date")
          or policy_row.get("effective")
        )
        expiration = (
          policy_row.get("expiration_date")
          or policy_row.get("policy_expiration_date")
          or policy_row.get("expiration")
          or policy_row.get("expiry_date")
        )

        if effective and not account_info.get("effective_date"):
          fixed_effective = lossq_section_csv_date(effective)
          account_info["effective_date"] = fixed_effective
          account_info["policy_effective_date"] = fixed_effective

        if expiration and not account_info.get("expiration_date"):
          fixed_expiration = lossq_section_csv_date(expiration)
          account_info["expiration_date"] = fixed_expiration
          account_info["policy_expiration_date"] = fixed_expiration

        if account_info.get("effective_date") and account_info.get("expiration_date"):
          break
  except Exception as exc:
    print("LOSSQ_PROFILE_DATES_FROM_POLICY_SCHEDULE_ERROR:", str(exc)[:200])


  # Parse policy schedule tables with columns like:
  # Policy Type / Coverage, Policy Number, Carrier, Effective, Expiration...
  for idx, raw_row in enumerate(rows):
    header = [lossq_section_csv_clean(cell) for cell in raw_row]
    header_keys = [lossq_section_csv_key(cell) for cell in header]

    if "policynumber" not in header_keys:
      continue

    if not any(key in header_keys for key in ["policytypecoverage", "lineofbusiness", "coverage", "policytype", "coverageline", "linecoverage"]):
      continue

    # Found a policy schedule header.
    for data_row in rows[idx + 1:]:
      row = [lossq_section_csv_clean(cell) for cell in data_row]

      if not any(row):
        break

      first_key = lossq_section_csv_key(row[0] if row else "")
      if first_key in {"claimdetail", "losssummary", "trailingexportnoise", "underwritingnotes"}:
        break

      mapped = {}

      for col_index, header_key in enumerate(header_keys):
        if col_index >= len(row):
          continue

        value = row[col_index]

        if header_key in {"policytypecoverage", "lineofbusiness", "coverage", "policytype", "coverageline", "linecoverage"}:
          mapped["line_of_business"] = value
          mapped["policy_type"] = value
          mapped["coverage"] = value

        elif header_key in {"policynumber", "policy", "policyno", "policynum"}:
          mapped["policy_number"] = value

        elif header_key in {"carrier", "carriername", "writingcarrier"} and lossq_section_csv_valid_carrier(value):
          mapped["carrier"] = value
          mapped["carrier_name"] = value
          mapped["writing_carrier"] = value

        elif header_key in {"effective", "effectivedate", "policyeffectivedate"}:
          fixed = lossq_section_csv_date(value)
          mapped["effective_date"] = fixed
          mapped["policy_effective_date"] = fixed

        elif header_key in {"expiration", "expirationdate", "expiry", "expirydate", "policyexpirationdate"}:
          fixed = lossq_section_csv_date(value)
          mapped["expiration_date"] = fixed
          mapped["policy_expiration_date"] = fixed

        # LOSSQ_POLICY_PERIOD_RANGE_SCHEDULE_DATES_V1
        elif header_key in {"policyperiod", "policyterm", "period", "coverageperiod", "policydates", "daterange"}:
          effective, expiration = lossq_policy_period_range_dates(value)

          if effective:
            mapped["effective_date"] = effective
            mapped["policy_effective_date"] = effective

          if expiration:
            mapped["expiration_date"] = expiration
            mapped["policy_expiration_date"] = expiration

          if effective or expiration:
            print("LOSSQ_POLICY_PERIOD_RANGE_SCHEDULE_DATES:", {
              "policy_number": mapped.get("policy_number"),
              "effective_date": mapped.get("effective_date"),
              "expiration_date": mapped.get("expiration_date"),
            })

        elif header_key in {"annualpremium", "currentpremium", "premium"}:
          mapped["current_premium"] = value

        elif header_key in {"exposurebasis", "exposure"}:
          mapped["exposure_basis"] = value

        elif header_key in {"claims", "claimcount"}:
          mapped["claim_count"] = value

        elif header_key in {"totalincurred", "incurred"}:
          mapped["total_incurred"] = value

      if mapped.get("policy_number"):
        policies.append(mapped)

    break

  # LOSSQ_DATES_AFTER_POLICY_SCHEDULE_PARSE_V1
  # Policy schedule rows are parsed above. Now copy the first valid policy
  # effective/expiration dates back to the account profile if missing.
  try:
    if policies:
      for policy_row in policies:
        if not isinstance(policy_row, dict):
          continue

        effective = (
          policy_row.get("effective_date")
          or policy_row.get("policy_effective_date")
          or policy_row.get("effective")
        )
        expiration = (
          policy_row.get("expiration_date")
          or policy_row.get("policy_expiration_date")
          or policy_row.get("expiration")
          or policy_row.get("expiry_date")
        )

        fixed_effective = lossq_profile_date_or_blank(effective)
        fixed_expiration = lossq_profile_date_or_blank(expiration)

        if fixed_effective and not account_info.get("effective_date"):
          account_info["effective_date"] = fixed_effective
          account_info["policy_effective_date"] = fixed_effective

        if fixed_expiration and not account_info.get("expiration_date"):
          account_info["expiration_date"] = fixed_expiration
          account_info["policy_expiration_date"] = fixed_expiration

        if account_info.get("effective_date") and account_info.get("expiration_date"):
          print("LOSSQ_DATES_AFTER_POLICY_SCHEDULE_PARSE:", {
            "effective_date": account_info.get("effective_date"),
            "expiration_date": account_info.get("expiration_date"),
          })
          break
  except Exception as exc:
    print("LOSSQ_DATES_AFTER_POLICY_SCHEDULE_PARSE_ERROR:", str(exc)[:200])


  # LOSSQ_FINAL_CSV_ACCOUNT_DATES_AFTER_POLICY_PARSE_V2
  # Policy schedule rows have already been parsed into `policies`.
  # If account-level effective/expiration dates are blank, copy the first valid
  # policy schedule dates back onto the account profile before final merge/save.
  try:
    if policies:
      for policy_row in policies:
        if not isinstance(policy_row, dict):
          continue

        effective = (
          policy_row.get("effective_date")
          or policy_row.get("policy_effective_date")
          or policy_row.get("effective")
          or policy_row.get("Effective Date")
          or policy_row.get("Policy Effective Date")
        )

        expiration = (
          policy_row.get("expiration_date")
          or policy_row.get("policy_expiration_date")
          or policy_row.get("expiration")
          or policy_row.get("expiry_date")
          or policy_row.get("Expiration Date")
          or policy_row.get("Policy Expiration Date")
        )

        fixed_effective = lossq_profile_date_or_blank(effective)
        fixed_expiration = lossq_profile_date_or_blank(expiration)

        if fixed_effective and not account_info.get("effective_date"):
          account_info["effective_date"] = fixed_effective
          account_info["policy_effective_date"] = fixed_effective

        if fixed_expiration and not account_info.get("expiration_date"):
          account_info["expiration_date"] = fixed_expiration
          account_info["policy_expiration_date"] = fixed_expiration

        if account_info.get("effective_date") and account_info.get("expiration_date"):
          print("LOSSQ_FINAL_CSV_ACCOUNT_DATES_AFTER_POLICY_PARSE:", {
            "effective_date": account_info.get("effective_date"),
            "expiration_date": account_info.get("expiration_date"),
          })
          break
  except Exception as exc:
    print("LOSSQ_FINAL_CSV_ACCOUNT_DATES_AFTER_POLICY_PARSE_ERROR:", str(exc)[:200])


  for key, value in account_info.items():
    if value:
      parsed_profile[key] = value

  if policies:
    parsed_profile["policies"] = policies
    parsed_profile["policy_schedule"] = policies

    if not parsed_profile.get("policy_number"):
      parsed_profile["policy_number"] = policies[0].get("policy_number", "")

    if not parsed_profile.get("effective_date"):
      parsed_profile["effective_date"] = policies[0].get("effective_date", "")
      parsed_profile["policy_effective_date"] = policies[0].get("effective_date", "")

    if not parsed_profile.get("expiration_date"):
      parsed_profile["expiration_date"] = policies[0].get("expiration_date", "")
      parsed_profile["policy_expiration_date"] = policies[0].get("expiration_date", "")

    if not parsed_profile.get("carrier_name"):
      for policy in policies:
        if lossq_section_csv_valid_carrier(policy.get("carrier_name")):
          parsed_profile["carrier_name"] = policy.get("carrier_name")
          break

    if not parsed_profile.get("writing_carrier"):
      parsed_profile["writing_carrier"] = parsed_profile.get("carrier_name", "")

  return parsed_profile


# LOSSQ_PDF_PROFILE_CLEANUP_V1
def lossq_pdf_profile_bad_value(value):
  clean = lossq_section_csv_clean(value)
  low = clean.lower()

  if not clean:
    return True

  bad = {
    "effective",
    "effective date",
    "expiration",
    "expiration date",
    "expiry",
    "expiry date",
    "policy",
    "policy number",
    "carrier",
    "writing carrier",
    "insured",
    "named insured",
    "producer",
    "agency",
    "not set",
    "-",
  }

  if low in bad:
    return True

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", clean):
    return True

  if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", clean):
    return True

  return False


def lossq_pdf_profile_extract_date_after_label(raw_text, labels):
  text_value = str(raw_text or "")
  if not text_value:
    return ""

  for label in labels:
    pattern = rf"{label}\s*[:#-]?\s*(\d{{1,2}}[/-]\d{{1,2}}[/-]\d{{2,4}}|\d{{4}}[/-]\d{{1,2}}[/-]\d{{1,2}})"
    match = re.search(pattern, text_value, flags=re.IGNORECASE)
    if match:
      return lossq_section_csv_date(match.group(1))

  return ""


def lossq_pdf_profile_extract_policy_period(raw_text):
  text_value = str(raw_text or "")
  if not text_value:
    return "", ""

  compact = re.sub(r"[ \t]+", " ", text_value)
  compact = re.sub(r"\r\n|\r", "\n", compact)

  # LOSSQ_UNIVERSAL_PDF_POLICY_PERIOD_EXTRACTION_V2
  # Universal policy period patterns commonly found in carrier loss runs.
  patterns = [
    r"policy\s*period\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(?:-|to|through|thru|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
    r"policy\s*term\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(?:-|to|through|thru|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
    r"coverage\s*period\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(?:-|to|through|thru|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
    r"effective\s*(?:date)?\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}).{0,160}?expir(?:ation|y)?\s*(?:date)?\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
    r"\beff\.?\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}).{0,160}?\bexp\.?\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
    r"\bfrom\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(?:-|to|through|thru|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})",
  ]

  for pattern in patterns:
    match = re.search(pattern, compact, flags=re.IGNORECASE | re.DOTALL)
    if match:
      first = lossq_section_csv_date(match.group(1))
      second = lossq_section_csv_date(match.group(2))
      if first and second and first != second:
        return first, second

  # Fallback: find two dates near policy/term/effective/expiration wording.
  lines = [line.strip() for line in compact.split("\n") if line.strip()]
  date_pattern = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})"

  for idx, line in enumerate(lines):
    window = " ".join(lines[max(0, idx - 3): min(len(lines), idx + 4)])
    low = window.lower()

    if not any(term in low for term in ["policy", "period", "effective", "expiration", "expiry", "coverage", "term", "eff", "exp"]):
      continue

    dates = re.findall(date_pattern, window)
    cleaned_dates = []
    for d in dates:
      fixed = lossq_section_csv_date(d)
      if fixed and fixed not in cleaned_dates:
        cleaned_dates.append(fixed)

    if len(cleaned_dates) >= 2:
      return cleaned_dates[0], cleaned_dates[1]

  effective = lossq_pdf_profile_extract_date_after_label(
    compact,
    [
      r"effective\s*date",
      r"policy\s*effective\s*date",
      r"coverage\s*effective\s*date",
      r"\beff\.?",
      r"effective",
    ],
  )

  expiration = lossq_pdf_profile_extract_date_after_label(
    compact,
    [
      r"expiration\s*date",
      r"expiry\s*date",
      r"policy\s*expiration\s*date",
      r"coverage\s*expiration\s*date",
      r"\bexp\.?",
      r"expiration",
      r"expiry",
    ],
  )

  return effective, expiration


def lossq_pdf_profile_extract_evaluation_date(raw_text):
  text_value = str(raw_text or "")

  return lossq_pdf_profile_extract_date_after_label(
    text_value,
    [
      r"valuation\s*date",
      r"evaluation\s*date",
      r"loss\s*run\s*valuation\s*date",
      r"loss\s*run\s*date",
      r"as\s*of\s*date",
      r"report\s*date",
    ],
  )


def lossq_pdf_profile_repair(file_path, parsed_profile):
  parsed_profile = parsed_profile or {}

  raw_text = (
    parsed_profile.get("raw_text")
    or parsed_profile.get("raw_text_preview")
    or parsed_profile.get("text")
    or parsed_profile.get("ocr_text")
    or ""
  )

  # LOSSQ_PDF_RAW_TEXT_REPAIR_RUNS_ON_RAW_TEXT_V1
  # Run this repair whenever extracted raw text exists. Temp upload paths may not preserve.pdf extension.
  if not raw_text:
    return parsed_profile

  # Clean fake carrier values.
  for key in ["carrier_name", "writing_carrier", "carrier"]:
    if lossq_pdf_profile_bad_value(parsed_profile.get(key)):
      parsed_profile[key] = ""

  # Never use today's date as evaluation date unless the document actually supplied it.
  extracted_eval = lossq_pdf_profile_extract_evaluation_date(raw_text)
  if extracted_eval:
    parsed_profile["evaluation_date"] = extracted_eval
    parsed_profile["valuation_date"] = extracted_eval
    parsed_profile["loss_run_valuation_date"] = extracted_eval
  else:
    # If a parser supplied today's date as a fallback, remove it so frontend can warn accurately.
    parsed_profile["evaluation_date"] = parsed_profile.get("evaluation_date") or ""
    parsed_profile["valuation_date"] = parsed_profile.get("valuation_date") or ""
    parsed_profile["loss_run_valuation_date"] = parsed_profile.get("loss_run_valuation_date") or ""

  effective, expiration = lossq_pdf_profile_extract_policy_period(raw_text)

  if effective and not parsed_profile.get("effective_date"):
    parsed_profile["effective_date"] = effective
    parsed_profile["policy_effective_date"] = effective

  if expiration and not parsed_profile.get("expiration_date"):
    parsed_profile["expiration_date"] = expiration
    parsed_profile["policy_expiration_date"] = expiration

  # If profile dates exist but policy schedule rows are missing dates/carrier, fill the schedule rows.
  policies = parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []
  if isinstance(policies, list):
    cleaned_policies = []
    for policy in policies:
      if not isinstance(policy, dict):
        continue

      next_policy = dict(policy)

      if lossq_pdf_profile_bad_value(next_policy.get("carrier")):
        next_policy["carrier"] = parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier") or ""
      if lossq_pdf_profile_bad_value(next_policy.get("carrier_name")):
        next_policy["carrier_name"] = parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier") or ""

      if not next_policy.get("effective_date") and parsed_profile.get("effective_date"):
        next_policy["effective_date"] = parsed_profile.get("effective_date")
        next_policy["policy_effective_date"] = parsed_profile.get("effective_date")

      if not next_policy.get("expiration_date") and parsed_profile.get("expiration_date"):
        next_policy["expiration_date"] = parsed_profile.get("expiration_date")
        next_policy["policy_expiration_date"] = parsed_profile.get("expiration_date")

      cleaned_policies.append(next_policy)

    parsed_profile["policies"] = cleaned_policies
    parsed_profile["policy_schedule"] = cleaned_policies

  return parsed_profile


# LOSSQ_LINE_OF_BUSINESS_FROM_POLICY_PREFIX_V1
def lossq_line_of_business_from_policy_prefix(value):
  """
  Universal line-of-business correction from policy/claim prefixes.
  Prevents CARGO, BOP, and UMB rows from being displayed as generic Commercial Auto or GL.
  """
  try:
    import re

    token = str(value or "").upper().strip()
    if not token:
      return ""

    # Match full tokens only, not random letters inside company names.
    parts = set(re.split(r"[^A-Z0-9]+", token))

    if "CARGO" in parts or "MTC" in parts or "TRUCKCARGO" in parts:
      return "Motor Truck Cargo"

    if "BOP" in parts or "BP" in parts:
      return "Businessowners Policy"

    if "UMB" in parts or "UMBRELLA" in parts or "EXCESS" in parts:
      return "Umbrella / Excess"

    if "WC" in parts or "WORKERS" in parts or "COMP" in parts:
      return "Workers Compensation"

    if "GL" in parts or "GENERAL" in parts:
      return "General Liability"

    if "AUTO" in parts or "CA" in parts or "AL" in parts:
      return "Commercial Auto"

    if "CY" in parts or "CYBER" in parts:
      return "Cyber Liability"

    if "LIAB" in parts or "LL" in parts or "LIQUOR" in parts:
      return "Liquor Liability"

    if "PL" in parts or "PROF" in parts or "PROFESSIONAL" in parts:
      return "Professional Liability"

    if "CP" in parts or "PROPERTY" in parts:
      return "Commercial Property"

    if "EPLI" in parts:
      return "Employment Practices Liability"

    if "DO" in parts or "DNO" in parts:
      return "Directors & Officers"

    return ""
  except Exception as exc:
    print("LOSSQ_LINE_OF_BUSINESS_FROM_POLICY_PREFIX_ERROR:", str(exc)[:200])
    return ""


# LOSSQ_APPLY_LINE_OF_BUSINESS_FROM_POLICY_PREFIX_V1
def lossq_apply_line_of_business_from_policy_prefix(parsed_claims, parsed_profile=None):
  """
  Correct parsed claim and policy schedule line names using policy/claim prefixes.
  """
  try:
    parsed_claims = parsed_claims or []
    parsed_profile = parsed_profile or {}

    for claim in parsed_claims:
      if not isinstance(claim, dict):
        continue

      policy_number = (
        claim.get("policy_number")
        or claim.get("Policy Number")
        or claim.get("policy_no")
        or ""
      )

      claim_number = (
        claim.get("claim_number")
        or claim.get("Claim Number")
        or claim.get("claim_no")
        or ""
      )

      detected_line = (
        lossq_line_of_business_from_policy_prefix(policy_number)
        or lossq_line_of_business_from_policy_prefix(claim_number)
      )

      if detected_line:
        claim["line_of_business"] = detected_line
        claim["claim_type"] = detected_line
        claim["coverage"] = detected_line
        claim["policy_type"] = detected_line

    policies = parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []
    if isinstance(policies, list):
      for policy in policies:
        if not isinstance(policy, dict):
          continue

        policy_number = (
          policy.get("policy_number")
          or policy.get("Policy Number")
          or policy.get("policy_no")
          or ""
        )

        detected_line = lossq_line_of_business_from_policy_prefix(policy_number)

        if detected_line:
          policy["line_of_business"] = detected_line
          policy["policy_type"] = detected_line
          policy["coverage"] = detected_line
          policy["line"] = detected_line

      parsed_profile["policies"] = policies
      parsed_profile["policy_schedule"] = policies

    return parsed_claims, parsed_profile
  except Exception as exc:
    print("LOSSQ_APPLY_LINE_OF_BUSINESS_FROM_POLICY_PREFIX_ERROR:", str(exc)[:200])
    return parsed_claims, parsed_profile


# LOSSQ_CLEAN_PROFILE_POLICY_SCHEDULE_ROWS_V1
def lossq_clean_profile_policy_schedule_rows(parsed_profile, parsed_claims=None):
  """
  Remove fake policy schedule rows created from claim-table text.
  Keeps real policy numbers like FPS-GL-2025-8801, but removes claim-looking
  or partial rows like GL-250012, GL-2025, WC-2025-8802, CARGO-250052.
  """
  try:
    import re

    parsed_profile = parsed_profile or {}
    parsed_claims = parsed_claims or []

    policies = parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []
    if not isinstance(policies, list):
      return parsed_profile

    claim_numbers = set()
    claim_policy_numbers = set()

    for claim in parsed_claims:
      if not isinstance(claim, dict):
        continue

      claim_number = str(
        claim.get("claim_number")
        or claim.get("Claim Number")
        or claim.get("claim_no")
        or ""
      ).strip().upper()

      policy_number = str(
        claim.get("policy_number")
        or claim.get("Policy Number")
        or claim.get("policy_no")
        or ""
      ).strip().upper()

      if claim_number:
        claim_numbers.add(claim_number)
      if policy_number:
        claim_policy_numbers.add(policy_number)

    def clean(value):
      return str(value or "").strip().upper()

    def looks_like_claim_number(value):
      value = clean(value)
      if not value:
        return True

      if value in claim_numbers:
        return True

      # Examples: GL-250012, WC-250026, BOP-250039, CARGO-250052, UMB-250067
      if re.match(r"^(GL|WC|BOP|AUTO|AU|CARGO|MTC|UMB|CY|CP|PROP|EPLI|DO|DNO)-\d{5,7}$", value):
        return True

      # Examples: GL-2025, WC-2025, BOP-2025, UMB-2025
      if re.match(r"^(GL|WC|BOP|AUTO|AU|CARGO|MTC|UMB|CY|CP|PROP|EPLI|DO|DNO)-20\d{2}$", value):
        return True

      # Examples: GL-2025-8801 or WC-2025-8802 can be claim/table fragments when
      # the same upload already has stronger real policies like FPS-GL-2025-8801.
      has_prefixed_real_policy = any(
        real_policy.endswith("-" + value) or real_policy.endswith(value)
        for real_policy in claim_policy_numbers
        if real_policy and real_policy != value
      )
      if has_prefixed_real_policy:
        return True

      return False

    cleaned_policies = []
    removed_policies = []

    for policy in policies:
      if not isinstance(policy, dict):
        continue

      policy_number = clean(
        policy.get("policy_number")
        or policy.get("Policy Number")
        or policy.get("policy")
        or policy.get("policy_no")
      )

      if looks_like_claim_number(policy_number):
        removed_policies.append(policy_number)
        continue

      cleaned_policies.append(policy)

    parsed_profile["policies"] = cleaned_policies
    parsed_profile["policy_schedule"] = cleaned_policies

    if removed_policies:
      print("LOSSQ_CLEAN_PROFILE_POLICY_SCHEDULE_REMOVED:", removed_policies[:25])

    return parsed_profile
  except Exception as exc:
    print("LOSSQ_CLEAN_PROFILE_POLICY_SCHEDULE_ERROR:", str(exc)[:200])
    return parsed_profile


# LOSSQ_FINAL_PROFILE_DATES_FROM_POLICIES_V1
def lossq_final_profile_dates_from_policies(parsed_profile):
  """
  Final universal profile repair:
  If account-level effective/expiration dates are missing, use the first valid
  effective/expiration dates from parsed policy schedule rows.
  """
  parsed_profile = parsed_profile or {}

  try:
    policy_rows = (
      parsed_profile.get("policies")
      or parsed_profile.get("policy_schedule")
      or parsed_profile.get("policySchedule")
      or []
    )

    if not isinstance(policy_rows, list):
      return parsed_profile

    for policy in policy_rows:
      if not isinstance(policy, dict):
        continue

      effective = (
        policy.get("effective_date")
        or policy.get("policy_effective_date")
        or policy.get("effective")
        or policy.get("policyEffectiveDate")
      )

      expiration = (
        policy.get("expiration_date")
        or policy.get("policy_expiration_date")
        or policy.get("expiration")
        or policy.get("expiry_date")
        or policy.get("policyExpirationDate")
      )

      fixed_effective = lossq_profile_date_or_blank(effective)
      fixed_expiration = lossq_profile_date_or_blank(expiration)

      if fixed_effective and not parsed_profile.get("effective_date"):
        parsed_profile["effective_date"] = fixed_effective
        parsed_profile["policy_effective_date"] = fixed_effective

      if fixed_expiration and not parsed_profile.get("expiration_date"):
        parsed_profile["expiration_date"] = fixed_expiration
        parsed_profile["policy_expiration_date"] = fixed_expiration

      if parsed_profile.get("effective_date") and parsed_profile.get("expiration_date"):
        break

    return parsed_profile
  except Exception as exc:
    print("LOSSQ_FINAL_PROFILE_DATES_FROM_POLICIES_ERROR:", str(exc)[:200])
    return parsed_profile


# LOSSQ_GLOBAL_PROFILE_CLEANUP_V1
def lossq_global_profile_bad_text(value):
  clean = lossq_section_csv_clean(value)
  low = clean.lower()

  if not clean:
    return True

  bad_values = {
    "effective",
    "effective date",
    "expiration",
    "expiration date",
    "expiry",
    "expiry date",
    "policy",
    "policy number",
    "carrier",
    "writing carrier",
    "insured",
    "named insured",
    "producer",
    "agency",
    "not set",
    "none",
    "null",
    "-",
  }

  if low in bad_values:
    return True

  if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", clean):
    return True

  if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", clean):
    return True

  return False


def lossq_profile_has_policy_dates(profile):
  profile = profile or {}

  if profile.get("effective_date") and profile.get("expiration_date"):
    return True

  policies = profile.get("policies") or profile.get("policy_schedule") or []
  if isinstance(policies, list):
    for policy in policies:
      if not isinstance(policy, dict):
        continue
      if (
        policy.get("effective_date")
        or policy.get("policy_effective_date")
        or policy.get("Effective Date")
      ) and (
        policy.get("expiration_date")
        or policy.get("policy_expiration_date")
        or policy.get("Expiration Date")
      ):
        return True

  return False


def lossq_global_profile_cleanup(parsed_profile):
  parsed_profile = parsed_profile or {}

  # Clean impossible carrier values.
  for key in ["carrier_name", "writing_carrier", "carrier"]:
    if lossq_global_profile_bad_text(parsed_profile.get(key)):
      parsed_profile[key] = ""

  # Backfill carrier from writing carrier if one is real.
  if not parsed_profile.get("carrier_name") and not lossq_global_profile_bad_text(parsed_profile.get("writing_carrier")):
    parsed_profile["carrier_name"] = parsed_profile.get("writing_carrier")

  if not parsed_profile.get("writing_carrier") and not lossq_global_profile_bad_text(parsed_profile.get("carrier_name")):
    parsed_profile["writing_carrier"] = parsed_profile.get("carrier_name")

  policies = parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []
  cleaned_policies = []

  if isinstance(policies, list):
    for policy in policies:
      if not isinstance(policy, dict):
        continue

      item = dict(policy)

      for key in ["carrier", "carrier_name", "writing_carrier"]:
        if lossq_global_profile_bad_text(item.get(key)):
          item[key] = ""

      if not item.get("carrier") and parsed_profile.get("carrier_name"):
        item["carrier"] = parsed_profile.get("carrier_name")

      if not item.get("carrier_name") and parsed_profile.get("carrier_name"):
        item["carrier_name"] = parsed_profile.get("carrier_name")

      if not item.get("writing_carrier") and parsed_profile.get("writing_carrier"):
        item["writing_carrier"] = parsed_profile.get("writing_carrier")

      cleaned_policies.append(item)

    parsed_profile["policies"] = cleaned_policies
    parsed_profile["policy_schedule"] = cleaned_policies

  # LOSSQ_GLOBAL_PROFILE_DATE_VALUE_CLEANUP_V1
  parsed_profile["effective_date"] = lossq_profile_date_or_blank(parsed_profile.get("effective_date"))
  parsed_profile["policy_effective_date"] = lossq_profile_date_or_blank(parsed_profile.get("policy_effective_date") or parsed_profile.get("effective_date"))
  parsed_profile["expiration_date"] = lossq_profile_date_or_blank(parsed_profile.get("expiration_date"))
  parsed_profile["policy_expiration_date"] = lossq_profile_date_or_blank(parsed_profile.get("policy_expiration_date") or parsed_profile.get("expiration_date"))

  # If the file has no policy dates, do not let a fallback "today" evaluation date make it appear current.
  if not lossq_profile_has_policy_dates(parsed_profile):
    parsed_profile["evaluation_date"] = ""
    parsed_profile["valuation_date"] = ""
    parsed_profile["loss_run_valuation_date"] = ""

  return parsed_profile


# LOSSQ_UNIVERSAL_PROFILE_IDENTITY_POLICY_CLEANUP_V1
def lossq_universal_profile_identity_policy_cleanup(profile):
  """
  Universal cleanup for section-style, carrier-style, CSV, XLSX, and PDF profiles.
  Does not hardcode any customer, carrier, file, or demo case.
  """
  if not isinstance(profile, dict):
    return profile

  profile = dict(profile)

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").strip())

  def norm_key(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

  def first_value(*keys):
    normalized = {norm_key(k): v for k, v in profile.items()}
    for key in keys:
      value = profile.get(key)
      if clean(value):
        return clean(value)

      value = normalized.get(norm_key(key))
      if clean(value):
        return clean(value)

    return ""

  def looks_like_policy(value):
    value = clean(value).upper()
    if not value:
      return False
    # LOSSQ_TRUE_ACCOUNT_NUMBER_FROM_UPLOAD_CSV_V1
    if lossq_true_account_number_value(value):
      return False
    return lossq_looks_like_policy_but_not_account(value)

  def split_policy_values(value):
    value = clean(value)
    if not value:
      return []

    pieces = re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", value, flags=re.IGNORECASE)
    results = []

    for piece in pieces:
      piece = clean(piece)
      if not piece:
        continue

      matches = re.findall(r"\b[A-Z]{1,8}[- ]?\d{2,6}[- ]?[A-Z0-9]{2,12}\b", piece.upper())

      if matches:
        for match in matches:
          match = clean(match).replace(" ", "-")
          if match and match not in results:
            results.append(match)
      elif looks_like_policy(piece):
        piece = piece.upper().replace(" ", "-")
        if piece not in results:
          results.append(piece)

    return results

  # 1) Business name / insured name universal mapping.
  business_name = first_value(
    "business_name",
    "insured_name",
    "insured",
    "named_insured",
    "named insured",
    "applicant",
    "account_name",
    "account name",
    "company_name",
    "company",
  )

  if business_name:
    profile["business_name"] = business_name
    profile["insured_name"] = profile.get("insured_name") or business_name
    profile["named_insured"] = profile.get("named_insured") or business_name

  # 2) Policy values may appear in policy_number, account_number, main_policy, or raw headings.
  raw_policy_value = first_value(
    "policy_number",
    "policy number",
    "policy_numbers",
    "policy numbers",
    "main_policy",
    "main policy",
    "account_number",
    "account number",
  )

  policies = split_policy_values(raw_policy_value)

  existing_policies = profile.get("policies")
  if isinstance(existing_policies, list):
    for item in existing_policies:
      if isinstance(item, dict):
        policies.extend(split_policy_values(item.get("policy_number") or item.get("policy") or item.get("number")))
      else:
        policies.extend(split_policy_values(item))

  # De-dupe while preserving order.
  deduped_policies = []
  for policy in policies:
    policy = clean(policy).upper()
    if policy and policy not in deduped_policies:
      deduped_policies.append(policy)

  if deduped_policies:
    profile["policy_number"] = deduped_policies[0]
    profile["main_policy"] = deduped_policies[0]
    profile["policy_numbers"] = deduped_policies

    existing_policy_rows = profile.get("policies") if isinstance(profile.get("policies"), list) else []
    rebuilt = []

    seen = set()
    for policy in deduped_policies:
      rebuilt.append({"policy_number": policy})
      seen.add(policy)

    for item in existing_policy_rows:
      if isinstance(item, dict):
        policy = clean(item.get("policy_number") or item.get("policy") or item.get("number")).upper()
        if policy and policy not in seen:
          rebuilt.append(item)
          seen.add(policy)

    profile["policies"] = rebuilt

  # 3) Account number should not be a policy-number bundle.
  account_number = first_value("account_number", "account number", "customer_number", "customer number")
  true_account_number = first_value(
    "account_id",
    "account id",
    "customer_id",
    "customer id",
    "insured_id",
    "insured id",
    "client_id",
    "client id",
  )

  if account_number and looks_like_policy(account_number):
    if true_account_number and not looks_like_policy(true_account_number):
      profile["account_number"] = true_account_number
    else:
      profile["account_number"] = ""

  return profile


# LOSSQ_UNIVERSAL_SECTION_CSV_CLAIMS_PROFILE_REPAIR_V1
def lossq_universal_section_csv_claims_profile_repair(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal section-based CSV parser for loss runs that contain profile sections,
  claims sections, and summary sections in one file.

  Supports generic layouts such as:
  - EXPOSURE / POLICY INFORMATION
  - POLICY INFORMATION
  - ACCOUNT INFORMATION
  - CLAIMS DETAIL
  - CLAIM DETAIL
  - CLAIMS
  - LOSS SUMMARY

  No customer, carrier, or demo-file hardcoding.
  """
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  # LOSSQ_PROFILE_AGENCY_FIELD_NORMALIZE_V1
  # Normalize any file-parsed producer/agency value into agency_name so the
  # account profile save and dashboard display use the uploaded file as source.
  def _lossq_profile_agency_clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  profile_agency_value = _lossq_profile_agency_clean(
    parsed_profile.get("agency_name")
    or parsed_profile.get("producing_agency")
    or parsed_profile.get("producingAgency")
    or parsed_profile.get("producer")
    or parsed_profile.get("producer_name")
    or parsed_profile.get("agency")
    or parsed_profile.get("agencyName")
    or parsed_profile.get("broker")
    or parsed_profile.get("brokerage")
    or parsed_profile.get("agent")
    or parsed_profile.get("agent_name")
    or parsed_profile.get("prepared_by")
    or parsed_profile.get("contact_name")
  )

  if profile_agency_value:
    parsed_profile["agency_name"] = profile_agency_value
    parsed_profile["producing_agency"] = profile_agency_value


  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return parsed_claims, parsed_profile

  if not rows:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm(value):
    return re.sub(r"[^a-z0-9]+", "_", clean(value).lower()).strip("_")

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", "-", ".", "-."}:
      return 0.0
    try:
      amount = float(raw)
      return -amount if neg else amount
    except Exception:
      return 0.0

  def looks_like_policy(value):
    value = clean(value).upper()
    if not value:
      return False
    return lossq_looks_like_policy_but_not_account(value)

  def split_policies(value):
    value = clean(value)
    if not value:
      return []

    pieces = re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", value, flags=re.IGNORECASE)
    found = []

    for piece in pieces:
      piece = clean(piece).upper()
      matches = re.findall(r"\b[A-Z]{1,8}[- ]?\d{2,6}[- ]?[A-Z0-9]{2,12}\b", piece)
      for match in matches:
        match = clean(match).upper().replace(" ", "-")
        if match and match not in found:
          found.append(match)

    return found

  def infer_line_from_policy(policy_number, fallback=""):
    value = clean(policy_number).upper()
    fallback = clean(fallback)

    if value.startswith("WC"):
      return "Workers Compensation"
    if value.startswith(("GL", "CGL")):
      return "General Liability"
    if value.startswith(("PL", "PROD")):
      return "Products Liability"
    if value.startswith(("AL", "AUTO", "CA")):
      return "Commercial Auto"
    if value.startswith(("BOP", "CP", "PROP")):
      return "Property / BOP"
    if value.startswith(("UMB", "UM")):
      return "Umbrella"
    if value.startswith(("CY", "CYBER")):
      return "Cyber Liability"

    return fallback or "Unknown"

  def header_index(headers, candidates):
    normalized = [norm(header) for header in headers]
    wanted = {norm(candidate) for candidate in candidates}

    for index, header in enumerate(normalized):
      if header in wanted:
        return index

    for index, header in enumerate(normalized):
      for candidate in wanted:
        if candidate and candidate in header:
          return index

    return None

  def value_at(row, index):
    if index is None:
      return ""
    if index < 0 or index >= len(row):
      return ""
    return clean(row[index])

  profile_labels = {}

  claims_header_index = None

  for index, row in enumerate(rows):
    first = norm(row[0] if row else "")

    if first in {
      "claims_detail",
      "claim_detail",
      "claim_details",
      "claims",
      "loss_detail",
      "loss_details",
      "claim_listing",
      "claim_list",
    }:
      claims_header_index = index + 1
      break

    if len(row) >= 2:
      label = norm(row[0])
      value = clean(row[1])
      if label and value:
        profile_labels[label] = value

  if profile_labels:
    label_map = {
      "insured_name": ["insured_name", "named_insured", "insured", "applicant", "account_name", "company_name", "company"],
      "policy_number": ["policy_number", "policy_numbers", "policy_no", "policy"],
      "policy_period": ["policy_period", "policy_term", "effective_expiration", "coverage_period"],
      "line_of_business": ["line_of_business", "lines_of_business", "coverage", "coverage_line"],
      "carrier": ["carrier", "insurance_carrier", "writing_carrier", "company"],
      "annual_revenue": ["annual_revenue", "revenue", "sales", "gross_sales"],
      "payroll": ["total_payroll_annual", "payroll", "annual_payroll"],
      "employee_count": ["full_time_employees", "employees", "employee_count"],
      "operations": ["operations", "business_operations", "description_of_operations"],
      "account_number": ["account_number", "account_no", "customer_number", "client_number"],
      "evaluation_date": ["evaluation_date", "valuation_date", "as_of_date", "loss_run_date"],
    }

    def first_label(keys):
      for key in keys:
        if key in profile_labels and clean(profile_labels[key]):
          return clean(profile_labels[key])
      return ""

    insured_name = first_label(label_map["insured_name"])
    if insured_name:
      parsed_profile["business_name"] = insured_name
      parsed_profile["insured_name"] = insured_name
      parsed_profile["named_insured"] = insured_name

    carrier = first_label(label_map["carrier"])
    if carrier:
      parsed_profile["carrier"] = carrier
      parsed_profile["carrier_name"] = carrier
      parsed_profile["writing_carrier"] = carrier

    policy_value = first_label(label_map["policy_number"])
    policy_numbers = split_policies(policy_value)

    line_value = first_label(label_map["line_of_business"])
    line_parts = [
      clean(part)
      for part in re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", line_value, flags=re.IGNORECASE)
      if clean(part)
    ]

    if policy_numbers:
      parsed_profile["policy_number"] = policy_numbers[0]
      parsed_profile["main_policy"] = policy_numbers[0]
      parsed_profile["policy_numbers"] = policy_numbers
      parsed_profile["policies"] = [
        {
          "policy_number": policy_number,
          "line_of_business": infer_line_from_policy(
            policy_number,
            line_parts[min(i, len(line_parts) - 1)] if line_parts else "",
          ),
          "carrier": carrier or parsed_profile.get("carrier") or "",
        }
        for i, policy_number in enumerate(policy_numbers)
      ]

    policy_period = first_label(label_map["policy_period"])
    dates = re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", policy_period)

    if len(dates) >= 1:
      parsed_profile["effective_date"] = parsed_profile.get("effective_date") or dates[0]
    if len(dates) >= 2:
      parsed_profile["expiration_date"] = parsed_profile.get("expiration_date") or dates[1]

    account_number = first_label(label_map["account_number"])
    if account_number and not looks_like_policy(account_number):
      parsed_profile["account_number"] = account_number
    elif looks_like_policy(parsed_profile.get("account_number", "")):
      parsed_profile["account_number"] = ""

    for target, keys in [
      ("annual_revenue", label_map["annual_revenue"]),
      ("payroll", label_map["payroll"]),
      ("employee_count", label_map["employee_count"]),
      ("operations", label_map["operations"]),
      ("evaluation_date", label_map["evaluation_date"]),
    ]:
      value = first_label(keys)
      if value:
        parsed_profile[target] = value

  repaired_claims = []

  if claims_header_index is not None and claims_header_index < len(rows):
    headers = [clean(value) for value in rows[claims_header_index]]

    claim_number_i = header_index(headers, ["claim #", "claim no", "claim number", "claim_number", "claim"])
    dol_i = header_index(headers, ["date of loss", "loss date", "date_of_loss"])
    reported_i = header_index(headers, ["date reported", "reported date", "date_reported"])
    claimant_i = header_index(headers, ["claimant", "claimant name", "injured worker", "party"])
    line_i = header_index(headers, ["line", "line of business", "coverage", "lob"])
    desc_i = header_index(headers, ["description", "loss description", "cause", "claim description"])
    status_i = header_index(headers, ["status", "claim status", "open closed"])
    incurred_i = header_index(headers, ["total incurred", "incurred", "total"])
    paid_i = header_index(headers, ["paid", "total paid"])
    reserve_i = header_index(headers, ["reserve", "total reserve", "outstanding reserve"])
    subro_i = header_index(headers, ["subrogation", "subro"])

    for row in rows[claims_header_index + 1:]:
      first = norm(row[0] if row else "")

      if first in {
        "loss_summary",
        "summary",
        "totals",
        "exposure_summary",
        "premium_summary",
        "policy_year",
      }:
        break

      if not any(clean(cell) for cell in row):
        continue

      claim_number = value_at(row, claim_number_i)
      if not claim_number:
        continue

      line = value_at(row, line_i)
      status = value_at(row, status_i)
      paid = money(value_at(row, paid_i))
      reserve = money(value_at(row, reserve_i))
      incurred = money(value_at(row, incurred_i))

      if incurred <= 0 and (paid or reserve):
        incurred = paid + reserve

      claim = {
        "claim_number": claim_number,
        "date_of_loss": value_at(row, dol_i),
        "date_reported": value_at(row, reported_i),
        "claimant": value_at(row, claimant_i),
        "line_of_business": line,
        "description": value_at(row, desc_i),
        "claim_status": status,
        "status": status,
        "total_incurred": incurred,
        "paid": paid,
        "reserve": reserve,
        "subrogation": value_at(row, subro_i),
      }

      policy_numbers = parsed_profile.get("policy_numbers") if isinstance(parsed_profile.get("policy_numbers"), list) else []
      line_upper = line.upper()

      if policy_numbers:
        if "WC" in line_upper:
          claim["policy_number"] = next((p for p in policy_numbers if str(p).upper().startswith("WC")), policy_numbers[0])
        elif "GL" in line_upper or "LIAB" in line_upper:
          claim["policy_number"] = next((p for p in policy_numbers if str(p).upper().startswith(("GL", "CGL", "PL"))), policy_numbers[0])
        else:
          claim["policy_number"] = policy_numbers[0]

      repaired_claims.append(claim)

  if repaired_claims and len(repaired_claims) >= len(parsed_claims):
    parsed_claims = repaired_claims

  if parsed_profile.get("policy_numbers") and not parsed_profile.get("account_number"):
    parsed_profile["account_number"] = ""

  return parsed_claims, parsed_profile


# LOSSQ_UNIVERSAL_SECTION_CSV_CLAIMS_PROFILE_REPAIR_V2
def lossq_universal_section_csv_claims_profile_repair_v2(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal section-based CSV loss run repair.

  Handles CSV files that include sections such as:
  - EXPOSURE / POLICY INFORMATION
  - ACCOUNT INFORMATION
  - POLICY INFORMATION
  - CLAIMS DETAIL
  - LOSS SUMMARY

  This does not hardcode carrier, insured, file name, or demo data.
  """
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return parsed_claims, parsed_profile

  if not rows:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm(value):
    return re.sub(r"[^a-z0-9]+", "_", clean(value).lower()).strip("_")

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0

    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]", "", raw)

    if raw in {"", "-", ".", "-."}:
      return 0.0

    try:
      amount = float(raw)
      return -amount if neg else amount
    except Exception:
      return 0.0

  def header_index(headers, candidates):
    normalized_headers = [norm(header) for header in headers]
    normalized_candidates = [norm(candidate) for candidate in candidates]

    for i, header in enumerate(normalized_headers):
      if header in normalized_candidates:
        return i

    for i, header in enumerate(normalized_headers):
      for candidate in normalized_candidates:
        if candidate and (candidate in header or header in candidate):
          return i

    return None

  def value_at(row, index):
    if index is None:
      return ""
    if index < 0 or index >= len(row):
      return ""
    return clean(row[index])

  def looks_like_policy(value):
    value = clean(value).upper()
    if not value:
      return False
    return lossq_looks_like_policy_but_not_account(value)

  def split_policy_numbers(value):
    raw = clean(value)
    if not raw:
      return []

    found = []
    pieces = re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", raw, flags=re.IGNORECASE)

    for piece in pieces:
      piece = clean(piece).upper()
      matches = re.findall(r"\b[A-Z]{1,8}[- ]?\d{2,6}[- ]?[A-Z0-9]{2,12}\b", piece)

      for match in matches:
        match = clean(match).upper().replace(" ", "-")
        if match and match not in found:
          found.append(match)

    return found

  def infer_policy_line(policy_number, fallback_line=""):
    policy = clean(policy_number).upper()
    fallback = clean(fallback_line)

    if policy.startswith("WC"):
      return "Workers Compensation"
    if policy.startswith(("GL", "CGL")):
      return "General Liability"
    if policy.startswith(("PL", "PROD")):
      return "Products Liability"
    if policy.startswith(("AL", "AUTO", "CA")):
      return "Commercial Auto"
    if policy.startswith(("BOP", "CP", "PROP")):
      return "Property / BOP"
    if policy.startswith(("UMB", "UM")):
      return "Umbrella"
    if policy.startswith(("CY", "CYBER")):
      return "Cyber Liability"

    return fallback or "Unknown"

  def choose_policy_for_claim(line_of_business, policy_numbers):
    if not policy_numbers:
      return ""

    line = clean(line_of_business).upper()

    if "WC" in line or "WORK" in line:
      return next((p for p in policy_numbers if str(p).upper().startswith("WC")), policy_numbers[0])

    if "PROD" in line or "LIAB" in line or "GL" in line or "GENERAL" in line:
      return next(
        (
          p for p in policy_numbers
          if str(p).upper().startswith(("GL", "CGL", "PL", "PROD"))
        ),
        policy_numbers[0],
      )

    return policy_numbers[0]

  section_headers = {
    "claims_detail",
    "claim_detail",
    "claim_details",
    "claims",
    "loss_detail",
    "loss_details",
    "claim_listing",
    "claim_list",
  }

  stop_sections = {
    "loss_summary",
    "summary",
    "totals",
    "total",
    "exposure_summary",
    "premium_summary",
    "policy_year",
    "underwriting_notes",
    "notes",
  }

  profile_labels = {}
  claims_header_row_index = None

  for i, row in enumerate(rows):
    first = norm(row[0] if row else "")

    if first in section_headers:
      claims_header_row_index = i + 1
      break

    if len(row) >= 2:
      label = norm(row[0])
      value = clean(row[1])
      if label and value:
        profile_labels[label] = value

  def profile_first(*labels):
    for label in labels:
      value = profile_labels.get(norm(label))
      if clean(value):
        return clean(value)
    return ""

  # Universal profile extraction from label/value section.
  insured_name = profile_first(
    "insured name",
    "named insured",
    "insured",
    "applicant",
    "account name",
    "company name",
    "company",
    "business name",
  )

  if insured_name:
    parsed_profile["business_name"] = insured_name
    parsed_profile["insured_name"] = insured_name
    parsed_profile["named_insured"] = insured_name

  carrier = profile_first("carrier", "writing carrier", "insurance carrier", "company")
  if carrier:
    parsed_profile["carrier"] = carrier
    parsed_profile["carrier_name"] = carrier
    parsed_profile["writing_carrier"] = carrier

  policy_value = profile_first("policy number", "policy numbers", "policy no", "policy")
  policy_numbers = split_policy_numbers(policy_value)

  line_value = profile_first("line of business", "lines of business", "coverage", "lob")
  line_parts = [
    clean(part)
    for part in re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", line_value, flags=re.IGNORECASE)
    if clean(part)
  ]

  if policy_numbers:
    parsed_profile["policy_number"] = policy_numbers[0]
    parsed_profile["main_policy"] = policy_numbers[0]
    parsed_profile["policy_numbers"] = policy_numbers
    parsed_profile["policies"] = [
      {
        "policy_number": policy_number,
        "line_of_business": infer_policy_line(
          policy_number,
          line_parts[min(index, len(line_parts) - 1)] if line_parts else "",
        ),
        "policy_type": infer_policy_line(
          policy_number,
          line_parts[min(index, len(line_parts) - 1)] if line_parts else "",
        ),
        "carrier": carrier or parsed_profile.get("carrier") or "",
        "writing_carrier": carrier or parsed_profile.get("writing_carrier") or "",
      }
      for index, policy_number in enumerate(policy_numbers)
    ]

  policy_period = profile_first("policy period", "policy term", "coverage period", "effective expiration")
  period_dates = re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", policy_period)

  if len(period_dates) >= 1:
    parsed_profile["effective_date"] = parsed_profile.get("effective_date") or period_dates[0]
  if len(period_dates) >= 2:
    parsed_profile["expiration_date"] = parsed_profile.get("expiration_date") or period_dates[1]

  account_number = profile_first("account number", "account no", "customer number", "client number")
  if account_number and not looks_like_policy(account_number):
    parsed_profile["account_number"] = account_number
  elif parsed_profile.get("account_number") and looks_like_policy(parsed_profile.get("account_number")):
    parsed_profile["account_number"] = ""

  exposure_map = {
    "annual_revenue": ["annual revenue", "revenue", "sales", "gross sales"],
    "payroll": ["total payroll annual", "payroll", "annual payroll"],
    "employee_count": ["full time employees", "employees", "employee count"],
    "operations": ["operations", "business operations", "description of operations"],
    "facilities": ["facilities", "locations", "premises"],
    "sic_code": ["primary sic code", "sic code"],
    "experience_modifier": ["experience modifier", "emod", "e mod", "experience modifier e mod"],
    "safety_program": ["safety program", "risk control", "safety"],
    "evaluation_date": ["evaluation date", "valuation date", "as of date", "loss run date"],
  }

  for target, labels in exposure_map.items():
    value = profile_first(*labels)
    if value:
      parsed_profile[target] = value

  # Universal claims extraction from claims section.
  repaired_claims = []

  if claims_header_row_index is not None and claims_header_row_index < len(rows):
    headers = [clean(value) for value in rows[claims_header_row_index]]

    claim_number_i = header_index(headers, ["claim #", "claim no", "claim number", "claim_number", "claim"])
    dol_i = header_index(headers, ["date of loss", "loss date", "date_of_loss", "dol"])
    reported_i = header_index(headers, ["date reported", "reported date", "date_reported"])
    claimant_i = header_index(headers, ["claimant", "claimant name", "injured worker", "party"])
    line_i = header_index(headers, ["line", "line of business", "coverage", "lob"])
    description_i = header_index(headers, ["description", "loss description", "cause", "claim description"])
    status_i = header_index(headers, ["status", "claim status", "open closed", "open/closed"])
    incurred_i = header_index(headers, ["total incurred", "incurred", "total"])
    paid_i = header_index(headers, ["paid", "total paid"])
    reserve_i = header_index(headers, ["reserve", "total reserve", "outstanding reserve"])
    subro_i = header_index(headers, ["subrogation", "subro", "recovery"])

    for row in rows[claims_header_row_index + 1:]:
      first = norm(row[0] if row else "")

      if first in stop_sections:
        break

      if not any(clean(cell) for cell in row):
        continue

      claim_number = value_at(row, claim_number_i)

      if not claim_number:
        continue

      line = value_at(row, line_i)
      status = value_at(row, status_i)
      paid = money(value_at(row, paid_i))
      reserve = money(value_at(row, reserve_i))
      incurred = money(value_at(row, incurred_i))

      if incurred <= 0 and (paid or reserve):
        incurred = paid + reserve

      claim = {
        "claim_number": claim_number,
        "date_of_loss": value_at(row, dol_i),
        "date_reported": value_at(row, reported_i),
        "claimant": value_at(row, claimant_i),
        "line_of_business": line,
        "description": value_at(row, description_i),
        "claim_status": status,
        "status": status,
        "total_incurred": incurred,
        "paid": paid,
        "reserve": reserve,
        "subrogation": value_at(row, subro_i),
      }

      claim_policy = choose_policy_for_claim(line, policy_numbers)
      if claim_policy:
        claim["policy_number"] = claim_policy

      repaired_claims.append(claim)

  # If this was truly a section CSV and claims were found, trust the section parser.
  if repaired_claims:
    parsed_claims = repaired_claims

  # LOSSQ_CSV_PRODUCING_AGENCY_PROFILE_EXTRACTION_V1
  # Pull Producing Agency / Producer from the uploaded CSV/loss run itself.
  # Do not use the user's company profile, organization name, carrier name, or demo fallback.
  try:
    agency_labels = [
      "producing agency",
      "producing agent",
      "producer",
      "producer name",
      "agency",
      "agency name",
      "broker",
      "broker name",
      "brokerage",
      "broker of record",
      "agent",
      "agent name",
      "prepared by",
      "contact",
      "contact name",
    ]

    csv_producing_agency = ""

    if callable(locals().get("profile_first")):
      csv_producing_agency = profile_first(*agency_labels)

    def _lossq_csv_agency_clean(value):
      return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

    def _lossq_csv_agency_key(value):
      return re.sub(r"[^a-z0-9]+", "", _lossq_csv_agency_clean(value).lower())

    agency_keys = {_lossq_csv_agency_key(label) for label in agency_labels}

    blocked_values = {
      "",
      "claimsdetail",
      "losssummary",
      "exposurepolicyinformation",
      "policyinformation",
      "claim",
      "claims",
      "claimnumber",
      "policy",
      "policynumber",
      "carrier",
      "writingcarrier",
      "insured",
      "insuredname",
      "namedinsured",
      "date",
      "value",
      "field",
    }

    def _lossq_csv_agency_good_value(value):
      clean_value = _lossq_csv_agency_clean(value)
      value_key = _lossq_csv_agency_key(clean_value)
      if not clean_value:
        return False
      if value_key in agency_keys or value_key in blocked_values:
        return False
      if len(clean_value) < 2:
        return False
      return True

    if not csv_producing_agency and isinstance(rows, list):
      # Same-row label/value, example: Producing Agency, ABC Agency
      for row in rows[:120]:
        cleaned_row = [_lossq_csv_agency_clean(cell) for cell in row]
        for idx, cell in enumerate(cleaned_row):
          if _lossq_csv_agency_key(cell) in agency_keys:
            for value in cleaned_row[idx + 1:]:
              if _lossq_csv_agency_good_value(value):
                csv_producing_agency = value
                break
          if csv_producing_agency:
            break
        if csv_producing_agency:
          break

    if not csv_producing_agency and isinstance(rows, list):
      # Header row style, example:
      # Producing Agency, Carrier, Policy Number
      # ABC Agency, Zenith, WC-123
      for row_index, row in enumerate(rows[:80]):
        header_keys = [_lossq_csv_agency_key(cell) for cell in row]
        if not any(key in agency_keys for key in header_keys):
          continue

        for idx, header_key in enumerate(header_keys):
          if header_key not in agency_keys:
            continue

          for next_row in rows[row_index + 1: row_index + 6]:
            if idx < len(next_row):
              candidate = _lossq_csv_agency_clean(next_row[idx])
              if _lossq_csv_agency_good_value(candidate):
                csv_producing_agency = candidate
                break
          if csv_producing_agency:
            break
        if csv_producing_agency:
          break

    if csv_producing_agency:
      parsed_profile["agency_name"] = csv_producing_agency
      parsed_profile["producing_agency"] = csv_producing_agency
      parsed_profile["producer"] = csv_producing_agency
      print("LOSSQ_CSV_PRODUCING_AGENCY_EXTRACTED", {
        "agency_name": csv_producing_agency[:120]
      })
  except Exception as exc:
    print("LOSSQ_CSV_PRODUCING_AGENCY_EXTRACTION_ERROR", str(exc)[:200])

  # Never let account number become a policy bundle.
  if parsed_profile.get("account_number") and looks_like_policy(parsed_profile.get("account_number")) and not lossq_final_account_like_v3(parsed_profile.get("account_number")):
    parsed_profile["account_number"] = ""

  return parsed_claims, parsed_profile


# LOSSQ_UNIVERSAL_CSV_SECTION_OVERLAY_V2
def lossq_universal_csv_section_overlay_v2(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal section-table overlay for CSV loss runs.

  Fixes:
  - Account Number / Customer Number stays account number, not policy number.
  - CLAIMS DETAIL rows are preserved, including specialty lines like Liquor Liability.
  - Claimant, Jurisdiction/State, Adjuster/Examiner are carried into saved claim rows.
  - Exposure / policy rows populate profile exposures and policy schedule.
  """
  import csv
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def k(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def good(value):
    raw = clean(value)
    return bool(raw and raw.lower() not in {"-", "na", "n/a", "none", "null", "unknown"})

  def is_account_like(value):
    raw = clean(value).upper()
    return bool(re.search(r"\b(ACCT|ACCOUNT|CUSTOMER|CLIENT|CUST)\b", raw) or "ACCT" in raw)

  def is_policy_like(value):
    raw = clean(value).upper()
    if not raw or is_account_like(raw):
      return False
    if not re.search(r"\d", raw):
      return False
    if re.search(r"\b(GL|BOP|WC|AUTO|CA|AL|LIQ|LIQUOR|PROP|CP|UMB|UM|IM|CARGO|GAR|DOL|CY|EPL|DO|PL)\b", raw):
      return True
    if "-" in raw and len(raw) >= 6:
      return True
    return False

  def to_money(value):
    raw = clean(value)
    if not raw:
      return ""
    temp = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
      return float(temp)
    except Exception:
      return raw

  def to_bool(value):
    raw = clean(value).lower()
    if raw in {"yes", "y", "true", "1", "litigated", "attorney", "suit"}:
      return True
    if raw in {"no", "n", "false", "0", "none", "-", "na", "n/a", ""}:
      return False
    return bool(raw)

  def map_row(headers, row):
    mapped = {}
    for idx, header in enumerate(headers):
      header_key = k(header)
      if header_key:
        mapped[header_key] = clean(row[idx]) if idx < len(row) else ""
    return mapped

  def first(row_map, *labels):
    for label in labels:
      value = row_map.get(k(label), "")
      if good(value):
        return clean(value)
    return ""

  def first_money(row_map, *labels):
    value = first(row_map, *labels)
    return to_money(value) if value != "" else ""

  # -----------------------------
  # Account information label/value rows.
  # -----------------------------
  profile_labels = {
    "businessname": "business_name",
    "accountname": "business_name",
    "namedinsured": "business_name",
    "insured": "business_name",
    "insuredname": "business_name",
    "accountnumber": "account_number",
    "accountno": "account_number",
    "accountid": "account_number",
    "customernumber": "customer_number",
    "customerno": "customer_number",
    "customerid": "customer_number",
    "clientnumber": "customer_number",
    "clientno": "customer_number",
    "producingagency": "agency_name",
    "producingagent": "agency_name",
    "agency": "agency_name",
    "agencyname": "agency_name",
    "broker": "agency_name",
    "brokerage": "agency_name",
    "producer": "producer",
    "producername": "producer",
    "carrier": "carrier_name",
    "writingcarrier": "writing_carrier",
    "evaluationdate": "evaluation_date",
    "valuationdate": "evaluation_date",
    "asofdate": "evaluation_date",
    "industry": "industry",
    "state": "state",
  }

  for row in rows[:200]:
    if len(row) < 2:
      continue

    label_key = k(row[0])
    value = clean(row[1])
    field = profile_labels.get(label_key)

    if field and good(value):
      parsed_profile[field] = value

  if parsed_profile.get("account_number"):
    parsed_profile["customer_number"] = parsed_profile.get("customer_number") or parsed_profile.get("account_number")

  if parsed_profile.get("customer_number") and not parsed_profile.get("account_number"):
    parsed_profile["account_number"] = parsed_profile.get("customer_number")

  if parsed_profile.get("agency_name"):
    parsed_profile["producing_agency"] = parsed_profile.get("producing_agency") or parsed_profile.get("agency_name")
    parsed_profile["producer"] = parsed_profile.get("producer") or parsed_profile.get("agency_name")

  if parsed_profile.get("producer") and not parsed_profile.get("agency_name"):
    parsed_profile["agency_name"] = parsed_profile.get("producer")
    parsed_profile["producing_agency"] = parsed_profile.get("producer")

  for row in rows[:200]:
    if len(row) >= 2 and k(row[0]) in {"policyperiod", "policyterm", "coverageperiod"}:
      period = clean(row[1])
      dates = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", period)
      if len(dates) >= 2:
        parsed_profile["effective_date"] = parsed_profile.get("effective_date") or dates[0]
        parsed_profile["expiration_date"] = parsed_profile.get("expiration_date") or dates[1]

  # -----------------------------
  # Table finder.
  # -----------------------------
  def find_header(required_groups, preferred_sections=None):
    preferred_sections = preferred_sections or []
    section_seen = not preferred_sections

    for idx, row in enumerate(rows):
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))
      row_keys = {k(cell) for cell in row if clean(cell)}

      if preferred_sections and any(token in row_text for token in preferred_sections):
        section_seen = True
        continue

      if not section_seen:
        continue

      if all(any(option in row_keys for option in group) for group in required_groups):
        return idx, row

    return None, []

  # -----------------------------
  # Exposure / policy table.
  # -----------------------------
  exposure_idx, exposure_headers = find_header(
    [
      {"policynumber", "policyno", "policy"},
      {"lineofbusiness", "coverage", "policytype", "lob", "currentpremium", "exposurebasis"},
    ],
    ["exposure", "policy information", "policy schedule"],
  )

  exposure_rows = []
  if exposure_idx is not None:
    for row in rows[exposure_idx + 1:]:
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))

      if not any(clean(cell) for cell in row):
        break

      if any(stop in row_text for stop in ["claims detail", "claim detail", "loss summary", "underwriting notes"]):
        break

      rm = map_row(exposure_headers, row)
      policy_number = first(rm, "Policy Number", "Policy No", "Policy")
      line = first(rm, "Line of Business", "Coverage", "Policy Type", "LOB")

      if not is_policy_like(policy_number) and not line:
        continue

      exposure = {
        "policy_number": policy_number,
        "policy_type": line,
        "line_of_business": line,
        "carrier": first(rm, "Carrier", "Writing Carrier") or parsed_profile.get("carrier_name", ""),
        "effective_date": first(rm, "Effective Date", "Policy Effective Date"),
        "expiration_date": first(rm, "Expiration Date", "Policy Expiration Date"),
        "exposure_basis": first(rm, "Exposure Basis", "Basis"),
        "exposure_value": first(rm, "Exposure Value", "Exposure"),
        "payroll": first_money(rm, "Payroll"),
        "revenue": first_money(rm, "Revenue", "Sales", "Gross Sales"),
        "employee_count": first_money(rm, "Employee Count", "Employees"),
        "vehicle_count": first_money(rm, "Vehicle Count", "Vehicles", "Autos"),
        "driver_count": first_money(rm, "Driver Count", "Drivers"),
        "property_tiv": first_money(rm, "Property TIV", "TIV", "Total Insured Value"),
        "current_premium": first_money(rm, "Current Premium"),
        "expiring_premium": first_money(rm, "Expiring Premium"),
        "target_renewal_premium": first_money(rm, "Target Renewal Premium"),
      }

      exposure_rows.append({key: value for key, value in exposure.items() if value not in ("", None)})

  # -----------------------------
  # Claims detail table.
  # -----------------------------
  claim_idx, claim_headers = find_header(
    [
      {"claimnumber", "claimno", "claim", "claimid"},
      {"policynumber", "policyno", "policy", "totalincurred", "paid", "reserve"},
    ],
    ["claims detail", "claim detail", "claims"],
  )

  csv_claims = []
  if claim_idx is not None:
    for row in rows[claim_idx + 1:]:
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))

      if not any(clean(cell) for cell in row):
        break

      if any(stop in row_text for stop in ["underwriting notes", "loss summary", "exposure / policy", "account information"]):
        break

      rm = map_row(claim_headers, row)
      claim_number = first(rm, "Claim Number", "Claim #", "Claim No", "Claim ID", "Claim")
      policy_number = first(rm, "Policy Number", "Policy No", "Policy")

      if not good(claim_number):
        continue

      claim_key = k(claim_number)
      if claim_key in {"claimnumber", "claimno", "claimid", "claim"}:
        continue

      claim = {
        "claim_number": claim_number,
        "policy_number": policy_number,
        "line_of_business": first(rm, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claim_type": first(rm, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claimant": first(rm, "Claimant", "Claimant Name", "Injured Worker", "Injured Party", "Employee Name", "Plaintiff", "Customer Name", "Third Party Name"),
        "jurisdiction_state": first(rm, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "venue_state": first(rm, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "adjuster": first(rm, "Adjuster", "Adjuster/Examiner", "Examiner", "Claim Adjuster", "Claim Examiner", "File Handler"),
        "examiner": first(rm, "Examiner", "Adjuster/Examiner", "Adjuster", "Claim Examiner", "File Handler"),
        "date_of_loss": first(rm, "Date of Loss", "Loss Date"),
        "date_reported": first(rm, "Date Reported", "Reported Date"),
        "date_closed": first(rm, "Date Closed", "Closed Date"),
        "status": first(rm, "Status", "Claim Status"),
        "cause_of_loss": first(rm, "Cause of Loss", "Loss Cause", "Cause"),
        "description": first(rm, "Description", "Loss Description", "Narrative"),
        "paid_amount": first_money(rm, "Paid", "Paid Amount", "Total Paid"),
        "reserve_amount": first_money(rm, "Reserve", "Reserve Amount", "Outstanding Reserve"),
        "total_incurred": first_money(rm, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred"),
        "litigation": to_bool(first(rm, "Litigation", "Litigated")),
        "attorney_assigned": to_bool(first(rm, "Attorney Assigned", "Attorney", "Counsel")),
      }

      csv_claims.append({key: value for key, value in claim.items() if value not in ("", None)})

  # -----------------------------
  # Merge parsed claims + CSV section claims.
  # -----------------------------
  merged = []
  seen = {}

  def merge_key(claim):
    claim_number = clean(claim.get("claim_number") or claim.get("Claim Number") or claim.get("claim #")).upper()
    policy_number = clean(claim.get("policy_number") or claim.get("Policy Number") or claim.get("policy")).upper()
    return f"{claim_number}|{policy_number}"

  for claim in parsed_claims:
    if isinstance(claim, dict):
      target = dict(claim)
      merged.append(target)
      seen[merge_key(target)] = target

  overlay_fields = [
    "policy_number",
    "line_of_business",
    "claim_type",
    "claimant",
    "jurisdiction_state",
    "venue_state",
    "adjuster",
    "examiner",
    "date_of_loss",
    "date_reported",
    "date_closed",
    "status",
    "cause_of_loss",
    "description",
    "paid_amount",
    "reserve_amount",
    "total_incurred",
    "litigation",
    "attorney_assigned",
  ]

  added_claims = 0
  updated_claims = 0

  for csv_claim in csv_claims:
    mk = merge_key(csv_claim)

    if mk in seen:
      target = seen[mk]
      for field in overlay_fields:
        value = csv_claim.get(field)
        if value not in ("", None):
          if field in {"claimant", "jurisdiction_state", "venue_state", "adjuster", "examiner"} or not target.get(field):
            target[field] = value
      updated_claims += 1
    else:
      target = dict(csv_claim)
      merged.append(target)
      seen[mk] = target
      added_claims += 1

  parsed_claims = merged

  # -----------------------------
  # Exposures and policy schedule.
  # -----------------------------
  if exposure_rows:
    parsed_profile["exposures"] = exposure_rows
    parsed_profile["exposure_inputs"] = {"exposure_rows": exposure_rows}

    def nums(field):
      values = []
      for exposure in exposure_rows:
        try:
          value = exposure.get(field)
          if value not in ("", None):
            values.append(float(value))
        except Exception:
          pass
      return values

    for field in ["current_premium", "expiring_premium", "target_renewal_premium"]:
      values = nums(field)
      if values:
        parsed_profile[field] = sum(values)
        parsed_profile["exposure_inputs"][field] = sum(values)

    for field in ["payroll", "revenue", "employee_count", "vehicle_count", "driver_count", "property_tiv"]:
      values = nums(field)
      if values:
        parsed_profile[field] = max(values)
        parsed_profile["exposure_inputs"][field] = max(values)

  claim_counts = {}
  claim_totals = {}
  claim_lines = {}

  for claim in parsed_claims:
    policy_number = clean(claim.get("policy_number")).upper()
    if not is_policy_like(policy_number):
      continue

    claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1

    try:
      claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + float(claim.get("total_incurred") or 0)
    except Exception:
      pass

    line = clean(claim.get("line_of_business") or claim.get("claim_type"))
    if line:
      claim_lines[policy_number] = line

  policies_by_number = {}

  existing_policies = parsed_profile.get("policies") if isinstance(parsed_profile.get("policies"), list) else []
  for policy in existing_policies:
    if not isinstance(policy, dict):
      continue

    policy_number = clean(policy.get("policy_number") or policy.get("Policy Number")).upper()
    if is_policy_like(policy_number):
      policies_by_number[policy_number] = dict(policy)

  for exposure in exposure_rows:
    policy_number = clean(exposure.get("policy_number")).upper()
    if not is_policy_like(policy_number):
      continue

    policy = policies_by_number.get(policy_number, {})
    policy.update({
      "policy_number": exposure.get("policy_number"),
      "policy_type": exposure.get("policy_type") or exposure.get("line_of_business") or claim_lines.get(policy_number) or policy.get("policy_type"),
      "line_of_business": exposure.get("line_of_business") or exposure.get("policy_type") or claim_lines.get(policy_number) or policy.get("line_of_business"),
      "carrier": exposure.get("carrier") or policy.get("carrier") or parsed_profile.get("carrier_name"),
      "effective_date": exposure.get("effective_date") or policy.get("effective_date") or parsed_profile.get("effective_date"),
      "expiration_date": exposure.get("expiration_date") or policy.get("expiration_date") or parsed_profile.get("expiration_date"),
      "claim_count": claim_counts.get(policy_number, policy.get("claim_count", 0)),
      "total_incurred": claim_totals.get(policy_number, policy.get("total_incurred", 0)),
      "current_premium": exposure.get("current_premium") or policy.get("current_premium"),
      "expiring_premium": exposure.get("expiring_premium") or policy.get("expiring_premium"),
      "target_renewal_premium": exposure.get("target_renewal_premium") or policy.get("target_renewal_premium"),
    })
    policies_by_number[policy_number] = policy

  # Add policy schedule rows from claims even if an exposure row is absent.
  for policy_number, count in claim_counts.items():
    if policy_number not in policies_by_number:
      policies_by_number[policy_number] = {
        "policy_number": policy_number,
        "policy_type": claim_lines.get(policy_number, ""),
        "line_of_business": claim_lines.get(policy_number, ""),
        "carrier": parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier"),
        "effective_date": parsed_profile.get("effective_date"),
        "expiration_date": parsed_profile.get("expiration_date"),
        "claim_count": count,
        "total_incurred": claim_totals.get(policy_number, 0),
      }

  if policies_by_number:
    parsed_profile["policies"] = list(policies_by_number.values())
    parsed_profile["policy_schedule"] = parsed_profile["policies"]

    first_policy = parsed_profile["policies"][0].get("policy_number")
    current_main = parsed_profile.get("policy_number") or parsed_profile.get("main_policy")

    if not is_policy_like(current_main):
      parsed_profile["policy_number"] = first_policy
      parsed_profile["main_policy"] = first_policy

  # Never let account/customer number overwrite policy fields.
  for field in ["policy_number", "main_policy"]:
    if parsed_profile.get(field) and is_account_like(parsed_profile.get(field)):
      first_policy = ""
      for policy in parsed_profile.get("policies") or []:
        candidate = policy.get("policy_number") if isinstance(policy, dict) else ""
        if is_policy_like(candidate):
          first_policy = candidate
          break
      parsed_profile[field] = first_policy

  print("LOSSQ_UNIVERSAL_CSV_SECTION_OVERLAY_V2", {
    "csv_claims": len(csv_claims),
    "added_claims": added_claims,
    "updated_claims": updated_claims,
    "exposure_rows": len(exposure_rows),
    "account_number": str(parsed_profile.get("account_number") or "")[:80],
    "policy_number": str(parsed_profile.get("policy_number") or "")[:80],
  })

  return parsed_claims, parsed_profile


# LOSSQ_CLAIM_DETAIL_FIELDS_FROM_UPLOAD_ROW_V2
def lossq_claim_detail_clean_v2(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def lossq_claim_detail_key_v2(value):
  return re.sub(r"[^a-z0-9]+", "", lossq_claim_detail_clean_v2(value).lower())


def lossq_claim_detail_first_v2(raw_claim, *labels):
  if not isinstance(raw_claim, dict):
    return ""

  label_keys = {lossq_claim_detail_key_v2(label) for label in labels}

  for key, value in raw_claim.items():
    if lossq_claim_detail_key_v2(key) in label_keys:
      clean_value = lossq_claim_detail_clean_v2(value)
      if clean_value and clean_value.lower() not in {"-", "na", "n/a", "none", "null", "unknown"}:
        return clean_value

  return ""


def lossq_apply_claim_detail_fields_to_normalized_claim_v2(normalized_claim, raw_claim):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  claimant = lossq_claim_detail_first_v2(
    raw_claim,
    "claimant",
    "claimant name",
    "injured worker",
    "injured party",
    "employee name",
    "plaintiff",
    "customer name",
    "third party name",
  )

  jurisdiction_state = lossq_claim_detail_first_v2(
    raw_claim,
    "jurisdiction/state",
    "jurisdiction",
    "state",
    "venue state",
    "loss state",
  )

  adjuster = lossq_claim_detail_first_v2(
    raw_claim,
    "adjuster",
    "adjuster/examiner",
    "examiner",
    "claim adjuster",
    "claim examiner",
    "file handler",
  )

  if claimant:
    normalized_claim["claimant"] = claimant

  if jurisdiction_state:
    normalized_claim["jurisdiction_state"] = jurisdiction_state
    normalized_claim["venue_state"] = normalized_claim.get("venue_state") or jurisdiction_state

  if adjuster:
    normalized_claim["adjuster"] = adjuster
    normalized_claim["examiner"] = normalized_claim.get("examiner") or adjuster

  normalized_claim.pop("claimant_name", None)
  normalized_claim.pop("jurisdiction", None)
  normalized_claim.pop("state", None)
  normalized_claim.pop("adjuster_examiner", None)

  return normalized_claim


# LOSSQ_FRENCH_CSV_SECTION_PARSER_V1
def lossq_parse_french_section_csv_v1(file_path):
  """
  Universal French / Quebec CSV parser.
  Normalizes French loss-run labels into the standard LossQ schema.
  This is language/terminology support, not company-specific logic.
  """
  try:
    import csv as _lossq_fr_csv
    import unicodedata as _lossq_fr_unicodedata
  except Exception:
    return [], {}

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(_lossq_fr_csv.reader(handle))
  except Exception:
    return [], {}

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    text = clean(value).lower()
    text = _lossq_fr_unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not _lossq_fr_unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text)

  all_text = "\n".join(" | ".join(clean(cell) for cell in row) for row in rows).lower()
  french_signals = [
    "numéro de sinistre",
    "numero de sinistre",
    "date de survenance",
    "date de déclaration",
    "montant payé",
    "provision",
    "responsabilité civile",
    "biens commerciaux",
    "détails des sinistres",
    "details des sinistres",
    "réclamant",
    "reclamant",
  ]
  if not any(signal in all_text for signal in french_signals):
    return [], {}

  def value_at(row, index):
    if index is None or index < 0 or index >= len(row):
      return ""
    return clean(row[index])

  def header_index(headers, aliases):
    alias_keys = {key(alias) for alias in aliases}
    for index, header in enumerate(headers):
      if key(header) in alias_keys:
        return index
    return None

  def first_value(*labels):
    label_keys = {key(label) for label in labels}
    for row in rows:
      if len(row) >= 3 and key(row[1]) in label_keys:
        return clean(row[2])
    return ""

  def money(value):
    text = clean(value)
    text = re.sub(r"(?i)\b(?:cad|cdn|cnd|usd)\b", "", text)
    text = text.replace("CA$", "").replace("C$", "").replace("$", "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else ""

  def count(value):
    match = re.search(r"\d+", clean(value).replace(",", ""))
    return match.group(0) if match else ""

  def date_fr(value):
    text = clean(value)
    match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", text)
    if not match:
      return text
    day, month, year = match.groups()
    if len(year) == 2:
      year = "20" + year
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

  def status_fr(value):
    text = key(value)
    if text in {"ferme", "fermee", "clos", "cloture", "cloturee", "closed"}:
      return "Closed"
    if text in {"ouvert", "ouverte", "open"}:
      return "Open"
    return clean(value) or "Open"

  def yes_no(value):
    text = key(value)
    if text in {"oui", "yes", "y", "true", "1", "o"}:
      return "Yes"
    if text in {"non", "no", "n", "false", "0"}:
      return "No"
    return clean(value)

  def is_yes(value):
    return key(value) in {"oui", "yes", "y", "true", "1", "o"}

  def line_fr(value):
    text = key(value)
    mapping = {
      "responsabilitecivilecommerciale": "General Liability",
      "responsabilitecivile": "General Liability",
      "bienscommerciaux": "Commercial Property",
      "proprietecommerciale": "Commercial Property",
      "automobilecommerciale": "Commercial Auto",
      "flotteautomobile": "Commercial Auto",
      "indemnisationdestravailleurs": "Workers Compensation",
      "accidentsdutravail": "Workers Compensation",
      "erreursetomissions": "Professional Liability",
      "responsabiliteprofessionnelle": "Professional Liability",
      "cyberresponsabilite": "Cyber",
      "responsabilitelieealalcool": "Liquor Liability",
      "interruptiondesaffaires": "Business Interruption",
      "assuranceexcedentaire": "Umbrella / Excess",
    }
    return mapping.get(text, clean(value))

  profile = {
    "business_name": first_value("Nom de l'assuré", "Nom assuré", "Assuré", "Entreprise"),
    "named_insured": first_value("Nom de l'assuré", "Nom assuré", "Assuré", "Entreprise"),
    "account_number": first_value("Numéro de compte", "No de compte", "Compte"),
    "customer_number": first_value("Numéro client", "No client", "Client"),
    "country": first_value("Pays"),
    "state": first_value("Province", "Territoire"),
    "province": first_value("Province", "Territoire"),
    "province_code": first_value("Province", "Territoire"),
    "postal_code": first_value("Code postal", "Code postal canadien"),
    "market_currency": first_value("Devise", "Monnaie") or "CAD",
    "currency": first_value("Devise", "Monnaie") or "CAD",
    "producing_agency": first_value("Courtier producteur", "Agence productrice", "Courtier", "Maison de courtage"),
    "agency_name": first_value("Courtier producteur", "Agence productrice", "Courtier", "Maison de courtage"),
    "broker_name": first_value("Courtier responsable", "Courtier", "Producteur"),
    "carrier_name": first_value("Compagnie d'assurance", "Assureur", "Transporteur", "Marché"),
    "writing_carrier": first_value("Compagnie d'assurance", "Assureur", "Transporteur", "Marché"),
    "underwriter": first_value("Souscripteur", "Souscription"),
    "evaluation_date": date_fr(first_value("Date d'évaluation", "Date de valorisation", "Date d'analyse")),
    "market_country": "Canada",
    "market_language": "fr",
    "market_currency": "CAD",
  }

  period = first_value("Période de police", "Terme de police", "Période d'assurance")
  period_dates = re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", period)
  if len(period_dates) >= 2:
    profile["effective_date"] = date_fr(period_dates[0])
    profile["expiration_date"] = date_fr(period_dates[1])

  # Exposure terms.
  profile["current_premium"] = money(first_value("Prime actuelle", "Prime annuelle"))
  profile["revenue"] = money(first_value("Chiffre d'affaires", "Revenu annuel", "Ventes brutes"))
  profile["payroll"] = money(first_value("Paie assurable", "Masse salariale", "Rémunération"))
  profile["employee_count"] = count(first_value("Nombre d'employés", "Employés"))
  profile["property_tiv"] = money(first_value("SOV / Valeur totale assurée", "Valeur totale assurée", "SOV", "TIV"))
  profile["liquor_sales"] = money(first_value("Ventes d'alcool", "Ventes alcool"))

  fleet_text = first_value("Véhicules / conducteurs", "Vehicules / conducteurs", "Véhicules", "Conducteurs")
  if fleet_text:
    vehicle_match = re.search(r"(\d+)\s*(?:véhicules?|vehicules?|autos?)", fleet_text, re.I)
    driver_match = re.search(r"(\d+)\s*(?:conducteurs?|drivers?)", fleet_text, re.I)
    if vehicle_match:
      profile["vehicle_count"] = vehicle_match.group(1)
    if driver_match:
      profile["driver_count"] = driver_match.group(1)

  policies = []
  claims = []

  for row_index, row in enumerate(rows):
    headers = [clean(cell) for cell in row]
    header_keys = {key(cell) for cell in headers}

    if "numerodepolice" in header_keys and "lignedaffaires" in header_keys:
      policy_i = header_index(headers, ["Numéro de police", "No de police", "Police"])
      line_i = header_index(headers, ["Ligne d'affaires", "Garantie", "Couverture"])
      eff_i = header_index(headers, ["Date d'effet", "Date effective"])
      exp_i = header_index(headers, ["Date d'expiration", "Expiration"])
      limit_i = header_index(headers, ["Limite", "Limite de garantie"])
      deductible_i = header_index(headers, ["Franchise", "Déductible"])

      for data in rows[row_index + 1:]:
        if not any(clean(cell) for cell in data):
          break
        if key(data[0]) not in {"horairedespolices", "calendrierdespolices", "polices"}:
          break
        policy_number = value_at(data, policy_i)
        if not policy_number:
          continue
        line = line_fr(value_at(data, line_i))
        policies.append({
          "policy_number": policy_number,
          "policy": policy_number,
          "line_of_business": line,
          "policy_type": line,
          "coverage": line,
          "effective_date": date_fr(value_at(data, eff_i)),
          "expiration_date": date_fr(value_at(data, exp_i)),
          "limit": money(value_at(data, limit_i)),
          "deductible": money(value_at(data, deductible_i)),
          "carrier": profile.get("carrier_name", ""),
          "writing_carrier": profile.get("writing_carrier", ""),
        })

    if "numerodesinistre" in header_keys:
      claim_i = header_index(headers, ["Numéro de sinistre", "No de sinistre", "Sinistre"])
      policy_i = header_index(headers, ["Numéro de police", "No de police", "Police"])
      line_i = header_index(headers, ["Ligne d'affaires", "Garantie", "Couverture"])
      status_i = header_index(headers, ["Statut", "État", "Etat"])
      loss_i = header_index(headers, ["Date de survenance", "Date du sinistre", "Date de perte"])
      reported_i = header_index(headers, ["Date de déclaration", "Date rapportée"])
      closed_i = header_index(headers, ["Date de fermeture", "Date fermée"])
      claimant_i = header_index(headers, ["Réclamant", "Reclamant", "Demandeur"])
      province_i = header_index(headers, ["Province", "Territoire", "Juridiction"])
      adjuster_i = header_index(headers, ["Expert en sinistres", "Expert", "Ajusteur", "Régleur"])
      paid_i = header_index(headers, ["Montant payé", "Payé", "Indemnité payée"])
      reserve_i = header_index(headers, ["Provision", "Réserve", "Montant réservé"])
      total_i = header_index(headers, ["Total encouru", "Total engagé", "Incurred"])
      attorney_i = header_index(headers, ["Avocat assigné", "Avocat", "Conseiller juridique"])
      litigation_i = header_index(headers, ["Litige", "Poursuite", "En litige"])
      desc_i = header_index(headers, ["Description", "Description du sinistre", "Cause"])

      for data in rows[row_index + 1:]:
        if not any(clean(cell) for cell in data):
          break
        if key(data[0]) not in {"detailsdessinistres", "detaildessinistres", "sinistres"}:
          break
        claim_number = value_at(data, claim_i)
        if not claim_number:
          continue

        paid = money(value_at(data, paid_i))
        reserve = money(value_at(data, reserve_i))
        total = money(value_at(data, total_i))
        if not total and (paid or reserve):
          try:
            total = str(float(paid or 0) + float(reserve or 0))
          except Exception:
            total = ""

        attorney_value = yes_no(value_at(data, attorney_i))
        litigation_value = yes_no(value_at(data, litigation_i))
        flag = "Attorney" if is_yes(attorney_value) else "Litigation" if is_yes(litigation_value) else ""

        line = line_fr(value_at(data, line_i))
        status = status_fr(value_at(data, status_i))
        jurisdiction = value_at(data, province_i) or profile.get("state", "")
        adjuster = value_at(data, adjuster_i)

        claims.append({
          "claim_number": claim_number,
          "claim_no": claim_number,
          "policy_number": value_at(data, policy_i),
          "policy": value_at(data, policy_i),
          "line_of_business": line,
          "claim_type": line,
          "coverage": line,
          "status": status,
          "claim_status": status,
          "date_of_loss": date_fr(value_at(data, loss_i)),
          "loss_date": date_fr(value_at(data, loss_i)),
          "date_reported": date_fr(value_at(data, reported_i)),
          "reported_date": date_fr(value_at(data, reported_i)),
          "date_closed": date_fr(value_at(data, closed_i)),
          "closed_date": date_fr(value_at(data, closed_i)),
          "claimant": value_at(data, claimant_i),
          "jurisdiction_state": jurisdiction,
          "venue_state": jurisdiction,
          "adjuster": adjuster,
          "examiner": adjuster,
          "paid_amount": paid,
          "paid": paid,
          "reserve_amount": reserve,
          "reserve": reserve,
          "total_incurred": total,
          "incurred": total,
          "attorney_assigned": attorney_value,
          "attorney_involved": attorney_value,
          "litigation": litigation_value,
          "suit_filed": litigation_value,
          "flag": flag,
          "description": value_at(data, desc_i),
          "loss_description": value_at(data, desc_i),
        })

  if policies:
    profile["policies"] = policies
    profile["policy_schedule"] = policies
    profile["policy_number"] = policies[0].get("policy_number", "")
    profile["effective_date"] = profile.get("effective_date") or policies[0].get("effective_date", "")
    profile["expiration_date"] = profile.get("expiration_date") or policies[0].get("expiration_date", "")
    profile["line_of_business"] = "Multi-line: " + ", ".join(dict.fromkeys([p.get("line_of_business", "") for p in policies if p.get("line_of_business")]))
    profile["primary_line_of_business"] = profile["line_of_business"]

  if lossq_normalize_market_profile is not None:
    try:
      profile = lossq_normalize_market_profile(profile, all_text)
    except Exception as exc:
      print("LOSSQ_FRENCH_MARKET_PROFILE_NORMALIZE_FAILED_V1", str(exc)[:300])

  profile = {k: v for k, v in profile.items() if v not in (None, "")}
  print("LOSSQ_FRENCH_CSV_SECTION_PARSER_V1", {"claims": len(claims), "policies": len(policies), "business_name": profile.get("business_name"), "state": profile.get("state")})
  return claims, profile

# LOSSQ_UNIVERSAL_CSV_ACCOUNT_EXPOSURE_CLAIM_DETAIL_OVERLAY_V1
def lossq_universal_csv_account_exposure_claim_detail_overlay(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal CSV section overlay.

  Purpose:
  - Pull true account/customer number from uploaded file.
  - Pull exposure/premium rows from uploaded file.
  - Pull every claim row from CLAIMS DETAIL, including Claimant, Jurisdiction/State,
   Adjuster/Examiner, and specialty lines such as Liquor Liability.
  - Merge missing rows/details into parsed_claims without hardcoding any account,
   business, carrier, or sample file.
  """
  import csv
  import re

  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

  def money(value):
    raw = clean(value)
    if not raw:
      return ""
    raw = raw.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
      return float(raw)
    except Exception:
      return clean(value)

  def bool_value(value):
    raw = clean(value).lower()
    if raw in {"yes", "y", "true", "1", "litigated", "attorney", "suit"}:
      return True
    if raw in {"no", "n", "false", "0", "none"}:
      return False
    return bool(raw and raw not in {"-", "na", "n/a"})

  def good_text(value):
    raw = clean(value)
    if not raw:
      return False
    lowered = raw.lower()
    if lowered in {"-", "na", "n/a", "none", "null", "unknown"}:
      return False
    return True

  def first(row_map, *labels):
    for label in labels:
      value = row_map.get(key(label), "")
      if good_text(value):
        return clean(value)
    return ""

  def first_money(row_map, *labels):
    value = first(row_map, *labels)
    return money(value) if value != "" else ""

  def row_map_from_headers(headers, row):
    mapped = {}
    for idx, header in enumerate(headers):
      header_key = key(header)
      if not header_key:
        continue
      mapped[header_key] = clean(row[idx]) if idx < len(row) else ""
    return mapped

  profile_label_map = {
    "businessname": "business_name",
    "accountname": "business_name",
    "namedinsured": "business_name",
    "insured": "business_name",
    "insuredname": "business_name",
    "accountnumber": "account_number",
    "accountno": "account_number",
    "account": "account_number",
    "customernumber": "customer_number",
    "customerno": "customer_number",
    "clientnumber": "customer_number",
    "producingagency": "agency_name",
    "producingagent": "agency_name",
    "agency": "agency_name",
    "agencyname": "agency_name",
    "broker": "agency_name",
    "brokerage": "agency_name",
    "producer": "producer",
    "producername": "producer",
    "carrier": "carrier_name",
    "writingcarrier": "writing_carrier",
    "policynumber": "policy_number",
    "mainpolicy": "policy_number",
    "mainpolicynumber": "policy_number",
    "evaluationdate": "evaluation_date",
    "valuationdate": "evaluation_date",
    "asofdate": "evaluation_date",
    "industry": "industry",
    "state": "state",
  }

  # Label/value account section.
  for row in rows[:150]:
    if len(row) < 2:
      continue
    label_key = key(row[0])
    value = clean(row[1])
    field = profile_label_map.get(label_key)
    if field and good_text(value):
      parsed_profile[field] = value

  if parsed_profile.get("account_number"):
    parsed_profile["customer_number"] = parsed_profile.get("customer_number") or parsed_profile.get("account_number")

  if parsed_profile.get("agency_name"):
    parsed_profile["producing_agency"] = parsed_profile.get("producing_agency") or parsed_profile.get("agency_name")
    parsed_profile["producer"] = parsed_profile.get("producer") or parsed_profile.get("agency_name")

  if parsed_profile.get("producer") and not parsed_profile.get("agency_name"):
    parsed_profile["agency_name"] = parsed_profile.get("producer")
    parsed_profile["producing_agency"] = parsed_profile.get("producer")

  # Policy period label.
  for row in rows[:150]:
    if len(row) >= 2 and key(row[0]) in {"policyperiod", "policyterm", "coverageperiod"}:
      period = clean(row[1])
      dates = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", period)
      if len(dates) >= 2:
        parsed_profile["effective_date"] = parsed_profile.get("effective_date") or dates[0]
        parsed_profile["expiration_date"] = parsed_profile.get("expiration_date") or dates[1]

  def find_table_header(required_any, preferred_section_tokens=None):
    preferred_section_tokens = preferred_section_tokens or []
    section_seen = not preferred_section_tokens

    for idx, row in enumerate(rows):
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))
      row_keys = {key(cell) for cell in row if clean(cell)}

      if preferred_section_tokens and any(token in row_text for token in preferred_section_tokens):
        section_seen = True
        continue

      if not section_seen:
        continue

      if all(any(req in row_keys for req in req_group) for req_group in required_any):
        return idx, row

    return None, []

  # Exposure / policy table.
  exposure_header_idx, exposure_headers = find_table_header(
    required_any=[
      {"policynumber", "policyno", "policy"},
      {"lineofbusiness", "coverage", "policytype", "lob", "exposurebasis", "currentpremium"},
    ],
    preferred_section_tokens=["exposure", "policy information", "policy schedule"],
  )

  exposure_rows = []
  if exposure_header_idx is not None:
    for row in rows[exposure_header_idx + 1:]:
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))
      if not any(clean(cell) for cell in row):
        break
      if any(stop in row_text for stop in ["claims detail", "claim detail", "loss summary", "underwriting notes"]):
        break

      row_map = row_map_from_headers(exposure_headers, row)
      policy_number = first(row_map, "Policy Number", "Policy No", "Policy")
      line_of_business = first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB")
      if not policy_number and not line_of_business:
        continue

      exposure = {
        "policy_number": policy_number,
        "policy_type": line_of_business,
        "line_of_business": line_of_business,
        "carrier": first(row_map, "Carrier", "Writing Carrier") or parsed_profile.get("carrier_name", ""),
        "effective_date": first(row_map, "Effective Date", "Policy Effective Date"),
        "expiration_date": first(row_map, "Expiration Date", "Policy Expiration Date"),
        "exposure_basis": first(row_map, "Exposure Basis", "Basis"),
        "exposure_value": first(row_map, "Exposure Value", "Exposure"),
        "payroll": first_money(row_map, "Payroll"),
        "revenue": first_money(row_map, "Revenue", "Sales", "Gross Sales"),
        "employee_count": first_money(row_map, "Employee Count", "Employees"),
        "vehicle_count": first_money(row_map, "Vehicle Count", "Vehicles", "Autos"),
        "driver_count": first_money(row_map, "Driver Count", "Drivers"),
        "property_tiv": first_money(row_map, "Property TIV", "TIV", "Total Insured Value"),
        "current_premium": first_money(row_map, "Current Premium"),
        "expiring_premium": first_money(row_map, "Expiring Premium"),
        "target_renewal_premium": first_money(row_map, "Target Renewal Premium"),
      }

      exposure_rows.append({k: v for k, v in exposure.items() if v not in ("", None)})

  # Claims detail table.
  claim_header_idx, claim_headers = find_table_header(
    required_any=[
      {"claimnumber", "claimno", "claim", "claimid"},
      {"policynumber", "policy", "policyno", "totalincurred", "paid", "reserve"},
    ],
    preferred_section_tokens=["claims detail", "claim detail", "claims"],
  )

  csv_claims = []
  if claim_header_idx is not None:
    for row in rows[claim_header_idx + 1:]:
      row_text = " ".join(clean(cell).lower() for cell in row if clean(cell))
      if not any(clean(cell) for cell in row):
        break
      if any(stop in row_text for stop in ["underwriting notes", "loss summary", "exposure / policy", "account information"]):
        break

      row_map = row_map_from_headers(claim_headers, row)

      claim_number = first(row_map, "Claim Number", "Claim #", "Claim No", "Claim ID", "Claim")
      policy_number = first(row_map, "Policy Number", "Policy No", "Policy")

      if not claim_number:
        continue

      claim = {
        "claim_number": claim_number,
        "policy_number": policy_number,
        "line_of_business": first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claim_type": first(row_map, "Line of Business", "Coverage", "Policy Type", "LOB"),
        "claimant": first(row_map, "Claimant", "Claimant Name", "Injured Worker", "Injured Party", "Employee Name", "Plaintiff", "Customer Name", "Third Party Name"),
        "jurisdiction_state": first(row_map, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "venue_state": first(row_map, "Jurisdiction/State", "Jurisdiction", "State", "Venue State", "Loss State"),
        "adjuster": first(row_map, "Adjuster", "Adjuster/Examiner", "Examiner", "Claim Adjuster", "Claim Examiner", "File Handler"),
        "examiner": first(row_map, "Examiner", "Adjuster/Examiner", "Adjuster", "Claim Examiner", "File Handler"),
        "date_of_loss": first(row_map, "Date of Loss", "Loss Date"),
        "date_reported": first(row_map, "Date Reported", "Reported Date"),
        "date_closed": first(row_map, "Date Closed", "Closed Date"),
        "status": first(row_map, "Status", "Claim Status"),
        "cause_of_loss": first(row_map, "Cause of Loss", "Loss Cause", "Cause"),
        "description": first(row_map, "Description", "Loss Description", "Narrative"),
        "paid_amount": first_money(row_map, "Paid", "Paid Amount", "Total Paid"),
        "reserve_amount": first_money(row_map, "Reserve", "Reserve Amount", "Outstanding Reserve"),
        "total_incurred": first_money(row_map, "Total Incurred", "Incurred", "Gross Incurred", "Net Incurred"),
        "litigation": bool_value(first(row_map, "Litigation", "Litigated")),
        "attorney_assigned": bool_value(first(row_map, "Attorney Assigned", "Attorney", "Counsel")),
      }

      csv_claims.append({k: v for k, v in claim.items() if v not in ("", None)})

  # Merge CSV claims into parsed claims.
  merged_claims = []
  seen = {}

  def claim_key(claim):
    claim_number = clean(claim.get("claim_number") or claim.get("Claim Number") or claim.get("claim #")).upper()
    policy_number = clean(claim.get("policy_number") or claim.get("Policy Number") or claim.get("policy")).upper()
    return f"{claim_number}|{policy_number}"

  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue
    merged = dict(claim)
    merged_claims.append(merged)
    seen[claim_key(merged)] = merged

  overlay_fields = [
    "policy_number",
    "line_of_business",
    "claim_type",
    "claimant",
    "jurisdiction_state",
    "venue_state",
    "adjuster",
    "examiner",
    "date_of_loss",
    "date_reported",
    "date_closed",
    "status",
    "cause_of_loss",
    "description",
    "paid_amount",
    "reserve_amount",
    "total_incurred",
    "litigation",
    "attorney_assigned",
  ]

  added_claims = 0
  updated_claims = 0

  for csv_claim in csv_claims:
    ck = claim_key(csv_claim)
    if ck in seen:
      target = seen[ck]
      for field in overlay_fields:
        value = csv_claim.get(field)
        if value not in ("", None):
          if field in {"claimant", "jurisdiction_state", "venue_state", "adjuster", "examiner"} or not target.get(field):
            target[field] = value
      updated_claims += 1
    else:
      merged_claims.append(dict(csv_claim))
      seen[ck] = merged_claims[-1]
      added_claims += 1

  parsed_claims = merged_claims

  # Aggregate exposure inputs.
  if exposure_rows:
    parsed_profile["exposures"] = exposure_rows
    parsed_profile["exposure_inputs"] = {
      "exposure_rows": exposure_rows,
    }

    def numeric_values(field):
      values = []
      for exposure in exposure_rows:
        value = exposure.get(field)
        try:
          if value not in ("", None):
            values.append(float(value))
        except Exception:
          pass
      return values

    for field in ["current_premium", "expiring_premium", "target_renewal_premium"]:
      values = numeric_values(field)
      if values:
        parsed_profile[field] = sum(values)
        parsed_profile["exposure_inputs"][field] = sum(values)

    for field in ["payroll", "revenue", "employee_count", "vehicle_count", "driver_count", "property_tiv"]:
      values = numeric_values(field)
      if values:
        parsed_profile[field] = max(values)
        parsed_profile["exposure_inputs"][field] = max(values)

  # Rebuild/merge policy schedule from exposure rows and claim rows.
  if exposure_rows:
    claim_counts = {}
    claim_totals = {}

    for claim in parsed_claims:
      policy_number = clean(claim.get("policy_number")).upper()
      if not policy_number:
        continue
      claim_counts[policy_number] = claim_counts.get(policy_number, 0) + 1
      try:
        claim_totals[policy_number] = claim_totals.get(policy_number, 0.0) + float(claim.get("total_incurred") or 0)
      except Exception:
        pass

    existing_policies = parsed_profile.get("policies") if isinstance(parsed_profile.get("policies"), list) else []
    policies_by_number = {}

    for policy in existing_policies:
      if not isinstance(policy, dict):
        continue
      policy_number = clean(policy.get("policy_number") or policy.get("Policy Number")).upper()
      if policy_number:
        policies_by_number[policy_number] = dict(policy)

    for exposure in exposure_rows:
      policy_number = clean(exposure.get("policy_number")).upper()
      if not policy_number:
        continue

      policy = policies_by_number.get(policy_number, {})
      policy.update({
        "policy_number": exposure.get("policy_number"),
        "policy_type": exposure.get("policy_type") or exposure.get("line_of_business") or policy.get("policy_type"),
        "line_of_business": exposure.get("line_of_business") or exposure.get("policy_type") or policy.get("line_of_business"),
        "carrier": exposure.get("carrier") or policy.get("carrier") or parsed_profile.get("carrier_name"),
        "effective_date": exposure.get("effective_date") or policy.get("effective_date") or parsed_profile.get("effective_date"),
        "expiration_date": exposure.get("expiration_date") or policy.get("expiration_date") or parsed_profile.get("expiration_date"),
        "claim_count": claim_counts.get(policy_number, policy.get("claim_count", 0)),
        "total_incurred": claim_totals.get(policy_number, policy.get("total_incurred", 0)),
        "current_premium": exposure.get("current_premium") or policy.get("current_premium"),
        "expiring_premium": exposure.get("expiring_premium") or policy.get("expiring_premium"),
        "target_renewal_premium": exposure.get("target_renewal_premium") or policy.get("target_renewal_premium"),
      })
      policies_by_number[policy_number] = policy

    parsed_profile["policies"] = list(policies_by_number.values())
    parsed_profile["policy_schedule"] = parsed_profile["policies"]

    if parsed_profile["policies"] and not parsed_profile.get("policy_number"):
      parsed_profile["policy_number"] = parsed_profile["policies"][0].get("policy_number")

  print("LOSSQ_UNIVERSAL_CSV_ACCOUNT_EXPOSURE_CLAIM_DETAIL_OVERLAY", {
    "csv_claims": len(csv_claims),
    "added_claims": added_claims,
    "updated_claims": updated_claims,
    "exposure_rows": len(exposure_rows),
    "account_number": str(parsed_profile.get("account_number") or "")[:80],
  })

  return parsed_claims, parsed_profile


# LOSSQ_CLAIM_DETAIL_FIELDS_FROM_UPLOAD_ROW_V1
def lossq_claim_detail_clean_value(value):
  return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def lossq_claim_detail_key(value):
  return re.sub(r"[^a-z0-9]+", "", lossq_claim_detail_clean_value(value).lower())


def lossq_claim_detail_value(raw_claim, *labels):
  if not isinstance(raw_claim, dict):
    return ""

  label_keys = {lossq_claim_detail_key(label) for label in labels}

  for key, value in raw_claim.items():
    if lossq_claim_detail_key(key) in label_keys:
      clean_value = lossq_claim_detail_clean_value(value)
      if clean_value and clean_value.lower() not in {"-", "na", "n/a", "none", "null", "unknown"}:
        return clean_value

  return ""


def lossq_apply_claim_detail_fields_to_normalized_claim(normalized_claim, raw_claim):
  if not isinstance(normalized_claim, dict):
    return normalized_claim

  claimant = lossq_claim_detail_value(
    raw_claim,
    "claimant",
    "claimant name",
    "injured worker",
    "injured party",
    "employee name",
    "plaintiff",
    "customer name",
    "third party name",
  )

  jurisdiction_state = lossq_claim_detail_value(
    raw_claim,
    "jurisdiction/state",
    "jurisdiction",
    "state",
    "venue state",
    "loss state",
  )

  adjuster = lossq_claim_detail_value(
    raw_claim,
    "adjuster",
    "adjuster/examiner",
    "examiner",
    "claim adjuster",
    "claim examiner",
    "file handler",
  )

  if claimant:
    normalized_claim["claimant"] = claimant

  if jurisdiction_state:
    normalized_claim["jurisdiction_state"] = jurisdiction_state
    normalized_claim["venue_state"] = normalized_claim.get("venue_state") or jurisdiction_state

  if adjuster:
    normalized_claim["adjuster"] = adjuster
    normalized_claim["examiner"] = normalized_claim.get("examiner") or adjuster

  return normalized_claim


# LOSSQ_UNIVERSAL_PROFILE_CLAIM_FINAL_NORMALIZER_V1
def lossq_universal_profile_claim_final_normalizer(parsed_claims=None, parsed_profile=None):
  """
  Final universal normalizer after all CSV/PDF/XLSX repairs.
  Keeps profile identity, producing agency, account number, policies, and claim rows consistent.
  No carrier, account, customer, or demo-file hardcoding.
  """
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm(value):
    return re.sub(r"[^a-z0-9]+", "_", clean(value).lower()).strip("_")

  def looks_like_policy(value):
    value = clean(value).upper()
    if not value:
      return False
    return lossq_looks_like_policy_but_not_account(value)

  def split_policy_numbers(value):
    raw = clean(value)
    if not raw:
      return []

    found = []
    pieces = re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", raw, flags=re.IGNORECASE)

    for piece in pieces:
      piece = clean(piece).upper()
      matches = re.findall(r"\b[A-Z]{1,8}[- ]?\d{2,6}[- ]?[A-Z0-9]{2,12}\b", piece)
      for match in matches:
        match = clean(match).upper().replace(" ", "-")
        if match and match not in found:
          found.append(match)

    return found

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0
    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", "-", ".", "-."}:
      return 0.0
    try:
      amount = float(raw)
      return -amount if neg else amount
    except Exception:
      return 0.0

  def first_profile_value(*keys):
    normalized = {norm(k): v for k, v in parsed_profile.items()}
    for key in keys:
      value = parsed_profile.get(key)
      if clean(value):
        return clean(value)

      value = normalized.get(norm(key))
      if clean(value):
        return clean(value)

    return ""

  # Business / insured name.
  insured_name = first_profile_value(
    "business_name",
    "insured_name",
    "named_insured",
    "insured",
    "applicant",
    "account_name",
    "company_name",
    "company",
  )

  if insured_name and insured_name.lower() not in {"business name not set", "not set", "unknown"}:
    parsed_profile["business_name"] = insured_name
    parsed_profile["insured_name"] = insured_name
    parsed_profile["named_insured"] = insured_name

  # Producing agency / broker.
  producing_agency = first_profile_value(
    "producing_agency",
    "agency_name",
    "agency",
    "broker",
    "brokerage",
    "producer",
    "producer_name",
  )

  if producing_agency and producing_agency.lower() not in {"agency not set", "not set", "unknown"}:
    parsed_profile["producing_agency"] = producing_agency
    parsed_profile["agency_name"] = producing_agency
    parsed_profile["producer"] = producing_agency

  # Policy numbers from policy fields and policy schedule.
  policy_candidates = []
  for key in [
    "policy_number",
    "main_policy",
    "policy_numbers",
    "account_number",
    "customer_number",
  ]:
    value = parsed_profile.get(key)
    if isinstance(value, list):
      for item in value:
        policy_candidates.extend(split_policy_numbers(item))
    else:
      policy_candidates.extend(split_policy_numbers(value))

  policies = parsed_profile.get("policies")
  if isinstance(policies, list):
    for item in policies:
      if isinstance(item, dict):
        policy_candidates.extend(
          split_policy_numbers(
            item.get("policy_number")
            or item.get("policy")
            or item.get("number")
          )
        )

  policy_numbers = []
  for policy in policy_candidates:
    policy = clean(policy).upper()
    if policy and policy not in policy_numbers:
      policy_numbers.append(policy)

  if policy_numbers:
    parsed_profile["policy_number"] = policy_numbers[0]
    parsed_profile["main_policy"] = policy_numbers[0]
    parsed_profile["policy_numbers"] = policy_numbers

  # Account number must not be a policy number.
  account_number = clean(parsed_profile.get("account_number"))
  if account_number and looks_like_policy(account_number):
    parsed_profile["account_number"] = ""

  customer_number = clean(parsed_profile.get("customer_number"))
  if customer_number and looks_like_policy(customer_number):
    parsed_profile["customer_number"] = ""

  # Rebuild policy rows while preserving line/carrier/dates/counts already found.
  if policy_numbers:
    existing_rows = parsed_profile.get("policies") if isinstance(parsed_profile.get("policies"), list) else []
    rebuilt = []

    for policy in policy_numbers:
      existing = None
      for row in existing_rows:
        if isinstance(row, dict) and clean(row.get("policy_number")).upper() == policy:
          existing = dict(row)
          break

      if existing is None:
        existing = {"policy_number": policy}

      existing["policy_number"] = policy
      existing["carrier"] = existing.get("carrier") or parsed_profile.get("carrier") or parsed_profile.get("carrier_name") or ""
      existing["writing_carrier"] = existing.get("writing_carrier") or parsed_profile.get("writing_carrier") or parsed_profile.get("carrier_name") or ""
      existing["effective_date"] = existing.get("effective_date") or parsed_profile.get("effective_date") or ""
      existing["expiration_date"] = existing.get("expiration_date") or parsed_profile.get("expiration_date") or ""
      rebuilt.append(existing)

    parsed_profile["policies"] = rebuilt

  # Normalize claim rows so DB save logic sees them as real claims.
  normalized_claims = []

  for claim in parsed_claims:
    if not isinstance(claim, dict):
      continue

    claim = dict(claim)

    claim_number = clean(
      claim.get("claim_number")
      or claim.get("claim #")
      or claim.get("claim_no")
      or claim.get("claim")
      or claim.get("claim_id")
    )

    if not claim_number:
      continue

    paid = money(claim.get("paid") or claim.get("total_paid"))
    reserve = money(claim.get("reserve") or claim.get("total_reserve"))
    incurred = money(
      claim.get("total_incurred")
      or claim.get("incurred")
      or claim.get("total")
      or claim.get("gross_incurred")
    )

    if incurred <= 0 and (paid or reserve):
      incurred = paid + reserve

    claim["claim_number"] = claim_number
    claim["paid"] = paid
    claim["reserve"] = reserve
    claim["total_incurred"] = incurred
    claim["incurred"] = incurred

    claim["date_of_loss"] = clean(
      claim.get("date_of_loss")
      or claim.get("loss_date")
      or claim.get("date of loss")
    )

    claim["date_reported"] = clean(
      claim.get("date_reported")
      or claim.get("reported_date")
      or claim.get("date reported")
    )

    claim["claim_status"] = clean(
      claim.get("claim_status")
      or claim.get("status")
      or claim.get("open_closed")
    )

    claim["status"] = claim["claim_status"]

    claim["line_of_business"] = clean(
      claim.get("line_of_business")
      or claim.get("line")
      or claim.get("coverage")
      or claim.get("lob")
    )

    if not clean(claim.get("policy_number")) and policy_numbers:
      line = claim["line_of_business"].upper()
      if "WC" in line or "WORK" in line:
        claim["policy_number"] = next((p for p in policy_numbers if p.startswith("WC")), policy_numbers[0])
      elif "GL" in line or "GENERAL" in line or "LIAB" in line:
        claim["policy_number"] = next((p for p in policy_numbers if p.startswith(("GL", "CGL", "PL"))), policy_numbers[0])
      else:
        claim["policy_number"] = policy_numbers[0]

    if insured_name:
      claim["business_name"] = claim.get("business_name") or insured_name
      claim["named_insured"] = claim.get("named_insured") or insured_name

    if parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier"):
      claim["carrier_name"] = claim.get("carrier_name") or parsed_profile.get("carrier_name") or parsed_profile.get("writing_carrier")
      claim["writing_carrier"] = claim.get("writing_carrier") or parsed_profile.get("writing_carrier") or parsed_profile.get("carrier_name")

    normalized_claims.append(claim)

  if normalized_claims:
    parsed_claims = normalized_claims

  return parsed_claims, parsed_profile


# LOSSQ_FORCE_SECTION_CSV_CLAIMS_BEFORE_SAVE_V1
def lossq_force_section_csv_claims_before_save(file_path, parsed_claims=None, parsed_profile=None):
  """
  Universal section CSV extraction that runs immediately after parse_file,
  before the DB save loop.

  This is for files where profile rows appear first, then a CLAIMS DETAIL section.
  It does not hardcode insured, carrier, agency, or file names.
  """
  parsed_claims = parsed_claims if isinstance(parsed_claims, list) else []
  parsed_profile = parsed_profile if isinstance(parsed_profile, dict) else {}

  if not str(file_path or "").lower().endswith(".csv"):
    return parsed_claims, parsed_profile

  try:
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
      rows = list(csv.reader(handle))
  except Exception:
    return parsed_claims, parsed_profile

  if not rows:
    return parsed_claims, parsed_profile

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def norm(value):
    return re.sub(r"[^a-z0-9]+", "_", clean(value).lower()).strip("_")

  def money(value):
    raw = clean(value)
    if not raw:
      return 0.0

    neg = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    raw = re.sub(r"[^0-9.\-]", "", raw)

    if raw in {"", "-", ".", "-."}:
      return 0.0

    try:
      amount = float(raw)
      return -amount if neg else amount
    except Exception:
      return 0.0

  def split_policy_numbers(value):
    raw = clean(value)
    if not raw:
      return []

    found = []
    for piece in re.split(r"\s*(?:/|,|;|\||\band\b|\+)\s*", raw, flags=re.IGNORECASE):
      piece = clean(piece).upper()
      for match in re.findall(r"\b[A-Z]{1,8}[- ]?\d{2,6}[- ]?[A-Z0-9]{2,12}\b", piece):
        match = clean(match).upper().replace(" ", "-")
        if match and match not in found:
          found.append(match)

    return found

  def header_index(headers, candidates):
    normalized_headers = [norm(header) for header in headers]
    normalized_candidates = [norm(candidate) for candidate in candidates]

    for index, header in enumerate(normalized_headers):
      if header in normalized_candidates:
        return index

    for index, header in enumerate(normalized_headers):
      for candidate in normalized_candidates:
        if candidate and (candidate in header or header in candidate):
          return index

    return None

  def value_at(row, index):
    if index is None:
      return ""
    if index < 0 or index >= len(row):
      return ""
    return clean(row[index])

  def line_name(value):
    value = clean(value).upper()

    if value in {"WC", "WORKERS COMP", "WORKERS COMPENSATION"} or "WORK" in value:
      return "Workers Compensation"

    if value in {"GL", "GENERAL LIABILITY"} or "GENERAL" in value:
      return "General Liability"

    if "PROD" in value or "PRODUCT" in value:
      return "Products Liability"

    if "AUTO" in value:
      return "Commercial Auto"

    return clean(value) or "Unknown"

  def choose_policy(line, policy_numbers):
    if not policy_numbers:
      return ""

    upper_line = clean(line).upper()

    if "WC" in upper_line or "WORK" in upper_line:
      return next((p for p in policy_numbers if str(p).upper().startswith("WC")), policy_numbers[0])

    if "GL" in upper_line or "GENERAL" in upper_line:
      return next((p for p in policy_numbers if str(p).upper().startswith(("GL", "CGL"))), policy_numbers[0])

    if "PROD" in upper_line or "PRODUCT" in upper_line or "LIAB" in upper_line:
      return next((p for p in policy_numbers if str(p).upper().startswith(("GL", "CGL", "PL", "PROD"))), policy_numbers[0])

    return policy_numbers[0]

  section_claim_headers = {
    "claims_detail",
    "claim_detail",
    "claim_details",
    "claims",
    "loss_detail",
    "loss_details",
    "claim_listing",
    "claim_list",
  }

  stop_sections = {
    "loss_summary",
    "summary",
    "totals",
    "total",
    "exposure_summary",
    "premium_summary",
    "underwriting_notes",
    "notes",
  }

  labels = {}
  claims_header_index = None

  for index, row in enumerate(rows):
    first = norm(row[0] if row else "")

    if first in section_claim_headers:
      claims_header_index = index + 1
      break

    if len(row) >= 2:
      label = norm(row[0])
      value = clean(row[1])
      if label and value:
        labels[label] = value

  if claims_header_index is None or claims_header_index >= len(rows):
    return parsed_claims, parsed_profile

  def first_label(*names):
    for name in names:
      value = labels.get(norm(name))
      if clean(value):
        return clean(value)
    return ""

  profile = dict(parsed_profile)

  insured_name = first_label(
    "insured name",
    "named insured",
    "insured",
    "applicant",
    "account name",
    "company name",
    "business name",
  )

  if insured_name:
    profile["business_name"] = insured_name
    profile["insured_name"] = insured_name
    profile["named_insured"] = insured_name

  carrier = first_label("carrier", "writing carrier", "insurance carrier")
  if carrier:
    profile["carrier"] = carrier
    profile["carrier_name"] = carrier
    profile["writing_carrier"] = carrier

  producing_agency = first_label(
    "producing agency",
    "agency",
    "agency name",
    "broker",
    "brokerage",
    "producer",
    "producer name",
  )

  if producing_agency:
    profile["producing_agency"] = producing_agency
    profile["agency_name"] = producing_agency
    profile["producer"] = producing_agency

  policy_numbers = split_policy_numbers(first_label("policy number", "policy numbers", "policy no", "policy"))

  if policy_numbers:
    profile["policy_number"] = policy_numbers[0]
    profile["main_policy"] = policy_numbers[0]
    profile["policy_numbers"] = policy_numbers

  period = first_label("policy period", "policy term", "coverage period")
  period_dates = re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", period)

  if len(period_dates) >= 1:
    profile["effective_date"] = profile.get("effective_date") or period_dates[0]
  if len(period_dates) >= 2:
    profile["expiration_date"] = profile.get("expiration_date") or period_dates[1]

  evaluation_date = first_label("evaluation date", "valuation date", "as of date", "loss run date")
  if evaluation_date:
    profile["evaluation_date"] = evaluation_date

  account_number = first_label("account number", "account no", "customer number", "client number")
  if account_number and (lossq_is_true_account_identifier(account_number) or not split_policy_numbers(account_number)):
    profile["account_number"] = account_number
  else:
    profile["account_number"] = ""

  headers = [clean(value) for value in rows[claims_header_index]]

  claim_i = header_index(headers, ["claim #", "claim number", "claim no", "claim"])
  dol_i = header_index(headers, ["date of loss", "loss date", "dol"])
  reported_i = header_index(headers, ["date reported", "reported date", "date_reported"])
  claimant_i = header_index(headers, ["claimant", "claimant name", "injured worker", "party"])
  line_i = header_index(headers, ["line", "line of business", "coverage", "lob"])
  desc_i = header_index(headers, ["description", "loss description", "cause", "claim description"])
  status_i = header_index(headers, ["status", "claim status", "open closed", "open/closed"])
  incurred_i = header_index(headers, ["total incurred", "incurred", "total"])
  paid_i = header_index(headers, ["paid", "total paid"])
  reserve_i = header_index(headers, ["reserve", "total reserve", "outstanding reserve"])
  subro_i = header_index(headers, ["subrogation", "subro", "recovery"])

  repaired_claims = []

  for row in rows[claims_header_index + 1:]:
    first = norm(row[0] if row else "")

    if first in stop_sections:
      break

    if not any(clean(cell) for cell in row):
      continue

    claim_number = value_at(row, claim_i)
    if not claim_number:
      continue

    raw_line = value_at(row, line_i)
    clean_line = line_name(raw_line)
    paid = money(value_at(row, paid_i))
    reserve = money(value_at(row, reserve_i))
    incurred = money(value_at(row, incurred_i))

    if incurred <= 0 and (paid or reserve):
      incurred = paid + reserve

    policy_number = choose_policy(raw_line, policy_numbers)

    claim = {
      "claim_number": claim_number,
      "policy_number": policy_number,
      "date_of_loss": value_at(row, dol_i),
      "date_reported": value_at(row, reported_i),
      "claimant": value_at(row, claimant_i),
      "line_of_business": clean_line,
      "policy_type": clean_line,
      "coverage": clean_line,
      "description": value_at(row, desc_i),
      "loss_description": value_at(row, desc_i),
      "claim_status": value_at(row, status_i),
      "status": value_at(row, status_i),
      "total_incurred": incurred,
      "incurred": incurred,
      "paid": paid,
      "total_paid": paid,
      "reserve": reserve,
      "total_reserve": reserve,
      "subrogation": value_at(row, subro_i),
      "business_name": insured_name,
      "named_insured": insured_name,
      "carrier_name": carrier,
      "writing_carrier": carrier,
    }

    repaired_claims.append(claim)

  if repaired_claims:
    profile["claims"] = repaired_claims
    profile["parsed_claims"] = repaired_claims

    line_groups = {}
    for claim in repaired_claims:
      policy_number = claim.get("policy_number") or ""
      if not policy_number:
        continue

      if policy_number not in line_groups:
        line_groups[policy_number] = {
          "policy_number": policy_number,
          "line_of_business": claim.get("line_of_business") or "Unknown",
          "policy_type": claim.get("line_of_business") or "Unknown",
          "carrier": carrier,
          "writing_carrier": carrier,
          "effective_date": profile.get("effective_date") or "",
          "expiration_date": profile.get("expiration_date") or "",
          "claim_count": 0,
          "total_incurred": 0.0,
        }

      line_groups[policy_number]["claim_count"] += 1
      line_groups[policy_number]["total_incurred"] += money(claim.get("total_incurred"))

    if line_groups:
      profile["policies"] = list(line_groups.values())

    print("LOSSQ_FORCE_SECTION_CSV_CLAIMS_BEFORE_SAVE:", {"claims": len(repaired_claims), "policies": len(profile.get("policies") or [])})
    return repaired_claims, profile

  return parsed_claims, profile


# LOSSQ_UPLOAD_ROOT_CAUSE_DEBUG_V1
def lossq_debug_upload_snapshot(stage, parsed_claims=None, parsed_profile=None, extra=None):
  try:
    claims = parsed_claims if isinstance(parsed_claims, list) else []
    profile = parsed_profile if isinstance(parsed_profile, dict) else {}

    sample_claims = []
    for claim in claims[:5]:
      if isinstance(claim, dict):
        sample_claims.append({
          "claim_number": claim.get("claim_number") or claim.get("claim #") or claim.get("claim"),
          "policy_number": claim.get("policy_number"),
          "line_of_business": claim.get("line_of_business") or claim.get("line") or claim.get("coverage"),
          "status": claim.get("claim_status") or claim.get("status"),
          "paid": claim.get("paid") or claim.get("total_paid"),
          "reserve": claim.get("reserve") or claim.get("total_reserve"),
          "total_incurred": claim.get("total_incurred") or claim.get("incurred"),
          "keys": sorted(list(claim.keys()))[:30],
        })

    print("LOSSQ_UPLOAD_DEBUG_SNAPSHOT:", {
      "stage": stage,
      "claim_count": len(claims),
      "sample_claims": sample_claims,
      "profile": {
        "id": profile.get("id"),
        "business_name": profile.get("business_name"),
        "insured_name": profile.get("insured_name"),
        "named_insured": profile.get("named_insured"),
        "carrier_name": profile.get("carrier_name"),
        "writing_carrier": profile.get("writing_carrier"),
        "producing_agency": profile.get("producing_agency"),
        "agency_name": profile.get("agency_name"),
        "account_number": profile.get("account_number"),
        "customer_number": profile.get("customer_number"),
        "policy_number": profile.get("policy_number"),
        "main_policy": profile.get("main_policy"),
        "policy_numbers": profile.get("policy_numbers"),
        "policies": profile.get("policies"),
      },
      "extra": extra or {},
    })
  except Exception as exc:
    print("LOSSQ_UPLOAD_DEBUG_SNAPSHOT_FAILED:", str(exc)[:500])


# LOSSQ_TRUE_ACCOUNT_IDENTIFIER_HELPER_V1
def lossq_is_true_account_identifier(value):
  text = str(value or "").strip().upper()
  if not text:
    return False

  # Universal account/customer/client identifiers should never be treated as policies.
  return bool(
    re.search(r"\b(ACCT|ACCOUNT|CUST|CUSTOMER|CLIENT)\b", text)
    or re.search(r"[-_ ](ACCT|ACCOUNT|CUST|CUSTOMER|CLIENT)[-_ ]", text)
  )


def lossq_looks_like_policy_but_not_account(value):
  # LOSSQ_POLICY_ACCOUNT_RECURSION_FIX_V2
  # This helper must never call itself. It only decides whether a value
  # looks like a policy number while protecting true account/customer IDs.
  import re

  text = str(value or "").strip().upper()
  if not text:
    return False

  compact = re.sub(r"[^A-Z0-9]", "", text)

  if lossq_is_true_account_identifier(text):
    return False

  if re.search(r"\b(ACCT|ACCOUNT|CUST|CUSTOMER|CLIENT)\b", text):
    return False

  if compact.startswith(("ACCT", "ACCOUNT", "CUST", "CUSTOMER", "CLIENT")):
    return False

  policy_prefix_pattern = r"^(GL|WC|CA|AUTO|BOP|PROP|CP|LIQ|UMB|UM|PL|EPLI|CY|CARGO|IM|DO|DNO|BPP)[-_/ ]?\d"

  if re.match(policy_prefix_pattern, text):
    return True

  if re.match(r"^(POL|POLICY)[-_/ ]?\d", text):
    return True

  if re.search(r"[A-Z]", text) and re.search(r"\d", text) and re.search(r"[-_/]", text):
    return True

  return False


# LOSSQ_CSV_TRUE_INSURED_NAME_HELPER_V1
def lossq_extract_true_insured_name_from_upload_csv(file_path):
  """
  Reads the actual insured/business name from a clean flat CSV row.
  Prevents header labels like 'Account Number' from being saved as insured.
  """
  import csv
  import re

  def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())

  def key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())

  insured_aliases = {
    "businessname",
    "namedinsured",
    "insured",
    "insuredname",
    "accountname",
    "clientname",
    "customername",
  }

  blocked_values = {
    "account number",
    "customer number",
    "client number",
    "policy number",
    "line of business",
    "carrier",
    "producing agency",
    "business name",
    "named insured",
    "insured",
  }

  try:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
      try:
        with open(file_path, "r", newline="", encoding=encoding, errors="ignore") as handle:
          reader = csv.DictReader(handle)
          for row in reader:
            lookup = {key(k): clean(v) for k, v in (row or {}).items()}

            for alias in insured_aliases:
              value = clean(lookup.get(alias))
              if not value:
                continue

              if value.lower() in blocked_values:
                continue

              if value.upper().startswith(("ACCT", "ACCOUNT", "POLICY")):
                continue

              if len(value) >= 3:
                return value
        break
      except Exception:
        continue
  except Exception:
    return ""

  return ""


async def save_uploaded_files(files, policy_number, db, current_user):
  # LOSSQ_SAVE_UPLOADED_FILES_SAFE_DEFAULTS_V1
  # Initialize shared parse variables inside the actual upload save loop.
  # This protects PDF, CSV, XLSX, and XLS from undefined rescue variables.
  parsed_claims = []
  parsed_profile = {}
  rescue_claims = []
  rescue_profile = {}

  ensure_claim_timeline_columns(db)
  ensure_claim_detail_columns(db)
  ensure_account_profile_columns(db)
  lossq_beta_upload_usage_guard(db, current_user, len(files or []))

  total_saved = 0
  total_duplicates_skipped = 0
  uploaded_files = []
  all_parsed_claims = []
  direct_profile = {}

  upload_session_id = datetime.now().strftime("%Y%m%d%H%M%S")
  clean_input_policy = str(policy_number or "").strip()

  for file in files:
    # LOSSQ_SAFE_UPLOAD_FILENAME_IN_SAVE_LOOP_V1
    safe_upload_filename = await validate_upload_file_security(file)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_filename = (safe_upload_filename or "loss_run.pdf").replace(" ", "_")
    file_path = os.path.join(UPLOAD_DIR, f"{timestamp}_{safe_filename}")

    with open(file_path, "wb") as buffer:
      shutil.copyfileobj(file.file, buffer)

    try:
      parsed_claims, parsed_profile = parse_file(file_path, safe_upload_filename or safe_filename)
      lossq_debug_upload_snapshot(
        "after_parse_file",
        parsed_claims,
        parsed_profile,
        {"filename": safe_upload_filename or safe_filename},
      )
      # LOSSQ_FORCE_SECTION_CSV_CLAIMS_BEFORE_SAVE_CALL_V1
      parsed_claims, parsed_profile = lossq_force_section_csv_claims_before_save(
        file_path,
        parsed_claims,
        parsed_profile,
      )
      lossq_debug_upload_snapshot(
        "after_force_section_csv_claims_before_save",
        parsed_claims,
        parsed_profile,
        {"filename": safe_upload_filename or safe_filename},
      )

      # LOSSQ_DIRECT_FILE_EXPOSURE_CAPTURE_V1
      direct_exposure_inputs = lossq_extract_exposure_inputs_directly_from_file(file_path)
      if direct_exposure_inputs:
        if not isinstance(parsed_profile, dict):
          parsed_profile = {}
        parsed_profile.update({k: v for k, v in direct_exposure_inputs.items() if v not in ("", None, [], {})})
        parsed_profile["exposure_inputs"] = direct_exposure_inputs
        parsed_profile["exposures"] = direct_exposure_inputs
      parsed_profile = lossq_section_csv_apply_profile_date_repair(file_path, parsed_profile)
      parsed_profile = lossq_csv_label_pair_profile_repair(file_path, parsed_profile)
      parsed_profile = lossq_pdf_profile_repair(file_path, parsed_profile)
      parsed_profile = lossq_global_profile_cleanup(parsed_profile)
      parsed_profile = lossq_global_profile_cleanup(parsed_profile)
    except Exception as exc:
      # LOSSQ_UPLOAD_LOOP_CLEAN_FLAT_CSV_RESCUE_V1
      # If the old/general parser fails on a clean flat CSV, rescue it here
      # instead of failing the entire upload.
      rescue_claims = []
      rescue_profile = {}

      try:
        # LOSSQ_UPLOAD_LOOP_PARSED_CLAIMS_ALIAS_V1
        parsed_claims = locals().get("parsed_claims", locals().get("claims", [])) if isinstance(locals().get("parsed_claims", locals().get("claims", [])), list) else []
        lower_upload_name = str(safe_upload_filename or safe_filename or "").lower()
        if lower_upload_name.endswith(".csv"):
          if "lossq_parse_clean_flat_csv_v1" in globals():
            rescue_claims, rescue_profile = lossq_parse_clean_flat_csv_v1(file_path)

          if not rescue_claims:
            # LOSSQ_UPLOAD_LOOP_PARSED_CLAIMS_ALIAS_V2
            # LOSSQ_UPLOAD_LOOP_CSV_RESCUE_ONLY_FOR_CSV_V5
            _lossq_upload_file_obj = locals().get("file", None)
            _lossq_original_filename = getattr(_lossq_upload_file_obj, "filename", "")
            _lossq_current_upload_name = str(
              locals().get("safe_upload_filename")
              or locals().get("safe_filename")
              or _lossq_original_filename
              or locals().get("filename", "")
              or locals().get("file_path", "")
            ).lower()

            if not _lossq_current_upload_name.endswith(".csv"):
              print("LOSSQ_UPLOAD_LOOP_SKIP_CSV_RESCUE_FOR_NON_CSV_V5:", {
                "filename": _lossq_current_upload_name,
                "error": str(exc)[:500],
              })
              raise exc

            parsed_claims = locals().get("parsed_claims", locals().get("claims", []))
            if not isinstance(parsed_claims, list):
              parsed_claims = []
            parsed_profile = locals().get("parsed_profile", locals().get("profile", {}))
            if not isinstance(parsed_profile, dict):
              parsed_profile = {}
            # LOSSQ_UPLOAD_LOOP_CLEAN_FLAT_CSV_RESCUE_SAFE_ARGS_V3
            rescue_claims, rescue_profile = lossq_clean_standard_csv_override(
              file_path,
              parsed_claims,
              parsed_profile,
            )

          if rescue_claims or rescue_profile:
            parsed_claims = rescue_claims or []
            parsed_profile = rescue_profile or parsed_profile or {}
            print("LOSSQ_UPLOAD_LOOP_CLEAN_FLAT_CSV_RESCUED:", {
              "claims": len(parsed_claims or []),
              "profile_keys": list((parsed_profile or {}).keys())[:20],
            })
          else:
            print("LOSSQ_UPLOAD_LOOP_CLEAN_FLAT_CSV_RESCUE_EMPTY:", str(exc)[:500])
            raise exc
        else:
          raise exc
      except Exception as rescue_exc:
        print("LOSSQ_UPLOAD_LOOP_CLEAN_FLAT_CSV_RESCUE_FAILED:", str(rescue_exc)[:1000])
        raise HTTPException(
          status_code=400,
          detail={
            "message": "Loss run could not be parsed cleanly. Please upload a valid PDF, Excel, or CSV loss run.",
            "error": str(rescue_exc)[:300],
            "stage": "parse_file",
          },
        )

    # LOSSQ_FINAL_UPLOAD_CLAIM_PROFILE_CLEANUP_CALL_V1
    parsed_claims, parsed_profile = lossq_final_upload_claim_profile_cleanup_v1(
      file_path,
      parsed_claims,
      parsed_profile,
    )

    # LOSSQ_UNIVERSAL_SECTION_CSV_CLAIMS_PROFILE_REPAIR_CALL_V1
    parsed_claims, parsed_profile = lossq_universal_section_csv_claims_profile_repair(
      file_path,
      parsed_claims,
      parsed_profile,
    )

    # LOSSQ_UNIVERSAL_SECTION_CSV_CLAIMS_PROFILE_REPAIR_CALL_V2
    parsed_claims, parsed_profile = lossq_universal_section_csv_claims_profile_repair_v2(
      file_path,
      parsed_claims,
      parsed_profile,
    )
    lossq_debug_upload_snapshot(
      "after_universal_section_csv_v2",
      parsed_claims,
      parsed_profile,
      {"filename": safe_upload_filename or safe_filename if "safe_upload_filename" in locals() else ""},
    )

    # LOSSQ_APPLY_LIVE_SECTION_BASED_CSV_REPAIR_V1
    parsed_claims, parsed_profile = lossq_live_repair_section_csv_upload(
      file_path,
      parsed_claims,
      parsed_profile,
    )

    # LOSSQ_CLEAN_STANDARD_CSV_ROW_POLICY_OVERRIDE_V1
    parsed_claims, parsed_profile = lossq_clean_standard_csv_override(
      file_path,
      parsed_claims,
      parsed_profile,
    )


    # LOSSQ_FINAL_FLAT_CSV_SAVE_PRESERVE_V1
    # Final save-level CSV pass. Some later upload/profile steps can rehydrate
    # filename fallback profile values or old claim defaults. Re-read clean flat
    # CSV values immediately before final exposure/profile work so account name,
    # claim status, and zero-claim policy rows survive to the dashboard.
    try:
      final_csv_claims, final_csv_profile = lossq_clean_standard_csv_override(
        file_path,
        parsed_claims,
        parsed_profile,
      )

      if isinstance(final_csv_profile, dict) and final_csv_profile:
        for key in [
          "business_name",
          "insured_name",
          "named_insured",
          "account_name",
          "carrier_name",
          "writing_carrier",
          "carrier",
          "producing_agency",
          "agency_name",
          "producer",
          "account_number",
          "customer_number",
          "effective_date",
          "expiration_date",
          "evaluation_date",
          "valuation_date",
          "policy_number",
          "main_policy",
          "policy_numbers",
          "policies",
          "policy_schedule",
        ]:
          value = final_csv_profile.get(key)
          if value not in (None, "", [], {}):
            parsed_profile[key] = value

      if isinstance(final_csv_claims, list) and final_csv_claims:
        parsed_claims = final_csv_claims
        parsed_profile["claims"] = final_csv_claims
        parsed_profile["parsed_claims"] = final_csv_claims

      print("LOSSQ_FINAL_FLAT_CSV_SAVE_PRESERVE_V1:", {
        "business_name": parsed_profile.get("business_name"),
        "claims": len(parsed_claims or []),
        "policies": len(parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []),
        "policy_numbers": parsed_profile.get("policy_numbers"),
        "statuses": [c.get("status") for c in (parsed_claims or [])[:10] if isinstance(c, dict)],
      })
    except Exception as final_csv_exc:
      print("LOSSQ_FINAL_FLAT_CSV_SAVE_PRESERVE_V1_ERROR:", str(final_csv_exc)[:500])

    # LOSSQ_REAPPLY_DIRECT_EXPOSURE_AFTER_CSV_REPAIRS_V1
    # Re-apply direct CSV/XLSX exposure values after CSV claim/profile repairs so they are not lost.
    direct_exposure_inputs_after_csv_repairs = lossq_extract_exposure_inputs_directly_from_file(file_path)
    if direct_exposure_inputs_after_csv_repairs:
      if not isinstance(parsed_profile, dict):
        parsed_profile = {}
      parsed_profile.update({
        k: v for k, v in direct_exposure_inputs_after_csv_repairs.items()
        if v not in ("", None, [], {})
      })
      parsed_profile["exposure_inputs"] = direct_exposure_inputs_after_csv_repairs
      parsed_profile["exposures"] = direct_exposure_inputs_after_csv_repairs
      print("LOSSQ_DIRECT_EXPOSURE_REAPPLIED_AFTER_CSV_REPAIRS:", direct_exposure_inputs_after_csv_repairs)

    # LOSSQ_UNIVERSAL_PRODUCING_AGENCY_EXTRACTION_V1
    upload_agency_name = lossq_header_agency_from_csv(file_path) or lossq_universal_agency_from_csv(file_path)
    if upload_agency_name:
      print("LOSSQ_AGENCY_SELECTED_FROM_UPLOAD:", upload_agency_name)

    # LOSSQ_CLEAN_PROFILE_POLICY_SCHEDULE_ROWS_V1
    parsed_profile = lossq_clean_profile_policy_schedule_rows(parsed_profile, parsed_claims)

    # LOSSQ_APPLY_LINE_OF_BUSINESS_FROM_POLICY_PREFIX_V1
    parsed_claims, parsed_profile = lossq_apply_line_of_business_from_policy_prefix(parsed_claims, parsed_profile)

    # LOSSQ_UNIVERSAL_CSV_ACCOUNT_EXPOSURE_CLAIM_DETAIL_OVERLAY_V1
    parsed_claims, parsed_profile = lossq_universal_csv_account_exposure_claim_detail_overlay(
      file_path=file_path,
      parsed_claims=parsed_claims,
      parsed_profile=parsed_profile,
    )


    # LOSSQ_FINAL_PROFILE_DATES_FROM_POLICIES_V1
    parsed_profile = lossq_final_profile_dates_from_policies(parsed_profile)
    # LOSSQ_UNIVERSAL_PROFILE_CLAIM_FINAL_NORMALIZER_CALL_V1
    parsed_claims, parsed_profile = lossq_universal_profile_claim_final_normalizer(
      parsed_claims,
      parsed_profile,
    )
    lossq_debug_upload_snapshot(
      "after_final_profile_claim_normalizer",
      parsed_claims,
      parsed_profile,
      {"filename": safe_upload_filename or safe_filename if "safe_upload_filename" in locals() else ""},
    )
    parsed_profile = lossq_universal_profile_identity_policy_cleanup(parsed_profile)
    if upload_agency_name:
      parsed_profile = parsed_profile or {}
      parsed_profile["agency_name"] = upload_agency_name
      parsed_profile["producing_agency"] = upload_agency_name
      parsed_profile["producer"] = upload_agency_name

    # LOSSQ_EXCEL_ACCOUNT_CUSTOMER_PRECEDENCE_CALL_V1
    parsed_profile, direct_profile = lossq_excel_account_customer_precedence_v1(
      file_path,
      parsed_profile,
      direct_profile if "direct_profile" in locals() and isinstance(direct_profile, dict) else {},
    )

    if "profile_data" in locals() and isinstance(profile_data, dict):
      for _lossq_account_key in ("account_number", "account_no", "account", "customer_number"):
        _lossq_account_value = parsed_profile.get(_lossq_account_key)
        if _lossq_account_value not in ("", None, [], {}):
          profile_data[_lossq_account_key] = _lossq_account_value
    file_policy_number = clean_input_policy
    file_account_key_for_claims = ""

    claim_policy_number = ""
    # LOSSQ_UNIVERSAL_CSV_SECTION_OVERLAY_V2
    parsed_claims, parsed_profile = lossq_universal_csv_section_overlay_v2(
      file_path=file_path,
      parsed_claims=parsed_claims,
      parsed_profile=parsed_profile,
    )

    # LOSSQ_BETA_FILTER_AND_PURGE_BEFORE_SAVE_V1
    parsed_claims, lossq_beta_removed_rows = lossq_beta_filter_claim_rows(parsed_claims)

    # LOSSQ_UNIVERSAL_CSV_SECTION_OVERLAY_V2_REAPPLY_AFTER_BETA
    parsed_claims, parsed_profile = lossq_universal_csv_section_overlay_v2(
      file_path=file_path,
      parsed_claims=parsed_claims,
      parsed_profile=parsed_profile,
    )


  # LOSSQ_UNIVERSAL_CLAIM_NUMBER_FILTER_V1
    if lossq_beta_removed_rows:
      print("LOSSQ_BETA_FILTER_REMOVED_ROWS:", lossq_beta_removed_rows[:10])

    lossq_beta_policy_keys = lossq_beta_collect_upload_policy_keys(
      parsed_profile,
      parsed_claims,
      file_policy_number,
    )
    lossq_beta_cleanup = lossq_beta_purge_prior_upload_data(
      db,
      current_user,
      lossq_beta_policy_keys,
    )

    # LOSSQ_FINAL_CSV_ACCOUNT_AND_MISSING_CLAIMS_V4_BEFORE_SAVE_LOOP
    if str(file_path or "").lower().endswith(".csv"):
      parsed_claims, parsed_profile = lossq_v4_merge_csv_sections_before_save(
        file_path=file_path,
        parsed_claims=parsed_claims,
        parsed_profile=parsed_profile,
      )

    for claim_data in parsed_claims:
      claim_policy = clean_profile_value(claim_data.get("policy_number"))
      if claim_policy:
        claim_policy_number = claim_policy
        break

    if parsed_profile:
      parsed_policy = clean_profile_value(parsed_profile.get("policy_number"))
      parsed_account = clean_profile_value(
        parsed_profile.get("account_number") or parsed_profile.get("customer_number")
      )

      # LOSSQ_DUPLICATE_REHOME_TO_PARSED_ACCOUNT_KEY_V1
      # Use the extracted account/customer key to re-home duplicate claims during upload.
      if parsed_account and not is_bad_policy_key_for_upload(parsed_account):
        file_account_key_for_claims = parsed_account

      # Important:
      # Prefer the actual policy number found on claim rows.
      # Do not let customer/account number override real claim policy number.
      if parsed_policy:
        file_policy_number = parsed_policy
      elif claim_policy_number:
        file_policy_number = claim_policy_number
      elif parsed_account:
        file_policy_number = parsed_account

      parsed_profile = lossq_clean_exposure_limits_field(parsed_profile)

      for key, value in parsed_profile.items():
        if key in ["policies", "validation", "raw_text_preview"]:
          direct_profile[key] = value
          continue

        if value and not direct_profile.get(key):
          direct_profile[key] = value

    if not file_policy_number and claim_policy_number:
      file_policy_number = claim_policy_number

    if not file_policy_number:
      file_policy_number = (
        direct_profile.get("policy_number")
        or direct_profile.get("account_number")
        or f"UPLOAD-{upload_session_id}-{len(uploaded_files) + 1}"
      )

    if not direct_profile.get("policy_number"):
      direct_profile["policy_number"] = file_policy_number

    # LOSSQ_AUTHORITATIVE_FLAT_CSV_BEFORE_CLAIM_SAVE_CALL_V1
    try:
      parsed_claims, parsed_profile, _lossq_unused_profile_data = lossq_apply_authoritative_flat_csv_snapshot_v1(
        file_path,
        parsed_claims,
        parsed_profile,
        None,
      )
    except Exception as authoritative_csv_claim_exc:
      print("LOSSQ_AUTHORITATIVE_FLAT_CSV_BEFORE_CLAIM_SAVE_CALL_V1_ERROR:", str(authoritative_csv_claim_exc)[:500])

    lossq_debug_upload_snapshot(
      "before_all_parsed_claims_extend",
      parsed_claims,
      parsed_profile,
      {"filename": safe_upload_filename or safe_filename if "safe_upload_filename" in locals() else ""},
    )
    # LOSSQ_FINAL_REAPPLY_CSV_OVERLAY_BEFORE_SAVE_V3
    if str(file_path or "").lower().endswith(".csv"):
      try:
        if callable(globals().get("lossq_universal_csv_section_overlay_v2")):
          parsed_claims, parsed_profile = lossq_universal_csv_section_overlay_v2(
            file_path=file_path,
            parsed_claims=parsed_claims,
            parsed_profile=parsed_profile,
          )
        parsed_profile = lossq_final_repair_profile_account_and_exposures_v3(parsed_profile)
      except Exception as exc:
        print("LOSSQ_FINAL_REAPPLY_CSV_OVERLAY_BEFORE_SAVE_V3_ERROR", str(exc)[:200])

    # LOSSQ_FINAL_CSV_ACCOUNT_AND_MISSING_CLAIMS_V4_BEFORE_EXTEND
    if str(file_path or "").lower().endswith(".csv"):
      parsed_claims, parsed_profile = lossq_v4_merge_csv_sections_before_save(
        file_path=file_path,
        parsed_claims=parsed_claims,
        parsed_profile=parsed_profile,
      )

    # LOSSQ_PDF_FINAL_BUSINESS_NAME_REPAIR_CALL_V1
    parsed_profile = lossq_pdf_final_business_name_repair_v1(file_path, parsed_profile)

    # LOSSQ_PDF_FULL_CLAIM_BLOCK_EXTRACT_BEFORE_SAVE_CALL_V1
    parsed_claims, parsed_profile = lossq_pdf_full_claim_block_extract_before_save_v1(
      file_path,
      parsed_claims,
      parsed_profile,
    )


    # LOSSQ_PDF_CLEAN_TABLE_CLAIM_REPAIR_CALL_V1
    parsed_claims, parsed_profile = lossq_pdf_clean_table_claim_repair_v1(
      file_path,
      parsed_claims,
      parsed_profile,
    )

    # LOSSQ_PDF_WIDE_CLAIMS_DETAIL_TABLE_REPAIR_CALL_V1
    parsed_claims, parsed_profile, direct_profile = lossq_pdf_wide_claims_detail_table_repair_v1(
      file_path,
      parsed_claims,
      parsed_profile,
      direct_profile,
    )


    # LOSSQ_PDF_SAVE_TIME_BUSINESS_NAME_REPAIR_CALL_V2
    parsed_profile, direct_profile = lossq_pdf_save_time_business_name_repair_v2(
      file_path,
      parsed_profile,
      parsed_claims,
      direct_profile,
    )


    # LOSSQ_PDF_FINAL_CARRIER_LABEL_REPAIR_CALL_V1
    parsed_profile, direct_profile = lossq_pdf_final_carrier_label_repair_v1(
      file_path,
      parsed_profile,
      direct_profile,
    )

    # LOSSQ_PDF_ACCOUNT_PROFILE_GRID_REPAIR_CALL_V1
    parsed_profile, direct_profile = lossq_pdf_account_profile_grid_repair_v1(
      file_path,
      parsed_profile,
      direct_profile,
    )


    # LOSSQ_PDF_MESSY_BLOCK_MINI_REPAIR_CALL_V1
    parsed_claims, parsed_profile, direct_profile = lossq_pdf_messy_block_mini_repair_v1(
      file_path,
      parsed_claims,
      parsed_profile,
      direct_profile,
    )

    # LOSSQ_PDF_MESSY_BLOCK_MINI_REPAIR_PROFILE_SYNC_V1
    if "profile_data" in locals() and isinstance(profile_data, dict) and isinstance(parsed_profile, dict):
      for _lossq_pdf_key in (
        "business_name",
        "insured_name",
        "named_insured",
        "account_name",
        "account_number",
        "customer_number",
        "carrier_name",
        "writing_carrier",
        "carrier",
        "producing_agency",
        "agency_name",
        "producer",
        "effective_date",
        "expiration_date",
        "evaluation_date",
        "valuation_date",
        "current_premium",
        "expiring_premium",
        "target_renewal_premium",
        "policies",
        "policy_schedule",
        "policy_numbers",
        "claim_count",
        "total_claims",
        "open_claims",
        "closed_claims",
        "total_paid",
        "total_reserve",
        "total_incurred",
      ):
        _lossq_pdf_value = parsed_profile.get(_lossq_pdf_key)
        if _lossq_pdf_value not in ("", None, [], {}):
          profile_data[_lossq_pdf_key] = _lossq_pdf_value
    # LOSSQ_PDF_VERTICAL_POLICY_SECTION_REPAIR_CALL_V1
    if "direct_profile" not in locals() or not isinstance(direct_profile, dict):
      direct_profile = {}
    parsed_claims, parsed_profile, direct_profile = lossq_pdf_vertical_policy_section_repair_v1(
      file_path,
      parsed_claims,
      parsed_profile,
      direct_profile,
    )

    if "profile_data" in locals() and isinstance(profile_data, dict) and isinstance(parsed_profile, dict):
      for _lossq_vertical_key in (
        "business_name",
        "insured_name",
        "named_insured",
        "account_name",
        "account_number",
        "customer_number",
        "carrier_name",
        "writing_carrier",
        "carrier",
        "effective_date",
        "expiration_date",
        "evaluation_date",
        "valuation_date",
        "current_premium",
        "expiring_premium",
        "target_renewal_premium",
        "payroll",
        "revenue",
        "annual_revenue",
        "employee_count",
        "location_count",
        "locations",
        "locationCount",
        "bed_count",
        "beds",
        "policies",
        "policy_schedule",
        "policy_numbers",
        "policy_number",
        "main_policy",
        "claim_count",
        "total_claims",
        "open_claims",
        "closed_claims",
        "total_paid",
        "total_reserve",
        "total_incurred",
      ):
        _lossq_vertical_value = parsed_profile.get(_lossq_vertical_key)
        if _lossq_vertical_value not in ("", None, [], {}):
          profile_data[_lossq_vertical_key] = _lossq_vertical_value


    # LOSSQ_PDF_WIDE_CLAIMS_TABLE_FINAL_SAVE_RESCUE_CALL_V2
    parsed_claims, parsed_profile, direct_profile = lossq_pdf_wide_claims_table_final_save_rescue_v2(
      file_path,
      parsed_claims,
      parsed_profile,
      direct_profile,
    )

    all_parsed_claims.extend(parsed_claims)

    # LOSSQ_CANONICAL_UPLOAD_CLAIM_PURGE_V1
    # Before saving this upload, remove stale rows tied to the same uploaded claim numbers
    # or policy numbers. This prevents old bad rows from surviving after parser repairs.
    upload_claim_numbers = []
    upload_policy_keys = []

    for purge_claim in parsed_claims or []:
      if not isinstance(purge_claim, dict):
        continue

      purge_claim_number = str(
        purge_claim.get("claim_number")
        or purge_claim.get("Claim Number")
        or purge_claim.get("claim_no")
        or purge_claim.get("Claim No")
        or ""
      ).strip().upper()

      purge_policy_number = str(
        purge_claim.get("policy_number")
        or purge_claim.get("Policy Number")
        or purge_claim.get("policy_no")
        or purge_claim.get("Policy No")
        or purge_claim.get("policy")
        or ""
      ).strip().upper()

      if purge_claim_number and purge_claim_number != "UNKNOWN":
        upload_claim_numbers.append(purge_claim_number)

      if purge_policy_number and not is_bad_policy_key_for_upload(purge_policy_number):
        upload_policy_keys.append(purge_policy_number)

    upload_claim_numbers = sorted(set(upload_claim_numbers))
    upload_policy_keys = sorted(set(upload_policy_keys))

    purged_by_claim_number = 0
    purged_by_policy_number = 0

    if upload_claim_numbers:
      purged_by_claim_number = (
        db.query(Claim)
       .filter(Claim.organization_id == current_user["organization_id"])
       .filter(func.upper(func.trim(Claim.claim_number)).in_(upload_claim_numbers))
       .delete(synchronize_session=False)
      )

    if upload_policy_keys:
      purged_by_policy_number = (
        db.query(Claim)
       .filter(Claim.organization_id == current_user["organization_id"])
       .filter(func.upper(func.trim(Claim.policy_number)).in_(upload_policy_keys))
       .delete(synchronize_session=False)
      )

    if purged_by_claim_number or purged_by_policy_number:
      db.flush()
      print(
        "LOSSQ_CANONICAL_UPLOAD_CLAIM_PURGE:",
        {
          "claim_numbers": len(upload_claim_numbers),
          "policy_keys": len(upload_policy_keys),
          "deleted_by_claim_number": int(purged_by_claim_number or 0),
          "deleted_by_policy_number": int(purged_by_policy_number or 0),
        },
      )

    file_saved = 0
    file_duplicates = 0

    # LOSSQ_CANONICAL_INSERT_ONLY_SAVE_LOOP_V1
    for claim_data in parsed_claims:
      normalized = normalize_claim_data(
        raw=claim_data,
        fallback_policy_number=file_policy_number,
        current_user=current_user,
      )

      normalized = lossq_preserve_row_policy_before_save(
        normalized=normalized,
        raw_claim=claim_data,
        fallback_policy_number=file_policy_number,
      )

      normalized = lossq_apply_row_values_at_final_save(
        normalized=normalized,
        raw_claim=claim_data,
      )

      # LOSSQ_CLAIMANT_FROM_UPLOAD_ROW_V1
      normalized = lossq_apply_claimant_to_normalized_claim(
        normalized_claim=normalized,
        raw_claim=claim_data,
      )

      normalized.pop("claimant_name", None)

      # LOSSQ_CLAIM_DETAIL_FIELDS_FROM_UPLOAD_ROW_V1
      normalized = lossq_apply_claim_detail_fields_to_normalized_claim(normalized, claim_data)

      normalized.pop("claimant_name", None)
      normalized.pop("jurisdiction", None)
      normalized.pop("state", None)
      normalized.pop("adjuster_examiner", None)

      # LOSSQ_CLAIM_DETAIL_FIELDS_FROM_UPLOAD_ROW_V2
      normalized = lossq_apply_claim_detail_fields_to_normalized_claim_v2(normalized, claim_data)

      # LOSSQ_FINAL_SAVE_CLAIM_DETAIL_REPAIR_V3
      normalized = lossq_final_fix_claim_detail_v3(normalized, claim_data)

      # LOSSQ_ATTORNEY_FLAGS_BEFORE_SAVE_CALL_V1
      normalized = lossq_apply_attorney_flags_before_save_v1(normalized, claim_data)

      claim_number = str(normalized.get("claim_number") or "").strip().upper()
      policy_value = str(normalized.get("policy_number") or file_policy_number or "").strip().upper()

      normalized["claim_number"] = claim_number
      normalized["policy_number"] = policy_value

      if not claim_number or claim_number == "UNKNOWN":
        print("LOSSQ_CANONICAL_SAVE_SKIPPED_NO_CLAIM_NUMBER:", claim_data)
        continue

      if not policy_value or is_bad_policy_key_for_upload(policy_value):
        print("LOSSQ_CANONICAL_SAVE_SKIPPED_BAD_POLICY:", {"claim_number": claim_number, "policy": policy_value})
        continue

      # Normalize fallback line fields so the frontend does not show all rows as one line.
      if not normalized.get("line_of_business") and normalized.get("claim_type"):
        normalized["line_of_business"] = normalized.get("claim_type")

      if not normalized.get("claim_type") and normalized.get("line_of_business"):
        normalized["claim_type"] = normalized.get("line_of_business")

      # Safe total fallback: if total is blank/zero but paid/reserve exist, use paid + reserve.
      try:
        paid_value = float(normalized.get("paid_amount") or 0)
      except Exception:
        paid_value = 0.0

      try:
        reserve_value = float(normalized.get("reserve_amount") or 0)
      except Exception:
        reserve_value = 0.0

      try:
        total_value = float(normalized.get("total_incurred") or 0)
      except Exception:
        total_value = 0.0

      if total_value <= 0 and (paid_value or reserve_value):
        normalized["total_incurred"] = paid_value + reserve_value

      # LOSSQ_FINAL_CLAIM_AMOUNT_DATE_COERCE_CORRECT_SAVE_CALL_V1
      normalized = lossq_final_claim_amount_date_coerce_before_save_v1(normalized)
      # LOSSQ_CLAIM_LEGAL_SIGNAL_STRICT_SAVE_CALL_V1
      normalized = lossq_claim_legal_signal_strict_save_v1(normalized, claim_data)
      clean_claim_payload = lossq_filter_claim_model_fields(normalized)

      db.add(Claim(**clean_claim_payload))
      file_saved += 1
      total_saved += 1

      print(
        "LOSSQ_CANONICAL_CLAIM_SAVED:",
        {
          "claim_number": clean_claim_payload.get("claim_number"),
          "policy_number": clean_claim_payload.get("policy_number"),
          "line_of_business": clean_claim_payload.get("line_of_business"),
          "status": clean_claim_payload.get("status"),
          "paid": clean_claim_payload.get("paid_amount"),
          "reserve": clean_claim_payload.get("reserve_amount"),
          "total": clean_claim_payload.get("total_incurred"),
        },
      )

    upload_record = UploadHistory(
      filename=safe_upload_filename,
      stored_path=file_path,
      content_type=file.content_type,
      claims_saved=file_saved,
      uploaded_at=datetime.now().isoformat(),
      uploaded_by_user_id=current_user["user_id"],
      organization_id=current_user["organization_id"],
    )

    db.add(upload_record)

    uploaded_files.append(
      {
        "filename": safe_upload_filename,
        "claims_saved": file_saved,
        "duplicates_skipped": file_duplicates,
        "policy_number": file_policy_number,
      }
    )

  profile_data = extract_profile_data(
    parsed_claims=all_parsed_claims,
    fallback_policy_number=direct_profile.get("policy_number")
    or direct_profile.get("account_number")
    or clean_input_policy
    or f"UPLOAD-{upload_session_id}",
    direct_profile=direct_profile,
  )

  # LOSSQ_FINAL_PROFILE_DATA_BUSINESS_NAME_REPAIR_CALL_V3
  profile_data, direct_profile = lossq_final_profile_data_business_name_repair_v3(
    file_path,
    profile_data,
    direct_profile,
  )

  # LOSSQ_PDF_ACCOUNT_PROFILE_GRID_REPAIR_PROFILE_DATA_CALL_V1
  profile_data, direct_profile = lossq_pdf_account_profile_grid_repair_v1(
    file_path,
    profile_data,
    direct_profile,
  )

  # LOSSQ_EXCEL_ONLY_PROFILE_REPAIR_CALL_V2
  profile_data, direct_profile = lossq_excel_only_profile_repair_v2(
    file_path,
    profile_data,
    direct_profile,
    locals().get("all_parsed_claims", locals().get("parsed_claims", [])),
  )

  # LOSSQ_EXCEL_ONLY_POLICY_SCHEDULE_REPAIR_CALL_V1
  profile_data, direct_profile = lossq_excel_only_policy_schedule_repair_v1(
    file_path,
    profile_data,
    direct_profile,
    locals().get("all_parsed_claims", locals().get("parsed_claims", [])),
  )

  # LOSSQ_EXCEL_ZERO_CLAIM_POLICY_REPAIR_CALL_V3
  profile_data = lossq_excel_zero_claim_policy_repair_v3(
    file_path,
    profile_data,
    locals().get("all_parsed_claims", locals().get("parsed_claims", [])),
  )

  if not profile_data.get("policy_number"):
    profile_data["policy_number"] = (
      profile_data.get("account_number")
      or direct_profile.get("policy_number")
      or f"UPLOAD-{upload_session_id}"
    )

  primary_claim_policy_number = ""
  for claim_data in all_parsed_claims:
    claim_policy_number = clean_profile_value(claim_data.get("policy_number"))
    if claim_policy_number and not is_bad_policy_key_for_upload(claim_policy_number):
      primary_claim_policy_number = claim_policy_number
      break

  # LOSSQ_TABULAR_UPLOAD_POLICY_SCHEDULE_SAVE_V1
  claim_policy_schedule = build_policy_schedule_from_claims_for_upload(all_parsed_claims)
  existing_policy_schedule = profile_data.get("policies") if isinstance(profile_data.get("policies"), list) else []
  profile_data["policies"] = merge_policy_lists_for_upload(
    existing_policy_schedule,
    claim_policy_schedule,
  )

  # LOSSQ_EXCEL_MULTISHEET_POLICY_SCHEDULE_AUTHORITY_CALL_V1
  profile_data = lossq_excel_multisheet_policy_schedule_authority_v1(
    file_path,
    profile_data,
    locals().get("all_parsed_claims", locals().get("parsed_claims", [])),
  )

  profile_account_key = choose_upload_account_key(profile_data, direct_profile)
  # LOSSQ_TRUE_ACCOUNT_NUMBER_FROM_UPLOAD_CSV_V1
  upload_true_account_number = lossq_extract_true_account_number_from_upload_csv(file_path)
  if upload_true_account_number:
    profile_data["account_number"] = upload_true_account_number
    profile_data["customer_number"] = profile_data.get("customer_number") or upload_true_account_number
    profile_account_key = upload_true_account_number


  # LOSSQ_UPLOAD_ACCOUNT_NUMBER_MUST_NOT_BE_POLICY_V1
  def _lossq_upload_value_looks_like_policy(value):
    value = str(value or "").strip().upper()

    # LOSSQ_TRUE_ACCOUNT_NUMBER_LOCAL_POLICY_CHECK_V1
    # Account/customer identifiers like WISC-ACCT-2026 are real account
    # numbers and must not be blanked as policy-like values.
    if lossq_true_account_number_value(value):
      return False

    return lossq_looks_like_policy_but_not_account(value)

  if profile_account_key and not _lossq_upload_value_looks_like_policy(profile_account_key):
    profile_data["account_number"] = profile_data.get("account_number") or profile_account_key

  if _lossq_upload_value_looks_like_policy(profile_data.get("account_number")) and not lossq_true_account_number_value(profile_data.get("account_number")):
    profile_data["account_number"] = ""

  if _lossq_upload_value_looks_like_policy(profile_data.get("customer_number")) and not lossq_true_account_number_value(profile_data.get("customer_number")):
    profile_data["customer_number"] = ""
    profile_data["customer_number"] = (
      profile_data.get("customer_number")
      or profile_data.get("account_number")
      or profile_account_key
    )

  # LOSSQ_TRUE_ACCOUNT_NUMBER_FROM_UPLOAD_CSV_V1_REAPPLY_AFTER_CLEANUP
  upload_true_account_number_after_cleanup = lossq_extract_true_account_number_from_upload_csv(file_path)
  if upload_true_account_number_after_cleanup:
    profile_data["account_number"] = upload_true_account_number_after_cleanup
    profile_data["customer_number"] = profile_data.get("customer_number") or upload_true_account_number_after_cleanup
    profile_account_key = upload_true_account_number_after_cleanup

  # Main saved profile key should be the stable account key.
  # Real policy numbers stay in profile_data["policies"].
  if is_bad_policy_key_for_upload(profile_data.get("policy_number")):
    profile_data["policy_number"] = profile_account_key or primary_claim_policy_number or f"UPLOAD-{upload_session_id}"


  # LOSSQ_DO_NOT_USE_POLICY_AS_ACCOUNT_NUMBER_V2
  def _lossq_upload_policy_like(value):
    value = str(value or "").strip().upper()

    # LOSSQ_TRUE_ACCOUNT_NUMBER_FINAL_POLICY_LIKE_GUARD_V1
    # Account/customer identifiers like WISC-ACCT-2026 are valid account
    # numbers. Do not blank them just because they contain letters, dashes,
    # and digits.
    if lossq_true_account_number_value(value):
      return False

    return lossq_looks_like_policy_but_not_account(value)

  if _lossq_upload_policy_like(profile_data.get("account_number")) and not lossq_true_account_number_value(profile_data.get("account_number")):
    profile_data["account_number"] = ""

  if _lossq_upload_policy_like(profile_data.get("customer_number")) and not lossq_true_account_number_value(profile_data.get("customer_number")):
    profile_data["customer_number"] = ""

  # LOSSQ_TRUE_ACCOUNT_NUMBER_FINAL_REAPPLY_AFTER_LAST_POLICY_CLEANUP_V1
  upload_true_account_number_final_cleanup = lossq_extract_true_account_number_from_upload_csv(file_path)
  if upload_true_account_number_final_cleanup:
    profile_data["account_number"] = upload_true_account_number_final_cleanup
    profile_data["customer_number"] = upload_true_account_number_final_cleanup

  # LOSSQ_TRUE_INSURED_NAME_FINAL_REAPPLY_AFTER_HEADER_CLEANUP_V1
  # Protect clean flat CSV uploads from saving header labels like "Account Number" as insured.
  def _lossq_bad_insured_name_value(value):
    bad = str(value or "").strip().lower()
    return bad in {
      "",
      "account number",
      "customer number",
      "client number",
      "policy number",
      "line of business",
      "carrier",
      "producing agency",
      "business name",
      "named insured",
      "insured",
    }

  upload_true_insured_name_final_cleanup = lossq_extract_true_insured_name_from_upload_csv(file_path)
  if upload_true_insured_name_final_cleanup and (
    _lossq_bad_insured_name_value(profile_data.get("business_name"))
    or _lossq_bad_insured_name_value(profile_data.get("insured_name"))
    or _lossq_bad_insured_name_value(profile_data.get("named_insured"))
  ):
    profile_data["business_name"] = upload_true_insured_name_final_cleanup
    profile_data["insured_name"] = upload_true_insured_name_final_cleanup
    profile_data["named_insured"] = upload_true_insured_name_final_cleanup

    profile_account_key = upload_true_account_number_final_cleanup

  # LOSSQ_AUTHORITATIVE_FLAT_CSV_BEFORE_PROFILE_UPSERT_CALL_V1
  try:
    parsed_claims, parsed_profile, profile_data = lossq_apply_authoritative_flat_csv_snapshot_v1(
      file_path,
      parsed_claims,
      parsed_profile,
      profile_data,
    )
  except Exception as authoritative_csv_profile_exc:
    print("LOSSQ_AUTHORITATIVE_FLAT_CSV_BEFORE_PROFILE_UPSERT_CALL_V1_ERROR:", str(authoritative_csv_profile_exc)[:500])

  # LOSSQ_FLAT_CSV_PREAMBLE_BEFORE_PROFILE_UPSERT_V2
  # Re-apply preamble-aware CSV truth immediately before profile upsert so
  # filename fallback and later profile repairs cannot overwrite Account Name,
  # Claim Status, or zero-claim policy schedule rows.
  try:
    final_preamble_csv_claims, final_preamble_csv_profile = lossq_clean_standard_csv_override(
      file_path,
      parsed_claims,
      parsed_profile,
    )

    if isinstance(final_preamble_csv_profile, dict) and final_preamble_csv_profile:
      parsed_profile = final_preamble_csv_profile
      if isinstance(profile_data, dict):
        for csv_key, csv_value in final_preamble_csv_profile.items():
          if csv_value not in (None, "", [], {}):
            profile_data[csv_key] = csv_value

    if isinstance(final_preamble_csv_claims, list) and final_preamble_csv_claims:
      parsed_claims = final_preamble_csv_claims
      parsed_profile["claims"] = final_preamble_csv_claims
      parsed_profile["parsed_claims"] = final_preamble_csv_claims

    print("LOSSQ_FLAT_CSV_PREAMBLE_BEFORE_PROFILE_UPSERT_V2:", {
      "business_name": parsed_profile.get("business_name") if isinstance(parsed_profile, dict) else None,
      "claims": len(parsed_claims or []),
      "policies": len(parsed_profile.get("policies") or parsed_profile.get("policy_schedule") or []) if isinstance(parsed_profile, dict) else 0,
      "policy_numbers": parsed_profile.get("policy_numbers") if isinstance(parsed_profile, dict) else None,
      "statuses": [c.get("status") for c in (parsed_claims or [])[:10] if isinstance(c, dict)],
    })
  except Exception as preamble_profile_exc:
    print("LOSSQ_FLAT_CSV_PREAMBLE_BEFORE_PROFILE_UPSERT_V2_ERROR:", str(preamble_profile_exc)[:500])


  # LOSSQ_EXCEL_ACCOUNT_PROFILE_FINAL_BEFORE_UPSERT_V1
  # Final Excel-only account snapshot repair placed directly before profile save.
  # This prevents later filename/carrier fallback from overwriting the account
  # profile fields after the policy schedule has already been repaired.
  try:
    import re as _lossq_excel_re
    import datetime as _lossq_excel_dt

    def _lossq_excel_clean_v1(value):
      if isinstance(value, (_lossq_excel_dt.datetime, _lossq_excel_dt.date)):
        return value.strftime("%m/%d/%Y")
      return _lossq_excel_re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip()).strip(" :-|")

    def _lossq_excel_key_v1(value):
      return _lossq_excel_re.sub(r"[^a-z0-9]+", "", _lossq_excel_clean_v1(value).lower())

    def _lossq_excel_good_business_v1(value):
      raw = _lossq_excel_clean_v1(value)
      low = raw.lower()
      if not raw or len(raw) < 3:
        return ""
      bad = [
        "lossq",
        ".xlsx",
        ".xls",
        "field",
        "value",
        "policy schedule",
        "claim detail",
        "exposure inputs",
        "loss summary",
        "account number",
        "policy number",
        "claim number",
        "writing carrier",
        "carrier",
        "effective date",
        "expiration date",
        "current premium",
      ]
      if any(item in low for item in bad):
        return ""
      return raw

    def _lossq_excel_good_account_number_v1(value):
      raw = _lossq_excel_clean_v1(value)
      low = raw.lower()
      upper = raw.upper()

      if not raw:
        return ""

      if any(item in low for item in ["insurance", "carrier", "company", "co.", "mutual", "policy schedule", "claim detail"]):
        return ""

      if not any(char.isdigit() for char in raw):
        return ""

      # LOSSQ_EXCEL_ACCOUNT_EXPOSURE_FINAL_BEFORE_UPSERT_V2
      # Account numbers commonly end in 4-6 digits, so do not reject values that
      # clearly identify account/customer/client numbers.
      account_tokens = ["ACCT", "ACCOUNT", "CUST", "CUSTOMER", "CLIENT", "CLNT"]
      if any(token in upper for token in account_tokens):
        return raw

      if _lossq_excel_re.search(r"-\d{5,}$", upper) and not _lossq_excel_re.search(r"-\d{4}-", upper):
        return ""

      return raw

    def _lossq_excel_normal_date_v1(value):
      raw = _lossq_excel_clean_v1(value)
      if not raw:
        return ""
      match = _lossq_excel_re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", raw)
      if match:
        month, day, year = match.groups()
        if len(year) == 2:
          year = "20" + year
        return f"{int(month):02d}/{int(day):02d}/{int(year):04d}"
      match = _lossq_excel_re.search(r"\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b", raw)
      if match:
        year, month, day = match.groups()
        return f"{int(month):02d}/{int(day):02d}/{int(year):04d}"
      return ""

    def _lossq_excel_period_v1(value):
      raw = _lossq_excel_clean_v1(value)
      dates = _lossq_excel_re.findall(
        r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b",
        raw,
      )
      if len(dates) >= 2:
        return _lossq_excel_normal_date_v1(dates[0]), _lossq_excel_normal_date_v1(dates[1])
      return "", ""

    if isinstance(profile_data, dict) and str(file_path or "").lower().endswith((".xlsx", ".xls")):
      excel_profile_values = {}

      alias_map = {
        "namedinsured": "business_name",
        "insured": "business_name",
        "insuredname": "business_name",
        "accountname": "business_name",
        "businessname": "business_name",
        "companyname": "business_name",
        "accountnumber": "account_number",
        "accountno": "account_number",
        "account": "account_number",
        "customernumber": "account_number",
        "clientnumber": "account_number",
        "writingcarrier": "carrier_name",
        "carrier": "carrier_name",
        "carriername": "carrier_name",
        "insurancecarrier": "carrier_name",
        "policyperiod": "policy_period",
        "policyterm": "policy_period",
        "coverageperiod": "policy_period",
        "effectivedate": "effective_date",
        "expirationdate": "expiration_date",
        "evaluationdate": "evaluation_date",
        "valuationdate": "evaluation_date",
        "asofdate": "evaluation_date",
        "lossrundate": "evaluation_date",
      }

      if str(file_path or "").lower().endswith(".xlsx"):
        from openpyxl import load_workbook as _lossq_excel_load_workbook
        workbook = _lossq_excel_load_workbook(file_path, data_only=True, read_only=True)

        for worksheet in workbook.worksheets:
          sheet_name_key = _lossq_excel_key_v1(worksheet.title)

          # Prefer Account/Profile sheets, but allow field/value rows anywhere.
          prefer_sheet = ("account" in sheet_name_key or "profile" in sheet_name_key)

          for row in worksheet.iter_rows(values_only=True):
            values = [_lossq_excel_clean_v1(cell) for cell in row]
            if len(values) < 2:
              continue

            left = values[0]
            right = values[1]
            field = alias_map.get(_lossq_excel_key_v1(left))

            if not field or not right:
              continue

            if not prefer_sheet and field in {"business_name", "account_number"}:
              continue

            if field == "business_name":
              good_value = _lossq_excel_good_business_v1(right)
              if good_value:
                excel_profile_values[field] = good_value

            elif field == "account_number":
              good_value = _lossq_excel_good_account_number_v1(right)
              if good_value:
                excel_profile_values[field] = good_value

            elif field == "carrier_name":
              good_value = _lossq_excel_clean_v1(right)
              if good_value and not _lossq_excel_good_account_number_v1(good_value):
                excel_profile_values[field] = good_value

            elif field == "policy_period":
              start_date, end_date = _lossq_excel_period_v1(right)
              if start_date and end_date:
                excel_profile_values["effective_date"] = start_date
                excel_profile_values["expiration_date"] = end_date

            elif field in {"effective_date", "expiration_date", "evaluation_date"}:
              good_value = _lossq_excel_normal_date_v1(right)
              if good_value:
                excel_profile_values[field] = good_value

      final_business_name = _lossq_excel_good_business_v1(excel_profile_values.get("business_name"))
      if final_business_name:
        profile_data["business_name"] = final_business_name
        profile_data["insured_name"] = final_business_name
        profile_data["named_insured"] = final_business_name
        profile_data["account_name"] = final_business_name
        profile_data["insured"] = final_business_name

      final_account_number = _lossq_excel_good_account_number_v1(excel_profile_values.get("account_number"))
      if final_account_number:
        profile_data["account_number"] = final_account_number
        profile_data["customer_number"] = final_account_number

      final_carrier_name = _lossq_excel_clean_v1(excel_profile_values.get("carrier_name"))
      if final_carrier_name:
        profile_data["carrier_name"] = final_carrier_name
        profile_data["writing_carrier"] = final_carrier_name
        profile_data["carrier"] = final_carrier_name

      # LOSSQ_EXCEL_EXPOSURE_INPUTS_FINAL_BEFORE_UPSERT_V2
      # Pull Excel two-column Exposure Inputs sheet directly into top-level
      # profile exposure fields and exposure_inputs aliases.
      excel_exposure_values = {}

      exposure_alias_map = {
        "annualrevenue": "revenue",
        "revenue": "revenue",
        "sales": "revenue",
        "receipts": "revenue",
        "payroll": "payroll",
        "employeecount": "employee_count",
        "employees": "employee_count",
        "vehiclecount": "vehicle_count",
        "vehicles": "vehicle_count",
        "drivercount": "driver_count",
        "drivers": "driver_count",
        "warehouselocations": "location_count",
        "locationcount": "location_count",
        "locations": "location_count",
        "currentpremiumtotal": "current_premium",
        "currentpremium": "current_premium",
        "premiumtotal": "current_premium",
        "targetrenewalpremium": "target_renewal_premium",
        "targetpremium": "target_renewal_premium",
        "installcrews": "unit_count",
        "unitcount": "unit_count",
        "annualstops": "annual_stops",
      }

      exposure_label_map = {
        "revenue": "Revenue / Sales",
        "payroll": "Payroll",
        "employee_count": "Employee Count",
        "vehicle_count": "Vehicle Count",
        "driver_count": "Driver Count",
        "location_count": "Location Count",
        "current_premium": "Current Premium",
        "target_renewal_premium": "Target Renewal Premium",
        "unit_count": "Unit Count",
        "annual_stops": "Annual Stops",
      }

      def _lossq_excel_money_or_count_v2(value):
        raw = _lossq_excel_clean_v1(value)
        if not raw:
          return ""
        cleaned = raw.replace("$", "").replace(",", "").strip()
        return cleaned

      for worksheet in workbook.worksheets:
        sheet_name_key = _lossq_excel_key_v1(worksheet.title)

        if "exposure" not in sheet_name_key:
          continue

        for row in worksheet.iter_rows(values_only=True):
          values = [_lossq_excel_clean_v1(cell) for cell in row]
          if len(values) < 2:
            continue

          left = values[0]
          right = values[1]
          field = exposure_alias_map.get(_lossq_excel_key_v1(left))

          if not field:
            continue

          value = _lossq_excel_money_or_count_v2(right)
          if value:
            excel_exposure_values[field] = value

      if excel_exposure_values:
        exposure_inputs = profile_data.get("exposure_inputs")
        if not isinstance(exposure_inputs, dict):
          exposure_inputs = {}

        exposure_rows = exposure_inputs.get("exposure_rows")
        if not isinstance(exposure_rows, list):
          exposure_rows = []

        for exposure_field, exposure_value in excel_exposure_values.items():
          if not exposure_value:
            continue

          profile_data[exposure_field] = exposure_value

          if exposure_field == "revenue":
            profile_data["sales"] = exposure_value
            profile_data["annual_revenue"] = exposure_value
            profile_data["receipts"] = exposure_value
            exposure_inputs["Revenue / Sales"] = exposure_value
            exposure_inputs["Sales"] = exposure_value
            exposure_inputs["Annual Revenue"] = exposure_value
            exposure_inputs["Receipts"] = exposure_value

          elif exposure_field == "location_count":
            profile_data["locations"] = exposure_value
            profile_data["locationCount"] = exposure_value
            exposure_inputs["Location Count"] = exposure_value
            exposure_inputs["Locations"] = exposure_value

          else:
            exposure_label = exposure_label_map.get(exposure_field)
            if exposure_label:
              exposure_inputs[exposure_label] = exposure_value

          exposure_rows.append({
            "field": exposure_field,
            "label": exposure_label_map.get(exposure_field, exposure_field),
            "value": exposure_value,
            "source": "excel_exposure_inputs_sheet",
          })

        exposure_inputs["exposure_rows"] = exposure_rows
        profile_data["exposure_inputs"] = exposure_inputs

        basis_parts = []
        for basis_field in ["revenue", "payroll", "employee_count", "vehicle_count", "driver_count", "location_count", "current_premium", "target_renewal_premium"]:
          basis_value = profile_data.get(basis_field)
          if basis_value:
            basis_parts.append(f"{basis_field}: {basis_value}")

        if basis_parts:
          profile_data["exposure_basis"] = " | ".join(basis_parts)

      print("LOSSQ_EXCEL_EXPOSURE_INPUTS_FINAL_BEFORE_UPSERT_V2:", {
        "account_number": profile_data.get("account_number"),
        "current_premium": profile_data.get("current_premium"),
        "target_renewal_premium": profile_data.get("target_renewal_premium"),
        "revenue": profile_data.get("revenue"),
        "payroll": profile_data.get("payroll"),
        "employee_count": profile_data.get("employee_count"),
        "vehicle_count": profile_data.get("vehicle_count"),
        "driver_count": profile_data.get("driver_count"),
        "location_count": profile_data.get("location_count"),
      })


      if excel_profile_values.get("effective_date"):
        profile_data["effective_date"] = excel_profile_values["effective_date"]

      if excel_profile_values.get("expiration_date"):
        profile_data["expiration_date"] = excel_profile_values["expiration_date"]

      if excel_profile_values.get("evaluation_date"):
        profile_data["evaluation_date"] = excel_profile_values["evaluation_date"]
        profile_data["valuation_date"] = excel_profile_values["evaluation_date"]

      try:
        profile_account_key = profile_data.get("account_number") or profile_account_key
      except Exception:
        pass

      print("LOSSQ_EXCEL_ACCOUNT_PROFILE_FINAL_BEFORE_UPSERT_V1:", {
        "business_name": profile_data.get("business_name"),
        "account_number": profile_data.get("account_number"),
        "carrier_name": profile_data.get("carrier_name"),
        "effective_date": profile_data.get("effective_date"),
        "expiration_date": profile_data.get("expiration_date"),
        "evaluation_date": profile_data.get("evaluation_date"),
      })

  except Exception as excel_final_profile_exc:
    print("LOSSQ_EXCEL_ACCOUNT_PROFILE_FINAL_BEFORE_UPSERT_V1_ERROR:", str(excel_final_profile_exc)[:500])

  lossq_debug_upload_snapshot(
    "before_profile_upsert",
    all_parsed_claims if "all_parsed_claims" in locals() else [],
    profile_data if "profile_data" in locals() else {},
    {
      "total_saved": total_saved if "total_saved" in locals() else None,
      "total_duplicates_skipped": total_duplicates_skipped if "total_duplicates_skipped" in locals() else None,
    },
  )
  profile_data = derive_exposure_inputs_from_policy_schedule(profile_data)

  # LOSSQ_PROFILE_DATA_EXPOSURE_SAVE_DEBUG_V1
  debug_exposure_payload = {
    key: profile_data.get(key)
    for key in [
      "current_premium",
      "expiring_premium",
      "target_renewal_premium",
      "payroll",
      "revenue",
      "sales",
      "employee_count",
      "vehicle_count",
      "driver_count",
      "property_tiv",
      "coverage_limit",
      "deductible",
      "umbrella_limit",
      "cyber_revenue",
      "experience_mod",
      "exposure_basis",
    ]
    if profile_data.get(key) not in ("", None, [], {})
  }
  if debug_exposure_payload:
    print("LOSSQ_PROFILE_DATA_EXPOSURE_BEFORE_SAVE:", debug_exposure_payload)

  # LOSSQ_FORCE_LOCATION_LIQUOR_PROFILE_DATA_SAVE_V1
  # Final exposure save guard. The preamble CSV parser may place restaurant
  # exposure values inside exposure_basis/exposure_inputs, but profile upsert
  # needs top-level fields for the dashboard inputs.
  try:
    import re as _lossq_exposure_re

    def _lossq_clean_exposure_value_v1(value):
      return str(value or "").strip()

    def _lossq_exposure_from_basis_v1(label, source):
      raw = str(source or "")
      if not raw:
        return ""

      match = _lossq_exposure_re.search(
        rf"{_lossq_exposure_re.escape(label)}\s*:\s*([^|]+)",
        raw,
        _lossq_exposure_re.IGNORECASE,
      )
      if not match:
        return ""

      return match.group(1).strip().replace(",", "")

    def _lossq_first_exposure_value_v1(*values):
      for value in values:
        clean_value = _lossq_clean_exposure_value_v1(value)
        if clean_value:
          return clean_value.replace(",", "")
      return ""

    if isinstance(profile_data, dict):
      exposure_inputs = profile_data.get("exposure_inputs")
      if not isinstance(exposure_inputs, dict):
        exposure_inputs = {}

      exposure_basis_text = profile_data.get("exposure_basis") or ""

      # LOSSQ_LOCATION_LIQUOR_EXPOSURE_SPLIT_GUARD_V1
      def _lossq_labeled_count_from_exposure_basis_v1(label, source):
        raw = str(source or "")
        match = _lossq_exposure_re.search(
          rf"{_lossq_exposure_re.escape(label)}\s*[:#-]?\s*(\d{{1,6}})(?=\s*[;,\.\n]|$)",
          raw,
          _lossq_exposure_re.IGNORECASE,
        )
        return match.group(1) if match else ""

      def _lossq_labeled_money_from_exposure_basis_v1(label, source):
        raw = str(source or "")
        match = _lossq_exposure_re.search(
          rf"{_lossq_exposure_re.escape(label)}\s*[:#-]?\s*\$?\s*([0-9][0-9,]*(?:\.\d{{2}})?)(?=\s*[;,\.\n]|$)",
          raw,
          _lossq_exposure_re.IGNORECASE,
        )
        return match.group(1).replace(',', '') if match else ""

      def _lossq_first_count_only_v1(value):
        match = _lossq_exposure_re.search(r"\b(\d{1,6})(?:\.\d+)?\b", str(value or ""))
        return match.group(1) if match else ""

      def _lossq_first_money_only_v1(value):
        raw = str(value or "")
        money_match = _lossq_exposure_re.search(r"\$\s*([0-9][0-9,]*(?:\.\d{2})?)", raw)
        if money_match:
          return money_match.group(1).replace(',', '')
        large_number_match = _lossq_exposure_re.search(r"\b([0-9][0-9,]{3,}(?:\.\d{2})?)\b", raw)
        return large_number_match.group(1).replace(',', '') if large_number_match else ""

      raw_location_value = _lossq_first_exposure_value_v1(
        profile_data.get("location_count"),
        profile_data.get("locations"),
        profile_data.get("locationCount"),
        exposure_inputs.get("Location Count"),
        exposure_inputs.get("Locations"),
        _lossq_exposure_from_basis_v1("Locations", exposure_basis_text),
        _lossq_exposure_from_basis_v1("Location Count", exposure_basis_text),
      )

      raw_liquor_value = _lossq_first_exposure_value_v1(
        profile_data.get("liquor_sales"),
        profile_data.get("liquorSales"),
        profile_data.get("alcohol_sales"),
        exposure_inputs.get("Liquor Sales"),
        exposure_inputs.get("Alcohol Sales"),
        _lossq_exposure_from_basis_v1("Liquor Sales", exposure_basis_text),
        _lossq_exposure_from_basis_v1("Alcohol Sales", exposure_basis_text),
      )

      location_value = _lossq_labeled_count_from_exposure_basis_v1("location count", exposure_basis_text) or _lossq_first_count_only_v1(raw_location_value)
      liquor_value = _lossq_labeled_money_from_exposure_basis_v1("liquor sales", exposure_basis_text) or _lossq_first_money_only_v1(raw_liquor_value) or _lossq_labeled_money_from_exposure_basis_v1("liquor sales", raw_location_value)


      if location_value:
        profile_data["location_count"] = location_value
        profile_data["locations"] = location_value
        profile_data["locationCount"] = location_value
        exposure_inputs["Location Count"] = location_value
        exposure_inputs["Locations"] = location_value

      if liquor_value:
        profile_data["liquor_sales"] = liquor_value
        profile_data["liquorSales"] = liquor_value
        profile_data["alcohol_sales"] = liquor_value
        exposure_inputs["Liquor Sales"] = liquor_value
        exposure_inputs["Alcohol Sales"] = liquor_value

      if exposure_inputs:
        profile_data["exposure_inputs"] = exposure_inputs

      print("LOSSQ_FORCE_LOCATION_LIQUOR_PROFILE_DATA_SAVE_V1:", {
        "location_count": profile_data.get("location_count"),
        "liquor_sales": profile_data.get("liquor_sales"),
        "exposure_inputs": profile_data.get("exposure_inputs"),
      })
  except Exception as location_liquor_save_exc:
    print("LOSSQ_FORCE_LOCATION_LIQUOR_PROFILE_DATA_SAVE_V1_ERROR:", str(location_liquor_save_exc)[:500])

    print("LOSSQ_PROFILE_DATA_EXPOSURE_BEFORE_SAVE:", debug_exposure_payload)

  # LOSSQ_CLEAN_EXPOSURE_LIMITS_FIELD_V1
  profile_data = lossq_clean_exposure_limits_field(profile_data)

  # LOSSQ_CANADA_PROFILE_ENHANCEMENT_CALL_V3
  profile_data = lossq_canada_profile_hook_v3(profile_data, all_parsed_claims)
  # LOSSQ_CANADA_UPLOAD_CLEANUP_CALL_V1_1
  profile_data = lossq_canada_upload_cleanup_v11(profile_data, all_parsed_claims, file_path)

  profile = upsert_account_profile(db, profile_data, current_user)

  record_audit_event(
    db,
    current_user=current_user,
    action="loss_run_uploaded",
    resource_type="upload",
    resource_id=profile_data.get("policy_number"),
    details={
      "policy_number": profile_data.get("policy_number"),
      "account_number": profile_data.get("account_number"),
      "saved_claims": total_saved,
      "duplicates_skipped": total_duplicates_skipped,
      "profile_auto_populated": bool(profile),
      "policy_count": len(profile_data.get("policies") or []),
      "validation": profile_data.get("validation") or {},
      "uploaded_files": uploaded_files,
    },
  )

  db.commit()

  account_profile_id = None
  if profile is not None:
    try:
      db.refresh(profile)
      account_profile_id = getattr(profile, "id", None)
    except Exception:
      print("LOSSQ_UPLOAD_ERROR_TRACE:", traceback.format_exc())
      account_profile_id = getattr(profile, "id", None)

  profile_response = dict(profile_data or {})
  profile_response["id"] = account_profile_id
  profile_response["account_profile_id"] = account_profile_id
  profile_response["selected_profile_id"] = account_profile_id
  profile_response["selected_policy_number"] = profile_data.get("policy_number")

  return {
    "message": "Loss run file(s) uploaded successfully",
    "saved_claims": total_saved,
    "duplicates_skipped": total_duplicates_skipped,
    "policy_number": profile_data.get("policy_number"),
    "account_number": profile_data.get("account_number"),
    "account_profile_id": account_profile_id,
    "selected_profile_id": account_profile_id,
    "selected_policy_number": profile_data.get("policy_number"),
    "profile_auto_populated": bool(profile),
    "profile": profile_response,
    "account_profile": profile_response,
    "policies": profile_data.get("policies") or [],
    "claims": all_parsed_claims,
    "parsed_claims": all_parsed_claims,
    "saved_claim_rows": all_parsed_claims,
    "parsed_claim_count": len(all_parsed_claims),
    "saved_claim_count": total_saved,
    "validation": profile_data.get("validation") or {},
    "uploaded_files": uploaded_files,
  }

# LOSSQ_DEPLOY_TRIGGER_20260614152009
